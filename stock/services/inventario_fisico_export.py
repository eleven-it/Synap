# -*- coding: utf-8 -*-
"""Exportación Excel del informe de auditoría/gerencia de inventario físico."""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable, List, Optional, Sequence, Tuple

from django.http import HttpResponse
from django.utils import timezone

from core.utils.administranet_types import str_or_default, to_decimal_or_none, to_int_or_none
from stock.services.inventario_fisico import (
    MOTIVO_FALTANTE,
    MOTIVO_SOBRANTE,
    cantidad_mstock_por_diferencia,
    contar_conflictos_sync,
    etiquetar_contadores,
    listar_auditoria_ajuste_campana,
    listar_contadores_candidatos,
    listar_depositos_elegibles,
    listar_eventos_campana,
    listar_lineas_analizador,
    listar_movimientos_post_snapshot_campana,
    motivo_mstock_por_diferencia,
    obtener_campana,
    obtener_progreso_campana,
)

_ACCION_AUDITORIA_ETIQUETA = {
    "override_guardado": "Override guardado",
    "override_quitado": "Override quitado",
    "override_pisado": "Override pisado",
    "autorizacion": "Autorización",
    "contado_cero_masivo": "Contado cero masivo",
}

_MOTIVO_MSTOCK_ETIQUETA = {
    MOTIVO_FALTANTE: "Faltante",
    MOTIVO_SOBRANTE: "Sobrante",
}


def _fmt_fecha(valor: Any) -> str:
    if valor is None:
        return "—"
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    texto = str_or_default(valor, "").strip()
    if not texto:
        return "—"
    if len(texto) >= 10 and texto[4:5] == "-" and texto[7:8] == "-":
        return f"{texto[8:10]}/{texto[5:7]}/{texto[0:4]}"
    return texto[:10]


def _fmt_fecha_hora(valor: Any) -> str:
    if valor is None:
        return "—"
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    texto = str_or_default(valor, "").strip()
    if not texto:
        return "—"
    return texto


def _num(valor: Any) -> Any:
    dec = to_decimal_or_none(valor)
    if dec is None:
        return ""
    if dec == dec.to_integral_value():
        return int(dec)
    return float(dec)


def _si_no(valor: Any) -> str:
    return "Sí" if valor else "No"


def _auto_width(ws) -> None:
    from openpyxl.utils import get_column_letter

    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _escribir_headers(ws, row: int, headers: Sequence[str], estilos: dict) -> int:
    for col_num, label in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_num, value=label)
        c.fill = estilos["header_fill"]
        c.font = estilos["header_font"]
        c.alignment = estilos["center"]
        c.border = estilos["border"]
    return row + 1


def _escribir_filas(
    ws,
    row: int,
    filas: Iterable[Sequence[Any]],
    estilos: dict,
) -> int:
    for fila in filas:
        for col_num, valor in enumerate(fila, 1):
            c = ws.cell(row=row, column=col_num, value="" if valor is None else valor)
            c.border = estilos["border"]
            c.alignment = estilos["left"]
        row += 1
    return row


def _tipo_diferencia(linea: dict) -> str:
    if linea.get("cantidad_contada") is None:
        return "Sin contar"
    diff = to_decimal_or_none(linea.get("diferencia_real"))
    if diff is None:
        return "Sin contar"
    if diff < 0:
        return "Faltante"
    if diff > 0:
        return "Sobrante"
    return "Sin diferencia"


def _etiqueta_motivo_mstock(motivo: Optional[int]) -> str:
    if motivo is None:
        return "—"
    return _MOTIVO_MSTOCK_ETIQUETA.get(motivo, str(motivo))


def _nombre_archivo(id_campana: int, campana: dict) -> str:
    ref = campana.get("fecha_snapshot") or campana.get("fecha")
    if isinstance(ref, datetime):
        sufijo = ref.strftime("%d%m%Y")
    elif isinstance(ref, date):
        sufijo = ref.strftime("%d%m%Y")
    else:
        texto = _fmt_fecha(ref)
        sufijo = texto.replace("/", "") if texto != "—" else timezone.localdate().strftime("%d%m%Y")
    estado = str_or_default(campana.get("estado"), "SinEstado").replace(" ", "")
    return f"InventarioFisico_Campana_{id_campana}_{estado}_{sufijo}.xlsx"


def _naturaleza_informe(estado: Any) -> str:
    if str_or_default(estado, "") == "Aplicado":
        return "Definitivo (campaña autorizada / MSTOCK aplicado)"
    return "Preliminar (campaña aún no autorizada — saldos previstos, no impacto contable final)"


def _mapa_depositos(base_empresa: str) -> dict[int, str]:
    return {
        to_int_or_none(d.get("id_deposito")): str_or_default(d.get("nombre"), "")
        for d in listar_depositos_elegibles(base_empresa)
        if to_int_or_none(d.get("id_deposito")) is not None
    }


def _construir_resumen(
    base_empresa: str,
    campana: dict,
    *,
    usuario_exportador: str,
    lineas: List[dict],
    progreso: dict,
    conflictos: int,
) -> List[Tuple[str, Any]]:
    depositos_map = _mapa_depositos(base_empresa)
    nombres_dep = [
        depositos_map.get(did, f"Depósito #{did}")
        for did in (campana.get("depositos") or [])
    ]
    candidatos = listar_contadores_candidatos(base_empresa)
    contadores = etiquetar_contadores(campana.get("contadores") or [], candidatos)
    etiquetas_cont = ", ".join(c["etiqueta"] for c in contadores) or "—"

    lineas_diff = sum(
        1
        for l in lineas
        if (to_decimal_or_none(l.get("diferencia_real")) or Decimal("0")) != 0
    )
    overrides = sum(1 for l in lineas if l.get("ajuste_manual") is not None)
    descuadres = sum(1 for l in lineas if l.get("descuadre"))
    id_mstock = campana.get("id_movimiento_mstock")
    ahora = timezone.localtime(timezone.now())
    estado = campana.get("estado")

    return [
        ("Inventario físico — informe de auditoría", ""),
        ("ESTADO DE LA CAMPAÑA", str_or_default(estado, "—")),
        ("Naturaleza del informe", _naturaleza_informe(estado)),
        ("Empresa", base_empresa),
        ("ID campaña", campana.get("id_campana")),
        ("Fecha campaña", _fmt_fecha(campana.get("fecha"))),
        ("Fecha snapshot", _fmt_fecha_hora(campana.get("fecha_snapshot"))),
        ("Fecha exportación", ahora.strftime("%d/%m/%Y %H:%M")),
        ("Usuario exportador", usuario_exportador or "—"),
        ("Depósitos incluidos", ", ".join(nombres_dep) or "—"),
        ("Contadores asignados", etiquetas_cont),
        ("Usuario alta", campana.get("id_usuario_alta") or "—"),
        ("Catálogo versión", campana.get("catalogo_version") or "—"),
        ("Código MSTOCK", id_mstock if id_mstock is not None else "—"),
        ("Total líneas", progreso.get("total", 0)),
        ("Contadas", progreso.get("contados", 0)),
        ("Pendientes", progreso.get("pendientes", 0)),
        ("% avance", f"{progreso.get('porcentaje', 0)}%"),
        ("Líneas con diferencia real ≠ 0", lineas_diff),
        ("Overrides manuales", overrides),
        ("Líneas con descuadre", descuadres),
        ("Conflictos sync", conflictos),
    ]


def exportar_campana_xlsx(
    base_empresa: str,
    id_campana: int,
    *,
    usuario_exportador: str = "",
) -> HttpResponse | None:
    """Genera un .xlsx multi-hoja para auditoría/gerencia. None si la campaña no existe."""
    campana = obtener_campana(base_empresa, id_campana)
    if not campana:
        return None

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    lineas = listar_lineas_analizador(base_empresa, id_campana)
    progreso = obtener_progreso_campana(base_empresa, id_campana)
    conflictos = contar_conflictos_sync(base_empresa, id_campana)
    eventos = listar_eventos_campana(base_empresa, id_campana)
    movimientos = listar_movimientos_post_snapshot_campana(base_empresa, id_campana)
    auditoria = listar_auditoria_ajuste_campana(base_empresa, id_campana)
    depositos_map = _mapa_depositos(base_empresa)

    estilos = {
        "header_fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "title_font": Font(bold=True, size=14, color="1E40AF"),
        "label_font": Font(bold=True, size=10),
        "value_font": Font(size=10),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
    }

    wb = openpyxl.Workbook()

    # ── Hoja Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    row = 1
    estado_txt = str_or_default(campana.get("estado"), "—")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(
        row=row,
        column=1,
        value=f"Inventario físico — informe de auditoría · ESTADO: {estado_txt}",
    )
    cell.font = estilos["title_font"]
    cell.alignment = estilos["left"]
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    nat = ws.cell(row=row, column=1, value=_naturaleza_informe(estado_txt))
    nat.font = Font(bold=True, size=11, color="B45309" if estado_txt != "Aplicado" else "166534")
    nat.alignment = estilos["left"]
    row += 1
    for etiqueta, valor in _construir_resumen(
        base_empresa,
        campana,
        usuario_exportador=usuario_exportador,
        lineas=lineas,
        progreso=progreso,
        conflictos=conflictos,
    )[1:]:
        lc = ws.cell(row=row, column=1, value=etiqueta)
        lc.font = estilos["label_font"]
        vc = ws.cell(row=row, column=2, value=valor)
        vc.font = estilos["value_font"]
        vc.alignment = estilos["left"]
        if etiqueta == "ESTADO DE LA CAMPAÑA":
            vc.font = Font(bold=True, size=12, color="1E3A8A")
        row += 1
    _auto_width(ws)

    # ── Hoja Líneas ──
    headers_lineas = [
        "ID línea",
        "Código",
        "Artículo",
        "Marca",
        "ID depósito",
        "Depósito",
        "Disponible (snapshot)",
        "Ajuste sistema",
        "Ajuste manual",
        "Override manual",
        "Cargado después (efectivo)",
        "Disponible ajustado",
        "Contado",
        "Diferencia vs snapshot",
        "Diferencia real",
        "Saldo final",
        "Tipo",
        "Contador",
        "Estado línea",
        "Descuadre",
        "Saldo actual ref",
        "Motivo MSTOCK",
        "Cantidad MSTOCK",
    ]
    ws_lineas = wb.create_sheet("Lineas")
    row = 1
    row = _escribir_headers(ws_lineas, row, headers_lineas, estilos)
    filas_lineas = []
    for l in lineas:
        id_dep = to_int_or_none(l.get("id_deposito"))
        diff = to_decimal_or_none(l.get("diferencia_real"))
        motivo = motivo_mstock_por_diferencia(diff)
        cant_mstock = cantidad_mstock_por_diferencia(diff) if motivo else None
        filas_lineas.append(
            (
                l.get("id_linea"),
                l.get("codigo"),
                l.get("nombre"),
                l.get("nombre_marca") or "",
                id_dep,
                depositos_map.get(id_dep, ""),
                _num(l.get("saldo_snapshot")),
                _num(l.get("ajuste_sistema")),
                _num(l.get("ajuste_manual")),
                _si_no(l.get("ajuste_manual") is not None),
                _num(l.get("ajuste_efectivo")),
                _num(l.get("disponible_ajustado")),
                _num(l.get("cantidad_contada")) if l.get("cantidad_contada") is not None else "—",
                _num(l.get("diferencia")),
                _num(diff) if diff is not None else "—",
                _num(l.get("saldo_final")) if l.get("saldo_final") is not None else "—",
                _tipo_diferencia(l),
                l.get("contador_etiqueta") or "",
                l.get("estado_linea"),
                _si_no(l.get("descuadre")),
                _num(l.get("saldo_actual_ref")),
                _etiqueta_motivo_mstock(motivo),
                _num(cant_mstock) if cant_mstock is not None else "—",
            )
        )
    _escribir_filas(ws_lineas, row, filas_lineas, estilos)
    _auto_width(ws_lineas)

    # ── Hoja Eventos ──
    headers_eventos = [
        "ID evento",
        "client_event_id",
        "Código",
        "Artículo",
        "ID depósito",
        "Depósito",
        "Contador",
        "Cantidad",
        "Fecha cliente",
        "Fecha servidor",
        "Resultado",
        "Motivo",
    ]
    ws_eventos = wb.create_sheet("Eventos")
    row = 1
    row = _escribir_headers(ws_eventos, row, headers_eventos, estilos)
    filas_eventos = []
    for ev in eventos:
        id_dep = to_int_or_none(ev.get("id_deposito"))
        filas_eventos.append(
            (
                ev.get("id_evento"),
                ev.get("client_event_id"),
                ev.get("codigo"),
                ev.get("nombre"),
                id_dep,
                depositos_map.get(id_dep, ""),
                ev.get("contador_etiqueta") or "",
                _num(ev.get("cantidad")),
                _fmt_fecha_hora(ev.get("client_ts")),
                _fmt_fecha_hora(ev.get("server_ts")),
                ev.get("resultado"),
                ev.get("motivo") or "",
            )
        )
    _escribir_filas(ws_eventos, row, filas_eventos, estilos)
    _auto_width(ws_eventos)

    # ── Hoja Movimientos post-snapshot ──
    headers_mov = [
        "Código",
        "Artículo",
        "ID depósito",
        "Depósito",
        "Fecha control",
        "Motivo",
        "Comprobante",
        "Entrada",
        "Salida",
        "Neto",
        "Detalle",
    ]
    ws_mov = wb.create_sheet("Movimientos post-snapshot")
    row = 1
    row = _escribir_headers(ws_mov, row, headers_mov, estilos)
    filas_mov = []
    for m in movimientos:
        id_dep = to_int_or_none(m.get("id_deposito"))
        filas_mov.append(
            (
                m.get("codigo"),
                m.get("nombre"),
                id_dep,
                depositos_map.get(id_dep, ""),
                _fmt_fecha_hora(m.get("fecha_control")),
                m.get("motivo"),
                m.get("nro"),
                _num(m.get("entrada")),
                _num(m.get("salida")),
                _num(m.get("neto")),
                m.get("detalle"),
            )
        )
    _escribir_filas(ws_mov, row, filas_mov, estilos)
    _auto_width(ws_mov)

    # ── Hoja Auditoría ajustes ──
    headers_audit = [
        "Fecha",
        "Acción",
        "Código",
        "Artículo",
        "ID depósito",
        "Depósito",
        "Usuario",
        "Ajuste sistema",
        "Ajuste anterior",
        "Ajuste nuevo",
        "Diferencia real",
        "Código movimiento",
    ]
    ws_audit = wb.create_sheet("Auditoria ajustes")
    row = 1
    row = _escribir_headers(ws_audit, row, headers_audit, estilos)
    filas_audit = []
    for a in auditoria:
        id_dep = to_int_or_none(a.get("id_deposito"))
        accion = _ACCION_AUDITORIA_ETIQUETA.get(
            str_or_default(a.get("accion"), ""),
            str_or_default(a.get("accion"), ""),
        )
        filas_audit.append(
            (
                _fmt_fecha_hora(a.get("created_at")),
                accion,
                a.get("codigo"),
                a.get("nombre"),
                id_dep,
                depositos_map.get(id_dep, ""),
                a.get("id_usuario"),
                _num(a.get("ajuste_sistema")),
                _num(a.get("ajuste_anterior")),
                _num(a.get("ajuste_nuevo")),
                _num(a.get("diferencia_real")),
                a.get("codigo_movimiento") or "",
            )
        )
    _escribir_filas(ws_audit, row, filas_audit, estilos)
    _auto_width(ws_audit)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = _nombre_archivo(id_campana, campana)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response
