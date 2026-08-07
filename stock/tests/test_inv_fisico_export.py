# -*- coding: utf-8 -*-
"""Tests exportación Excel inventario físico."""
from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from django.test import SimpleTestCase

from stock.services.inventario_fisico_export import exportar_campana_xlsx


class InvFisicoExportTests(SimpleTestCase):
    """Exportación xlsx multi-hoja con mocks."""

    @patch("stock.services.inventario_fisico_export.obtener_campana")
    def test_export_none_si_sin_campana(self, mock_campana):
        mock_campana.return_value = None
        self.assertIsNone(exportar_campana_xlsx("administranet1", 99))

    @patch("stock.services.inventario_fisico_export.listar_auditoria_ajuste_campana")
    @patch("stock.services.inventario_fisico_export.listar_movimientos_post_snapshot_campana")
    @patch("stock.services.inventario_fisico_export.listar_eventos_campana")
    @patch("stock.services.inventario_fisico_export.contar_conflictos_sync")
    @patch("stock.services.inventario_fisico_export.obtener_progreso_campana")
    @patch("stock.services.inventario_fisico_export.listar_lineas_analizador")
    @patch("stock.services.inventario_fisico_export.listar_depositos_elegibles")
    @patch("stock.services.inventario_fisico_export.listar_contadores_candidatos")
    @patch("stock.services.inventario_fisico_export.obtener_campana")
    def test_export_genera_xlsx_con_hojas(
        self,
        mock_campana,
        mock_contadores,
        mock_depositos,
        mock_lineas,
        mock_progreso,
        mock_conflictos,
        mock_eventos,
        mock_movimientos,
        mock_auditoria,
    ):
        mock_campana.return_value = {
            "id_campana": 5,
            "fecha": "2026-03-15",
            "estado": "EnRevision",
            "depositos": [1],
            "catalogo_version": "abc123",
            "contadores": [10],
            "id_usuario_alta": 1,
            "fecha_snapshot": "2026-03-15 08:00:00",
            "id_movimiento_mstock": None,
        }
        mock_contadores.return_value = [
            {"id_usuario": 10, "cod_usuario": "op1", "nombre_completo": "Operario Uno"}
        ]
        mock_depositos.return_value = [{"id_deposito": 1, "nombre": "Depósito Central"}]
        mock_lineas.return_value = [
            {
                "id_linea": 1,
                "codigo": "ART-1",
                "nombre": "Artículo prueba",
                "nombre_marca": "Marca X",
                "id_deposito": 1,
                "saldo_snapshot": Decimal("10"),
                "ajuste_sistema": Decimal("2"),
                "ajuste_manual": None,
                "ajuste_efectivo": Decimal("2"),
                "disponible_ajustado": Decimal("12"),
                "cantidad_contada": Decimal("11"),
                "diferencia": Decimal("1"),
                "diferencia_real": Decimal("-1"),
                "id_contador": 10,
                "contador_etiqueta": "op1 · Operario Uno",
                "estado_linea": "Contado",
                "descuadre": False,
                "saldo_actual_ref": Decimal("12"),
                "saldo_final": Decimal("11"),
            }
        ]
        mock_progreso.return_value = {
            "total": 1,
            "contados": 1,
            "pendientes": 0,
            "porcentaje": 100.0,
        }
        mock_conflictos.return_value = 0
        mock_eventos.return_value = []
        mock_movimientos.return_value = []
        mock_auditoria.return_value = []

        response = exportar_campana_xlsx("administranet1", 5, usuario_exportador="Supervisor Test")
        self.assertIsNotNone(response)
        assert response is not None
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("InventarioFisico", response["Content-Disposition"])

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            set(wb.sheetnames),
            {
                "Resumen",
                "Lineas",
                "Eventos",
                "Movimientos post-snapshot",
                "Auditoria ajustes",
            },
        )
        ws_resumen = wb["Resumen"]
        valores = [c.value for row in ws_resumen.iter_rows(max_col=2) for c in row if c.value]
        self.assertTrue(any("ESTADO" in str(v) for v in valores))
        self.assertIn("EnRevision", valores)
        self.assertTrue(any("Preliminar" in str(v) for v in valores))
        self.assertIn("EnRevision", response["Content-Disposition"])

    def test_analizador_html_tiene_boton_exportar(self):
        ruta = (
            Path(__file__).resolve().parents[1]
            / "templates"
            / "stock"
            / "inventario_fisico"
            / "analizador.html"
        )
        contenido = ruta.read_text(encoding="utf-8")
        self.assertIn("Exportar Excel", contenido)
        self.assertIn("inventario_fisico_export_xlsx", contenido)
        self.assertIn("puede_exportar", contenido)

    def test_puede_exportar_estados_operativos(self):
        from stock.services.inventario_fisico import (
            ESTADO_APLICADO,
            ESTADO_BORRADOR,
            ESTADO_EN_CONTEO,
            ESTADO_EN_REVISION,
            ESTADO_ANULADO,
            puede_exportar_informe_campana,
        )

        self.assertTrue(puede_exportar_informe_campana(ESTADO_APLICADO))
        self.assertTrue(puede_exportar_informe_campana(ESTADO_EN_CONTEO))
        self.assertTrue(puede_exportar_informe_campana(ESTADO_EN_REVISION))
        self.assertFalse(puede_exportar_informe_campana(ESTADO_BORRADOR))
        self.assertFalse(puede_exportar_informe_campana(ESTADO_ANULADO))
        self.assertFalse(puede_exportar_informe_campana(None))
