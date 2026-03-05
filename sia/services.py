"""
Servicios para el módulo Strategic Insights & Alignment (SIA)
Centraliza la lógica de agregación de datos para dashboards y APIs
"""
import io
from datetime import datetime
from django.db.models import Avg, Min, Max, Count, StdDev
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, BaseDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    PageBreak, PageTemplate, Frame
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sia.models import EvaluationCycle, StrategicSurveyResponse, FodaItem, Rating, OpenAnswer, CameAction


class DashboardDataService:
    """
    Servicio para consolidar datos del dashboard SIA.
    Centraliza la lógica de agregación para evitar duplicación entre views y APIs.
    """
    
    @staticmethod
    def get_consolidated_data(empresa_id=None, cycle_id=None):
        """
        Obtiene datos consolidados para el dashboard.
        
        Args:
            empresa_id: ID de la empresa (opcional)
            cycle_id: ID del ciclo de evaluación (opcional, si no se proporciona usa el más reciente)
        
        Returns:
            dict con:
                - ratings: lista de estadísticas por dimensión
                - foda: dict con items por cuadrante
                - total_responses: número total de respuestas
                - cycle_info: información del ciclo seleccionado
        """
        # Obtener el ciclo de evaluación
        evaluation_cycle = None
        
        if cycle_id:
            try:
                evaluation_cycle = EvaluationCycle.objects.select_related('empresa').get(
                    id=cycle_id,
                    is_active=True
                )
                # Validar que pertenezca a la empresa si se especificó
                if empresa_id and evaluation_cycle.empresa_id != empresa_id:
                    evaluation_cycle = None
            except EvaluationCycle.DoesNotExist:
                evaluation_cycle = None
        
        # Si no se especificó ciclo o no se encontró, buscar el más reciente activo
        if not evaluation_cycle:
            queryset = EvaluationCycle.objects.filter(is_active=True)
            if empresa_id:
                queryset = queryset.filter(empresa_id=empresa_id)
            evaluation_cycle = queryset.order_by('-start_date').first()
        
        if not evaluation_cycle:
            return {
                'ratings': [],
                'foda': {
                    'strength': [],
                    'weakness': [],
                    'opportunity': [],
                    'threat': []
                },
                'total_responses': 0,
                'cycle_info': None
            }
        
        # Ratings consolidados por dimensión
        ratings_data = Rating.objects.filter(
            survey_response__evaluation_cycle=evaluation_cycle,
            survey_response__status='submitted'
        ).values('dimension').annotate(
            average=Avg('value'),
            min_value=Min('value'),
            max_value=Max('value'),
            std_dev=StdDev('value'),
            count=Count('id')
        ).order_by('dimension')
        
        # FODA consolidado por cuadrante
        foda_data = {}
        for quadrant in ['strength', 'weakness', 'opportunity', 'threat']:
            items = FodaItem.objects.filter(
                survey_response__evaluation_cycle=evaluation_cycle,
                survey_response__status='submitted',
                quadrant=quadrant
            ).values('description').annotate(
                count=Count('id')
            ).order_by('-count')[:10]  # Top 10 más mencionados
            
            foda_data[quadrant] = [
                {
                    'description': item['description'],
                    'count': item['count']
                }
                for item in items
            ]
        
        # Total de respuestas en el ciclo
        total_responses = StrategicSurveyResponse.objects.filter(
            evaluation_cycle=evaluation_cycle,
            status='submitted'
        ).count()
        
        return {
            'ratings': [
                {
                    'dimension': item['dimension'],
                    'average': float(item['average']) if item['average'] else 0.0,
                    'min_value': int(item['min_value']) if item['min_value'] else 0,
                    'max_value': int(item['max_value']) if item['max_value'] else 0,
                    'std_dev': float(item['std_dev']) if item['std_dev'] else 0.0,
                    'count': int(item['count'])
                }
                for item in ratings_data
            ],
            'foda': foda_data,
            'total_responses': total_responses,
            'cycle_info': {
                'id': evaluation_cycle.id,
                'name': evaluation_cycle.name,
                'start_date': evaluation_cycle.start_date.isoformat() if evaluation_cycle.start_date else None,
                'end_date': evaluation_cycle.end_date.isoformat() if evaluation_cycle.end_date else None,
                'empresa_id': evaluation_cycle.empresa_id,
                'empresa_name': evaluation_cycle.empresa.nombre if evaluation_cycle.empresa else None
            }
        }


def generate_cycle_report_pdf(empresa, evaluation_cycle) -> bytes:
    """
    Genera un PDF ejecutivo de nivel consultora con el resumen consolidado de un ciclo de evaluación SIA.
    
    El PDF incluye:
    - Portada profesional con branding
    - Headers y footers con marca y paginación
    - Textos narrativos interpretativos en cada sección
    - Diseño consistente con colores corporativos de administraNET ANALYTICS
    
    Args:
        empresa: instancia de core.models.Empresa
        evaluation_cycle: instancia de sia.models.EvaluationCycle
    
    Returns:
        bytes: contenido del PDF generado
    
    Raises:
        ValueError: si evaluation_cycle.empresa != empresa
    """
    # Validar que el ciclo pertenece a la empresa
    if evaluation_cycle.empresa_id != empresa.id:
        raise ValueError(
            f"El ciclo de evaluación {evaluation_cycle.id} no pertenece a la empresa {empresa.id}"
        )
    
    # Crear buffer en memoria
    buffer = io.BytesIO()
    
    # Colores corporativos
    COLOR_ORANGE = colors.HexColor('#f97316')
    COLOR_DARK = colors.HexColor('#0f172a')
    COLOR_LIGHT_GRAY = colors.HexColor('#e5e7eb')
    COLOR_TEXT_GRAY = colors.HexColor('#4B5563')
    
    # Función para dibujar header/footer (se llama en cada página)
    def draw_header_footer(canvas_obj, doc):
        """Dibuja header y footer en cada página (excepto portada)."""
        canvas_obj.saveState()
        width, height = A4
        
        # HEADER: Banda superior con branding
        canvas_obj.setFillColor(COLOR_DARK)
        canvas_obj.rect(0, height - 2*cm, width, 0.8*cm, fill=1, stroke=0)
        
        canvas_obj.setFillColor(colors.white)
        canvas_obj.setFont("Helvetica-Bold", 10)
        canvas_obj.drawString(2*cm, height - 1.6*cm, "administraNET ANALYTICS — Módulo SIA")
        
        # Subtítulo del header (nombre del ciclo o empresa)
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.drawString(2*cm, height - 2.2*cm, evaluation_cycle.name)
        
        # FOOTER: Pie de página con paginación
        canvas_obj.setFillColor(COLOR_TEXT_GRAY)
        canvas_obj.setFont("Helvetica", 8)
        
        # Número de página
        page_num = canvas_obj.getPageNumber()
        canvas_obj.drawRightString(width - 2*cm, 1.2*cm, f"Página {page_num}")
        
        # Texto de generación automática
        canvas_obj.setFont("Helvetica-Oblique", 7)
        canvas_obj.drawString(2*cm, 1.2*cm, "Informe generado automáticamente por administraNET ANALYTICS")
        
        canvas_obj.restoreState()
    
    # Función para portada sin header/footer
    def draw_cover_page(canvas_obj, doc):
        """Dibuja la portada sin header/footer."""
        pass  # La portada se maneja directamente en el story
    
    # Crear documento PDF con BaseDocTemplate para headers/footers
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=3*cm,  # Más espacio arriba para el header
        bottomMargin=2.5*cm  # Más espacio abajo para el footer
    )
    
    # Frame para el contenido (excluyendo header/footer)
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id='normal'
    )
    
    # PageTemplate para páginas normales (con header/footer)
    normal_template = PageTemplate(
        id='normal_template',
        frames=[frame],
        onPage=draw_header_footer
    )
    
    # PageTemplate para portada (sin header/footer)
    # La portada usa un frame más grande para aprovechar todo el espacio
    cover_frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height + 3*cm,
        id='cover'
    )
    cover_template = PageTemplate(
        id='cover_template',
        frames=[cover_frame],
        onPage=draw_cover_page
    )
    
    # Agregar templates: primero el de portada (se usa primero), luego el normal
    doc.addPageTemplates([cover_template, normal_template])
    
    # Estilos base
    styles = getSampleStyleSheet()
    
    # ============================================================
    # ESTILOS PERSONALIZADOS
    # ============================================================
    # Portada: Título principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=COLOR_ORANGE,
        spaceAfter=40,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Portada: Subtítulo del módulo
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=COLOR_TEXT_GRAY,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    # Portada: Nombre de empresa
    company_name_style = ParagraphStyle(
        'CompanyName',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=COLOR_DARK,
        spaceAfter=25,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Portada: Nombre del ciclo
    cycle_name_style = ParagraphStyle(
        'CycleName',
        parent=styles['Heading2'],
        fontSize=18,
        textColor=COLOR_DARK,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Secciones: Título de sección
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=COLOR_DARK,
        spaceAfter=12,
        spaceBefore=25,
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderPadding=0
    )
    
    # Texto normal
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        fontName='Helvetica',
        leading=14,
        alignment=TA_LEFT
    )
    
    # Texto pequeño gris (para notas, fechas, etc.)
    small_gray_style = ParagraphStyle(
        'SmallGray',
        parent=styles['Normal'],
        fontSize=9,
        textColor=COLOR_TEXT_GRAY,
        fontName='Helvetica',
        alignment=TA_CENTER
    )
    
    # Contenedor de elementos del PDF
    story = []
    
    # Contenedor de elementos del PDF
    story = []
    
    # ============================================================
    # PORTADA MEJORADA CON BRANDING
    # ============================================================
    # Banda superior simulada (en ReportLab se hace con tabla con fondo)
    cover_header_data = [[Paragraph(
        "<b>administraNET ANALYTICS — SIA</b>",
        ParagraphStyle('CoverHeader', parent=styles['Normal'], fontSize=12, 
                      textColor=colors.white, fontName='Helvetica-Bold', 
                      alignment=TA_CENTER)
    )]]
    cover_header_table = Table(cover_header_data, colWidths=[16.5*cm])
    cover_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COLOR_DARK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(cover_header_table)
    story.append(Spacer(1, 2*cm))
    
    # Título principal
    story.append(Paragraph(
        "Informe de Diagnóstico Estratégico",
        title_style
    ))
    
    # Subtítulo del módulo
    story.append(Paragraph(
        "Módulo SIA — Strategic Insights & Alignment",
        subtitle_style
    ))
    
    story.append(Spacer(1, 2*cm))
    
    # Nombre de la empresa (grande, destacado)
    story.append(Paragraph(
        empresa.nombre,
        company_name_style
    ))
    
    # Nombre del ciclo
    story.append(Paragraph(
        evaluation_cycle.name,
        cycle_name_style
    ))
    
    story.append(Spacer(1, 1*cm))
    
    # Fechas del ciclo
    if evaluation_cycle.start_date and evaluation_cycle.end_date:
        dates_text = (
            f"{evaluation_cycle.start_date.strftime('%d/%m/%Y')} - "
            f"{evaluation_cycle.end_date.strftime('%d/%m/%Y')}"
        )
    elif evaluation_cycle.start_date:
        dates_text = f"Desde {evaluation_cycle.start_date.strftime('%d/%m/%Y')}"
    else:
        dates_text = "Fechas no definidas"
    
    story.append(Paragraph(
        dates_text,
        ParagraphStyle(
            'Dates',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=30,
            fontName='Helvetica'
        )
    ))
    
    story.append(Spacer(1, 3*cm))
    
    # Pie de portada: Firma de la consultora
    story.append(Paragraph(
        "Preparado por: administraNET ANALYTICS",
        ParagraphStyle(
            'FooterText',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=8,
            fontName='Helvetica-Oblique',
            textColor=COLOR_TEXT_GRAY
        )
    ))
    
    story.append(Paragraph(
        "Uso interno — Confidencial",
        ParagraphStyle(
            'Confidential',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_CENTER,
            spaceAfter=20,
            fontName='Helvetica-Oblique',
            textColor=COLOR_TEXT_GRAY
        )
    ))
    
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
        small_gray_style
    ))
    
    # Cambiar a template normal (con header/footer) después de la portada
    story.append(PageBreak())
    
    # ============================================================
    # OBTENER DATOS CONSOLIDADOS
    # ============================================================
    consolidated_data = DashboardDataService.get_consolidated_data(
        empresa_id=empresa.id,
        cycle_id=evaluation_cycle.id
    )
    
    ratings = consolidated_data['ratings']
    foda_data = consolidated_data['foda']
    total_responses = consolidated_data['total_responses']
    
    # Obtener acciones CAME
    came_actions = CameAction.objects.filter(
        evaluation_cycle=evaluation_cycle
    ).select_related('assigned_to', 'created_by').order_by('action_type', 'priority')
    
    # Obtener respuestas abiertas
    open_answers = OpenAnswer.objects.filter(
        survey_response__evaluation_cycle=evaluation_cycle,
        survey_response__status='submitted'
    ).select_related('survey_response').order_by('question_type', 'created_at')[:30]
    
    # ============================================================
    # RESUMEN EJECUTIVO CON STORYTELLING
    # ============================================================
    story.append(Paragraph("Resumen Ejecutivo", section_title_style))
    
    if total_responses == 0:
        # Caso sin datos: texto narrativo explicativo
        story.append(Paragraph(
            "En este ciclo aún no se registran respuestas, por lo que no se dispone de información cuantitativa para este informe. Se recomienda activar la participación de los directivos para obtener una visión estratégica consolidada.",
            normal_style
        ))
    else:
        # Calcular métricas para el texto narrativo
        avg_ratings = []
        for rating in ratings:
            if rating['count'] > 0:
                avg_ratings.append(rating['average'])
        
        avg_general = sum(avg_ratings) / len(avg_ratings) if avg_ratings else 0.0
        
        # Encontrar dimensión con mejor y peor promedio
        dim_best = None
        dim_worst = None
        best_avg = 0.0
        worst_avg = 10.0
        
        dim_name_map = {
            'area_health': 'Salud del Área',
            'team_performance': 'Rendimiento del Equipo',
            'strategy_alignment': 'Alineación Estratégica',
            'process_maturity': 'Madurez del Proceso',
            'tech_maturity': 'Madurez Tecnológica'
        }
        
        for rating in ratings:
            if rating['average'] > best_avg:
                best_avg = rating['average']
                dim_best = dim_name_map.get(rating['dimension'], rating['dimension'].replace('_', ' ').title())
            if rating['average'] < worst_avg:
                worst_avg = rating['average']
                dim_worst = dim_name_map.get(rating['dimension'], rating['dimension'].replace('_', ' ').title())
        
        # Generar texto narrativo interpretativo
        narrative_parts = [
            f"En este ciclo participaron <b>{total_responses}</b> directivo(s).",
            f"El promedio general de las dimensiones evaluadas es <b>{avg_general:.2f}/10</b>."
        ]
        
        if dim_best:
            narrative_parts.append(
                f"La dimensión mejor valorada es '<b>{dim_best}</b>' con {best_avg:.2f}/10"
            )
        if dim_worst and dim_worst != dim_best:
            narrative_parts.append(
                f"mientras que la dimensión con mayor oportunidad de mejora es '<b>{dim_worst}</b>' con {worst_avg:.2f}/10."
            )
        elif dim_worst:
            narrative_parts.append(".")
        else:
            narrative_parts.append(".")
        
        narrative_text = " ".join(narrative_parts)
        story.append(Paragraph(narrative_text, normal_style))
        story.append(Spacer(1, 0.8*cm))
        
        # Tabla de métricas clave
        metrics_data = [['Métrica', 'Valor']]
        metrics_data.append(['Total de Respuestas', str(total_responses)])
        metrics_data.append(['Promedio General', f"{avg_general:.2f}/10"])
        if ratings:
            metrics_data.append(['Dimensiones Evaluadas', str(len(ratings))])
        
        metrics_table = Table(metrics_data, colWidths=[8*cm, 6*cm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_ORANGE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.8*cm))
        
        # Tabla de promedio por dimensión
        if ratings:
            story.append(Paragraph(
                "<b>Promedio por Dimensión</b>",
                ParagraphStyle(
                    'SubsectionTitle',
                    parent=styles['Normal'],
                    fontSize=11,
                    spaceAfter=10,
                    fontName='Helvetica-Bold'
                )
            ))
            
            dim_data = [['Dimensión', 'Promedio', 'Mín', 'Máx', '# Respuestas']]
            for rating in ratings:
                dim_name = dim_name_map.get(rating['dimension'], rating['dimension'].replace('_', ' ').title())
                dim_data.append([
                    dim_name,
                    f"{rating['average']:.2f}",
                    str(rating['min_value']),
                    str(rating['max_value']),
                    str(rating['count'])
                ])
            
            dim_table = Table(dim_data, colWidths=[5*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm])
            dim_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_DARK),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_GRAY]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(dim_table)
    
    story.append(PageBreak())
    
    # ============================================================
    # RATINGS CONSOLIDADOS CON TEXTO DE APOYO
    # ============================================================
    story.append(Paragraph("Ratings Consolidados", section_title_style))
    
    if not ratings:
        story.append(Paragraph(
            "No hay datos de ratings disponibles para este ciclo.",
            normal_style
        ))
    else:
        story.append(Paragraph(
            "A continuación se presentan las estadísticas detalladas por dimensión evaluada:",
            normal_style
        ))
        story.append(Spacer(1, 0.5*cm))
        
        # Calcular si hay dimensiones con promedio bajo (< 6) o todas altas (> 7)
        dims_low = [r for r in ratings if r['average'] < 6]
        dims_high = [r for r in ratings if r['average'] > 7]
        
        # Texto interpretativo después de la tabla
        ratings_data = [['Dimensión', 'Promedio', 'Desvío Est.', 'Mín', 'Máx', 'Cantidad']]
        dim_name_map = {
            'area_health': 'Salud del Área',
            'team_performance': 'Rendimiento del Equipo',
            'strategy_alignment': 'Alineación Estratégica',
            'process_maturity': 'Madurez del Proceso',
            'tech_maturity': 'Madurez Tecnológica'
        }
        
        for rating in ratings:
            dim_name = dim_name_map.get(rating['dimension'], rating['dimension'].replace('_', ' ').title())
            ratings_data.append([
                dim_name,
                f"{rating['average']:.2f}",
                f"{rating['std_dev']:.2f}" if rating['std_dev'] else "N/A",
                str(rating['min_value']),
                str(rating['max_value']),
                str(rating['count'])
            ])
        
        ratings_table = Table(ratings_data, colWidths=[4*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm])
        ratings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_DARK),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_GRAY]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(ratings_table)
        
        # Texto interpretativo después de la tabla
        story.append(Spacer(1, 0.6*cm))
        if dims_low:
            story.append(Paragraph(
                "<i>Nota: Las dimensiones con promedio inferior a 6 indican áreas que requieren atención prioritaria y deberían ser consideradas en el plan de acción estratégico.</i>",
                ParagraphStyle(
                    'InterpretiveNote',
                    parent=normal_style,
                    fontSize=9,
                    textColor=COLOR_TEXT_GRAY,
                    fontName='Helvetica-Oblique'
                )
            ))
        elif len(dims_high) == len(ratings) and len(ratings) > 0:
            story.append(Paragraph(
                "<i>Nota: En general, las percepciones sobre las dimensiones evaluadas son positivas, aunque se recomienda revisar en detalle los casos puntuales con menor valoración para mantener la mejora continua.</i>",
                ParagraphStyle(
                    'InterpretiveNote',
                    parent=normal_style,
                    fontSize=9,
                    textColor=COLOR_TEXT_GRAY,
                    fontName='Helvetica-Oblique'
                )
            ))
    
    story.append(PageBreak())
    
    # ============================================================
    # FODA CONSOLIDADO CON ENFOQUE EN TOP ITEMS
    # ============================================================
    story.append(Paragraph("FODA Consolidado", section_title_style))
    
    # Mapeo de cuadrantes
    quadrant_names = {
        'strength': 'Fortalezas',
        'weakness': 'Debilidades',
        'opportunity': 'Oportunidades',
        'threat': 'Amenazas'
    }
    
    quadrant_colors = {
        'strength': colors.HexColor('#10B981'),  # Verde
        'weakness': colors.HexColor('#EF4444'),  # Rojo
        'opportunity': colors.HexColor('#3B82F6'),  # Azul
        'threat': colors.HexColor('#F97316')  # Naranja
    }
    
    # Texto introductorio antes de las tablas
    story.append(Paragraph(
        "A continuación se resumen los principales elementos del FODA consolidado. Los ítems se ordenan por frecuencia de mención, destacando aquellos que aparecen con mayor recurrencia entre los directivos.",
        normal_style
    ))
    story.append(Spacer(1, 0.6*cm))
    
    has_foda = False
    for quadrant_key, quadrant_label in quadrant_names.items():
        items = foda_data.get(quadrant_key, [])
        if items:
            has_foda = True
            story.append(Paragraph(
                f"<b>{quadrant_label}</b>",
                ParagraphStyle(
                    'QuadrantTitle',
                    parent=styles['Heading3'],
                    fontSize=13,
                    textColor=quadrant_colors[quadrant_key],
                    spaceBefore=18,
                    spaceAfter=10,
                    fontName='Helvetica-Bold'
                )
            ))
            
            # Limitar a top 10
            top_items = items[:10]
            
            foda_table_data = [['Descripción', 'Veces mencionada']]
            for item in top_items:
                # Truncar descripción si es muy larga
                desc = item['description'][:150] + "..." if len(item['description']) > 150 else item['description']
                foda_table_data.append([desc, str(item['count'])])
            
            foda_table = Table(foda_table_data, colWidths=[11*cm, 3*cm])
            foda_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), quadrant_colors[quadrant_key]),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_GRAY]),
            ]))
            story.append(foda_table)
            story.append(Spacer(1, 0.6*cm))
        else:
            # Mensaje si un cuadrante no tiene datos
            story.append(Paragraph(
                f"No se registraron {quadrant_label.lower()} en este ciclo.",
                ParagraphStyle(
                    'NoDataQuadrant',
                    parent=normal_style,
                    fontSize=9,
                    textColor=COLOR_TEXT_GRAY,
                    fontName='Helvetica-Oblique',
                    spaceBefore=10,
                    spaceAfter=5
                )
            ))
    
    if not has_foda:
        story.append(Paragraph(
            "No hay elementos FODA registrados para este ciclo. Se recomienda incentivar a los directivos a completar el análisis FODA para obtener una visión estratégica completa.",
            normal_style
        ))
    
    story.append(PageBreak())
    
    # ============================================================
    # ACCIONES CAME CON STORYTELLING
    # ============================================================
    story.append(Paragraph("Acciones CAME", section_title_style))
    
    if not came_actions.exists():
        story.append(Paragraph(
            "Aún no se han definido acciones CAME para este ciclo. Se recomienda trabajar con el equipo directivo para traducir el diagnóstico en un plan de acción concreto que permita Corregir debilidades, Afrontar amenazas, Mantener fortalezas y Explotar oportunidades.",
            normal_style
        ))
    else:
        # Texto introductorio para acciones CAME
        story.append(Paragraph(
            "Las siguientes acciones CAME se derivan del análisis FODA y representan el puente entre diagnóstico y ejecución. Se recomienda priorizar las acciones con prioridad 1 y 2 y seguimiento cercano de sus responsables.",
            normal_style
        ))
        story.append(Spacer(1, 0.8*cm))
        
        # Agrupar por tipo de acción
        # Nota: Los labels se obtienen del modelo usando get_action_type_display()
        # pero mapeamos manualmente para asegurar consistencia en español
        action_type_labels = {
            'correct': 'Corregir',
            'address': 'Afrontar',
            'maintain': 'Mantener',
            'exploit': 'Explotar'
        }
        
        action_type_colors = {
            'correct': colors.HexColor('#EF4444'),  # Rojo
            'address': colors.HexColor('#F97316'),  # Naranja
            'maintain': colors.HexColor('#10B981'),  # Verde
            'exploit': colors.HexColor('#3B82F6')  # Azul
        }
        
        current_type = None
        for action in came_actions:
            if action.action_type != current_type:
                current_type = action.action_type
                type_label = action_type_labels.get(current_type, current_type)
                story.append(Spacer(1, 0.3*cm))
                story.append(Paragraph(
                    f"<b>{type_label}</b>",
                    ParagraphStyle(
                        'ActionTypeTitle',
                        parent=styles['Heading3'],
                        fontSize=12,
                        textColor=action_type_colors.get(current_type, colors.black),
                        spaceBefore=15,
                        spaceAfter=8
                    )
                ))
            
            # Título de la acción
            story.append(Paragraph(
                f"<b>{action.title}</b>",
                ParagraphStyle(
                    'ActionTitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    spaceAfter=4
                )
            ))
            
            # Detalles en una pequeña tabla
            # Mapeo de estados a español
            status_labels = {
                'planned': 'Planificado',
                'in_progress': 'En Progreso',
                'completed': 'Completado',
                'cancelled': 'Cancelado'
            }
            status_display = status_labels.get(action.status, action.get_status_display())
            
            details_data = []
            details_data.append(['Prioridad', str(action.priority)])
            details_data.append(['Estado', status_display])
            if action.assigned_to:
                details_data.append(['Responsable', str(action.assigned_to)])
            if action.due_date:
                details_data.append(['Fecha límite', action.due_date.strftime('%d/%m/%Y')])
            
            details_table = Table(details_data, colWidths=[3*cm, 11*cm])
            details_table.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(details_table)
            
            # Descripción
            if action.description:
                story.append(Paragraph(
                    action.description,
                    ParagraphStyle(
                        'ActionDescription',
                        parent=styles['Normal'],
                        fontSize=9,
                        leftIndent=0.5*cm,
                        spaceAfter=10,
                        textColor=colors.HexColor('#4B5563')
                    )
                ))
    
    story.append(PageBreak())
    
    # ============================================================
    # PREGUNTAS ABIERTAS CON ENFOQUE CUALITATIVO
    # ============================================================
    story.append(Paragraph("Síntesis de Respuestas Abiertas", section_title_style))
    
    if not open_answers.exists():
        story.append(Paragraph(
            "No se registraron respuestas abiertas en este ciclo. Este tipo de respuestas aporta contexto cualitativo valioso al diagnóstico estratégico.",
            normal_style
        ))
    else:
        story.append(Paragraph(
            "Las respuestas abiertas aportan contexto cualitativo al diagnóstico. A continuación se listan algunas de las percepciones más relevantes registradas en este ciclo (máximo 30 respuestas mostradas):",
            normal_style
        ))
        story.append(Spacer(1, 0.3*cm))
        
        # Tabla de respuestas abiertas
        open_data = [['Tipo', 'Pregunta', 'Respuesta']]
        
        for answer in open_answers:
            question_type = answer.get_question_type_display()
            question_text = answer.question_text[:100] + "..." if len(answer.question_text) > 100 else answer.question_text
            answer_text = answer.answer[:250] + "..." if len(answer.answer) > 250 else answer.answer
            
            open_data.append([question_type, question_text, answer_text])
        
        open_table = Table(open_data, colWidths=[3*cm, 5*cm, 6*cm])
        open_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_GRAY]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(open_table)
    
    # ============================================================
    # GENERAR PDF
    # ============================================================
    # Construir PDF con headers/footers automáticos
    # Nota: ReportLab usa el PageTemplate para aplicar headers/footers
    # El total de páginas se puede calcular después del build si es necesario
    doc.build(story)
    
    # Obtener bytes del buffer
    buffer.seek(0)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes


def generate_cycle_report_excel(empresa, evaluation_cycle) -> bytes:
    """
    Genera un archivo Excel (.xlsx) con datos analíticos detallados de un ciclo de evaluación SIA.
    
    El Excel contiene las siguientes hojas:
    - "Resumen": Encabezado y métricas consolidadas por dimensión
    - "Ratings": Datos detallados de cada rating individual
    - "FODA": Datos detallados de cada elemento FODA
    - "Preguntas Abiertas": Respuestas a preguntas abiertas
    - "CAME": Acciones CAME del ciclo
    
    Args:
        empresa: instancia de core.models.Empresa
        evaluation_cycle: instancia de sia.models.EvaluationCycle
    
    Returns:
        bytes: contenido del archivo Excel generado
    
    Raises:
        ValueError: si evaluation_cycle.empresa != empresa
    """
    # Validar que el ciclo pertenece a la empresa
    if evaluation_cycle.empresa_id != empresa.id:
        raise ValueError(
            f"El ciclo de evaluación {evaluation_cycle.id} no pertenece a la empresa {empresa.id}"
        )
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Estilos para headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ============================================================
    # HOJA "RESUMEN"
    # ============================================================
    ws_resumen = wb.create_sheet("Resumen")
    
    # Encabezado
    ws_resumen['A1'] = "Empresa"
    ws_resumen['B1'] = empresa.nombre
    ws_resumen['A2'] = "Ciclo"
    ws_resumen['B2'] = evaluation_cycle.name
    ws_resumen['A3'] = "Fecha de Inicio"
    ws_resumen['B3'] = evaluation_cycle.start_date.strftime('%d/%m/%Y') if evaluation_cycle.start_date else "N/A"
    ws_resumen['A4'] = "Fecha de Fin"
    ws_resumen['B4'] = evaluation_cycle.end_date.strftime('%d/%m/%Y') if evaluation_cycle.end_date else "N/A"
    ws_resumen['A5'] = "Fecha de Generación"
    ws_resumen['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # Obtener datos consolidados
    consolidated_data = DashboardDataService.get_consolidated_data(
        empresa_id=empresa.id,
        cycle_id=evaluation_cycle.id
    )
    
    ratings = consolidated_data['ratings']
    total_responses = consolidated_data['total_responses']
    
    ws_resumen['A7'] = "Total de Respuestas"
    ws_resumen['B7'] = total_responses
    
    # Tabla de métricas por dimensión
    row = 9
    if total_responses == 0:
        ws_resumen[f'A{row}'] = "No hay respuestas registradas para este ciclo."
        ws_resumen.merge_cells(f'A{row}:F{row}')
    else:
        # Headers de la tabla
        headers = ['Dimensión', 'Promedio', 'Mín', 'Máx', 'Desvío Estándar', 'Cantidad']
        for col, header in enumerate(headers, start=1):
            cell = ws_resumen.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for rating in ratings:
            row += 1
            ws_resumen.cell(row=row, column=1, value=rating['dimension'].replace('_', ' ').title())
            ws_resumen.cell(row=row, column=2, value=round(rating['average'], 2))
            ws_resumen.cell(row=row, column=3, value=rating['min_value'])
            ws_resumen.cell(row=row, column=4, value=rating['max_value'])
            ws_resumen.cell(row=row, column=5, value=round(rating['std_dev'], 2) if rating['std_dev'] else 0.0)
            ws_resumen.cell(row=row, column=6, value=rating['count'])
        
        # Activar filtros
        ws_resumen.auto_filter.ref = ws_resumen.dimensions
    
    # Ajustar anchos de columna
    ws_resumen.column_dimensions['A'].width = 20
    ws_resumen.column_dimensions['B'].width = 30
    ws_resumen.column_dimensions['C'].width = 12
    ws_resumen.column_dimensions['D'].width = 12
    ws_resumen.column_dimensions['E'].width = 18
    ws_resumen.column_dimensions['F'].width = 12
    
    # ============================================================
    # HOJA "Ratings"
    # ============================================================
    ws_ratings = wb.create_sheet("Ratings")
    
    # Headers
    headers = ['ID Respuesta', 'Usuario', 'Departamento', 'Dimensión', 'Valor', 'Notas', 'Fecha de Creación']
    for col, header in enumerate(headers, start=1):
        cell = ws_ratings.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener ratings con relaciones
    ratings_qs = Rating.objects.filter(
        survey_response__evaluation_cycle=evaluation_cycle,
        survey_response__evaluation_cycle__empresa=empresa
    ).select_related(
        'survey_response__user',
        'survey_response__department'
    ).order_by('survey_response', 'dimension')
    
    row = 2
    for rating in ratings_qs:
        ws_ratings.cell(row=row, column=1, value=rating.survey_response.id)
        ws_ratings.cell(row=row, column=2, value=str(rating.survey_response.user) if rating.survey_response.user else "N/A")
        ws_ratings.cell(row=row, column=3, value=rating.survey_response.department.name if rating.survey_response.department else "N/A")
        ws_ratings.cell(row=row, column=4, value=rating.get_dimension_display())
        ws_ratings.cell(row=row, column=5, value=rating.value)
        ws_ratings.cell(row=row, column=6, value=rating.notes or "")
        ws_ratings.cell(row=row, column=7, value=rating.created_at.strftime('%d/%m/%Y %H:%M') if rating.created_at else "")
        row += 1
    
    if row == 2:
        ws_ratings.cell(row=2, column=1, value="No hay ratings registrados para este ciclo.")
        ws_ratings.merge_cells(f'A2:G2')
    else:
        ws_ratings.auto_filter.ref = ws_ratings.dimensions
    
    # Ajustar anchos
    ws_ratings.column_dimensions['A'].width = 12
    ws_ratings.column_dimensions['B'].width = 25
    ws_ratings.column_dimensions['C'].width = 20
    ws_ratings.column_dimensions['D'].width = 25
    ws_ratings.column_dimensions['E'].width = 10
    ws_ratings.column_dimensions['F'].width = 40
    ws_ratings.column_dimensions['G'].width = 18
    
    # ============================================================
    # HOJA "FODA"
    # ============================================================
    ws_foda = wb.create_sheet("FODA")
    
    # Headers
    headers = ['ID Respuesta', 'Usuario', 'Departamento', 'Cuadrante', 'Descripción', 'Prioridad', 'Fecha de Creación']
    for col, header in enumerate(headers, start=1):
        cell = ws_foda.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener FODA items
    foda_qs = FodaItem.objects.filter(
        survey_response__evaluation_cycle=evaluation_cycle,
        survey_response__evaluation_cycle__empresa=empresa
    ).select_related(
        'survey_response__user',
        'survey_response__department'
    ).order_by('survey_response', 'quadrant', 'priority')
    
    row = 2
    for item in foda_qs:
        ws_foda.cell(row=row, column=1, value=item.survey_response.id)
        ws_foda.cell(row=row, column=2, value=str(item.survey_response.user) if item.survey_response.user else "N/A")
        ws_foda.cell(row=row, column=3, value=item.survey_response.department.name if item.survey_response.department else "N/A")
        ws_foda.cell(row=row, column=4, value=item.get_quadrant_display())
        ws_foda.cell(row=row, column=5, value=item.description)
        ws_foda.cell(row=row, column=6, value=item.priority)
        ws_foda.cell(row=row, column=7, value=item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else "")
        row += 1
    
    if row == 2:
        ws_foda.cell(row=2, column=1, value="No hay elementos FODA registrados para este ciclo.")
        ws_foda.merge_cells(f'A2:G2')
    else:
        ws_foda.auto_filter.ref = ws_foda.dimensions
    
    # Ajustar anchos
    ws_foda.column_dimensions['A'].width = 12
    ws_foda.column_dimensions['B'].width = 25
    ws_foda.column_dimensions['C'].width = 20
    ws_foda.column_dimensions['D'].width = 15
    ws_foda.column_dimensions['E'].width = 50
    ws_foda.column_dimensions['F'].width = 10
    ws_foda.column_dimensions['G'].width = 18
    
    # ============================================================
    # HOJA "Preguntas Abiertas"
    # ============================================================
    ws_open = wb.create_sheet("Preguntas Abiertas")
    
    # Headers
    headers = ['ID Respuesta', 'Usuario', 'Departamento', 'Tipo de Pregunta', 'Pregunta', 'Respuesta', 'Fecha de Creación']
    for col, header in enumerate(headers, start=1):
        cell = ws_open.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener respuestas abiertas
    open_qs = OpenAnswer.objects.filter(
        survey_response__evaluation_cycle=evaluation_cycle,
        survey_response__evaluation_cycle__empresa=empresa
    ).select_related(
        'survey_response__user',
        'survey_response__department'
    ).order_by('survey_response', 'question_type', 'created_at')
    
    row = 2
    for answer in open_qs:
        ws_open.cell(row=row, column=1, value=answer.survey_response.id)
        ws_open.cell(row=row, column=2, value=str(answer.survey_response.user) if answer.survey_response.user else "N/A")
        ws_open.cell(row=row, column=3, value=answer.survey_response.department.name if answer.survey_response.department else "N/A")
        ws_open.cell(row=row, column=4, value=answer.get_question_type_display())
        ws_open.cell(row=row, column=5, value=answer.question_text)
        ws_open.cell(row=row, column=6, value=answer.answer)
        ws_open.cell(row=row, column=7, value=answer.created_at.strftime('%d/%m/%Y %H:%M') if answer.created_at else "")
        row += 1
    
    if row == 2:
        ws_open.cell(row=2, column=1, value="No hay respuestas abiertas registradas para este ciclo.")
        ws_open.merge_cells(f'A2:G2')
    else:
        ws_open.auto_filter.ref = ws_open.dimensions
    
    # Ajustar anchos
    ws_open.column_dimensions['A'].width = 12
    ws_open.column_dimensions['B'].width = 25
    ws_open.column_dimensions['C'].width = 20
    ws_open.column_dimensions['D'].width = 20
    ws_open.column_dimensions['E'].width = 40
    ws_open.column_dimensions['F'].width = 50
    ws_open.column_dimensions['G'].width = 18
    
    # ============================================================
    # HOJA "CAME"
    # ============================================================
    ws_came = wb.create_sheet("CAME")
    
    # Headers
    headers = [
        'ID Acción', 'Tipo de Acción', 'Título', 'Descripción', 'Prioridad', 'Estado',
        'Responsable', 'Fecha Límite', 'Fecha Completado', 'FODA Relacionado', 'Fecha de Creación'
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws_came.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Mapeo de tipos de acción
    action_type_labels = {
        'correct': 'Corregir',
        'address': 'Afrontar',
        'maintain': 'Mantener',
        'exploit': 'Explotar'
    }
    
    # Mapeo de estados
    status_labels = {
        'planned': 'Planificado',
        'in_progress': 'En Progreso',
        'completed': 'Completado',
        'cancelled': 'Cancelado'
    }
    
    # Obtener acciones CAME
    came_qs = CameAction.objects.filter(
        evaluation_cycle=evaluation_cycle,
        evaluation_cycle__empresa=empresa
    ).select_related(
        'assigned_to',
        'related_foda_item'
    ).order_by('action_type', 'priority', 'created_at')
    
    row = 2
    for action in came_qs:
        ws_came.cell(row=row, column=1, value=action.id)
        ws_came.cell(row=row, column=2, value=action_type_labels.get(action.action_type, action.get_action_type_display()))
        ws_came.cell(row=row, column=3, value=action.title)
        ws_came.cell(row=row, column=4, value=action.description)
        ws_came.cell(row=row, column=5, value=action.priority)
        ws_came.cell(row=row, column=6, value=status_labels.get(action.status, action.get_status_display()))
        ws_came.cell(row=row, column=7, value=str(action.assigned_to) if action.assigned_to else "N/A")
        ws_came.cell(row=row, column=8, value=action.due_date.strftime('%d/%m/%Y') if action.due_date else "")
        ws_came.cell(row=row, column=9, value=action.completed_at.strftime('%d/%m/%Y %H:%M') if action.completed_at else "")
        foda_desc = ""
        if action.related_foda_item:
            foda_desc = action.related_foda_item.description[:100] + "..." if len(action.related_foda_item.description) > 100 else action.related_foda_item.description
        ws_came.cell(row=row, column=10, value=foda_desc)
        ws_came.cell(row=row, column=11, value=action.created_at.strftime('%d/%m/%Y %H:%M') if action.created_at else "")
        row += 1
    
    if row == 2:
        ws_came.cell(row=2, column=1, value="No hay acciones CAME registradas para este ciclo.")
        ws_came.merge_cells(f'A2:K2')
    else:
        ws_came.auto_filter.ref = ws_came.dimensions
    
    # Ajustar anchos
    ws_came.column_dimensions['A'].width = 12
    ws_came.column_dimensions['B'].width = 15
    ws_came.column_dimensions['C'].width = 30
    ws_came.column_dimensions['D'].width = 50
    ws_came.column_dimensions['E'].width = 10
    ws_came.column_dimensions['F'].width = 15
    ws_came.column_dimensions['G'].width = 25
    ws_came.column_dimensions['H'].width = 15
    ws_came.column_dimensions['I'].width = 18
    ws_came.column_dimensions['J'].width = 40
    ws_came.column_dimensions['K'].width = 18
    
    # ============================================================
    # GUARDAR WORKBOOK EN MEMORIA
    # ============================================================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes

