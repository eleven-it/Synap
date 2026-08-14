"""Tests reporte Inventario por depósito (hub MPR demanda/inventario_deposito)."""
from datetime import date
from decimal import Decimal
from unittest.mock import MagicMock, patch

from django.test import SimpleTestCase

from mpr.views import ReportesMPRView


class MedidasInventarioDocenasTest(SimpleTestCase):
    """REQ-INVDEP-04: divisores 12/6/4, pares pipeline, docenas float."""

    def test_pipeline_pares_divisor_doce(self):
        from mpr.inventario_docenas import medidas_inventario_excel

        m = medidas_inventario_excel(24, "SemiElaborado", None)
        self.assertEqual(m["stock_um"], 24.0)
        self.assertEqual(m["um_etiqueta"], "pares")
        self.assertEqual(m["divisor"], 12)
        self.assertEqual(m["docenas"], 2.0)

    def test_produccion_usa_pares(self):
        from mpr.inventario_docenas import divisor_docena_inventario

        self.assertEqual(divisor_docena_inventario("Produccion", 6), (12, "pares"))

    def test_terminado_divisor_seis(self):
        from mpr.inventario_docenas import medidas_inventario_excel

        m = medidas_inventario_excel(18, "Terminado", 6)
        self.assertEqual(m["stock_um"], 18.0)
        self.assertEqual(m["um_etiqueta"], "packs")
        self.assertEqual(m["divisor"], 6)
        self.assertEqual(m["docenas"], 3.0)

    def test_terminado_divisor_cuatro(self):
        from mpr.inventario_docenas import medidas_inventario_excel

        m = medidas_inventario_excel(8, "Terminado", 4)
        self.assertEqual(m["docenas"], 2.0)

    def test_terminado_divisor_doce_default(self):
        from mpr.inventario_docenas import medidas_inventario_excel

        m = medidas_inventario_excel(12, "Terminado", 0)
        self.assertEqual(m["divisor"], 12)
        self.assertEqual(m["docenas"], 1.0)

    def test_docenas_float_no_entero(self):
        from mpr.inventario_docenas import medidas_inventario_excel

        m = medidas_inventario_excel(7, "Terminado", 12)
        self.assertEqual(m["docenas"], 0.58)


class InventarioDepositoGranoTest(SimpleTestCase):
    """REQ-INVDEP-02: una fila por (id_deposito, id_articulo)."""

    @patch("mpr.services_inventario_deposito.mysql_cursor")
    def test_una_fila_por_deposito_articulo(self, mock_cursor):
        from mpr.services_inventario_deposito import (
            consultar_inventario_deposito,
            parse_filtros_inventario_deposito,
        )

        cursor = MagicMock()
        mock_cursor.return_value.__enter__.return_value = cursor
        cursor.fetchone.side_effect = [
            {"Tables_in_x": "stock_deposito"},
            {"Tables_in_x": "deposito"},
            {"Tables_in_x": "articulo"},
            {"Tables_in_x": "articulo_valor_ce"},
            {"Tables_in_x": "marca"},
        ]
        cursor.fetchall.return_value = [
            {
                "id_articulo": 100,
                "id_deposito": 1,
                "saldo": Decimal("12"),
                "codigo_manual": "A100",
                "codigo_articulo": "A100",
                "descripcion_articulo": "Art dep 1",
                "nombre_deposito": "Produccion",
                "tipo_mpr": "Produccion",
                "marca_nombre": "Marca X",
                "talle": "M",
                "cantidad_promedio_bulto": 12,
                "tipo_art_fab": "Produccion",
            },
            {
                "id_articulo": 100,
                "id_deposito": 3,
                "saldo": Decimal("24"),
                "codigo_manual": "A100",
                "codigo_articulo": "A100",
                "descripcion_articulo": "Art dep 3",
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca X",
                "talle": "M",
                "cantidad_promedio_bulto": 12,
                "tipo_art_fab": "Terminado",
            },
            {
                "id_articulo": 200,
                "id_deposito": 3,
                "saldo": Decimal("6"),
                "codigo_manual": "B200",
                "codigo_articulo": "B200",
                "descripcion_articulo": "Otro art",
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca Y",
                "talle": "L",
                "cantidad_promedio_bulto": 6,
                "tipo_art_fab": "Terminado",
            },
        ]

        with patch("mpr.services_inventario_deposito._nombre_tabla") as mock_tbl:
            mock_tbl.side_effect = lambda c, n: n
            with patch(
                "mpr.services_inventario_deposito.enriquecer_medidas_inventario",
                side_effect=lambda filas, base: [
                    {**f, "stock_um": float(f["saldo"]), "docenas": float(f["saldo"]) / 12, "um_etiqueta": "packs", "divisor": 12}
                    for f in filas
                ],
            ):
                filtros = parse_filtros_inventario_deposito({})
                resultado = consultar_inventario_deposito("empresa92", filtros)

        filas = resultado["filas"]
        self.assertEqual(len(filas), 3)
        pares = {(f["id_deposito"], f["id_articulo"]) for f in filas}
        self.assertEqual(len(pares), 3)
        self.assertIn((1, 100), pares)
        self.assertIn((3, 100), pares)
        self.assertIn((3, 200), pares)


class InventarioDepositoJerarquiaTest(SimpleTestCase):
    """REQ-INVDEP-03: Depósito→Marca con subtotales = SUM(docenas)."""

    def test_marcas_separadas_subtotales_suma_docenas(self):
        from mpr.services_inventario_deposito import agrupar_jerarquia_deposito_marca

        filas = [
            {
                "id_deposito": 3,
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca A",
                "codigo_manual": "A1",
                "talle": "M",
                "docenas": 2.0,
            },
            {
                "id_deposito": 3,
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca B",
                "codigo_manual": "B1",
                "talle": "L",
                "docenas": 3.0,
            },
            {
                "id_deposito": 3,
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca A",
                "codigo_manual": "A2",
                "talle": "S",
                "docenas": 1.0,
            },
        ]
        jerarquia, total = agrupar_jerarquia_deposito_marca(filas)
        self.assertEqual(len(jerarquia), 1)
        self.assertEqual(jerarquia[0]["total_docenas"], 6.0)
        self.assertEqual(len(jerarquia[0]["marcas"]), 2)
        marcas = {m["marca_nombre"]: m for m in jerarquia[0]["marcas"]}
        self.assertEqual(marcas["Marca A"]["subtotal_docenas"], 3.0)
        self.assertEqual(len(marcas["Marca A"]["filas"]), 2)
        self.assertEqual(marcas["Marca B"]["subtotal_docenas"], 3.0)
        self.assertEqual(total, 6.0)


class InventarioDepositoParidadExcelFixtureTest(SimpleTestCase):
    """REQ-INVDEP-13: stand-in automatizado — TOTAL=SUM(docenas) mix divisores 12/6/4.

    UAT live vs Inventarios.xlsx permanece documentado en docs/mpr/INVENTARIO_DEPOSITO_ARTICULO.md.
    """

    def test_total_docenas_fixture_ratios_excel_12_6_4(self):
        from mpr.inventario_docenas import medidas_inventario_excel
        from mpr.services_inventario_deposito import (
            agrupar_jerarquia_deposito_marca,
            calcular_total_docenas,
        )

        casos = [
            (24, "SemiElaborado", None),
            (18, "Terminado", 6),
            (8, "Terminado", 4),
        ]
        filas = []
        for i, (saldo, tipo_mpr, bulto) in enumerate(casos):
            medidas = medidas_inventario_excel(saldo, tipo_mpr, bulto)
            filas.append({
                "id_deposito": 3,
                "id_articulo": 100 + i,
                "nombre_deposito": "Terminado",
                "tipo_mpr": tipo_mpr,
                "marca_nombre": f"Marca {i + 1}",
                "codigo_manual": f"T00{i + 1}",
                "talle": "M",
                **medidas,
            })

        self.assertEqual(medidas_inventario_excel(24, "SemiElaborado", None)["docenas"], 2.0)
        self.assertEqual(medidas_inventario_excel(18, "Terminado", 6)["docenas"], 3.0)
        self.assertEqual(medidas_inventario_excel(8, "Terminado", 4)["docenas"], 2.0)

        total = calcular_total_docenas(filas)
        self.assertEqual(total, 7.0)
        jerarquia, total_jer = agrupar_jerarquia_deposito_marca(filas)
        self.assertEqual(total_jer, 7.0)
        self.assertEqual(jerarquia[0]["total_docenas"], 7.0)


class TotalesInventarioDepositoTest(SimpleTestCase):
    """REQ-INVDEP-05/06/09: TOTAL=SUM(docenas), 2da OFF, Tercero incluido."""

    def test_total_es_suma_docenas_mix(self):
        from mpr.services_inventario_deposito import calcular_total_docenas

        filas = [
            {"docenas": 2.0},
            {"docenas": 3.0},
        ]
        self.assertEqual(calcular_total_docenas(filas), 5.0)

    def test_filtro_2da_excluido_default(self):
        from mpr.services_inventario_deposito import parse_filtros_inventario_deposito

        f = parse_filtros_inventario_deposito({})
        self.assertFalse(f.incluir_2da)

    def test_filtro_2da_opt_in(self):
        from mpr.services_inventario_deposito import parse_filtros_inventario_deposito

        f = parse_filtros_inventario_deposito({"incluir_2da": "1"})
        self.assertTrue(f.incluir_2da)

    def test_sql_excluye_2da_por_default(self):
        from mpr.services_inventario_deposito import _sql_filtro_tipo_mpr

        sql = _sql_filtro_tipo_mpr(incluir_2da=False)
        self.assertIn("2daSeleccion", sql)
        self.assertIn("!=", sql)

    def test_sql_incluye_2da_cuando_opt_in(self):
        from mpr.services_inventario_deposito import _sql_filtro_tipo_mpr

        self.assertEqual(_sql_filtro_tipo_mpr(incluir_2da=True), "")

    def test_no_filtra_tercero_en_where(self):
        from mpr.services_inventario_deposito import _build_where_articulo

        clausula, _ = _build_where_articulo(busqueda=None, marcas_incluidos=[], id_articulo=None)
        self.assertNotIn("tipo_art_fab", clausula.lower())

    @patch("mpr.services_inventario_deposito.mysql_cursor")
    def test_tercero_incluido_en_resultado(self, mock_cursor):
        from mpr.services_inventario_deposito import (
            consultar_inventario_deposito,
            parse_filtros_inventario_deposito,
        )

        cursor = MagicMock()
        mock_cursor.return_value.__enter__.return_value = cursor
        cursor.execute.side_effect = [
            None,
            None,
        ]
        # Tablas
        cursor.fetchone.side_effect = [
            {"Tables_in_x": "stock_deposito"},
            {"Tables_in_x": "deposito"},
            {"Tables_in_x": "articulo"},
            {"Tables_in_x": "articulo_valor_ce"},
            {"Tables_in_x": "marca"},
        ]
        cursor.fetchall.return_value = [
            {
                "id_articulo": 100,
                "id_deposito": 3,
                "saldo": Decimal("12"),
                "codigo_manual": "T001",
                "codigo_articulo": "T001",
                "descripcion_articulo": "Art tercero",
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca X",
                "talle": "M",
                "cantidad_promedio_bulto": 12,
                "tipo_art_fab": "Tercero",
            }
        ]

        with patch("mpr.services_inventario_deposito._nombre_tabla") as mock_tbl:
            mock_tbl.side_effect = lambda c, n: n
            with patch("mpr.services_inventario_deposito.enriquecer_medidas_inventario") as mock_enr:
                mock_enr.side_effect = lambda filas, base: filas
                filtros = parse_filtros_inventario_deposito({})
                resultado = consultar_inventario_deposito("empresa92", filtros)
        self.assertEqual(len(resultado["filas"]), 1)
        self.assertEqual(resultado["filas"][0]["tipo_art_fab"], "Tercero")


class ReportesInventarioDepositoViewTest(SimpleTestCase):
    """REQ-INVDEP-08/12, REQ-SHELL-02: filtros, empty state, shell ignora período."""

    def _context(self, get_params, consulta_payload=None):
        view = ReportesMPRView()
        view.request = MagicMock()
        view.request.session = {"user": {"base_empresa": "empresa92"}}
        view.request.GET = get_params
        payload = consulta_payload or {
            "filas": [],
            "depositos_jerarquia": [],
            "total_docenas": 0.0,
            "kpis": {"total_docenas": 0, "depositos": 0, "filas": 0},
        }
        with patch("mpr.views._get_base_empresa", return_value="empresa92"):
            with patch("mpr.views.listar_depositos_config", return_value=[]):
                with patch(
                    "stock.services.inventario_tabla.listar_marcas_catalogo",
                    return_value=[],
                ):
                    with patch(
                        "mpr.services_inventario_deposito.consultar_inventario_deposito",
                        return_value=payload,
                    ):
                        return view.get_context_data()

    def test_ruta_inventario_deposito(self):
        ctx = self._context({"grupo": "demanda", "reporte": "inventario_deposito"})
        self.assertEqual(ctx["grupo"], "demanda")
        self.assertEqual(ctx["reporte"], "inventario_deposito")
        self.assertEqual(ctx["titulo_reporte"], "Inventario por depósito")

    def test_filtro_marca_pasa_a_servicio(self):
        with patch(
            "mpr.services_inventario_deposito.consultar_inventario_deposito"
        ) as mock_consulta:
            mock_consulta.return_value = {
                "filas": [],
                "depositos_jerarquia": [],
                "total_docenas": 0.0,
                "kpis": {"total_docenas": 0, "depositos": 0, "filas": 0},
            }
            view = ReportesMPRView()
            view.request = MagicMock()
            view.request.session = {"user": {"base_empresa": "empresa92"}}
            view.request.GET = {
                "grupo": "demanda",
                "reporte": "inventario_deposito",
                "marcas_incluidos": "5",
            }
            with patch("mpr.views._get_base_empresa", return_value="empresa92"):
                with patch("mpr.views.listar_depositos_config", return_value=[]):
                    with patch(
                        "stock.services.inventario_tabla.listar_marcas_catalogo",
                        return_value=[],
                    ):
                        view.get_context_data()
            filtros = mock_consulta.call_args[0][1]
            self.assertEqual(filtros.marcas_incluidos, [5])

    @patch("mpr.services_inventario_deposito.mysql_cursor")
    def test_filtro_marca_aplica_where_y_recalcula_total(self, mock_cursor):
        from mpr.services_inventario_deposito import (
            InventarioDepositoFiltros,
            consultar_inventario_deposito,
        )

        cursor = MagicMock()
        mock_cursor.return_value.__enter__.return_value = cursor
        cursor.fetchone.side_effect = [
            {"Tables_in_x": "stock_deposito"},
            {"Tables_in_x": "deposito"},
            {"Tables_in_x": "articulo"},
            {"Tables_in_x": "articulo_valor_ce"},
            {"Tables_in_x": "marca"},
        ]
        cursor.fetchall.return_value = [
            {
                "id_articulo": 10,
                "id_deposito": 3,
                "saldo": Decimal("12"),
                "codigo_manual": "MA",
                "codigo_articulo": "MA",
                "descripcion_articulo": "Marca A art",
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca A",
                "talle": "M",
                "cantidad_promedio_bulto": 12,
                "tipo_art_fab": "Terminado",
            },
        ]

        with patch("mpr.services_inventario_deposito._nombre_tabla") as mock_tbl:
            mock_tbl.side_effect = lambda c, n: n
            with patch(
                "mpr.services_inventario_deposito.enriquecer_medidas_inventario",
                side_effect=lambda filas, base: [
                    {**f, "stock_um": 12.0, "docenas": 1.0, "um_etiqueta": "packs", "divisor": 12}
                    for f in filas
                ],
            ):
                filtros = InventarioDepositoFiltros(marcas_incluidos=[5])
                resultado = consultar_inventario_deposito("empresa92", filtros)

        sql = cursor.execute.call_args[0][0]
        params = cursor.execute.call_args[0][1]
        self.assertIn("CodigoMarca IN", sql)
        self.assertIn(5, params)
        self.assertEqual(len(resultado["filas"]), 1)
        self.assertEqual(resultado["filas"][0]["marca_nombre"], "Marca A")
        self.assertEqual(resultado["total_docenas"], 1.0)

    def test_fecha_corte_display_formato_dd_mm_yyyy(self):
        ctx = self._context(
            {
                "grupo": "demanda",
                "reporte": "inventario_deposito",
                "fecha_corte": "2025-08-14",
            },
            consulta_payload={
                "filas": [],
                "depositos_jerarquia": [],
                "total_docenas": 0.0,
                "kpis": {"total_docenas": 0, "depositos": 0, "filas": 0},
                "fecha_corte": date(2025, 8, 14),
                "usa_stock_deposito": True,
            },
        )
        self.assertEqual(ctx["fecha_corte_display"], "14/08/2025")

    def test_preparar_presentacion_fecha_corte_dd_mm_yyyy(self):
        from mpr.reportes_presentacion import preparar_inventario_deposito_presentacion

        ctx = preparar_inventario_deposito_presentacion(
            {
                "filas": [],
                "depositos_jerarquia": [],
                "total_docenas": 0.0,
                "fecha_corte": date(2025, 8, 14),
                "usa_stock_deposito": True,
            },
            "docenas",
        )
        self.assertEqual(ctx["fecha_corte_display"], "14/08/2025")

    def test_empty_state_espanol(self):
        ctx = self._context({"grupo": "demanda", "reporte": "inventario_deposito"})
        self.assertIn("empty_titulo", ctx)
        self.assertIn("depósito", ctx["empty_titulo"].lower())

    def test_shell_periodo_no_afecta_fecha_corte(self):
        hoy = date.today()
        with patch(
            "mpr.services_inventario_deposito.consultar_inventario_deposito"
        ) as mock_consulta:
            mock_consulta.return_value = {
                "filas": [],
                "depositos_jerarquia": [],
                "total_docenas": 0.0,
                "kpis": {"total_docenas": 0, "depositos": 0, "filas": 0},
            }
            view = ReportesMPRView()
            view.request = MagicMock()
            view.request.session = {"user": {"base_empresa": "empresa92"}}
            view.request.GET = {
                "grupo": "demanda",
                "reporte": "inventario_deposito",
                "desde": "2020-01-01",
                "hasta": "2020-01-31",
            }
            with patch("mpr.views._get_base_empresa", return_value="empresa92"):
                with patch("mpr.views.listar_depositos_config", return_value=[]):
                    with patch(
                        "stock.services.inventario_tabla.listar_marcas_catalogo",
                        return_value=[],
                    ):
                        view.get_context_data()
            filtros = mock_consulta.call_args[0][1]
            self.assertEqual(filtros.fecha_corte, hoy)

    def test_stock_demanda_sin_regresion(self):
        view = ReportesMPRView()
        view.request = MagicMock()
        view.request.session = {"user": {"base_empresa": "empresa92"}}
        view.request.GET = {"grupo": "demanda", "reporte": "stock"}
        with patch("mpr.views._get_base_empresa", return_value="empresa92"):
            with patch("mpr.views.reporte_mpr_stock", return_value=[]) as mock_stock:
                ctx = view.get_context_data()
        mock_stock.assert_called_once()
        self.assertEqual(ctx["reporte"], "stock")


class InventarioDepositoCorteHoyTest(SimpleTestCase):
    """REQ-INVDEP-07: corte=hoy usa stock_deposito, no stock_a_fecha."""

    def test_usa_stock_deposito_cuando_fecha_es_hoy(self):
        from mpr.services_inventario_deposito import _usa_stock_deposito

        self.assertTrue(_usa_stock_deposito(date.today()))

    @patch("mpr.services_inventario_deposito.mysql_cursor")
    @patch("mpr.services_inventario_deposito.saldos_stock_a_fecha")
    def test_corte_hoy_consulta_stock_deposito_no_historico(self, mock_saldos, mock_cursor):
        from mpr.services_inventario_deposito import (
            InventarioDepositoFiltros,
            consultar_inventario_deposito,
        )

        cursor = MagicMock()
        mock_cursor.return_value.__enter__.return_value = cursor
        cursor.fetchone.side_effect = [
            {"Tables_in_x": "stock_deposito"},
            {"Tables_in_x": "deposito"},
            {"Tables_in_x": "articulo"},
            {"Tables_in_x": "articulo_valor_ce"},
            {"Tables_in_x": "marca"},
        ]
        cursor.fetchall.return_value = [
            {
                "id_articulo": 100,
                "id_deposito": 3,
                "saldo": Decimal("12"),
                "codigo_manual": "A100",
                "codigo_articulo": "A100",
                "descripcion_articulo": "Art hoy",
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca",
                "talle": "M",
                "cantidad_promedio_bulto": 12,
                "tipo_art_fab": "Terminado",
            },
        ]

        with patch("mpr.services_inventario_deposito._nombre_tabla") as mock_tbl:
            mock_tbl.side_effect = lambda c, n: n
            with patch(
                "mpr.services_inventario_deposito.enriquecer_medidas_inventario",
                side_effect=lambda filas, base: [
                    {**f, "stock_um": 12.0, "docenas": 1.0, "um_etiqueta": "packs", "divisor": 12}
                    for f in filas
                ],
            ):
                filtros = InventarioDepositoFiltros(fecha_corte=date.today())
                resultado = consultar_inventario_deposito("empresa92", filtros)

        mock_saldos.assert_not_called()
        self.assertTrue(resultado["usa_stock_deposito"])
        sql = cursor.execute.call_args[0][0]
        self.assertIn("stock_deposito", sql.lower())
        self.assertEqual(len(resultado["filas"]), 1)


class InventarioDepositoHistoricoTest(SimpleTestCase):
    """REQ-INVDEP-07: corte pasado usa stock_a_fecha, sin advertencia PR-1."""

    @patch("mpr.services_inventario_deposito.mysql_cursor")
    @patch("mpr.services_inventario_deposito.saldos_stock_a_fecha")
    def test_corte_historico_usa_stock_a_fecha(self, mock_saldos, mock_cursor):
        from mpr.services_inventario_deposito import (
            InventarioDepositoFiltros,
            consultar_inventario_deposito,
        )

        corte = date(2020, 6, 30)
        mock_saldos.return_value = {(100, 3): Decimal("24")}

        cursor = MagicMock()
        mock_cursor.return_value.__enter__.return_value = cursor
        cursor.fetchone.side_effect = [
            {"Tables_in_x": "deposito"},
            {"Tables_in_x": "articulo"},
            {"Tables_in_x": "articulo_valor_ce"},
            {"Tables_in_x": "marca"},
        ]
        cursor.fetchall.return_value = [
            {
                "id_articulo": 100,
                "id_deposito": 3,
                "saldo": Decimal("24"),
                "codigo_manual": "A100",
                "codigo_articulo": "A100",
                "descripcion_articulo": "Art hist",
                "nombre_deposito": "Terminado",
                "tipo_mpr": "Terminado",
                "marca_nombre": "Marca",
                "talle": "M",
                "cantidad_promedio_bulto": 12,
                "tipo_art_fab": "Terminado",
            }
        ]

        with patch("mpr.services_inventario_deposito._nombre_tabla") as mock_tbl:
            mock_tbl.side_effect = lambda c, n: n
            with patch(
                "mpr.services_inventario_deposito.enriquecer_medidas_inventario",
                side_effect=lambda filas, base: [
                    {**f, "stock_um": 24.0, "docenas": 2.0, "um_etiqueta": "packs", "divisor": 12}
                    for f in filas
                ],
            ):
                filtros = InventarioDepositoFiltros(fecha_corte=corte)
                resultado = consultar_inventario_deposito("empresa92", filtros)

        mock_saldos.assert_called_once()
        self.assertFalse(resultado["usa_stock_deposito"])
        self.assertNotIn("advertencia_fecha", resultado)
        self.assertEqual(resultado["fecha_corte"], corte)
        self.assertEqual(len(resultado["filas"]), 1)


class InventarioDepositoExportXlsxTest(SimpleTestCase):
    """REQ-INVDEP-10 / REQ-SHELL-10: export xlsx columnas español + TOTAL docenas."""

    def test_export_xlsx_encabezados_y_total(self):
        from mpr.export import exportar_inventario_deposito_xlsx

        filas = [
            {
                "nombre_deposito": "Terminado",
                "marca_nombre": "Marca X",
                "codigo_manual": "T001",
                "descripcion_articulo": "Calcetín",
                "talle": "M",
                "stock_um": 12.0,
                "um_etiqueta": "packs",
                "docenas": 1.0,
            },
            {
                "nombre_deposito": "Semi elaborado",
                "marca_nombre": "Marca Y",
                "codigo_manual": "S002",
                "descripcion_articulo": "Semi",
                "talle": "",
                "stock_um": 24.0,
                "um_etiqueta": "pares",
                "docenas": 2.0,
            },
        ]
        fecha = date(2025, 8, 14)
        resp = exportar_inventario_deposito_xlsx(
            filas,
            total_docenas=3.0,
            fecha_corte=fecha,
            titulo="Inventario por depósito",
        )
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header_row = 2 if fecha else 1
        headers = [cell.value for cell in ws[header_row]]
        self.assertEqual(
            headers,
            ["Depósito", "Marca", "Artículo", "Talle", "Stock", "Docenas"],
        )
        self.assertEqual(ws.cell(row=header_row + 1, column=1).value, "Terminado")
        self.assertEqual(ws.cell(row=header_row + 1, column=6).value, 1.0)
        total_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "TOTAL":
                total_row = row
                break
        self.assertIsNotNone(total_row)
        self.assertEqual(total_row[5], 3.0)

    def test_hub_declara_soporte_excel(self):
        from mpr.reportes_hub import reporte_soporta_export_xlsx

        self.assertTrue(reporte_soporta_export_xlsx("demanda", "inventario_deposito"))
        self.assertFalse(reporte_soporta_export_xlsx("demanda", "stock"))

    def test_vista_responde_xlsx(self):
        view = ReportesMPRView()
        ctx = {
            "grupo": "demanda",
            "reporte": "inventario_deposito",
            "titulo_reporte": "Inventario por depósito",
            "filas": [
                {
                    "nombre_deposito": "Dep",
                    "marca_nombre": "Marca",
                    "codigo_manual": "X",
                    "descripcion_articulo": "Desc",
                    "talle": "L",
                    "stock_um": 6.0,
                    "um_etiqueta": "packs",
                    "docenas": 0.5,
                }
            ],
            "total_docenas": 0.5,
            "fecha_corte_iso": date.today().isoformat(),
            "modo_presentacion": "docenas",
        }
        response = view._respuesta_xlsx(ctx)
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
