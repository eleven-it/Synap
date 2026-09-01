"""Exportación CSV/Excel para reportes MPR."""
import csv
import io
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from django.http import HttpResponse

from core.utils.administranet_types import to_decimal_or_none


def filas_a_csv(
    filas: Iterable[Dict[str, Any]],
    columnas: Sequence[Tuple[str, str]],
) -> bytes:
    """
    Genera CSV UTF-8 con BOM.

    columnas: secuencia de (clave_dict, encabezado_español).
    """
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow([titulo for _, titulo in columnas])
    for fila in filas or []:
        writer.writerow([_celda_csv(fila.get(clave)) for clave, _ in columnas])
    return buf.getvalue().encode("utf-8")


def _celda_csv(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Sí" if val else "No"
    return str(val)


def _escribir_seccion_csv(
    writer: csv.writer,
    titulo: str,
    columnas: Sequence[Tuple[str, str]],
    filas: Iterable[Dict[str, Any]],
) -> None:
    writer.writerow([])
    writer.writerow([titulo])
    writer.writerow([titulo_col for _, titulo_col in columnas])
    for fila in filas or []:
        writer.writerow([_celda_csv(fila.get(clave)) for clave, _ in columnas])


def _cantidad_export(fila: Dict[str, Any], clave: str, modo: str) -> Any:
    if modo == "docenas":
        display = fila.get(f"{clave}_display")
        if display is not None and str(display).strip():
            return display
    return fila.get(clave)


def analisis_trazabilidad_a_csv(
    analisis: Dict[str, Any],
    *,
    modo: str = "docenas",
    fecha_desde_display: str = "",
    fecha_hasta_display: str = "",
) -> bytes:
    """
    CSV multi-sección del análisis trazabilidad artículo (UTF-8 BOM).

    Secciones: resumen, demanda PED, stock/brecha, movimientos, eventos MPR, a producir.
    """
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf, lineterminator="\r\n")

    articulo = analisis.get("articulo") or {}
    kpis = analisis.get("kpis") or {}
    demanda = analisis.get("demanda_ped") or {}
    stock = analisis.get("stock") or {}
    brechas = analisis.get("brechas") or {}
    a_producir = analisis.get("a_producir") or {}
    saldo_inicial = analisis.get("saldo_inicial") or {}
    movimientos = analisis.get("movimientos") or []
    eventos = analisis.get("eventos_mpr") or []

    writer.writerow(["Análisis trazabilidad artículo"])
    writer.writerow(["Código", articulo.get("codigo") or "-"])
    writer.writerow(["Descripción", articulo.get("descripcion") or "-"])
    writer.writerow(["ID artículo", articulo.get("id") or ""])
    periodo = ""
    if fecha_desde_display or fecha_hasta_display:
        periodo = f"{fecha_desde_display} — {fecha_hasta_display}".strip(" —")
    writer.writerow(["Período", periodo])
    writer.writerow(["Saldo inicial", saldo_inicial.get("valor", 0)])
    writer.writerow(["Saldo inicial calculado", "Sí" if saldo_inicial.get("calculado_ok") else "No"])

    writer.writerow([])
    writer.writerow(["Resumen KPI"])
    writer.writerow(["Concepto", "Valor"])
    writer.writerow(["Pedido (P_ped)", kpis.get("pedido", 0)])
    writer.writerow(["Terminado", kpis.get("terminado", 0)])
    writer.writerow(["PED Urgente", kpis.get("ped_urgente", 0)])
    writer.writerow(["TOT Urgente", kpis.get("tot_urgente", 0)])
    writer.writerow(["Saldo final", kpis.get("saldo_final", 0)])

    demanda_filas = []
    for ped in demanda.get("filas") or []:
        demanda_filas.append({
            "nro_pedido": ped.get("nro_pedido") or "-",
            "nombre_cliente": ped.get("nombre_cliente") or "-",
            "fecha": ped.get("fecha") or "-",
            "cantidad_pendiente_prod": _cantidad_export(
                ped, "cantidad_pendiente_prod", modo
            ),
        })
    _escribir_seccion_csv(
        writer,
        "DEMANDA PED",
        [
            ("nro_pedido", "Nº pedido"),
            ("nombre_cliente", "Cliente"),
            ("fecha", "Fecha"),
            ("cantidad_pendiente_prod", "Pendiente"),
        ],
        demanda_filas,
    )
    writer.writerow([])
    writer.writerow(["Total P_ped", demanda.get("totales", {}).get("p_ped", 0)])

    writer.writerow([])
    writer.writerow(["STOCK Y BRECHA"])
    writer.writerow(["Concepto", "Valor"])
    writer.writerow(["Terminado", stock.get("terminado", 0)])
    writer.writerow(["Terminado negativo", "Sí" if stock.get("negativo") else "No"])
    writer.writerow(["PED Urgente", brechas.get("ped_urgente", 0)])
    writer.writerow(["TOT Urgente", brechas.get("tot_urgente", 0)])
    writer.writerow(["Reserva", brechas.get("reserva", 0)])
    writer.writerow(["Texto explicativo", brechas.get("texto_explicativo") or ""])

    mov_filas = []
    for mov in movimientos:
        mov_filas.append({
            "fecha_display": mov.get("fecha_display") or "-",
            "tipo_mov": mov.get("tipo_mov") or mov.get("clase_ui") or "-",
            "nro_comprobante": mov.get("nro_comprobante") or "-",
            "detalle": mov.get("detalle") or "",
            "entrada": _cantidad_export(mov, "entrada", modo),
            "salida": _cantidad_export(mov, "salida", modo),
            "saldo_corrido": mov.get("saldo_corrido", ""),
            "afecta_deposito": mov.get("afecta_deposito"),
            "operario": mov.get("operario") or "-",
        })
    _escribir_seccion_csv(
        writer,
        "MOVIMIENTOS",
        [
            ("fecha_display", "Fecha"),
            ("tipo_mov", "Tipo"),
            ("nro_comprobante", "Comprobante"),
            ("detalle", "Detalle"),
            ("entrada", "Entrada"),
            ("salida", "Salida"),
            ("saldo_corrido", "Saldo corrido"),
            ("afecta_deposito", "Afecta depósito"),
            ("operario", "Operario"),
        ],
        mov_filas,
    )

    evento_filas = []
    for ev in eventos:
        evento_filas.append({
            "fecha_display": ev.get("fecha_display") or "-",
            "tipo_label": ev.get("tipo_label") or ev.get("tipo") or "-",
            "cantidad": _cantidad_export(ev, "cantidad", modo),
            "detalle": ev.get("detalle") or "",
            "operario": ev.get("operario") or "-",
        })
    _escribir_seccion_csv(
        writer,
        "EVENTOS MPR",
        [
            ("fecha_display", "Fecha"),
            ("tipo_label", "Tipo"),
            ("cantidad", "Cantidad"),
            ("detalle", "Detalle"),
            ("operario", "Operario"),
        ],
        evento_filas,
    )

    writer.writerow([])
    writer.writerow(["A PRODUCIR"])
    writer.writerow(["Concepto", "Valor"])
    writer.writerow(["Cantidad a producir", a_producir.get("cantidad", 0)])
    writer.writerow(["Capacidad Semi", a_producir.get("capacidad_semi", 0)])
    writer.writerow(["Alerta Semi cero", "Sí" if a_producir.get("alerta_semi_cero") else "No"])

    return buf.getvalue().encode("utf-8")


def _fmt_fecha_export(val: Any) -> str:
    if isinstance(val, datetime):
        val = val.date()
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    return str(val or "")


def _num_xlsx(val: Any) -> Any:
    dec = to_decimal_or_none(val)
    if dec is None:
        return ""
    if dec == dec.to_integral_value():
        return int(dec)
    return float(dec)


def exportar_inventario_deposito_xlsx(
    filas: Iterable[Dict[str, Any]],
    *,
    total_docenas: float,
    fecha_corte: Optional[date] = None,
    titulo: str = "Inventario por depósito",
) -> HttpResponse:
    """Export Excel inventario_deposito: Depósito, Marca, Artículo, Talle, Stock, Docenas + TOTAL."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Depósito", "Marca", "Artículo", "Talle", "Stock", "Docenas"]
    row = 1
    if fecha_corte:
        ws.cell(row=row, column=1, value=f"{titulo} · Corte {_fmt_fecha_export(fecha_corte)}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1E40AF")
        row += 1

    for col_num, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    row += 1

    for fila in filas or []:
        codigo = fila.get("codigo_manual") or fila.get("codigo_articulo") or "-"
        desc = (fila.get("descripcion_articulo") or "").strip()
        articulo = f"{codigo} — {desc}" if desc else str(codigo)
        valores = [
            fila.get("nombre_deposito") or "",
            fila.get("marca_nombre") or "",
            articulo,
            fila.get("talle") or "",
            _num_xlsx(fila.get("stock_um")),
            _num_xlsx(fila.get("docenas")),
        ]
        for col_num, valor in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col_num, value="" if valor is None else valor)
            cell.border = border
            cell.alignment = left if col_num <= 4 else center
        row += 1

    total_font = Font(bold=True)
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    ws.cell(row=row, column=5, value="").font = total_font
    total_cell = ws.cell(row=row, column=6, value=_num_xlsx(total_docenas))
    total_cell.font = total_font
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=row, column=col_num).border = border

    from openpyxl.utils import get_column_letter

    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_len = len(str(headers[col_num - 1]))
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    sufijo = _fmt_fecha_export(fecha_corte).replace("/", "") if fecha_corte else "hoy"
    nombre = f"inventario_deposito_{sufijo}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response
