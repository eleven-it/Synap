"""Tests mínimos de vistas de auditoría contable (URLs y mocks)."""
import io
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock, patch

from django.contrib.messages.storage.fallback import FallbackStorage
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.test import Client, RequestFactory, SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone

from contabilidad_audit.models import AprobacionREI, PlanCorreccion
from contabilidad_audit.views import (
    PERMISO_CORREGIR,
    PERMISO_LEER,
    _listar_detalle_lote,
    _listar_planes_diagnostico,
    _parse_alcance_dry_run,
    _purgar_planes_vencidos,
    auditoria_dry_run,
    auditoria_lote_detalle,
    auditoria_lote_rollback,
    auditoria_lotes,
)

_TEMPLATES = Path(__file__).resolve().parent.parent / "templates" / "contabilidad_audit"


def _session_user(base_empresa: str = "empresa_test"):
    return {
        "id_usuario": 1,
        "cod_usuario": "auditor",
        "base_empresa": base_empresa,
    }


class _UserConPermisoLeer:
    is_authenticated = True
    cod_usuario = "auditor"

    def is_admin(self):
        return False

    def tiene_permiso(self, code):
        return code == PERMISO_LEER


class _UserConPermisoCorregir:
    is_authenticated = True
    cod_usuario = "auditor"

    def is_admin(self):
        return False

    def tiene_permiso(self, code):
        return code in (PERMISO_LEER, PERMISO_CORREGIR)


class _UserSinPermiso:
    is_authenticated = True
    cod_usuario = "vendedor"

    def is_admin(self):
        return False

    def tiene_permiso(self, code):
        return False


class AuditoriaUrlsSmokeTestCase(SimpleTestCase):
    def test_reverse_urls_lotes(self):
        self.assertEqual(
            reverse("contabilidad_audit:auditoria_lotes"),
            "/contabilidad/auditoria/lotes/",
        )
        self.assertEqual(
            reverse("contabilidad_audit:auditoria_lote_rollback", kwargs={"lote_id": "L20260718-001"}),
            "/contabilidad/auditoria/lotes/L20260718-001/rollback/",
        )
        self.assertEqual(
            reverse("contabilidad_audit:auditoria_lote_detalle", kwargs={"lote_id": "L20260718-001"}),
            "/contabilidad/auditoria/lotes/L20260718-001/",
        )
        self.assertEqual(
            reverse("contabilidad_audit:auditoria_dry_run"),
            "/contabilidad/auditoria/dry-run/",
        )
        self.assertEqual(
            reverse("contabilidad_audit:manual_usuario"),
            "/contabilidad/manual/",
        )
        self.assertEqual(
            reverse("contabilidad_audit:auditoria_apply"),
            "/contabilidad/auditoria/apply/",
        )


class AuditoriaClaridadUiTemplatesTestCase(SimpleTestCase):
    """Smoke de textos/columnas para huérfanos FA/FC/OP y autorización de escritura."""

    def test_tablero_tiene_detalle_huerfanos_fa_fc_op(self):
        html = (_TEMPLATES / "auditoria_tablero.html").read_text(encoding="utf-8")
        self.assertIn("comprobante_compra_pago_sin_asiento", html)
        self.assertIn("comprobante_venta_cobranza_sin_asiento", html)
        self.assertIn("esCheckHuerfanosComprobante", html)
        self.assertIn("resumenHuerfanosPorTipo", html)
        self.assertIn("Nro comprobante", html)
        self.assertIn("CodigoMovimiento sin filas en cont_asiento", html)
        self.assertIn("Generar diagnóstico", html)
        self.assertIn("mostrarEsperaDiagnostico", html)
        self.assertIn("tituloDiagnostico", html)
        self.assertIn("Generando diagnóstico", html)
        self.assertIn("synapShowPostLoadingProgress", html)
        self.assertIn("Ejecutando auditoría", html)
        self.assertIn("partials/synap_post_loading_modal.html", html)
        # Terminología UI: nunca «dry-run» visible en el tablero.
        self.assertNotIn("Generar dry-run", html)
        self.assertNotIn("dry-run de regeneración", html)

    def test_tablero_cta_diagnostico_por_tarjeta_sin_cta_global(self):
        """El diagnóstico se abre desde la tarjeta del kanban, nunca desde el header."""
        html = (_TEMPLATES / "auditoria_tablero.html").read_text(encoding="utf-8")
        # Sin CTA global: el header sólo tiene Ejecutar, Ayuda, Configuración, Lotes y Excel.
        self.assertNotIn("puedeDryRun", html)
        self.assertNotIn("dryRunLink", html)
        # CTA por tarjeta acotado al check_id de esa tarjeta.
        self.assertIn("contabilidad-checks-corregibles", html)
        self.assertIn("checksCorregibles", html)
        self.assertIn("puedeAbrirDiagnostico(check)", html)
        self.assertIn("diagnosticoLink(check.check_id)", html)
        self.assertIn("p.append('check_ids', checkId)", html)
        # Con diferencias pero sin motor de corrección: texto discreto, sin CTA.
        self.assertIn("Sin corrección automática", html)
        # Con 0 diferencias no hay pie con CTA.
        self.assertIn("mostrarPieDiagnostico", html)

    def test_tablero_seleccion_diagnosticos_sin_default(self):
        html = (_TEMPLATES / "auditoria_tablero.html").read_text(encoding="utf-8")
        self.assertIn('aria-label="Selección de diagnósticos"', html)
        self.assertIn("Elegí al menos un diagnóstico para ejecutar la auditoría.", html)
        self.assertIn("Diagnósticos OK", html)
        self.assertIn("contabilidad-checks-seleccionados", html)
        self.assertIn("checksPorSeveridad", html)
        self.assertIn("seleccionarGrupo", html)
        self.assertIn("estaSeleccionado", html)
        self.assertIn(":aria-pressed=", html)
        # Ningún diagnóstico activo por defecto: la selección arranca vacía.
        self.assertNotIn("checksSeleccionados: checks.map(", html)
        self.assertIn("checksSeleccionados: seleccionPrevia", html)
        # Ejecutar y exportar exigen diagnósticos seleccionados.
        self.assertIn("puedeAuditar()", html)
        self.assertIn("checksSeleccionados.length", html)
        self.assertIn("p.append('check_ids', cid)", html)
        self.assertNotIn("puedeEjecutar", html)

    def test_tablero_franja_diagnosticos_compacta(self):
        """La selección vive en una franja cerrada + panel «Elegir», no en un muro de pills."""
        html = (_TEMPLATES / "auditoria_tablero.html").read_text(encoding="utf-8")
        self.assertIn("diagAbierto", html)
        self.assertIn("Filtrar diagnósticos…", html)
        self.assertIn("seleccionarSoloCriticos", html)
        self.assertIn("Solo críticos", html)
        self.assertIn("chipsSeleccion", html)
        self.assertIn("chipsRestantes", html)
        self.assertIn("gruposFiltrados", html)
        self.assertIn("Ninguno seleccionado", html)
        self.assertIn("max-h-64", html)

    def test_tablero_encabezado_denso_una_fila(self):
        """Chrome denso: barra slate-800 de una fila, sin hero con overlap."""
        html = (_TEMPLATES / "auditoria_tablero.html").read_text(encoding="utf-8")
        self.assertIn("rounded-lg border border-slate-700 bg-slate-800", html)
        self.assertIn("sticky top-14", html)
        self.assertIn("flex flex-wrap items-center gap-x-3 gap-y-2", html)
        self.assertIn('id="auditoria-ejercicio"', html)
        self.assertNotIn("-mt-12", html)
        self.assertNotIn("pt-8 pb-20", html)
        self.assertNotIn("Contabilidad · Auditoría solo lectura", html)
        # Empresa y período no se muestran en la UI del tablero.
        self.assertNotIn("Empresa (base)", html)
        self.assertNotIn("Período (opcional)", html)
        # Export solo Excel en el encabezado.
        self.assertNotIn("exportUrl('csv')", html)
        self.assertIn("exportUrl('xlsx')", html)

    def test_dry_run_chrome_denso_sin_empresa_periodo(self):
        """Diagnóstico alineado al tablero: chrome denso, ejercicio en barra, sin empresa/período."""
        html = (_TEMPLATES / "auditoria_dry_run.html").read_text(encoding="utf-8")
        self.assertIn("rounded-lg border border-slate-700 bg-slate-800", html)
        self.assertIn("sticky top-14", html)
        self.assertIn('id="auditoria-ejercicio"', html)
        self.assertIn("ejerciciosPeriodosUrl", html)
        self.assertIn("seleccionarEjercicio", html)
        self.assertNotIn("-mt-12", html)
        self.assertNotIn("Empresa (base)", html)
        self.assertNotIn("Período (opc.)", html)
        self.assertNotIn("name=\"id_periodo\"", html)
        self.assertNotIn("exportUrl('csv')", html)
        self.assertIn("exportUrl('xlsx')", html)
        self.assertIn("Diagnóstico", html)
        self.assertIn("Detalle de correcciones (muestra)", html)
        self.assertNotIn("Cambios a cambiar", html)
        self.assertIn("Nro asiento", html)
        self.assertIn("CodigoMovimiento", html)
        self.assertIn("Fecha", html)
        self.assertIn("Cuenta", html)
        self.assertIn("Debe", html)
        self.assertIn("Haber", html)
        self.assertIn("Descripción", html)
        self.assertNotIn('<th class="px-2 py-2 text-left">Valor anterior</th>', html)
        self.assertNotIn('<th class="px-2 py-2 text-left">Valor nuevo</th>', html)
        self.assertIn("valorTexto", html)
        self.assertIn("resumenValor", html)
        self.assertIn("campoAsiento", html)
        self.assertIn("campoCodigoMovimiento", html)
        self.assertIn("campoDescripcion", html)
        self.assertIn("filasTablaPlan", html)
        self.assertIn(":rowspan=", html)
        self.assertIn("align-middle", html)
        self.assertIn("style: 'currency'", html)
        self.assertIn("currency: 'ARS'", html)
        self.assertIn("abrirDetalleItem", html)
        self.assertIn("Datos técnicos", html)
        self.assertIn("Información técnica del plan", html)
        self.assertIn("tecnicoPlanAbierto", html)
        self.assertIn("diagnósticos que no entran en este plan", html)

    def test_dry_run_terminologia_diagnostico(self):
        """Copy UI: «Diagnóstico», nunca «dry-run» ni «simulación» visibles."""
        html = (_TEMPLATES / "auditoria_dry_run.html").read_text(encoding="utf-8")
        self.assertIn("Generar diagnóstico", html)
        self.assertIn("Diagnóstico de corrección contable", html)
        self.assertIn("Id diagnóstico", html)
        self.assertIn("Diagnóstico <strong>sin escritura</strong> en la base contable", html)
        # Guards / hashes / backups no van en la vista principal del contador.
        self.assertNotIn("Guards de validez del plan", html)
        self.assertIn("Información técnica del plan", html)
        self.assertNotIn("Generar dry-run", html)
        self.assertNotIn("Simulación de corrección", html)
        self.assertNotIn("Generando simulación", html)
        # El CTA de apply sólo aparece con ítems aplicables (plan no vacío) y abre modal local.
        self.assertIn('x-show="payload?.impacto?.total_aplicables > 0"', html)
        self.assertIn("abrirConfirmacionApply", html)
        self.assertIn("auditoria_apply_ejecutar", html)
        self.assertNotIn("/contabilidad/auditoria/apply/?dry_run_id=", html)

    def test_dry_run_titulo_pagina_es_diagnostico(self):
        from contabilidad_audit.views import _contexto_dry_run

        request = RequestFactory().get("/contabilidad/auditoria/dry-run/")
        request.session = {}
        with patch("contabilidad_audit.views._base_empresa_sesion", return_value="empresa_test"):
            ctx = _contexto_dry_run(request)
        self.assertEqual(ctx["titulo_pagina"], "Diagnóstico de corrección contable")

    def test_dry_run_plan_vacio_y_bloque_huerfanos(self):
        html = (_TEMPLATES / "auditoria_dry_run.html").read_text(encoding="utf-8")
        self.assertIn("Plan vacío", html)
        self.assertIn("Asientos huérfanos a regenerar", html)
        self.assertNotIn("comprobante_venta_cobranza_sin_asiento", html)
        self.assertIn("totalHuerfanosRegenerar", html)
        self.assertIn("no hay comprobantes sin asiento regenerables en el alcance", html)
        self.assertIn("synapShowPostLoadingProgress", html)
        self.assertIn("Generando diagnóstico", html)

    def test_dry_run_exige_diagnosticos_en_ui(self):
        html = (_TEMPLATES / "auditoria_dry_run.html").read_text(encoding="utf-8")
        self.assertIn('aria-label="Selección de diagnósticos"', html)
        self.assertIn("seleccionarSoloCriticos", html)
        self.assertIn("Elegí al menos un diagnóstico para generarlo.", html)
        self.assertIn("p.append('check_ids', cid)", html)
        self.assertIn("contabilidad-checks-seleccionados", html)

    def test_lotes_sin_cta_generar_dry_run(self):
        """El flujo correcto es tablero → kanban → tarjeta; lotes no genera diagnósticos."""
        html = (_TEMPLATES / "auditoria_lotes.html").read_text(encoding="utf-8")
        self.assertNotIn("Generar dry-run", html)
        self.assertIn("Tablero de auditoría", html)
        self.assertIn(">Tablero<", html)
        self.assertIn("Id diagnóstico", html)
        self.assertIn("Planes de diagnóstico", html)
        self.assertIn("rounded-lg border border-slate-700 bg-slate-800", html)
        self.assertIn("sticky top-14", html)
        self.assertNotIn("via-violet-900", html)
        self.assertNotIn("Log de correcciones", html)
        self.assertIn("Abrir", html)

    def test_lote_detalle_template_chrome_y_columnas(self):
        html = (_TEMPLATES / "auditoria_lote_detalle.html").read_text(encoding="utf-8")
        self.assertIn("rounded-lg border border-slate-700 bg-slate-800", html)
        self.assertIn("sticky top-14", html)
        self.assertIn("Cambios aplicados", html)
        self.assertIn("CodigoMovimiento", html)
        self.assertIn("Resumen del lote", html)
        self.assertIn("table_view", html)

    def test_lotes_lista_tiene_ver_y_excel(self):
        html = (_TEMPLATES / "auditoria_lotes.html").read_text(encoding="utf-8")
        self.assertIn("auditoria_lote_detalle", html)
        self.assertIn("?format=xlsx", html)
        self.assertIn("visibility", html)


class AuditoriaChecksCorregiblesTestCase(SimpleTestCase):
    """El tablero expone a Alpine los checks con corrección automática."""

    def test_checks_corregibles_coincide_con_motor(self):
        from legacy_db.services.cont_recalculo_service import CHECKS_INCLUIDOS

        from contabilidad_audit.views import _checks_corregibles

        self.assertEqual(_checks_corregibles(), list(CHECKS_INCLUIDOS))
        self.assertIn("comprobante_compra_pago_sin_asiento", _checks_corregibles())

    @patch("contabilidad_audit.views._base_empresa_sesion", return_value="empresa_test")
    def test_contexto_tablero_incluye_checks_corregibles(self, _mock_base):
        from legacy_db.services.cont_recalculo_service import CHECKS_INCLUIDOS

        from contabilidad_audit.views import _contexto_tablero

        request = RequestFactory().get("/contabilidad/auditoria/")
        request.session = {}
        ctx = _contexto_tablero(request)
        self.assertEqual(ctx["checks_corregibles"], list(CHECKS_INCLUIDOS))


class AuditoriaDryRunParseTestCase(SimpleTestCase):
    @patch("contabilidad_audit.views._base_empresa_sesion", return_value="empresa_test")
    def test_parse_alcance_dry_run_exige_check_ids(self, _mock_base):
        request = RequestFactory().get(
            "/contabilidad/auditoria/dry-run/",
            {"id_ejercicio": "1"},
        )
        with self.assertRaises(ValueError) as ctx:
            _parse_alcance_dry_run(request)
        self.assertIn("diagnóstico", str(ctx.exception))

    @patch("contabilidad_audit.views._base_empresa_sesion", return_value="empresa_test")
    def test_parse_alcance_dry_run_incluye_check_ids(self, _mock_base):
        request = RequestFactory().get(
            "/contabilidad/auditoria/dry-run/",
            [("id_ejercicio", "1"), ("check_ids", "asiento_balanceado")],
        )
        alcance = _parse_alcance_dry_run(request)
        self.assertEqual(alcance["check_ids"], ["asiento_balanceado"])

    def test_apply_banner_unica_escritura_y_resumen(self):
        html = (_TEMPLATES / "auditoria_apply.html").read_text(encoding="utf-8")
        self.assertIn("única acción que escribe en MySQL legacy", html)
        self.assertIn("Resumen del plan a aplicar", html)
        self.assertIn("Asientos huérfanos a regenerar (compra/venta)", html)
        self.assertIn("asientos_regenerar_por_tipo", html)
        self.assertIn("confirmacion_entiendo", html)
        self.assertNotIn("confirmacion_final", html)
        self.assertNotIn("APLICAR-", html)
        self.assertIn("synap-post-loading", html)
        self.assertIn("Aplicando corrección contable", html)

    def test_lotes_rollback_tiene_modal_espera(self):
        html = (_TEMPLATES / "auditoria_lotes.html").read_text(encoding="utf-8")
        self.assertIn("synap-post-loading", html)
        self.assertIn("Revirtiendo lote", html)
        self.assertIn('procesando = true', html)
        self.assertIn("partials/synap_post_loading_modal.html", html)


class AuditoriaLotesViewsTestCase(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.client = Client()

    def _request_get_lotes(self, user=None):
        request = self.factory.get("/contabilidad/auditoria/lotes/")
        request.session = self.client.session
        request.session["user"] = _session_user()
        request.session.save()
        request.user = user or _UserConPermisoLeer()
        return request

    def _request_post_rollback(self, lote_id: str, user=None):
        request = self.factory.post(
            f"/contabilidad/auditoria/lotes/{lote_id}/rollback/",
        )
        request.session = self.client.session
        request.session["user"] = _session_user()
        request.session.save()
        request.user = user or _UserConPermisoCorregir()
        messages = FallbackStorage(request)
        request._messages = messages
        return request

    def test_lotes_sin_permiso_403(self):
        request = self._request_get_lotes(_UserSinPermiso())
        with self.assertRaises(PermissionDenied):
            auditoria_lotes(request)

    @patch("contabilidad_audit.views.render")
    @patch("contabilidad_audit.views._listar_planes_diagnostico", return_value=[])
    @patch(
        "contabilidad_audit.views._listar_lotes_correccion",
        return_value=[
            {
                "lote_id": "L20260718-001",
                "fecha": "18/07/2026 14:30",
                "usuario": "auditor",
                "estado": "aplicado",
                "dry_run_id": "abc-123",
                "filas_correccion": 5,
            }
        ],
    )
    def test_lotes_con_permiso_lista(self, mock_listar, mock_planes, mock_render):
        mock_render.return_value = HttpResponse("ok")
        request = self._request_get_lotes()
        response = auditoria_lotes(request)
        self.assertEqual(response.status_code, 200)
        mock_listar.assert_called_once_with("empresa_test")
        mock_planes.assert_called_once_with("empresa_test")
        ctx = mock_render.call_args[0][2]
        self.assertEqual(len(ctx["lotes"]), 1)
        self.assertEqual(ctx["lotes"][0]["lote_id"], "L20260718-001")
        self.assertEqual(ctx["planes"], [])

    def test_rollback_sin_permiso_403(self):
        request = self._request_post_rollback("L20260718-001", _UserSinPermiso())
        with self.assertRaises(PermissionDenied):
            auditoria_lote_rollback(request, "L20260718-001")

    @patch("legacy_db.services.cont_recalculo_service.rollback_lote")
    def test_rollback_con_permiso_redirige(self, mock_rollback):
        mock_rollback.return_value = {"ok": True, "lote_id": "L20260718-001"}
        request = self._request_post_rollback("L20260718-001")
        response = auditoria_lote_rollback(request, "L20260718-001")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("contabilidad_audit:auditoria_lotes"))
        mock_rollback.assert_called_once()
        args, kwargs = mock_rollback.call_args
        self.assertEqual(args[0], "empresa_test")
        self.assertEqual(args[1], "L20260718-001")
        self.assertTrue(kwargs.get("tiene_permiso_corregir"))

    @patch("contabilidad_audit.views.get_mysql_pool")
    def test_listar_lotes_helper(self, mock_pool):
        from contabilidad_audit.views import _listar_lotes_correccion

        conn = MagicMock()
        cursor = MagicMock()
        cursor.fetchall.return_value = [
            ("L1", datetime(2026, 7, 18, 14, 30, 0), "u1", "aplicado", "dry-1", 3),
        ]
        conn.cursor.return_value = cursor
        mock_pool.return_value.get_connection.return_value.__enter__.return_value = conn

        lotes = _listar_lotes_correccion("empresa_test")
        self.assertEqual(len(lotes), 1)
        self.assertEqual(lotes[0]["lote_id"], "L1")
        self.assertEqual(lotes[0]["filas_correccion"], 3)
        self.assertIn("/", lotes[0]["fecha"])


class AuditoriaLoteDetalleTestCase(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.client = Client()

    def _request_get_detalle(self, lote_id: str, user=None, params=None):
        qs = params or {}
        request = self.factory.get(
            f"/contabilidad/auditoria/lotes/{lote_id}/",
            qs,
        )
        request.session = self.client.session
        request.session["user"] = _session_user()
        request.session.save()
        request.user = user or _UserConPermisoLeer()
        messages = FallbackStorage(request)
        request._messages = messages
        return request

    @patch("contabilidad_audit.views.get_mysql_pool")
    def test_listar_detalle_lote_helper(self, mock_pool):
        conn = MagicMock()
        cursor = MagicMock()
        cursor.fetchall.return_value = [
            (
                1,
                "asiento_balanceado",
                "cont_ejercicio_saldo_cta",
                '{"id_pc": 10}',
                "100.50",
                "200.75",
                "auditor",
                datetime(2026, 7, 18, 15, 0, 0),
            ),
        ]
        conn.cursor.return_value = cursor
        mock_pool.return_value.get_connection.return_value.__enter__.return_value = conn

        filas = _listar_detalle_lote("empresa_test", "L1")
        self.assertEqual(len(filas), 1)
        self.assertEqual(filas[0]["check_id"], "asiento_balanceado")
        self.assertIn("→", filas[0]["cambio_resumen"])
        self.assertIn("/", filas[0]["fecha"])

    @patch("contabilidad_audit.views.render")
    @patch("contabilidad_audit.views._listar_detalle_lote", return_value=[])
    @patch(
        "contabilidad_audit.views._obtener_lote",
        return_value={
            "lote_id": "L1",
            "base_empresa": "empresa_test",
            "dry_run_id": "abc-123",
            "dry_run_id_corto": "abc-123",
            "fecha": "18/07/2026 14:30",
            "usuario": "auditor",
            "estado": "aplicado",
            "filas_correccion": 0,
        },
    )
    def test_lote_detalle_con_permiso(self, mock_obtener, mock_detalle, mock_render):
        mock_render.return_value = HttpResponse("ok")
        request = self._request_get_detalle("L1")
        response = auditoria_lote_detalle(request, "L1")
        self.assertEqual(response.status_code, 200)
        mock_obtener.assert_called_once_with("empresa_test", "L1")
        ctx = mock_render.call_args[0][2]
        self.assertEqual(ctx["lote"]["lote_id"], "L1")
        self.assertIn("excel_url", ctx)

    @patch("contabilidad_audit.views._obtener_lote", return_value=None)
    def test_lote_detalle_inexistente_redirige(self, _mock_obtener):
        request = self._request_get_detalle("L-inexistente")
        response = auditoria_lote_detalle(request, "L-inexistente")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("contabilidad_audit:auditoria_lotes"))

    def test_exportar_lote_xlsx_smoke(self):
        from contabilidad_audit.services.export import exportar_lote_xlsx

        lote = {
            "lote_id": "L20260718-001",
            "base_empresa": "empresa_test",
            "dry_run_id": "dry-uuid-123",
            "fecha": "18/07/2026 14:30",
            "usuario": "auditor",
            "estado": "aplicado",
            "filas_correccion": 1,
        }
        filas = [
            {
                "check_id": "saldo_ejercicio_vs_diario",
                "titulo_check": "Saldo ejercicio vs diario",
                "tabla": "cont_ejercicio_saldo_cta",
                "clave": {"id_pc": 1},
                "valor_anterior": "100.00",
                "valor_nuevo": "200.00",
                "cambio_resumen": "$ 100,00 → $ 200,00",
                "usuario": "auditor",
                "fecha": "18/07/2026 14:30",
            },
            {
                "check_id": "comprobante_venta_cobranza_sin_asiento",
                "titulo_check": "Comprobante venta/cobranza sin asiento",
                "tabla": "cont_asiento",
                "clave": {"codigo_movimiento": "1", "id_pc": 13},
                "valor_anterior": None,
                "valor_nuevo": {
                    "nro_asiento": 100,
                    "fecha_asiento": "2026-01-08",
                    "codigo_movimiento": "1",
                    "id_pc": 13,
                    "debe_asiento": "1500.50",
                    "haber_asiento": "0.00",
                    "desc_asiento": "Venta - Nro Comp. 0001",
                    "desc_concepto_asiento": "Venta",
                },
                "cambio_resumen": "Asiento 100 · 08/01/2026 · Cta 13 · Debe $ 1.500,50",
                "usuario": "auditor",
                "fecha": "18/07/2026 14:30",
            },
        ]
        response = exportar_lote_xlsx(lote, filas)
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("lote_correccion_empresa_test", response["Content-Disposition"])
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(wb.sheetnames, ["Resumen", "Detalle"])
        detalle = wb["Detalle"]
        headers = [c.value for c in detalle[1]]
        self.assertIn("Diagnóstico", headers)
        self.assertIn("Cambios aplicados", headers)
        self.assertIn("CodigoMovimiento", headers)
        self.assertNotIn("Cambio a realizar", headers)
        self.assertNotIn("Cambios a cambiar", headers)
        self.assertIn("Debe", headers)
        self.assertNotIn("Clave", headers)
        self.assertNotIn("Check", headers)
        # Fila asiento: sin JSON, con monto y fecha legibles
        self.assertEqual(detalle.max_row, 3)
        fila_asiento = [c.value for c in detalle[3]]
        self.assertIn("Asiento insertado", fila_asiento)
        self.assertIn("08/01/2026", fila_asiento)
        self.assertTrue(any(isinstance(v, str) and "$" in v for v in fila_asiento))
        self.assertFalse(any(isinstance(v, str) and v.strip().startswith("{") for v in fila_asiento))

        # Segunda fila de asiento con mismo nro: merge vertical en columna Nro asiento
        filas.append(
            {
                "check_id": "comprobante_venta_cobranza_sin_asiento",
                "titulo_check": "Comprobante venta/cobranza sin asiento",
                "tabla": "cont_asiento",
                "clave": {"codigo_movimiento": "1", "id_pc": 14},
                "valor_anterior": None,
                "valor_nuevo": {
                    "nro_asiento": 100,
                    "fecha_asiento": "2026-01-08",
                    "codigo_movimiento": "1",
                    "id_pc": 14,
                    "debe_asiento": "0.00",
                    "haber_asiento": "1500.50",
                    "desc_asiento": "Venta - Nro Comp. 0002",
                    "desc_concepto_asiento": "Venta",
                },
                "cambio_resumen": "Asiento 100 · 08/01/2026 · Cta 14 · Haber $ 1.500,50",
                "usuario": "auditor",
                "fecha": "18/07/2026 14:30",
            },
        )
        lote["filas_correccion"] = 2
        response = exportar_lote_xlsx(lote, filas)
        wb2 = openpyxl.load_workbook(io.BytesIO(response.content))
        detalle2 = wb2["Detalle"]
        self.assertEqual(detalle2.max_row, 4)
        merged = list(detalle2.merged_cells.ranges)
        self.assertTrue(any(str(r).startswith("C") and ":C" in str(r) for r in merged))


class ExportMergeVerticalTestCase(SimpleTestCase):
    """Merge vertical de celdas iguales consecutivas en export contador."""

    def test_rangos_merge_vertical_mismo_asiento(self):
        from contabilidad_audit.services.export import _rangos_merge_vertical

        filas = [
            {"nro_asiento": 100, "diagnostico": "A", "debe": "", "haber": "$ 1,00"},
            {"nro_asiento": 100, "diagnostico": "A", "debe": "$ 1,00", "haber": ""},
        ]
        self.assertEqual(_rangos_merge_vertical(filas, "nro_asiento"), [(0, 1)])
        self.assertEqual(_rangos_merge_vertical(filas, "diagnostico"), [(0, 1)])
        self.assertEqual(_rangos_merge_vertical(filas, "debe"), [])
        self.assertEqual(_rangos_merge_vertical(filas, "haber"), [])

    def test_rangos_merge_no_cruza_asientos_distintos(self):
        """Mismo diagnóstico en asientos distintos no se combina."""
        from contabilidad_audit.services.export import _rangos_merge_vertical

        filas = [
            {"nro_asiento": 100, "diagnostico": "Compra sin asiento", "excluido": "No"},
            {"nro_asiento": 100, "diagnostico": "Compra sin asiento", "excluido": "No"},
            {"nro_asiento": 101, "diagnostico": "Compra sin asiento", "excluido": "No"},
            {"nro_asiento": 101, "diagnostico": "Compra sin asiento", "excluido": "No"},
        ]
        self.assertEqual(_rangos_merge_vertical(filas, "nro_asiento"), [(0, 1), (2, 3)])
        self.assertEqual(_rangos_merge_vertical(filas, "diagnostico"), [(0, 1), (2, 3)])
        self.assertEqual(_rangos_merge_vertical(filas, "excluido"), [(0, 1), (2, 3)])
        # Un solo rango (0,3) sería incorrecto: cruzaría asientos.
        self.assertNotIn((0, 3), _rangos_merge_vertical(filas, "diagnostico"))

    def test_rangos_merge_vertical_valores_vacios_no_generan_rango(self):
        from contabilidad_audit.services.export import _rangos_merge_vertical

        filas = [
            {"nro_asiento": 1, "debe": "", "haber": ""},
            {"nro_asiento": 1, "debe": "", "haber": ""},
            {"nro_asiento": 2, "debe": "—", "haber": "—"},
            {"nro_asiento": 2, "debe": "—", "haber": "—"},
        ]
        self.assertEqual(_rangos_merge_vertical(filas, "debe"), [])
        self.assertEqual(_rangos_merge_vertical(filas, "haber"), [])

    def test_exportar_dry_run_xlsx_merge_smoke(self):
        from contabilidad_audit.services.export import exportar_dry_run_xlsx
        import openpyxl

        payload = {
            "base_empresa": "empresa_test",
            "dry_run_id": "dry-merge-test",
            "creado_en": "26/07/2026",
            "expira_en": "26/07/2026",
            "alcance": {"id_ejercicio": 1},
            "impacto": {"total_items": 2, "total_aplicables": 2, "total_excluidos": 0},
            "plan": {
                "items": [
                    {
                        "check_id": "comprobante_venta_cobranza_sin_asiento",
                        "accion": "insert",
                        "tabla": "cont_asiento",
                        "excluido": False,
                        "valor_nuevo": {
                            "nro_asiento": 100,
                            "fecha_asiento": "2026-01-08",
                            "codigo_movimiento": "1",
                            "id_pc": 13,
                            "debe_asiento": "1500.50",
                            "haber_asiento": "0.00",
                            "desc_asiento": "Venta - renglón 1",
                        },
                    },
                    {
                        "check_id": "comprobante_venta_cobranza_sin_asiento",
                        "accion": "insert",
                        "tabla": "cont_asiento",
                        "excluido": False,
                        "valor_nuevo": {
                            "nro_asiento": 100,
                            "fecha_asiento": "2026-01-08",
                            "codigo_movimiento": "1",
                            "id_pc": 14,
                            "debe_asiento": "0.00",
                            "haber_asiento": "1500.50",
                            "desc_asiento": "Venta - renglón 2",
                        },
                    },
                ],
            },
        }
        response = exportar_dry_run_xlsx(payload)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        plan = wb["Plan"]
        merged = list(plan.merged_cells.ranges)
        self.assertTrue(any(str(r).startswith("C2:C3") for r in merged))


class AuditoriaPlanesDiagnosticoTestCase(TestCase):
    """Historial de planes: purge lazy, listado y reapertura por dry_run_id."""

    BASE = "empresa_test"

    def _crear_plan(
        self,
        *,
        estado="propuesto",
        expira_delta=None,
        check_ids=None,
        total_aplicables=3,
    ) -> PlanCorreccion:
        ahora = timezone.now()
        if expira_delta is None:
            expira_delta = timedelta(minutes=30)
        return PlanCorreccion.objects.create(
            base_empresa=self.BASE,
            alcance={"id_ejercicio": 1, "check_ids": check_ids or ["asiento_balanceado"]},
            config_hash="hash-test",
            data_fingerprint="fp-test",
            plan={"impacto": {"total_aplicables": total_aplicables}, "items": []},
            estado=estado,
            creado_por="auditor",
            creado_en=ahora,
            expira_en=ahora + expira_delta,
        )

    def test_purgar_planes_vencidos_borra_solo_expirado_e_invalidado(self):
        vencido = self._crear_plan(expira_delta=timedelta(minutes=-5))
        invalidado = self._crear_plan(estado="invalidado", expira_delta=timedelta(minutes=-10))
        aplicado = self._crear_plan(estado="aplicado", expira_delta=timedelta(minutes=-20))
        vigente = self._crear_plan(expira_delta=timedelta(minutes=20))

        AprobacionREI.objects.create(
            dry_run_id=vencido.dry_run_id,
            id_pc=1,
            id_ejercicio=1,
            rei_teorico="100.0000",
            rei_actual="90.0000",
        )
        AprobacionREI.objects.create(
            dry_run_id=invalidado.dry_run_id,
            id_pc=2,
            id_ejercicio=1,
            rei_teorico="50.0000",
            rei_actual="40.0000",
        )

        borrados = _purgar_planes_vencidos(self.BASE)

        self.assertEqual(borrados, 2)
        self.assertFalse(PlanCorreccion.objects.filter(dry_run_id=vencido.dry_run_id).exists())
        self.assertFalse(PlanCorreccion.objects.filter(dry_run_id=invalidado.dry_run_id).exists())
        self.assertTrue(PlanCorreccion.objects.filter(dry_run_id=aplicado.dry_run_id).exists())
        self.assertTrue(PlanCorreccion.objects.filter(dry_run_id=vigente.dry_run_id).exists())
        self.assertEqual(AprobacionREI.objects.filter(dry_run_id=vencido.dry_run_id).count(), 0)
        self.assertEqual(AprobacionREI.objects.filter(dry_run_id=invalidado.dry_run_id).count(), 0)

        vigente.refresh_from_db()
        self.assertEqual(vigente.estado, "propuesto")
        aplicado.refresh_from_db()
        self.assertEqual(aplicado.estado, "aplicado")

    def test_listar_planes_estados_ui(self):
        vigente = self._crear_plan(expira_delta=timedelta(minutes=15))
        aplicado = self._crear_plan(estado="aplicado", expira_delta=timedelta(minutes=-5))

        planes = _listar_planes_diagnostico(self.BASE)

        self.assertEqual(len(planes), 2)
        por_id = {p["dry_run_id"]: p for p in planes}
        self.assertEqual(por_id[str(vigente.dry_run_id)]["estado_ui"], "vigente")
        self.assertTrue(por_id[str(vigente.dry_run_id)]["abrir_url"])
        self.assertEqual(por_id[str(aplicado.dry_run_id)]["estado_ui"], "aplicado")
        self.assertEqual(por_id[str(aplicado.dry_run_id)]["abrir_url"], "")

    def _request_dry_run(self, dry_run_id=None, check_ids=None, refresh=None):
        params = {}
        if dry_run_id:
            params["dry_run_id"] = str(dry_run_id)
        if check_ids:
            params["check_ids"] = check_ids
        if refresh:
            params["refresh"] = refresh
        request = RequestFactory().get("/contabilidad/auditoria/dry-run/", params)
        request.session = Client().session
        request.session["user"] = _session_user(self.BASE)
        request.session.save()
        request.user = _UserConPermisoLeer()
        messages = FallbackStorage(request)
        request._messages = messages
        return request

    @patch("contabilidad_audit.views.render")
    def test_reopen_dry_run_id_vigente(self, mock_render):
        mock_render.return_value = HttpResponse("ok")
        plan = self._crear_plan(expira_delta=timedelta(minutes=20))
        request = self._request_dry_run(dry_run_id=plan.dry_run_id)

        response = auditoria_dry_run(request)

        self.assertEqual(response.status_code, 200)
        ctx = mock_render.call_args[0][2]
        self.assertEqual(ctx["payload"]["dry_run_id"], str(plan.dry_run_id))
        self.assertEqual(ctx["checks_seleccionados"], ["asiento_balanceado"])
        self.assertFalse(ctx["auto_ejecutar"])

    def test_reopen_dry_run_id_vencido_redirige_lotes(self):
        plan = self._crear_plan(expira_delta=timedelta(minutes=-5))
        request = self._request_dry_run(dry_run_id=plan.dry_run_id)

        response = auditoria_dry_run(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("contabilidad_audit:auditoria_lotes"))
        self.assertFalse(PlanCorreccion.objects.filter(dry_run_id=plan.dry_run_id).exists())

    @patch("legacy_db.services.cont_recalculo_service.dry_run")
    @patch("contabilidad_audit.views.render")
    def test_refresh_plan_vigente_llama_dry_run_con_id(self, mock_render, mock_dry_run):
        mock_render.return_value = HttpResponse("ok")
        plan = self._crear_plan(expira_delta=timedelta(minutes=20))
        mock_dry_run.return_value = {
            "dry_run_id": str(plan.dry_run_id),
            "base_empresa": self.BASE,
            "alcance": {"id_ejercicio": 1, "check_ids": ["asiento_balanceado"]},
            "config_hash": "hash-nuevo",
            "data_fingerprint": "fp-nuevo",
            "estado": "propuesto",
            "guards": {},
            "plan": {"items": [], "impacto": {"total_aplicables": 0}},
            "impacto": {"total_aplicables": 0},
        }
        request = self._request_dry_run(dry_run_id=plan.dry_run_id, refresh="1")

        response = auditoria_dry_run(request)

        self.assertEqual(response.status_code, 200)
        mock_dry_run.assert_called_once()
        _, kwargs = mock_dry_run.call_args
        self.assertEqual(kwargs["dry_run_id"], plan.dry_run_id)
        ctx = mock_render.call_args[0][2]
        self.assertEqual(ctx["payload"]["config_hash"], "hash-nuevo")

    def test_refresh_plan_vencido_redirige_lotes(self):
        plan = self._crear_plan(expira_delta=timedelta(minutes=-5))
        request = self._request_dry_run(dry_run_id=plan.dry_run_id, refresh="1")

        response = auditoria_dry_run(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("contabilidad_audit:auditoria_lotes"))
        self.assertFalse(PlanCorreccion.objects.filter(dry_run_id=plan.dry_run_id).exists())

    def test_listar_planes_incluye_actualizar_url(self):
        vigente = self._crear_plan(expira_delta=timedelta(minutes=15))
        aplicado = self._crear_plan(estado="aplicado", expira_delta=timedelta(minutes=-5))

        planes = _listar_planes_diagnostico(self.BASE)

        por_id = {p["dry_run_id"]: p for p in planes}
        self.assertIn("refresh=1", por_id[str(vigente.dry_run_id)]["actualizar_url"])
        self.assertEqual(por_id[str(aplicado.dry_run_id)]["actualizar_url"], "")

    def test_template_lotes_incluye_actualizar(self):
        html = (_TEMPLATES / "auditoria_lotes.html").read_text(encoding="utf-8")
        self.assertIn("Actualizar", html)
        self.assertIn("actualizarDiagnostico", html)
        self.assertIn("actualizar_url", html)
        self.assertIn("Actualizando diagnóstico", html)

    def test_template_dry_run_actualizar_diagnostico(self):
        html = (_TEMPLATES / "auditoria_dry_run.html").read_text(encoding="utf-8")
        self.assertIn("Actualizar diagnóstico", html)
        self.assertIn("Actualizando diagnóstico", html)
        self.assertIn("p.set('dry_run_id', this.payload.dry_run_id)", html)
