"""Exportación CSV/Excel de una corrida de auditoría contable.

Reutiliza las convenciones visuales del canon de reportes
(`reports/services/export_service.py`: openpyxl, estilos de cabecera,
bloque de metadatos) pero opera sobre el payload de `ejecutar_corrida()`,
que no es un `ReportDefinition`. Incluye `config_hash`, fecha dd/MM/yyyy
y detalle por check. No escribe nunca en el MySQL legacy.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from django.http import HttpResponse

SEVERIDAD_ETIQUETA = {
    "critico": "Crítico",
    "alto": "Alto",
    "medio": "Medio",
}

# Columnas del detalle de diferencias (orden estable, español).
_COLUMNAS_DIFERENCIA = [
    ("check_id", "Check"),
    ("titulo", "Título"),
    ("severidad", "Severidad"),
    ("referencia_hallazgo", "Referencia"),
    ("cod_pc", "Cód. cuenta"),
    ("id_pc", "ID cuenta"),
    ("id_ejercicio", "Ejercicio"),
    ("id_periodo", "Período"),
    ("nro_asiento", "N° asiento"),
    ("codigo_movimiento", "CodigoMovimiento"),
    ("valor_esperado", "Valor esperado"),
    ("valor_actual", "Valor actual"),
    ("delta", "Delta"),
]

_COLUMNAS_RESUMEN = [
    ("check_id", "Check"),
    ("titulo", "Título"),
    ("severidad", "Severidad"),
    ("estado", "Estado"),
    ("total_evaluado", "Evaluados"),
    ("total_diferencias", "Diferencias"),
    ("error", "Error"),
]


def _severidad_label(valor: str) -> str:
    return SEVERIDAD_ETIQUETA.get((valor or "").lower(), valor or "")


def _estado_label(check: dict) -> str:
    if check.get("error"):
        return "Error"
    return "OK" if check.get("ok") else "Con diferencias"


def _nombre_archivo(payload: dict, extension: str) -> str:
    base = (payload.get("base_empresa") or "empresa").replace("/", "_")
    filtros = payload.get("filtros") or {}
    ejercicio = filtros.get("id_ejercicio") or "ejercicio"
    return f"auditoria_contable_{base}_{ejercicio}.{extension}"


def _filas_resumen(payload: dict) -> list[dict[str, Any]]:
    filas: list[dict[str, Any]] = []
    for check in payload.get("checks", []):
        filas.append(
            {
                "check_id": check.get("check_id", ""),
                "titulo": check.get("titulo", ""),
                "severidad": _severidad_label(check.get("severidad", "")),
                "estado": _estado_label(check),
                "total_evaluado": check.get("total_evaluado", 0),
                "total_diferencias": check.get("total_diferencias", 0),
                "error": check.get("error") or "",
            }
        )
    return filas


def _filas_diferencias(payload: dict) -> list[dict[str, Any]]:
    filas: list[dict[str, Any]] = []
    for check in payload.get("checks", []):
        for dif in check.get("diferencias", []):
            fila = {
                "check_id": check.get("check_id", ""),
                "titulo": check.get("titulo", ""),
                "severidad": _severidad_label(check.get("severidad", "")),
            }
            for clave, _label in _COLUMNAS_DIFERENCIA:
                if clave in ("check_id", "titulo", "severidad"):
                    continue
                fila[clave] = dif.get(clave)
            filas.append(fila)
    return filas


def _metadatos(payload: dict) -> list[tuple[str, str]]:
    filtros = payload.get("filtros") or {}
    return [
        ("Auditoría de imputación contable", ""),
        ("Empresa", str(payload.get("base_empresa") or "—")),
        ("Ejercicio", str(filtros.get("id_ejercicio") or "—")),
        ("Período", str(filtros.get("id_periodo")) if filtros.get("id_periodo") else "Todos"),
        ("Fecha de corrida", str(payload.get("fecha_corrida") or "—")),
        ("config_hash", str(payload.get("config_hash") or "—")),
        ("Corrida", str(payload.get("corrida_id") or "—")),
    ]


def exportar_corrida_csv(payload: dict) -> HttpResponse:
    """CSV plano: bloque de metadatos + resumen por check + detalle de diferencias."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    for etiqueta, valor in _metadatos(payload):
        writer.writerow([etiqueta, valor])
    writer.writerow([])

    writer.writerow(["Resumen por check"])
    writer.writerow([label for _clave, label in _COLUMNAS_RESUMEN])
    for fila in _filas_resumen(payload):
        writer.writerow([fila.get(clave, "") for clave, _label in _COLUMNAS_RESUMEN])
    writer.writerow([])

    writer.writerow(["Detalle de diferencias"])
    writer.writerow([label for _clave, label in _COLUMNAS_DIFERENCIA])
    filas_dif = _filas_diferencias(payload)
    if filas_dif:
        for fila in filas_dif:
            writer.writerow(
                ["" if fila.get(clave) is None else fila.get(clave) for clave, _label in _COLUMNAS_DIFERENCIA]
            )
    else:
        writer.writerow(["Sin diferencias en la corrida."])

    contenido = buffer.getvalue()
    # BOM para que Excel abra correctamente los acentos en UTF-8.
    response = HttpResponse(
        "\ufeff" + contenido,
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{_nombre_archivo(payload, "csv")}"'
    return response


def exportar_corrida_xlsx(payload: dict) -> HttpResponse:
    """Excel con hoja Resumen (metadatos + por check) y hoja Diferencias."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E40AF")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = openpyxl.Workbook()

    # ── Hoja Resumen ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(_COLUMNAS_RESUMEN))
    cell = ws.cell(row=row, column=1, value="Auditoría de imputación contable")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    for etiqueta, valor in _metadatos(payload)[1:]:
        lc = ws.cell(row=row, column=1, value=etiqueta)
        lc.font = label_font
        vc = ws.cell(row=row, column=2, value=valor)
        vc.font = value_font
        vc.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Resumen por check").font = Font(bold=True, size=12)
    row += 1
    for col_num, (_clave, label) in enumerate(_COLUMNAS_RESUMEN, 1):
        c = ws.cell(row=row, column=col_num, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    row += 1
    for fila in _filas_resumen(payload):
        for col_num, (clave, _label) in enumerate(_COLUMNAS_RESUMEN, 1):
            c = ws.cell(row=row, column=col_num, value=fila.get(clave, ""))
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # ── Hoja Diferencias ──────────────────────────────────────────
    ws2 = wb.create_sheet("Diferencias")
    row = 1
    for col_num, (_clave, label) in enumerate(_COLUMNAS_DIFERENCIA, 1):
        c = ws2.cell(row=row, column=col_num, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    row += 1
    filas_dif = _filas_diferencias(payload)
    if filas_dif:
        for fila in filas_dif:
            for col_num, (clave, _label) in enumerate(_COLUMNAS_DIFERENCIA, 1):
                valor = fila.get(clave)
                c = ws2.cell(row=row, column=col_num, value="" if valor is None else valor)
                c.border = border
                c.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
    else:
        ws2.cell(row=row, column=1, value="Sin diferencias en la corrida.")

    for col_num in range(1, ws2.max_column + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for cell in ws2[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_nombre_archivo(payload, "xlsx")}"'
    return response


# ── Export dry-run Fase 2 ───────────────────────────────────────────────────

# Columnas del plan (diagnóstico): redacción en potencial / a realizar.
_COLUMNAS_PLAN_ITEM = [
    ("diagnostico", "Diagnóstico"),
    ("tipo_cambio", "Tipo de cambio a aplicar"),
    ("nro_asiento", "Nro asiento"),
    ("codigo_movimiento", "CodigoMovimiento"),
    ("fecha_asiento", "Fecha asiento"),
    ("cuenta", "Cuenta"),
    ("debe", "Debe"),
    ("haber", "Haber"),
    ("descripcion", "Descripción"),
    ("valor_anterior", "Valor anterior"),
    ("valor_nuevo", "Valor nuevo previsto"),
    ("cambio", "Cambios a realizar"),
    ("delta", "Delta"),
    ("excluido", "Excluido"),
]


def _nombre_archivo_dry_run(payload: dict, extension: str) -> str:
    base = (payload.get("base_empresa") or "empresa").replace("/", "_")
    alcance = payload.get("alcance") or {}
    ejercicio = alcance.get("id_ejercicio") or "ejercicio"
    dry_id = (payload.get("dry_run_id") or "plan")[:8]
    return f"dry_run_contable_{base}_{ejercicio}_{dry_id}.{extension}"


def _metadatos_dry_run(payload: dict) -> list[tuple[str, str]]:
    alcance = payload.get("alcance") or {}
    impacto = payload.get("impacto") or {}
    return [
        ("Diagnóstico de corrección contable", ""),
        ("Empresa", str(payload.get("base_empresa") or "—")),
        ("Ejercicio", str(alcance.get("id_ejercicio") or "—")),
        ("Fecha del diagnóstico", str(payload.get("creado_en") or "—")),
        ("Válido hasta", str(payload.get("expira_en") or "—")),
        ("Ítems del plan", str(impacto.get("total_items", 0))),
        ("Ítems aplicables", str(impacto.get("total_aplicables", 0))),
        ("Ítems excluidos", str(impacto.get("total_excluidos", 0))),
    ]


def _titulo_check_export(check_id: str, catalogo: dict[str, str] | None = None) -> str:
    if not check_id:
        return "—"
    if catalogo and check_id in catalogo:
        return catalogo[check_id]
    return check_id.replace("_", " ").capitalize()


def _es_numerico_export(valor) -> bool:
    if valor is None or valor == "":
        return False
    if isinstance(valor, (int, float)):
        return True
    try:
        float(str(valor).strip().replace(",", "."))
        return True
    except (TypeError, ValueError):
        return False


def _monto_export(valor) -> str:
    if valor is None or valor == "":
        return ""
    try:
        n = float(str(valor).strip().replace(",", "."))
    except (TypeError, ValueError):
        return str(valor)
    entero, dec = f"{n:,.2f}".split(".")
    entero = entero.replace(",", ".")
    return f"$ {entero},{dec}"


def _fecha_asiento_export(valor) -> str:
    if valor is None or valor == "":
        return ""
    if hasattr(valor, "strftime"):
        try:
            return valor.strftime("%d/%m/%Y")
        except (AttributeError, ValueError):
            pass
    txt = str(valor).strip()
    if len(txt) >= 10 and txt[4:5] == "-" and txt[7:8] == "-":
        try:
            from datetime import datetime as dt_cls

            return dt_cls.strptime(txt[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            pass
    return txt


def _filas_plan_items(payload: dict) -> list[dict[str, Any]]:
    """Filas del plan en columnas contables (sin JSON)."""
    plan = payload.get("plan") or {}
    # Catálogo opcional si el payload trae checks; si no, título derivado del id.
    catalogo = {}
    for c in payload.get("checks_disponibles") or []:
        if isinstance(c, dict) and c.get("check_id"):
            catalogo[c["check_id"]] = c.get("titulo") or c["check_id"]

    filas: list[dict[str, Any]] = []
    for item in plan.get("items") or []:
        clave = item.get("clave") or {}
        vn = item.get("valor_nuevo")
        va = item.get("valor_anterior")
        tabla = str(item.get("tabla") or "")
        accion = str(item.get("accion") or "")
        tipo = {
            "insert": "Insertar",
            "update": "Actualizar",
            "insert_marcador": "Insertar marcador",
            "marcar_original_anulado": "Marcar anulado",
            "insert_contra_asiento": "Contra-asiento",
            "bloqueado": "Bloqueado",
        }.get(accion, accion or "—")
        if tabla == "cont_asiento" and accion == "insert":
            tipo = "Asiento a insertar"
        elif "saldo" in tabla.lower():
            tipo = "Actualización de saldo" if accion == "update" else tipo

        fila = {
            "diagnostico": _titulo_check_export(str(item.get("check_id") or ""), catalogo),
            "tipo_cambio": tipo,
            "nro_asiento": "",
            "fecha_asiento": "",
            "codigo_movimiento": "",
            "cuenta": "",
            "debe": "",
            "haber": "",
            "descripcion": "",
            "valor_anterior": "",
            "valor_nuevo": "",
            "cambio": "",
            "delta": _monto_export(item.get("delta")) if _es_numerico_export(item.get("delta")) else (item.get("delta") or ""),
            "excluido": "Sí" if item.get("excluido") else "No",
        }

        if isinstance(vn, dict) and vn.get("nro_asiento") is not None:
            fila["nro_asiento"] = vn.get("nro_asiento", "")
            fila["fecha_asiento"] = _fecha_asiento_export(vn.get("fecha_asiento"))
            fila["codigo_movimiento"] = vn.get("codigo_movimiento") or clave.get("codigo_movimiento") or ""
            fila["cuenta"] = vn.get("id_pc") if vn.get("id_pc") is not None else clave.get("id_pc", "")
            fila["debe"] = _monto_export(vn.get("debe_asiento"))
            fila["haber"] = _monto_export(vn.get("haber_asiento"))
            fila["descripcion"] = vn.get("desc_asiento") or ""
        elif _es_numerico_export(va) or _es_numerico_export(vn):
            fila["cuenta"] = clave.get("id_pc", "")
            if _es_numerico_export(va):
                fila["valor_anterior"] = _monto_export(va)
            elif va not in (None, ""):
                fila["valor_anterior"] = str(va)
            if _es_numerico_export(vn):
                fila["valor_nuevo"] = _monto_export(vn)
            elif vn not in (None, "") and not isinstance(vn, (dict, list)):
                fila["valor_nuevo"] = str(vn)
        elif isinstance(vn, dict):
            fila["descripcion"] = (
                f"{vn.get('TipoComprobante') or ''} {vn.get('NroComprobante') or ''}".strip()
                or "Ver detalle en diagnóstico"
            )
            fila["codigo_movimiento"] = vn.get("codigo_movimiento") or clave.get("codigo_movimiento") or ""
        elif vn not in (None, ""):
            fila["valor_nuevo"] = str(vn)
            if va not in (None, ""):
                fila["valor_anterior"] = str(va)

        fila["cambio"] = _resumen_cambio_fila_export(fila)
        filas.append(fila)
    return filas


def _es_valor_mergeable(valor) -> bool:
    """True si el valor de celda puede participar en merge vertical."""
    if valor is None:
        return False
    texto = str(valor)
    return texto != "" and texto != "—"


def _rangos_merge_vertical(
    filas: list[dict],
    clave: str,
    clave_grupo: str = "nro_asiento",
) -> list[tuple[int, int]]:
    """Índices inclusivos (start, end) de rachas con el mismo valor no vacío.

    El merge solo ocurre **dentro del mismo nro de asiento** (rachas consecutivas
    con el mismo ``clave_grupo`` no vacío). Filas sin asiento no se combinan
    entre sí. Así no se fusionan Diagnóstico/Excluido/etc. de asientos distintos.
    """
    rangos: list[tuple[int, int]] = []
    n = len(filas)
    i = 0
    while i < n:
        grupo = filas[i].get(clave_grupo)
        if not _es_valor_mergeable(grupo):
            i += 1
            continue
        j = i + 1
        while j < n and str(filas[j].get(clave_grupo, "") or "") == str(grupo):
            j += 1

        # Dentro del bloque del mismo asiento [i, j).
        if clave == clave_grupo:
            if j - i >= 2:
                rangos.append((i, j - 1))
            i = j
            continue

        k = i
        while k < j:
            valor = filas[k].get(clave, "")
            if not _es_valor_mergeable(valor):
                k += 1
                continue
            m = k + 1
            while m < j:
                otro = filas[m].get(clave, "")
                if not _es_valor_mergeable(otro) or str(otro) != str(valor):
                    break
                m += 1
            if m - k >= 2:
                rangos.append((k, m - 1))
            k = m
        i = j
    return rangos


def _aplicar_merge_vertical_hoja(ws, filas, columnas, fila_inicio_datos=2, alignment=None):
    """Tras escribir filas, merge vertical por columna acotado al mismo nro asiento."""
    from openpyxl.styles import Alignment

    if alignment is None:
        alignment = Alignment(horizontal="left", vertical="center")

    for col_num, (clave, _label) in enumerate(columnas, 1):
        for start, end in _rangos_merge_vertical(filas, clave):
            row_start = fila_inicio_datos + start
            row_end = fila_inicio_datos + end
            ws.merge_cells(
                start_row=row_start,
                start_column=col_num,
                end_row=row_end,
                end_column=col_num,
            )
            ws.cell(row=row_start, column=col_num).alignment = alignment


def _resumen_cambio_fila_export(fila: dict) -> str:
    """Texto corto contable para la columna de cambios (diagnóstico o lote)."""
    if fila.get("nro_asiento") not in (None, ""):
        partes = [f"Asiento {fila['nro_asiento']}"]
        if fila.get("fecha_asiento"):
            partes.append(str(fila["fecha_asiento"]))
        if fila.get("cuenta") not in (None, ""):
            partes.append(f"Cta {fila['cuenta']}")
        if fila.get("debe"):
            partes.append(f"Debe {fila['debe']}")
        elif fila.get("haber"):
            partes.append(f"Haber {fila['haber']}")
        return " · ".join(partes)
    if fila.get("valor_anterior") or fila.get("valor_nuevo"):
        ant = fila.get("valor_anterior") or "—"
        nue = fila.get("valor_nuevo") or "—"
        return f"{ant} → {nue}"
    if fila.get("descripcion"):
        return str(fila["descripcion"])
    return str(fila.get("tipo_cambio") or "")


def exportar_dry_run_csv(payload: dict) -> HttpResponse:
    """CSV del plan dry-run con metadatos, impacto y detalle de items."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    for etiqueta, valor in _metadatos_dry_run(payload):
        writer.writerow([etiqueta, valor])
    writer.writerow([])

    impacto = payload.get("impacto") or {}
    writer.writerow(["Impacto por tipo de comprobante"])
    por_tipo = impacto.get("asientos_regenerar_por_tipo") or {}
    if por_tipo:
        writer.writerow(["Tipo", "Asientos a regenerar"])
        for tipo, cant in sorted(por_tipo.items()):
            writer.writerow([tipo, cant])
    else:
        writer.writerow(["Sin regeneración de asientos en el alcance."])
    writer.writerow([])

    writer.writerow(["Cuentas impactadas (delta saldo)"])
    writer.writerow(["ID cuenta", "Delta total"])
    for cuenta in impacto.get("cuentas_impactadas") or []:
        delta = cuenta.get("delta_total")
        writer.writerow([
            cuenta.get("id_pc"),
            _monto_export(delta) if _es_numerico_export(delta) else (delta or ""),
        ])
    writer.writerow([])

    writer.writerow(["Detalle del plan"])
    writer.writerow([label for _clave, label in _COLUMNAS_PLAN_ITEM])
    filas = _filas_plan_items(payload)
    if filas:
        for fila in filas:
            writer.writerow([fila.get(clave, "") for clave, _label in _COLUMNAS_PLAN_ITEM])
    else:
        writer.writerow(["Plan vacío: no hay cambios propuestos en el alcance."])

    response = HttpResponse(
        "\ufeff" + buffer.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{_nombre_archivo_dry_run(payload, "csv")}"'
    return response


def exportar_dry_run_xlsx(payload: dict) -> HttpResponse:
    """Excel del dry-run: Resumen + Plan."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E40AF")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Diagnóstico de corrección contable")
    cell.font = title_font
    row += 1

    for etiqueta, valor in _metadatos_dry_run(payload)[1:]:
        ws.cell(row=row, column=1, value=etiqueta).font = label_font
        ws.cell(row=row, column=2, value=valor).font = value_font
        row += 1
    row += 1

    impacto = payload.get("impacto") or {}
    ws.cell(row=row, column=1, value="Asientos a regenerar por tipo").font = Font(bold=True, size=12)
    row += 1
    for tipo, cant in sorted((impacto.get("asientos_regenerar_por_tipo") or {}).items()):
        ws.cell(row=row, column=1, value=tipo)
        ws.cell(row=row, column=2, value=cant)
        row += 1
    row += 1

    ws2 = wb.create_sheet("Plan")
    row = 1
    for col_num, (_clave, label) in enumerate(_COLUMNAS_PLAN_ITEM, 1):
        c = ws2.cell(row=row, column=col_num, value=label)
        c.fill = header_fill
        c.font = header_font
        c.border = border
    row += 1
    filas_plan = _filas_plan_items(payload)
    alignment_merge = Alignment(horizontal="left", vertical="center")
    for fila in filas_plan:
        for col_num, (clave, _label) in enumerate(_COLUMNAS_PLAN_ITEM, 1):
            c = ws2.cell(row=row, column=col_num, value=fila.get(clave, ""))
            c.border = border
            c.alignment = alignment_merge
        row += 1
    if filas_plan:
        _aplicar_merge_vertical_hoja(
            ws2,
            filas_plan,
            _COLUMNAS_PLAN_ITEM,
            fila_inicio_datos=2,
            alignment=alignment_merge,
        )

    for hoja in (ws, ws2):
        for col_num in range(1, hoja.max_column + 1):
            col_letter = get_column_letter(col_num)
            max_len = 0
            for cell_obj in hoja[col_letter]:
                if cell_obj.value is not None:
                    max_len = max(max_len, len(str(cell_obj.value)))
            hoja.column_dimensions[col_letter].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_nombre_archivo_dry_run(payload, "xlsx")}"'
    return response


# ── Export detalle lote Fase 3 (formato contador) ───────────────────────────

# Columnas del lote aplicado: redacción en pasado.
_COLUMNAS_LOTE_CONTADOR = [
    ("diagnostico", "Diagnóstico"),
    ("tipo_cambio", "Tipo de cambio aplicado"),
    ("nro_asiento", "Nro asiento"),
    ("codigo_movimiento", "CodigoMovimiento"),
    ("fecha_asiento", "Fecha asiento"),
    ("cuenta", "Cuenta"),
    ("debe", "Debe"),
    ("haber", "Haber"),
    ("descripcion", "Descripción"),
    ("concepto", "Concepto"),
    ("valor_anterior", "Valor anterior"),
    ("valor_nuevo", "Valor aplicado"),
    ("cambio", "Cambios aplicados"),
    ("fecha_aplicacion", "Fecha de aplicación"),
]


def _nombre_archivo_lote(lote: dict, extension: str) -> str:
    base = (lote.get("base_empresa") or "empresa").replace("/", "_")
    lote_id = (lote.get("lote_id") or "lote").replace("/", "_")
    corto = lote_id[:16] if len(lote_id) > 16 else lote_id
    return f"lote_correccion_{base}_{corto}.{extension}"


def _metadatos_lote(lote: dict) -> list[tuple[str, str]]:
    """Metadatos legibles para contador (sin hashes técnicos)."""
    estado = str(lote.get("estado") or "—")
    estado_ui = {
        "aplicado": "Aplicado",
        "revertido": "Revertido",
    }.get(estado, estado.capitalize() if estado != "—" else "—")
    return [
        ("Correcciones aplicadas — lote contable", ""),
        ("Empresa", str(lote.get("base_empresa") or "—")),
        ("Lote", str(lote.get("lote_id") or "—")),
        ("Fecha de aplicación", str(lote.get("fecha") or "—")),
        ("Usuario", str(lote.get("usuario") or "—")),
        ("Estado", estado_ui),
        ("Cantidad de cambios", str(lote.get("filas_correccion") or 0)),
        ("Reapertura de ejercicio", "Sí" if lote.get("reapertura_flag") else "No"),
        ("Autorizador", str(lote.get("autorizador") or "—")),
    ]


def _tipo_cambio_contador(tabla: str, valor_nuevo, valor_anterior) -> str:
    t = (tabla or "").lower()
    if t == "cont_asiento" and isinstance(valor_nuevo, dict):
        return "Asiento insertado"
    if t == "cont_asiento" and _renglones_asiento_eliminado(valor_anterior):
        return "Asiento eliminado"
    if "saldo" in t:
        if valor_anterior in (None, "") and valor_nuevo not in (None, ""):
            return "Saldo creado"
        return "Saldo actualizado"
    if t == "cuentaproveedor":
        return "Comprobante / anulación"
    if isinstance(valor_nuevo, dict) and (
        valor_nuevo.get("TipoComprobante") or valor_nuevo.get("renglones_preview")
    ):
        return "Anulación / contra-asiento"
    return "Corrección"


def _renglones_asiento_eliminado(valor_anterior) -> list[dict] | None:
    """Detecta el payload de eliminación (lista de renglones en valor_anterior)."""
    va = valor_anterior
    if isinstance(va, str):
        txt = va.strip()
        if not (txt.startswith("[") or txt.startswith("{")):
            return None
        try:
            va = json.loads(txt)
        except (TypeError, ValueError, json.JSONDecodeError):
            return None
    if not isinstance(va, list) or not va:
        return None
    if not isinstance(va[0], dict):
        return None
    sample = va[0]
    if any(k in sample for k in ("debe_asiento", "haber_asiento", "id_pc", "codigo_movimiento")):
        return [r for r in va if isinstance(r, dict)]
    return None


def _nro_asiento_fila_lote(fila: dict) -> Any:
    nro = fila.get("nro_asiento")
    if nro not in (None, "", "—"):
        return nro
    clave = fila.get("clave") if isinstance(fila.get("clave"), dict) else {}
    if clave.get("nro_asiento") is not None:
        return clave.get("nro_asiento")
    return ""


def _filas_export_eliminacion(fila: dict, renglones: list[dict]) -> list[dict[str, Any]]:
    """Una fila Excel por renglón eliminado (legible para contador)."""
    nro = _nro_asiento_fila_lote(fila)
    diagnostico = fila.get("titulo_check") or fila.get("check_id") or "Eliminación de asiento"
    if str(diagnostico) in ("eliminacion_asiento", ""):
        diagnostico = "Eliminación de asiento"
    fecha_aplicacion = fila.get("fecha") or ""
    out: list[dict[str, Any]] = []
    for r in renglones:
        anulado = str(r.get("anulado") or "No")
        desc = str(r.get("desc_renglon_asiento") or "").strip()
        if anulado == "Si" and desc:
            desc = f"{desc} (anulado)"
        elif anulado == "Si":
            desc = "Renglón anulado"
        out.append(
            {
                "diagnostico": diagnostico,
                "tipo_cambio": "Asiento eliminado",
                "nro_asiento": nro,
                "fecha_asiento": "",
                "codigo_movimiento": r.get("codigo_movimiento") or "",
                "cuenta": r.get("id_pc") if r.get("id_pc") is not None else "",
                "debe": _monto_export(r.get("debe_asiento")),
                "haber": _monto_export(r.get("haber_asiento")),
                "descripcion": desc,
                "concepto": "",
                "valor_anterior": "",
                "valor_nuevo": "",
                "cambio": "Renglón eliminado",
                "fecha_aplicacion": fecha_aplicacion,
            }
        )
    return out


def _fila_lote_contador(fila: dict) -> dict[str, Any]:
    """Aplana una fila del log a columnas interpretables por un contador."""
    va = fila.get("valor_anterior")
    vn = fila.get("valor_nuevo")
    tabla = str(fila.get("tabla") or "")
    cambio = fila.get("cambio_resumen") or ""
    # No volcar JSON técnico en «Cambios aplicados».
    if isinstance(cambio, str) and cambio.strip()[:1] in ("{", "["):
        cambio = ""

    out: dict[str, Any] = {
        "diagnostico": fila.get("titulo_check") or fila.get("check_id") or "—",
        "tipo_cambio": _tipo_cambio_contador(tabla, vn, va),
        "nro_asiento": "",
        "fecha_asiento": "",
        "codigo_movimiento": "",
        "cuenta": "",
        "debe": "",
        "haber": "",
        "descripcion": "",
        "concepto": "",
        "valor_anterior": "",
        "valor_nuevo": "",
        "cambio": cambio,
        "fecha_aplicacion": fila.get("fecha") or "",
    }

    clave = fila.get("clave") if isinstance(fila.get("clave"), dict) else {}

    if isinstance(vn, dict) and vn.get("nro_asiento") is not None:
        out["nro_asiento"] = vn.get("nro_asiento", "")
        out["fecha_asiento"] = _fecha_asiento_export(vn.get("fecha_asiento"))
        out["codigo_movimiento"] = vn.get("codigo_movimiento") or clave.get("codigo_movimiento") or ""
        out["cuenta"] = vn.get("id_pc") if vn.get("id_pc") is not None else clave.get("id_pc", "")
        out["debe"] = _monto_export(vn.get("debe_asiento"))
        out["haber"] = _monto_export(vn.get("haber_asiento"))
        out["descripcion"] = vn.get("desc_asiento") or ""
        out["concepto"] = vn.get("desc_concepto_asiento") or ""
        return out

    if isinstance(vn, dict):
        if vn.get("TipoComprobante") or vn.get("NroComprobante"):
            out["descripcion"] = (
                f"{vn.get('TipoComprobante') or ''} {vn.get('NroComprobante') or ''}".strip()
            )
            out["codigo_movimiento"] = vn.get("CodigoMovimiento") or vn.get("codigo_movimiento") or ""
            return out
        if isinstance(vn.get("renglones_preview"), list):
            out["descripcion"] = f"Contra-asiento · {len(vn['renglones_preview'])} renglones"
            out["codigo_movimiento"] = vn.get("codigo_movimiento") or ""
            return out

    if _es_numerico_export(va) or _es_numerico_export(vn):
        out["cuenta"] = clave.get("id_pc", "") if isinstance(clave, dict) else ""
        if _es_numerico_export(va):
            out["valor_anterior"] = _monto_export(va)
        elif va not in (None, "") and not isinstance(va, (dict, list)):
            out["valor_anterior"] = str(va)
        if _es_numerico_export(vn):
            out["valor_nuevo"] = _monto_export(vn)
        elif vn not in (None, "") and not isinstance(vn, (dict, list)):
            out["valor_nuevo"] = str(vn)
        if not cambio and out["valor_anterior"] and out["valor_nuevo"]:
            out["cambio"] = f"{out['valor_anterior']} → {out['valor_nuevo']}"
        return out

    if va not in (None, "") and not isinstance(va, (dict, list)):
        out["valor_anterior"] = str(va)
    if vn not in (None, "") and not isinstance(vn, (dict, list)):
        out["valor_nuevo"] = str(vn)
    if not out["cambio"] and out["tipo_cambio"] == "Asiento eliminado":
        out["cambio"] = "Asiento eliminado"
        out["nro_asiento"] = _nro_asiento_fila_lote(fila)
    return out


def _filas_lote_export(filas: list[dict]) -> list[dict[str, Any]]:
    """Expande eliminaciones a una fila por renglón; el resto se aplana 1:1."""
    resultado: list[dict[str, Any]] = []
    for fila in filas:
        renglones = _renglones_asiento_eliminado(fila.get("valor_anterior"))
        check_id = str(fila.get("check_id") or "")
        tabla = str(fila.get("tabla") or "").lower()
        es_eliminacion = check_id == "eliminacion_asiento" or (
            tabla == "cont_asiento"
            and renglones is not None
            and fila.get("valor_nuevo") in (None, "")
        )
        if es_eliminacion and renglones:
            resultado.extend(_filas_export_eliminacion(fila, renglones))
        else:
            resultado.append(_fila_lote_contador(fila))
    return resultado

def exportar_lote_xlsx(lote: dict, filas: list[dict]) -> HttpResponse:
    """Excel del lote en formato contador: Resumen + Detalle (sin JSON técnico)."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E40AF")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    filas_export = _filas_lote_export(filas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Correcciones aplicadas — lote contable")
    cell.font = title_font
    row += 1

    for etiqueta, valor in _metadatos_lote(lote)[1:]:
        ws.cell(row=row, column=1, value=etiqueta).font = label_font
        ws.cell(row=row, column=2, value=valor).font = value_font
        row += 1

    ws2 = wb.create_sheet("Detalle")
    row = 1
    for col_num, (_clave, label) in enumerate(_COLUMNAS_LOTE_CONTADOR, 1):
        c = ws2.cell(row=row, column=col_num, value=label)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    alignment_merge = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if filas_export:
        for fila in filas_export:
            for col_num, (clave, _label) in enumerate(_COLUMNAS_LOTE_CONTADOR, 1):
                valor = fila.get(clave, "")
                c = ws2.cell(row=row, column=col_num, value="" if valor is None else valor)
                c.border = border
                c.alignment = alignment_merge
            row += 1
        _aplicar_merge_vertical_hoja(
            ws2,
            filas_export,
            _COLUMNAS_LOTE_CONTADOR,
            fila_inicio_datos=2,
            alignment=alignment_merge,
        )
    else:
        ws2.cell(row=row, column=1, value="Sin cambios registrados en este lote.")

    anchos = {
        "A": 36, "B": 18, "C": 12, "D": 14, "E": 16, "F": 10,
        "G": 14, "H": 14, "I": 32, "J": 16, "K": 16, "L": 16, "M": 42, "N": 18,
    }
    for col_letter, width in anchos.items():
        ws2.column_dimensions[col_letter].width = width
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_nombre_archivo_lote(lote, "xlsx")}"'
    return response
