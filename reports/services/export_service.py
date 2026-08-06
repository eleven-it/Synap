from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set
import logging

from datetime import date, datetime

from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse

from .export_filter_labels import build_export_filter_lines
from .query_runner import QueryRunnerService
from ..models import ReportDefinition

logger = logging.getLogger(__name__)


def _vo_objetivos_vs_bo_sort_export_rows(
    filas: List[Dict[str, Any]],
    ordenar_por: str = "objetivo_meta",
    orden_forma: str = "desc",
) -> List[Dict[str, Any]]:
    """
    Misma lógica que la jerarquía web: vendedor por suma de objetivo descendente;
    dentro de cada vendedor, clientes por nombre alfabético.
    """
    if not filas:
        return []
    by_cv: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for r in filas:
        try:
            cv = int(r.get("cod_viajante") or 0)
        except (TypeError, ValueError):
            cv = 0
        if cv <= 0:
            continue
        by_cv[cv].append(r)

    metric_key = {
        "objetivo_meta": "objetivo",
        "objetivo_falta": "falta",
        "total_ventas_periodo": "total",
    }.get(ordenar_por, "objetivo")
    direction = 1 if str(orden_forma).lower() == "asc" else -1

    def _vendor_metric_total(cv: int) -> float:
        return sum(float(x.get(metric_key) or 0) for x in by_cv[cv])

    sorted_cvs = sorted(
        by_cv.keys(),
        key=lambda cv: (
            direction * _vendor_metric_total(cv),
            (by_cv[cv][0].get("nombre_vendedor") or "").strip().upper(),
            cv,
        ),
    )
    out: List[Dict[str, Any]] = []
    for cv in sorted_cvs:
        rows_v = sorted(
            by_cv[cv],
            key=lambda row: (
                direction * float(row.get(metric_key) or 0),
                0 if (row.get("estado_compra") or "") == "con_compra" else 1,
                (row.get("nombre_cliente") or "").strip().upper(),
                int(row.get("codigo_cliente") or 0),
            ),
        )
        out.extend(rows_v)
    return out


@dataclass
class ExportResult:
    """Resultado estandarizado de exportaciones."""

    path: str
    created_at: str
    expires_at: str | None = None
    filename: str | None = None


class ExportService:
    """Servicio para generar exportaciones PDF/XLSX."""

    def __init__(self, user):
        self.user = user

    def export(self, report_slug: str, payload: Dict, export_type: str) -> ExportResult:
        """
        Exporta un reporte a Excel.
        
        Args:
            report_slug: Slug del reporte a exportar
            payload: Payload con filtros y configuración
            export_type: Tipo de exportación ('xlsx' o 'xls')
        
        Returns:
            ExportResult con la ruta del archivo generado
        """
        if export_type not in ['xlsx', 'xls']:
            raise ValueError(f"Tipo de exportación no soportado: {export_type}")
        
        # Obtener la definición del reporte
        try:
            report = ReportDefinition.objects.get(slug=report_slug, is_active=True)
        except ReportDefinition.DoesNotExist:
            raise ValueError(f"Reporte no encontrado: {report_slug}")
        
        # Asegurar base_empresa dentro de filters para reportes que lo usan (ventas_netas, etc.)
        if payload.get("base_empresa"):
            payload.setdefault("filters", {})["base_empresa"] = payload["base_empresa"]
        
        # Ejecutar la consulta para obtener los datos
        query_service = QueryRunnerService(self.user)
        run_payload = dict(payload)
        if report_slug == "ventas-marcas-mensual":
            run_payload["_export_detalle"] = True
        query_result = query_service.run(report, run_payload)
        
        # Generar el archivo Excel
        export_dir = Path(settings.MEDIA_ROOT) / "reports" / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = self._resolve_export_filename(report_slug, payload, timestamp)
        file_path = export_dir / filename
        
        # Generar Excel usando openpyxl
        self._generate_excel(file_path, report, query_result, payload)

        return ExportResult(
            path=str(file_path.relative_to(settings.MEDIA_ROOT)),
            created_at=timezone.now().isoformat(),
            expires_at=None,
            filename=filename,
        )

    def _append_excel_filter_block(
        self,
        ws,
        start_row: int,
        filter_lines: List[tuple],
        label_font,
        value_font,
        max_merge_col: int = 8,
    ) -> int:
        """Escribe bloque «Filtros aplicados» y devuelve la siguiente fila libre."""
        if not filter_lines:
            return start_row

        from openpyxl.styles import Alignment

        title_cell = ws.cell(row=start_row, column=1, value="Filtros aplicados")
        title_cell.font = label_font
        row = start_row + 1
        for etiqueta, valor in filter_lines:
            lc = ws.cell(row=row, column=1, value=etiqueta)
            lc.font = label_font
            lc.alignment = Alignment(horizontal="left", vertical="top")
            if max_merge_col > 2:
                ws.merge_cells(
                    start_row=row,
                    start_column=2,
                    end_row=row,
                    end_column=max_merge_col,
                )
            vc = ws.cell(row=row, column=2, value=valor)
            vc.font = value_font
            vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row += 1
        return row + 1

    def _resolve_export_filename(self, report_slug: str, payload: Dict[str, Any], timestamp: str) -> str:
        """
        Nombre del .xlsx en disco y en Content-Disposition.
        Objetivos vs BO: Ventas_objetivo_vendedores_{fecha_inicio_fact}_{fecha_fin_fact}.xlsx
        """
        if report_slug == "documento-presupuesto-ventas":
            filters = payload.get("filters") or {}
            nro = (filters.get("nro_comprobante_archivo") or "PRE").strip() or "PRE"
            return f"Presupuesto_{nro}_{timestamp}.xlsx"

        if report_slug not in ("ventas-objetivos-vs-bo", "ventas-por-vendedor", "ventas-por-articulo", "ventas-marcas-mensual"):
            return f"{report_slug}_{timestamp}.xlsx"

        filters = payload.get("filters") or {}
        fi = filters.get("fecha_inicio_facturacion")
        ff = filters.get("fecha_fin_facturacion")

        def _segmento_fecha(v: Any) -> str:
            if v is None or v == "":
                return "sin_fecha"
            if isinstance(v, datetime):
                return v.date().isoformat()[:10]
            if isinstance(v, date):
                return v.isoformat()[:10]
            s = str(v).strip().replace("/", "-")[:10]
            s = re.sub(r"[^\d\-]", "_", s)
            return s if s else "sin_fecha"

        a = _segmento_fecha(fi)
        b = _segmento_fecha(ff)
        if report_slug == "ventas-por-vendedor":
            return f"Ventas_por_vendedor_{a}_{b}.xlsx"
        if report_slug == "ventas-por-articulo":
            return f"Ventas_por_articulo_{a}_{b}.xlsx"
        if report_slug == "ventas-marcas-mensual":
            return f"Ventas_marcas_mensual_{a}_{b}.xlsx"
        return f"Ventas_objetivo_vendedores_{a}_{b}.xlsx"

    def _declarative_export_headers(self, report: ReportDefinition, sample_row: Dict[str, Any]) -> Optional[List[str]]:
        """
        Orden de columnas alineado con WidgetEngine.renderTable: table_dimensions, table_metrics,
        y el mismo criterio cuando faltan esas claves (dimensiones del schema; sin métricas si kind=table
        y no hay table_metrics en options).
        """
        from .schema_service import ReportSchemaService

        try:
            schema = ReportSchemaService().build_schema(report)
        except Exception as exc:
            logger.warning("Export Excel: build_schema falló para %s: %s", report.slug, exc)
            return None

        table_widgets = [w for w in schema.default_widgets if w.kind == "table"]
        if not table_widgets:
            return None

        def _widget_sort_key(w):
            opts = w.options or {}
            o = opts.get("order")
            if o is None:
                o = 10**6
            return (o, str(w.id))

        table_widgets.sort(key=_widget_sort_key)
        w = table_widgets[0]
        opts = w.options or {}
        available = set(sample_row.keys())
        insertion_order = list(sample_row.keys())
        metric_names_schema: Set[str] = {m.name for m in schema.metrics}

        headers: List[str] = []

        if "table_dimensions" in opts and isinstance(opts["table_dimensions"], list):
            for name in opts["table_dimensions"]:
                if name in available:
                    headers.append(name)
        else:
            for d in schema.dimensions:
                if d.name in available:
                    headers.append(d.name)

        opts_tm = opts.get("table_metrics")
        explicit_tm = opts_tm is not None and isinstance(opts_tm, list)
        if explicit_tm:
            for name in opts_tm:
                if name in available:
                    headers.append(name)

        seen = set(headers)
        skip_metrics_tail: Set[str] = set()
        if w.kind == "table" and (not explicit_tm or (isinstance(opts_tm, list) and len(opts_tm) == 0)):
            skip_metrics_tail = set(metric_names_schema)

        for k in insertion_order:
            if k in seen:
                continue
            if k in skip_metrics_tail:
                continue
            headers.append(k)
            seen.add(k)

        return headers

    def _resolve_export_headers(self, report: ReportDefinition, sample_row: Dict[str, Any]) -> List[str]:
        """
        Define orden y columnas exportables (misma lógica que la tabla en pantalla cuando aplica).
        """
        if not sample_row:
            return []

        slug = report.slug
        cfg = report.config or {}
        insertion_order = list(sample_row.keys())
        available = set(sample_row.keys())

        if slug in ("ventas_netas", "ventas-netas"):
            preferred = [
                "mes_formato",
                "nombre_sucursal",
                "nro_punto_venta",
                "ventas_netas",
                "notas_credito",
                "ventas_brutas",
            ]
            return [h for h in preferred if h in available]

        if slug == "ventas-por-vendedor":
            preferred = [
                "cod_viajante",
                "nombre_vendedor",
                "codigo_cliente",
                "nombre_cliente",
                "cantidades_vendidas",
                "facturacion",
            ]
            return [h for h in preferred if h in available]

        if slug == "ventas-por-articulo":
            preferred = [
                "id_art",
                "nombre_articulo",
                "codigo_proveedor",
                "nombre_proveedor",
                "codigo_cliente",
                "nombre_cliente",
                "cantidades_vendidas",
                "facturacion",
            ]
            return [h for h in preferred if h in available]

        if slug == "ventas-marcas-mensual":
            # Sin Cód. vendedor / Cód. cliente en Excel (siguen en data[] para orden interno).
            if "unidades_a" in available:
                preferred = [
                    "nombre_vendedor",
                    "nombre_cliente",
                    "anio_mes",
                    "unidades_a",
                    "facturacion_a",
                    "unidades_b",
                    "facturacion_b",
                ]
                if "unidades_proy_a" in available:
                    preferred.extend(
                        ["unidades_proy_a", "facturacion_proy_a", "unidades_proy_b", "facturacion_proy_b"]
                    )
                return [h for h in preferred if h in available]
            preferred = [
                "nombre_vendedor",
                "nombre_cliente",
                "anio_mes",
                "unidades",
                "facturacion",
            ]
            if "unidades_proy" in available:
                preferred.extend(["unidades_proy", "facturacion_proy"])
            return [h for h in preferred if h in available]

        if slug == "ventas-objetivos-vs-bo":
            preferred = [
                "cod_viajante",
                "nombre_vendedor",
                "codigo_cliente",
                "nombre_cliente",
                "objetivo",
                "falta",
                "cantidades_vendidas",
                "facturacion",
                "remitos",
                "pedidos_en_armado",
                "total",
                "bo_con_stock",
                "bo_con_ingreso",
                "bo_sin_stock",
                "backorder_total",
            ]
            return [h for h in preferred if h in available]

        if slug == "uninvoiced_remitos":
            preferred = [
                "fecha",
                "nro_comprobante",
                "sucursal",
                "punto_venta",
                "subtotal_desc",
            ]
            headers = [h for h in preferred if h in available]
            seen = set(headers)
            for k in insertion_order:
                if k in seen or k in ("id_sucursal", "id_punto_venta"):
                    continue
                headers.append(k)
                seen.add(k)
            return headers

        if slug == "pedidos-pendientes":
            preferred = [
                "fecha",
                "nro_comprobante",
                "cliente",
                "nombre_cliente",
                "subtotal_desc",
            ]
            headers = [h for h in preferred if h in available]
            seen = set(headers)
            for k in insertion_order:
                if k in seen or k in ("tipo_comprobante", "estado"):
                    continue
                headers.append(k)
                seen.add(k)
            return headers

        if slug == "documento-presupuesto-ventas":
            preferred = [
                "orden",
                "codigo_articulo",
                "descripcion",
                "cantidad",
                "precio_unitario",
                "precio_neto_renglon",
                "precio_venta_renglon",
                "cod_deposito",
                "detalle_renglon",
            ]
            return [h for h in preferred if h in available]

        if cfg.get("version") == "declarative-v1":
            decl = self._declarative_export_headers(report, sample_row)
            if decl:
                return decl

        headers = list(insertion_order)
        if slug == "uninvoiced_remitos":
            headers = [h for h in headers if h not in ("id_sucursal", "id_punto_venta")]
        elif slug == "pedidos-pendientes":
            headers = [h for h in headers if h not in ("tipo_comprobante", "estado")]
        return headers

    def _generate_excel(self, file_path: Path, report: ReportDefinition, query_result, payload: Dict):
        """
        Genera un archivo Excel con los datos del reporte.
        
        Args:
            file_path: Ruta donde guardar el archivo
            report: Definición del reporte
            query_result: Resultado de la consulta (QueryResult)
            payload: Payload con filtros
        """
        if report.slug == "documento-presupuesto-ventas":
            self._generate_excel_documento_presupuesto_ventas(file_path, report, query_result, payload)
            return
        if report.slug == "ventas-marcas-mensual":
            self._generate_excel_ventas_marcas_mensual(file_path, report, query_result, payload)
            return
        if report.slug == "bo-stock-facturacion":
            self._generate_excel_bo(file_path, report, query_result, payload)
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report.name[:31]  # Excel limita a 31 caracteres
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1E40AF")
        filter_label_font = Font(bold=True, size=10)
        filter_value_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        base_empresa = None
        if isinstance(payload, dict):
            base_empresa = payload.get("base_empresa") or (payload.get("filters") or {}).get(
                "base_empresa"
            )
        filter_lines = build_export_filter_lines(
            report.slug,
            payload if isinstance(payload, dict) else {},
            (query_result.meta or {}).get("filters_applied"),
            base_empresa,
        )
        
        row = 1
        
        # 1. Título del reporte
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = report.name
        cell.font = title_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

        # 2. Filtros aplicados (nombres legibles)
        row = self._append_excel_filter_block(
            ws, row, filter_lines, filter_label_font, filter_value_font
        )
        
        # 3. Información del período (desde notes)
        if query_result.notes:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = query_result.notes[0] if query_result.notes else ""
            cell.font = Font(size=10, italic=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        
        row += 1  # Espacio
        
        # 4. Headers y datos según el tipo de reporte
        if report.slug == "sales_summary":
            # Resumen de Ventas: mostrar como tabla de totales
            headers = ["Concepto", "Valor"]
            ws.append(headers)
            
            # Aplicar estilo a headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            row += 1
            
            # Datos del resumen
            if query_result.data and len(query_result.data) > 0:
                data_row = query_result.data[0]
                summary_data = [
                    ["Ventas Netas", data_row.get("ventas_netas", 0)],
                    ["Remitos no Facturados", data_row.get("remitos_no_facturados", 0)],
                    ["Pedidos Pendientes", data_row.get("pedidos_pendientes", 0)],
                    ["Total Consolidado", data_row.get("total_consolidado", 0)],
                ]
                
                for concept, value in summary_data:
                    ws.append([concept, value])
                    # Formatear valor como moneda
                    value_cell = ws.cell(row=row, column=2)
                    value_cell.number_format = '"$"#,##0.00'
                    value_cell.alignment = Alignment(horizontal='right', vertical='center')
                    value_cell.border = border
                    # Aplicar borde a la celda de concepto
                    concept_cell = ws.cell(row=row, column=1)
                    concept_cell.border = border
                    concept_cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1
        else:
            # Otros reportes: mostrar datos en tabla
            if not query_result.data:
                ws.append(["Sin datos disponibles"])
                row += 1
            else:
                currency_headers_data = {
                    "ventas_brutas",
                    "notas_credito",
                    "ventas_netas",
                    "subtotal_desc",
                }
                if report.slug in ("ventas-objetivos-vs-bo", "ventas-por-vendedor", "ventas-por-articulo", "ventas-marcas-mensual"):
                    currency_headers_data.update(
                        {
                            "objetivo",
                            "facturacion",
                            "remitos",
                            "pedidos_en_armado",
                            "total",
                            "falta",
                            "backorder_total",
                            "bo_con_stock",
                            "bo_con_ingreso",
                            "bo_sin_stock",
                        }
                    )

                # Orden y columnas alineados con la tabla del dashboard / schema declarativo
                headers = self._resolve_export_headers(report, query_result.data[0])
                
                # Traducir headers al español si es necesario
                header_translations = {
                    "id_art": "ID sistema",
                    "codigo_articulo": "Código artículo",
                    "nombre": "Nombre",
                    "stock": "Stock",
                    "reservado": "Reservado",
                    "disponible": "Disponible",
                    "mes": "Mes",
                    "mes_formato": "Mes (Formato)",
                    "id_sucursal": "ID Sucursal",
                    "nombre_sucursal": "Sucursal",
                    "id_punto_venta": "ID Punto de Venta",
                    "nro_punto_venta": "Punto de Venta",
                    "ventas_brutas": "Ventas Brutas",
                    "notas_credito": "Notas Crédito",
                    "ventas_netas": "Ventas Netas",
                    "fecha": "Fecha",
                    "nro_comprobante": "N° Comprobante",
                    "tipo_comprobante": "Tipo Comprobante",
                    "subtotal_desc": "Subtotal",
                    "estado": "Estado",
                    "sucursal": "Sucursal",
                    "punto_venta": "Punto de Venta",
                    "cliente": "Cliente",
                    "nombre_cliente": "Cliente",
                    "cod_viajante": "Cód. vendedor",
                    "nombre_vendedor": "Vendedor",
                    "codigo_cliente": "Cód. cliente",
                    "id_art": "ID artículo",
                    "nombre_articulo": "Artículo",
                    "codigo_proveedor": "Cód. proveedor",
                    "nombre_proveedor": "Proveedor",
                    "objetivo": "Objetivo",
                    "facturacion": "Facturación",
                    "remitos": "Remitos",
                    "pedidos_en_armado": "Pedidos en armado",
                    "total": "Total consolidado",
                    "falta": "Falta",
                    "cantidades_vendidas": "Unidades",
                    "unidades": "Unidades",
                    "unidades_proy": "Unidades proy",
                    "facturacion_proy": "Facturación proy",
                    "anio_mes": "AñoMes",
                    "backorder_total": "BO total",
                    "bo_con_stock": "BO c/stock",
                    "bo_con_ingreso": "BO c/ingreso",
                    "bo_sin_stock": "BO s/stock",
                }
                # Ventas Netas: etiqueta "Mes" para la primera columna (mes_formato)
                if report.slug in ("ventas_netas", "ventas-netas"):
                    header_translations = {**header_translations, "mes_formato": "Mes"}
                
                translated_headers = [header_translations.get(h, h.replace("_", " ").title()) for h in headers]
                if report.slug in ("ventas-objetivos-vs-bo", "ventas-por-vendedor", "ventas-por-articulo", "ventas-marcas-mensual"):
                    translated_headers = [str(s).upper() for s in translated_headers]
                if report.slug == "ventas-marcas-mensual":
                    modo = str((payload.get("filters") or {}).get("modo_unidades") or "packs").lower()
                    unidad_label = "DOCENAS" if modo == "docenas" else "UNIDADES"
                    translated_headers = [
                        unidad_label if str(h).upper() == "UNIDADES" else h for h in translated_headers
                    ]
                    translated_headers = [
                        f"{unidad_label} PROY" if str(h).upper() == "UNIDADES PROY" else h
                        for h in translated_headers
                    ]
                
                # Escribir headers
                ws.append(translated_headers)
                
                # Aplicar estilo a headers
                for col_num, header in enumerate(translated_headers, 1):
                    cell = ws.cell(row=row, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                row += 1

                def _vo_build_row_values(data_row: Dict[str, Any]) -> List[Any]:
                    row_values: List[Any] = []
                    for h in headers:
                        value = data_row.get(h, "")
                        if h in currency_headers_data:
                            try:
                                if value == "" or value is None:
                                    row_values.append("")
                                else:
                                    row_values.append(float(value))
                            except (ValueError, TypeError):
                                row_values.append("")
                        elif h == "cantidades_vendidas":
                            try:
                                if value == "" or value is None:
                                    row_values.append(0.0)
                                else:
                                    row_values.append(float(value))
                            except (ValueError, TypeError):
                                row_values.append(0.0)
                        else:
                            row_values.append(value)
                    return row_values

                def _vo_format_data_row(ridx: int, row_values: List[Any]) -> None:
                    for col_num, (h, val) in enumerate(zip(headers, row_values), 1):
                        cell = ws.cell(row=ridx, column=col_num)
                        cell.border = border
                        if h in currency_headers_data:
                            if isinstance(val, (int, float)):
                                cell.number_format = '"$"#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif h == "cantidades_vendidas" and isinstance(val, (int, float)):
                            cell.number_format = "#,##0.00"
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                name_col_idx = headers.index("nombre_cliente") + 1 if "nombre_cliente" in headers else None

                def _vo_append_data_row(data_row: Dict[str, Any], outline_lvl: int = 0, name_indent: int = 0) -> None:
                    vals = _vo_build_row_values(data_row)
                    ws.append(vals)
                    ridx = ws.max_row
                    if outline_lvl > 0:
                        ws.row_dimensions[ridx].outline_level = outline_lvl
                    _vo_format_data_row(ridx, vals)
                    if name_col_idx:
                        ncell = ws.cell(row=ridx, column=name_col_idx)
                        ncell.alignment = Alignment(horizontal="left", vertical="center", indent=max(0, int(name_indent)))

                # Ventas por artículo: export plano (filas ya vienen artículo/proveedor/cliente).
                if report.slug == "ventas-por-articulo":
                    for data_row in query_result.data:
                        _vo_append_data_row(data_row, outline_lvl=0)
                # Escribir datos (objetivos vs BO: orden por cód. vendedor + cód. cliente, agrupación Excel)
                elif report.slug in ("ventas-objetivos-vs-bo", "ventas-por-vendedor"):
                    from openpyxl.worksheet.properties import Outline

                    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)
                    vendor_fill = PatternFill(start_color="DCE6F2", end_color="DCE6F2", fill_type="solid")
                    vendor_title_font = Font(bold=True, size=11, color="1E3A5F")
                    filters_for_order = payload.get("filters", {}) if isinstance(payload, dict) else {}
                    excel_scope = str(filters_for_order.get("excel_scope") or "resumen").strip().lower()
                    use_detailed_tree = excel_scope == "detallado"
                    tabs = (((query_result.meta or {}).get("extra") or {}).get("tabs") or {})
                    jerarquia = tabs.get("objetivos_jerarquia")
                    if not isinstance(jerarquia, list):
                        jerarquia = []

                    def _vo_row_from_node(
                        node: Dict[str, Any],
                        cod_viajante: int,
                        nombre_vendedor: str,
                        nombre_cliente_label: str,
                        include_codigo_cliente: Any = "",
                        use_detalle_rem_ped: bool = False,
                    ) -> Dict[str, Any]:
                        row_dict = {
                            "cod_viajante": cod_viajante,
                            "nombre_vendedor": nombre_vendedor,
                            "codigo_cliente": include_codigo_cliente,
                            "nombre_cliente": nombre_cliente_label,
                            "objetivo": node.get("objetivo"),
                            "falta": node.get("falta"),
                            "cantidades_vendidas": node.get("cantidades_vendidas"),
                            "facturacion": node.get("facturacion"),
                            "remitos": node.get("remitos"),
                            "pedidos_en_armado": node.get("pedidos_en_armado"),
                            "total": node.get("total"),
                            "bo_con_stock": node.get("bo_con_stock"),
                            "bo_con_ingreso": node.get("bo_con_ingreso"),
                            "bo_sin_stock": node.get("bo_sin_stock"),
                            "backorder_total": node.get("backorder_total"),
                        }
                        if use_detalle_rem_ped:
                            row_dict["remitos"] = node.get("remitos_lineas")
                            row_dict["pedidos_en_armado"] = node.get("pedidos_armado_lineas")
                            row_dict["objetivo"] = ""
                            row_dict["falta"] = ""
                            row_dict["total"] = ""
                        return row_dict

                    def _append_detalle_rows(
                        detalle_nodes: List[Dict[str, Any]],
                        cod_viajante: int,
                        nombre_vendedor: str,
                        outline_lvl: int,
                        name_indent: int,
                    ) -> None:
                        for rub in detalle_nodes or []:
                            if not isinstance(rub, dict):
                                continue
                            rub_name = str(rub.get("nombre_rubro") or "").strip() or "Rubro"
                            rub_label = f"RUBRO  {rub_name}"
                            _vo_append_data_row(
                                _vo_row_from_node(
                                    rub,
                                    cod_viajante,
                                    nombre_vendedor,
                                    rub_label,
                                    include_codigo_cliente="",
                                    use_detalle_rem_ped=True,
                                ),
                                outline_lvl=outline_lvl,
                                name_indent=name_indent,
                            )

                            for sub in rub.get("children") or []:
                                if not isinstance(sub, dict):
                                    continue
                                sub_name = str(sub.get("nombre_subrubro") or "").strip() or "Subrubro"
                                sub_label = f"SUBRUBRO  {sub_name}"
                                _vo_append_data_row(
                                    _vo_row_from_node(
                                        sub,
                                        cod_viajante,
                                        nombre_vendedor,
                                        sub_label,
                                        include_codigo_cliente="",
                                        use_detalle_rem_ped=True,
                                    ),
                                    outline_lvl=outline_lvl + 1,
                                    name_indent=name_indent + 1,
                                )

                                for art in sub.get("children") or []:
                                    if not isinstance(art, dict):
                                        continue
                                    art_name = str(art.get("nombre_articulo") or "").strip() or "Artículo"
                                    _vo_append_data_row(
                                        _vo_row_from_node(
                                            art,
                                            cod_viajante,
                                            nombre_vendedor,
                                            art_name,
                                            include_codigo_cliente="",
                                            use_detalle_rem_ped=True,
                                        ),
                                        outline_lvl=outline_lvl + 2,
                                        name_indent=name_indent + 2,
                                    )

                    if use_detailed_tree and jerarquia:
                        for vend in jerarquia:
                            if not isinstance(vend, dict):
                                continue
                            cv = int(vend.get("cod_viajante") or 0)
                            nv = (vend.get("nombre_vendedor") or "").strip() or f"Cód. {cv}"
                            label = f"Vendedor {cv} — {nv}"
                            ws.append([label] + [""] * (len(headers) - 1))
                            vr = ws.max_row
                            if len(headers) > 1:
                                ws.merge_cells(
                                    start_row=vr,
                                    start_column=1,
                                    end_row=vr,
                                    end_column=len(headers),
                                )
                            top = ws.cell(row=vr, column=1)
                            top.value = label
                            top.font = vendor_title_font
                            top.fill = vendor_fill
                            top.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                            for cx in range(1, len(headers) + 1):
                                ws.cell(row=vr, column=cx).border = border
                            ws.row_dimensions[vr].outline_level = 0

                            for estado in vend.get("children") or []:
                                if not isinstance(estado, dict):
                                    continue
                                estado_nombre = str(estado.get("nombre") or "").strip() or "Estado"
                                _vo_append_data_row(
                                    _vo_row_from_node(
                                        estado,
                                        cv,
                                        nv,
                                        estado_nombre,
                                        include_codigo_cliente="",
                                        use_detalle_rem_ped=False,
                                    ),
                                    outline_lvl=1,
                                    name_indent=1,
                                )
                                for cli in estado.get("children") or []:
                                    if not isinstance(cli, dict):
                                        continue
                                    cli_nombre = str(cli.get("nombre_cliente") or "").strip() or "Cliente"
                                    cli_codigo = cli.get("codigo_cliente") or ""
                                    _vo_append_data_row(
                                        _vo_row_from_node(
                                            cli,
                                            cv,
                                            nv,
                                            cli_nombre,
                                            include_codigo_cliente=cli_codigo,
                                            use_detalle_rem_ped=False,
                                        ),
                                        outline_lvl=2,
                                        name_indent=2,
                                    )
                                    _append_detalle_rows(
                                        cli.get("venta_detalle") or [],
                                        cv,
                                        nv,
                                        outline_lvl=3,
                                        name_indent=3,
                                    )
                    else:
                        sorted_data = _vo_objetivos_vs_bo_sort_export_rows(
                            query_result.data,
                            ordenar_por=str(filters_for_order.get("ordenar_por") or "objetivo_meta"),
                            orden_forma=str(filters_for_order.get("orden_forma") or "desc"),
                        )
                        last_cv = None
                        for data_row in sorted_data:
                            cv = int(data_row.get("cod_viajante") or 0)
                            if last_cv is None or cv != last_cv:
                                last_cv = cv
                                nv = (data_row.get("nombre_vendedor") or "").strip() or f"Cód. {cv}"
                                label = f"Vendedor {cv} — {nv}"
                                ws.append([label] + [""] * (len(headers) - 1))
                                vr = ws.max_row
                                if len(headers) > 1:
                                    ws.merge_cells(
                                        start_row=vr,
                                        start_column=1,
                                        end_row=vr,
                                        end_column=len(headers),
                                    )
                                top = ws.cell(row=vr, column=1)
                                top.value = label
                                top.font = vendor_title_font
                                top.fill = vendor_fill
                                top.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                                for cx in range(1, len(headers) + 1):
                                    ws.cell(row=vr, column=cx).border = border
                                ws.row_dimensions[vr].outline_level = 0
                            _vo_append_data_row(data_row, outline_lvl=1, name_indent=1)
                else:
                    for data_row in query_result.data:
                        _vo_append_data_row(data_row, outline_lvl=0)

                row = ws.max_row + 1

                # Agregar fila de totales si existe
                if query_result.totals:
                    # Mapeo de headers a nombres de totales en query_result.totals
                    # Algunos reportes usan "total_<campo>" en lugar de solo "<campo>"
                    header_to_total_map = {
                        "subtotal_desc": "total_subtotal_desc",
                        "ventas_netas": "total_ventas_netas",
                        "ventas_brutas": "total_ventas_brutas",
                        "notas_credito": "total_notas_credito",
                    }
                    
                    # Construir la fila de totales con los valores correctos
                    total_row = ["TOTALES"]
                    for header in headers[1:]:  # Saltar el primer header (ya pusimos "TOTALES")
                        # Buscar el total usando el nombre del header o el mapeo
                        total_key = header_to_total_map.get(header, header)
                        if total_key in query_result.totals:
                            total_value = query_result.totals[total_key]
                            if isinstance(total_value, (int, float)):
                                total_row.append(total_value)
                            else:
                                total_row.append("")
                        # Si no se encuentra con el mapeo, intentar con el nombre original
                        elif header in query_result.totals:
                            total_value = query_result.totals[header]
                            if isinstance(total_value, (int, float)):
                                total_row.append(total_value)
                            else:
                                total_row.append("")
                        else:
                            total_row.append("")
                    
                    # Escribir la fila completa (row ya apunta a la siguiente fila libre)
                    ws.append(total_row)
                    
                    # Aplicar estilo y formato a la fila de totales (la que acabamos de escribir en row)
                    for col_num in range(1, len(total_row) + 1):
                        if col_num <= len(total_row):
                            cell = ws.cell(row=row, column=col_num)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.border = border
                            
                            if col_num == 1:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                header = headers[col_num - 1]
                                # Mismo formato moneda que las filas de datos
                                if header in currency_headers_data:
                                    value = total_row[col_num - 1]
                                    if isinstance(value, (int, float)):
                                        cell.number_format = '"$"#,##0.00'
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                    else:
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                elif header == "cantidades_vendidas":
                                    value = total_row[col_num - 1]
                                    if isinstance(value, (int, float)):
                                        cell.number_format = "#,##0.00"
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                    else:
                                        cell.alignment = Alignment(horizontal='right', vertical='center')
                                else:
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 4. Ajustar anchos de columna
        for col_num in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 5. Guardar archivo
        wb.save(file_path)
        logger.info(f"✅ Archivo Excel generado: {file_path}")

    def _generate_excel_ventas_marcas_mensual(
        self, file_path: Path, report: ReportDefinition, query_result, payload: Dict
    ):
        """Excel dual hoja Matriz + Detalle para ventas-marcas-mensual."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        from reports.services.ventas_marcas_mensual_export import resolve_detalle_headers

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1E40AF")
        filter_label_font = Font(bold=True, size=10)
        filter_value_font = Font(size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        payload = payload if isinstance(payload, dict) else {}
        base_empresa = payload.get("base_empresa") or (payload.get("filters") or {}).get("base_empresa")
        filter_lines = build_export_filter_lines(
            report.slug,
            payload,
            (query_result.meta or {}).get("filters_applied"),
            base_empresa,
        )

        extra = (query_result.meta or {}).get("extra") or {}
        detalle_data = extra.get("detalle_rows") or []
        matriz_data = query_result.data or []

        wb = openpyxl.Workbook()
        sheets = [
            ("Matriz", matriz_data),
            ("Detalle", detalle_data),
        ]

        currency_headers = {"facturacion", "facturacion_a", "facturacion_b", "facturacion_proy", "facturacion_proy_a", "facturacion_proy_b"}
        modo = str((payload.get("filters") or {}).get("modo_unidades") or "packs").lower()
        unidad_label = "DOCENAS" if modo == "docenas" else "PACKS"

        header_translations = {
            "nombre_vendedor": "Vendedor",
            "nombre_cliente": "Cliente",
            "anio_mes": "AñoMes",
            "unidades": unidad_label,
            "facturacion": "Monto",
            "unidades_a": f"{unidad_label} A",
            "facturacion_a": "Monto A",
            "unidades_b": f"{unidad_label} B",
            "facturacion_b": "Monto B",
            "unidades_proy": f"{unidad_label} proy",
            "facturacion_proy": "Monto proy",
            "unidades_proy_a": f"{unidad_label} proy A",
            "facturacion_proy_a": "Monto proy A",
            "unidades_proy_b": f"{unidad_label} proy B",
            "facturacion_proy_b": "Monto proy B",
            "fecha": "Fecha",
            "tipo_comprobante": "Tipo",
            "nro_comprobante": "N° comprobante",
            "id_manual": "SuperArt",
            "nombre_articulo": "Artículo",
            "nombre_marca": "Marca",
        }

        first = True
        for sheet_title, data in sheets:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = sheet_title[:31]

            row = 1
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws[f"A{row}"]
            cell.value = f"{report.name} — {sheet_title}"
            cell.font = title_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            row = self._append_excel_filter_block(
                ws, row, filter_lines, filter_label_font, filter_value_font
            )
            if query_result.notes:
                for note_line in query_result.notes[:2]:
                    if not note_line:
                        continue
                    ws.merge_cells(f"A{row}:D{row}")
                    c = ws[f"A{row}"]
                    c.value = note_line
                    c.font = Font(size=10, italic=True)
                    row += 1
            row += 1

            if not data:
                ws.cell(row=row, column=1, value="Sin datos")
                continue

            if sheet_title == "Detalle":
                headers = resolve_detalle_headers(data[0])
            else:
                headers = self._resolve_export_headers(report, data[0])

            translated = [str(header_translations.get(h, h.replace("_", " ").title())).upper() for h in headers]
            # Escribir en `row` (no ws.append): el bloque de filtros deja max_row
            # desfasado respecto al contador y append + estilo en `row` desalineaba headers.
            for col_num, label in enumerate(translated, 1):
                c = ws.cell(row=row, column=col_num, value=label)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
            row += 1

            for data_row in data:
                for col_num, h in enumerate(headers, 1):
                    raw = data_row.get(h, "")
                    if h in currency_headers:
                        try:
                            val = float(raw) if raw not in ("", None) else ""
                        except (TypeError, ValueError):
                            val = ""
                    elif h == "unidades" or h.startswith("unidades"):
                        try:
                            val = float(raw) if raw not in ("", None) else 0.0
                        except (TypeError, ValueError):
                            val = 0.0
                    else:
                        val = raw
                    c = ws.cell(row=row, column=col_num, value=val)
                    c.border = border
                    if h in currency_headers and isinstance(val, (int, float)):
                        c.number_format = '"$"#,##0.00'
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    elif (h == "unidades" or h.startswith("unidades")) and isinstance(val, (int, float)):
                        c.number_format = "#,##0.00"
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")
                row += 1

            for col_num in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_num)
                max_len = 0
                for c in ws[col_letter]:
                    if c.value is not None:
                        max_len = max(max_len, len(str(c.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(file_path)
        logger.info("Excel ventas-marcas-mensual (Matriz + Detalle): %s", file_path)

    def _generate_excel_bo(
        self, file_path: Path, report: ReportDefinition, query_result, payload: Optional[Dict] = None
    ):
        """
        Genera un archivo Excel multi-hoja para el reporte BO vs Stock vs Facturación.
        Cada tab del reporte se exporta como una hoja.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        extra = query_result.meta.get("extra") or {}
        tabs = extra.get("tabs") or {}

        # Orden y nombre de hojas (key en tabs, título en Excel, clave de datos si difiere)
        sheets_config = [
            ("resumen", "Resumen", None),
            ("detalle_sin_stock", "Detalle sin stock", None),
            ("detalle_con_stock", "Detalle con stock", None),
            ("detalle_con_ingreso", "Detalle con ingreso", None),
            ("facturacion", "Facturación", None),
            ("remitos", "Remitos", None),
            ("backorder_detalle_rows", "Backorder detalle", None),
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1E40AF")
        filter_label_font = Font(bold=True, size=10)
        filter_value_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        payload = payload if isinstance(payload, dict) else {}
        base_empresa = payload.get("base_empresa") or (payload.get("filters") or {}).get("base_empresa")
        filter_lines = build_export_filter_lines(
            report.slug,
            payload,
            (query_result.meta or {}).get("filters_applied"),
            base_empresa,
        )
        currency_headers = {
            "importe", "sub_total", "subtotal_desc", "bo_importe", "con_stock_importe",
            "con_ingreso_importe", "sin_stock_importe", "precio_x_renglon", "costo", "saldo_valorizado",
        }

        def sanitize_sheet_name(name: str) -> str:
            invalid = set('\\/*?:[]')
            out = "".join(c if c not in invalid else " " for c in name)
            return (out or "Hoja")[:31]

        wb = openpyxl.Workbook()
        first_sheet = True

        for sheet_key, sheet_title, data_key in sheets_config:
            data = tabs.get(data_key or sheet_key)
            if data is None:
                data = []
            if not isinstance(data, list):
                data = []

            # Excluir columnas no exportables (listas/dicts)
            if data and isinstance(data[0], dict):
                all_keys = list(data[0].keys())
                headers = [k for k in all_keys if not isinstance(data[0].get(k), (list, dict))]
                # Backorder detalle: no exportar cantidad, estado, id_cliente (alineado con la UI)
                if sheet_key == "backorder_detalle_rows":
                    headers = [h for h in headers if h not in ("cantidad", "estado", "id_cliente")]
            elif sheet_key == "resumen":
                headers = ["concepto", "importe", "tipo"]
                if not data:
                    data = []
            else:
                headers = []

            if first_sheet:
                ws = wb.active
                first_sheet = False
            else:
                ws = wb.create_sheet()
            ws.title = sanitize_sheet_name(sheet_title)

            row = 1
            # Título, filtros y período solo en la primera hoja
            if ws == wb.active:
                ws.merge_cells(f"A{row}:D{row}")
                cell = ws[f"A{row}"]
                cell.value = report.name
                cell.font = title_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
                row = self._append_excel_filter_block(
                    ws, row, filter_lines, filter_label_font, filter_value_font
                )
                if query_result.notes:
                    for note_line in query_result.notes[:2]:
                        ws.merge_cells(f"A{row}:D{row}")
                        cell = ws[f"A{row}"]
                        cell.value = note_line
                        cell.font = Font(size=10, italic=True)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        row += 1
                row += 1

            if not headers:
                ws.cell(row=row, column=1).value = "Sin datos"
                row += 1
            else:
                header_labels = {
                    "concepto": "Concepto",
                    "importe": "Importe",
                    "tipo": "Tipo",
                    "codigo": "Código",
                    "articulo": "Artículo",
                    "categoria": "Categoría",
                    "bo_qty": "Cant. Pendientes",
                    "bo_importe": "Pendiente valorizado",
                    "stock_actual": "Stock actual",
                    "stock_reservado": "Stock reservado",
                    "disponible": "Disponible",
                    "costo": "Costo",
                    "saldo_valorizado": "Saldo valorizado",
                    "oc_pendiente": "OC pendiente",
                    "con_stock_qty": "Con stock cant.",
                    "con_stock_importe": "Con stock importe",
                    "con_ingreso_qty": "Con ingreso cant.",
                    "con_ingreso_importe": "Con ingreso importe",
                    "sin_stock_qty": "Sin stock cant.",
                    "sin_stock_importe": "Sin stock importe",
                    "nro": "Nº",
                    "id_cliente": "ID Cliente",
                    "cliente": "Cliente",
                    "sub_total": "Subtotal",
                    "porc_ventas": "% Ventas",
                    "ultima_compra": "Última compra",
                    "vendedor": "Vendedor",
                    "zona": "Zona",
                    "telefono": "Teléfono",
                    "email": "Email",
                    "cuit": "CUIT",
                    "fecha": "Fecha",
                    "nro_comprobante": "N° Comprobante",
                    "id_sucursal": "ID Sucursal",
                    "sucursal": "Sucursal",
                    "id_punto_venta": "ID Punto de venta",
                    "punto_venta": "Punto de venta",
                    "subtotal_desc": "Subtotal",
                    "nro_comp": "N° Comp.",
                    "descripcion": "Descripción",
                    "cod_manual": "Cód. manual",
                    "cantidad": "Cantidad",
                    "cant_pend": "Cant. pend.",
                    "estado": "Estado",
                    "precio_x_renglon": "Pendiente valorizado",
                    "nombre_rubro": "Rubro",
                    "nombre_sub_rubro": "Subrubro",
                    "nombre_vendedor": "Vendedor",
                    "id_art": "ID Art.",
                }
                translated = [header_labels.get(h, h.replace("_", " ").title()) for h in headers]
                ws.append(translated)
                for col_num, _ in enumerate(translated, 1):
                    c = ws.cell(row=row, column=col_num)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = border
                row += 1

                for data_row in data:
                    row_vals = []
                    for h in headers:
                        val = data_row.get(h, "")
                        if isinstance(val, (list, dict)):
                            val = ""
                        row_vals.append(val)
                    ws.append(row_vals)
                    for col_num, (header, value) in enumerate(zip(headers, row_vals), 1):
                        cell = ws.cell(row=row, column=col_num)
                        cell.border = border
                        if header in currency_headers and isinstance(value, (int, float)):
                            cell.number_format = '"$"#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1

            for col_num in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_num)
                max_len = 0
                for cell in ws[col_letter]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(file_path)
        logger.info(f"✅ Archivo Excel BO generado con {len(sheets_config)} hojas: {file_path}")

    def _generate_excel_documento_presupuesto_ventas(
        self, file_path: Path, report: ReportDefinition, query_result, payload: Dict
    ):
        """Excel con bloque de cabecera PRE y tabla de renglones (`stockp`)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presupuesto"[:31]

        title_font = Font(bold=True, size=14, color="1E40AF")
        label_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        meta = query_result.meta or {}
        cab = meta.get("cabecera") or {}
        row = 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1)
        c.value = report.name
        c.font = title_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        if query_result.notes:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row=row, column=1)
            c.value = query_result.notes[0]
            c.font = Font(size=10, italic=True)
            row += 1

        base_empresa = None
        if isinstance(payload, dict):
            base_empresa = payload.get("base_empresa") or (payload.get("filters") or {}).get(
                "base_empresa"
            )
        filter_lines = build_export_filter_lines(
            report.slug,
            payload if isinstance(payload, dict) else {},
            (query_result.meta or {}).get("filters_applied"),
            base_empresa,
        )
        filter_label_font = Font(bold=True, size=10)
        filter_value_font = Font(size=10)
        row = self._append_excel_filter_block(
            ws, row, filter_lines, filter_label_font, filter_value_font, max_merge_col=6
        )

        def escribe_kv(etiqueta: str, valor):
            nonlocal row
            ws.cell(row=row, column=1, value=etiqueta).font = label_font
            vcell = ws.cell(row=row, column=2, value=valor)
            vcell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1

        if cab:
            escribe_kv("N° comprobante", cab.get("nro_comprobante") or "—")
            escribe_kv("Código movimiento", cab.get("codigo_movimiento") or meta.get("codigo_movimiento") or "—")
            escribe_kv("Tipo", cab.get("tipo_comprobante") or "PRE")
            escribe_kv("Fecha", cab.get("fecha_fmt") or "—")
            escribe_kv("Vencimiento", cab.get("vencimiento_fmt") or "—")
            escribe_kv("Cliente", cab.get("nombre_cliente") or "—")
            escribe_kv("Cód. cliente", cab.get("codigo_cliente") if cab.get("codigo_cliente") is not None else "—")
            escribe_kv("Cond. venta", cab.get("cond_venta") or "—")
            escribe_kv("Estado", cab.get("estado") or "—")
            escribe_kv("Anulado", cab.get("anulado") or "No")
            escribe_kv("Sucursal", cab.get("cod_sucursal") if cab.get("cod_sucursal") is not None else "—")
            escribe_kv("Vendedor (cód.)", cab.get("cod_viajante") if cab.get("cod_viajante") is not None else "—")
            iv = cab.get("importe_venta")
            escribe_kv("Importe cabecera", iv if iv is not None else "—")
            det = (cab.get("detalle") or "").strip()
            if det:
                ws.cell(row=row, column=1, value="Observaciones").font = label_font
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                ws.cell(row=row, column=2, value=det).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                row += 1
        else:
            ws.cell(row=row, column=1, value="No hay datos de cabecera para mostrar.")
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Renglones").font = Font(bold=True, size=12)
        row += 1

        if not query_result.data:
            ws.cell(row=row, column=1, value="Sin renglones en stockp para este movimiento.")
            row += 1
        else:
            headers = self._resolve_export_headers(report, query_result.data[0])
            labels = {
                "orden": "Orden",
                "codigo_articulo": "Código artículo",
                "descripcion": "Descripción",
                "cantidad": "Cantidad",
                "precio_unitario": "P. unitario",
                "precio_neto_renglon": "Neto renglón",
                "precio_venta_renglon": "Importe venta renglón",
                "cod_deposito": "Depósito",
                "detalle_renglon": "Detalle",
            }
            translated = [labels.get(h, h.replace("_", " ").title()) for h in headers]
            for col_num, txt in enumerate(translated, 1):
                cell = ws.cell(row=row, column=col_num, value=txt)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            row += 1

            currency_headers = {
                "precio_unitario",
                "precio_neto_renglon",
                "precio_venta_renglon",
            }
            qty_headers = {"cantidad"}

            for data_row in query_result.data:
                vals = [data_row.get(h) for h in headers]
                for col_num, (header, value) in enumerate(zip(headers, vals), 1):
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.border = border
                    if header in currency_headers and value is not None:
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header in qty_headers and value is not None:
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                row += 1

        totals = query_result.totals or {}
        row += 1
        ws.cell(row=row, column=1, value="Totales").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="Suma neto renglones")
        v = totals.get("suma_precio_neto_renglon")
        c = ws.cell(row=row, column=2, value=v if v is not None else 0)
        c.number_format = '"$"#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value="Suma importe venta renglones")
        v = totals.get("suma_precio_venta_renglon")
        c = ws.cell(row=row, column=2, value=v if v is not None else 0)
        c.number_format = '"$"#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value="Importe cabecera comp_ped")
        v = totals.get("importe_cabecera_comp_ped")
        c = ws.cell(row=row, column=2, value=v if v is not None else 0)
        c.number_format = '"$"#,##0.00'

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        for col_letter in ("C", "D", "E", "F"):
            ws.column_dimensions[col_letter].width = 16

        wb.save(file_path)
        logger.info("Excel documento presupuesto ventas generado: %s", file_path)

    def get_file_response(self, export_result: ExportResult) -> HttpResponse:
        """
        Retorna una HttpResponse con el archivo para descarga.
        
        Args:
            export_result: Resultado de la exportación
        
        Returns:
            HttpResponse con el archivo
        """
        file_path = Path(settings.MEDIA_ROOT) / export_result.path
        
        if not file_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{export_result.filename}"'
            return response
