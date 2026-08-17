# -*- coding: utf-8 -*-
"""Genera plantillas xlsx mínimas para los 6 packs Monthly Reporting."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from reports.services.monthly_reporting_pack_seed import (
    MONTHLY_REPORTING_PACK_DEFINITIONS,
    MONTHLY_REPORTING_TEMPLATE_FILES,
)

TEMPLATE_DIR = (
    Path(__file__).resolve().parents[1]
    / "templates"
    / "reports"
    / "excel"
    / "monthly_reporting"
)


def _build_template(pack_id: str, product_group: str, unit_mode: str, royalty_rate: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    sales = wb.active
    sales.title = "input Licensee sales"
    headers = ["Customer", "City / Province", "Store Type", "Product group"]
    for idx, label in enumerate(headers, start=1):
        sales.cell(row=4, column=idx, value=label).font = Font(bold=True)
    for month in range(1, 13):
        col = 3 + (month * 2)
        month_cell = sales.cell(row=4, column=col, value=datetime(2026, month, 1))
        month_cell.number_format = "mmm-yy"
        sales.cell(row=3, column=col, value="units")
        sales.cell(row=3, column=col + 1, value="amounts")
        for offset in (0, 1):
            letter = get_column_letter(col + offset)
            sales.cell(row=2, column=col + offset, value=f"=SUM({letter}5:{letter}4931)")
    sales.cell(row=4, column=29, value="YTD_Units")
    sales.cell(row=4, column=30, value="YTD_Sales")
    sales["AC2"] = "=SUM(AC5:AC4931)"
    sales["AD2"] = "=SUM(AD5:AD4931)"

    monthly = wb.create_sheet("monthly")
    monthly["B2"] = "Best Sox"
    monthly["B4"] = "dozens" if unit_mode == "dozens" else "PACKS"
    monthly["C6"] = float(royalty_rate)
    monthly["D4"] = "='input Licensee sales'!E2"

    if pack_id == "lw_propia":
        ooh = wb.create_sheet("input Licensee ooh")
        ooh["A4"] = "Licencee"
        minimum = wb.create_sheet("minimum agreed")
        minimum["D13"] = "FY 2026"
    elif pack_id.startswith("puma_"):
        minimum = wb.create_sheet("minimum agreed")
        minimum["A1"] = "Plantilla Puma — minimum agreed vacía equivalente"

    return wb


def build_all_templates(target_dir: Path | None = None) -> list[Path]:
    target = target_dir or TEMPLATE_DIR
    target.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for definition in MONTHLY_REPORTING_PACK_DEFINITIONS:
        filename = MONTHLY_REPORTING_TEMPLATE_FILES[definition["pack_id"]]
        path = target / filename
        workbook = _build_template(
            definition["pack_id"],
            definition["product_group"],
            definition["unit_mode"],
            str(definition["royalty_rate"]),
        )
        workbook.save(path)
        written.append(path)
    return written


if __name__ == "__main__":
    for item in build_all_templates():
        print(item)
