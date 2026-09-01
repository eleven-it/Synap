# -*- coding: utf-8 -*-
"""Export Excel anual Monthly Reporting licenciatarios (openpyxl + plantillas)."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.utils.administranet_types import str_or_default
from reports.models import MonthlyReportingPack
from reports.services.monthly_reporting_pack_seed import MONTHLY_REPORTING_TEMPLATE_FILES
from reports.services.monthly_reporting_template_builder import TEMPLATE_DIR
from reports.services.ventas_mensuales_licenciatarios_merger import (
    MergeResult,
    MergedClientMonth,
)

SHEET_SALES = "input Licensee sales"
SHEET_MONTHLY = "monthly"
SHEET_OOH = "input Licensee ooh"
SHEET_MINIMUM = "minimum agreed"
QA_SHEET = "QA"
SHEET_FILTROS = "Filtros"

QA_HEADERS = ("SuperArt / tipo", "Detalle", "Cliente", "Estado match")
SALES_HEADERS_LEVIS = ("Customer", "City / Province", "Store Type", "Product group")
YTD_UNITS_HEADER = "YTD_Units"
YTD_SALES_HEADER = "YTD_Sales"
SUM_LAST_ROW = 4931
UNITS_FORMAT = "#,##0"
AMOUNTS_FORMAT = '"$"#,##0.00'
MONTH_DATE_FORMAT = "mmm-yy"

# Paridad visual plantilla julio (accent1 #4F81BD + texto blanco Tahoma 10).
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4F81BD")
HEADER_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
TOTALS_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
TOTALS_FONT = Font(name="Tahoma", size=10, bold=True, color="000000")
DATA_FONT = Font(name="Calibri", size=11, color="000000")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="B4B4B4")
HEADER_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Anchos mínimos tomados de la planilla LEVIS BW julio (autofit solo amplía).
MIN_COL_WIDTHS = {
    1: 31.0,
    2: 12.0,
    3: 20.0,
    4: 12.0,
}
MIN_UNITS_WIDTH = 10.0
MIN_AMOUNTS_WIDTH = 14.0
MIN_YTD_UNITS_WIDTH = 12.5
MIN_YTD_SALES_WIDTH = 15.5
MAX_COL_WIDTH = 48.0


def resolve_template_path(pack_id: str) -> Path:
    filename = MONTHLY_REPORTING_TEMPLATE_FILES[pack_id]
    return TEMPLATE_DIR / filename


def ensure_template_path(pack_id: str) -> Path:
    """Devuelve la plantilla del pack; si falta en disco, regenera las 6 plantillas mínimas."""
    path = resolve_template_path(pack_id)
    if path.exists():
        return path
    from reports.services.monthly_reporting_template_builder import build_all_templates

    build_all_templates(TEMPLATE_DIR)
    if not path.exists():
        raise FileNotFoundError(f"Plantilla no encontrada tras regenerar: {path}")
    return path


def _month_columns(year: int) -> List[tuple[int, date]]:
    """Enero siempre en columna E (5), 12 pares units|amounts — mismo eje que la plantilla Levi’s."""
    cols: List[tuple[int, date]] = []
    col = 5
    for month in range(1, 13):
        cols.append((col, date(year, month, 1)))
        col += 2
    return cols


def _write_row2_sums(ws, columns: Iterable[int]) -> None:
    for col_idx in columns:
        letter = get_column_letter(col_idx)
        cell = ws.cell(row=2, column=col_idx)
        cell.value = f"=SUM({letter}5:{letter}{SUM_LAST_ROW})"
        cell.font = TOTALS_FONT
        cell.fill = TOTALS_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER
        # Pares: units (impar desde E=5) / amounts (par). YTD sales = 30.
        if col_idx % 2 == 0 or col_idx == 30:
            cell.number_format = AMOUNTS_FORMAT
        else:
            cell.number_format = UNITS_FORMAT


def _unmerge_all(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def _style_header_cell(cell, *, number_format: Optional[str] = None) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = HEADER_BORDER
    if number_format:
        cell.number_format = number_format


def _apply_sales_header_layout(ws, *, year: int, ytd_units_col: int, ytd_sales_col: int) -> None:
    """
    Layout julio: fila 3 = units/amounts; fila 4 = Customer… + meses (merge E4:F4…).
    Colores azul #4F81BD y texto blanco.
    """
    _unmerge_all(ws)
    month_cols = _month_columns(year)

    for idx, label in enumerate(SALES_HEADERS_LEVIS, start=1):
        top = ws.cell(row=3, column=idx, value=None)
        _style_header_cell(top)
        bottom = ws.cell(row=4, column=idx, value=label)
        _style_header_cell(bottom)

    for col_idx, month_date in month_cols:
        units_cell = ws.cell(row=3, column=col_idx, value="units")
        amounts_cell = ws.cell(row=3, column=col_idx + 1, value="amounts")
        _style_header_cell(units_cell)
        _style_header_cell(amounts_cell)
        month_cell = ws.cell(row=4, column=col_idx, value=month_date)
        _style_header_cell(month_cell, number_format=MONTH_DATE_FORMAT)
        pair = ws.cell(row=4, column=col_idx + 1, value=None)
        _style_header_cell(pair)
        ws.merge_cells(
            start_row=4,
            start_column=col_idx,
            end_row=4,
            end_column=col_idx + 1,
        )

    ytd_u = ws.cell(row=4, column=ytd_units_col, value=YTD_UNITS_HEADER)
    ytd_s = ws.cell(row=4, column=ytd_sales_col, value=YTD_SALES_HEADER)
    _style_header_cell(ytd_u)
    _style_header_cell(ytd_s)
    for col in (ytd_units_col, ytd_sales_col):
        top = ws.cell(row=3, column=col, value=None)
        _style_header_cell(top)

    ws.row_dimensions[4].height = 28


def _display_width_for_cell(cell) -> float:
    value = cell.value
    if value is None:
        return 0.0
    if isinstance(value, datetime):
        return 9.0
    if isinstance(value, date):
        return 9.0
    if isinstance(value, (int, float)):
        nf = cell.number_format or ""
        if "$" in nf or "0.00" in nf:
            return float(len(f"{value:,.2f}") + 3)
        return float(len(f"{value:,.0f}") + 2)
    text = str(value)
    if text.startswith("="):
        return 12.0
    return float(len(text))


def _autosize_sales_columns(
    ws,
    *,
    last_data_row: int,
    ytd_units_col: int,
    ytd_sales_col: int,
) -> None:
    """Expande anchos para que Customer/montos queden legibles sin ajuste manual."""
    month_cols = _month_columns(2026)  # solo índices de columna; año irrelevante
    last_col = max(ytd_sales_col, 30)

    for col_idx in range(1, last_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0.0
        for row_idx in range(2, max(last_data_row, 5) + 1):
            max_len = max(max_len, _display_width_for_cell(ws.cell(row=row_idx, column=col_idx)))

        if col_idx in MIN_COL_WIDTHS:
            minimum = MIN_COL_WIDTHS[col_idx]
        elif col_idx == ytd_units_col:
            minimum = MIN_YTD_UNITS_WIDTH
        elif col_idx == ytd_sales_col:
            minimum = MIN_YTD_SALES_WIDTH
        elif any(col_idx == c for c, _ in month_cols):
            minimum = MIN_UNITS_WIDTH
        elif any(col_idx == c + 1 for c, _ in month_cols):
            minimum = MIN_AMOUNTS_WIDTH
        else:
            minimum = 10.0

        width = min(max(max_len + 2.5, minimum), MAX_COL_WIDTH)
        ws.column_dimensions[letter].width = width


def _client_meta_from_rows(rows: Iterable[MergedClientMonth]) -> Dict[str, dict]:
    meta: Dict[str, dict] = {}
    for row in rows:
        bucket = meta.setdefault(
            row.identity,
            {
                "display_name": row.display_name,
                "city": "",
                "store_type": "",
                "product_group": "",
                "match_estado": row.match_estado,
                "pending": row.pending,
            },
        )
        bucket["display_name"] = row.display_name
        bucket["match_estado"] = row.match_estado
        bucket["pending"] = row.pending
        if row.city:
            bucket["city"] = row.city
        if row.store_type:
            bucket["store_type"] = row.store_type
        if row.product_group:
            bucket["product_group"] = row.product_group
    return meta


def _write_levis_sales_sheet(
    ws,
    *,
    rows: List[MergedClientMonth],
    year: int,
    month_from: int,
    month_to: int,
    product_group: str,
) -> None:
    month_cols = _month_columns(year)
    ytd_units_col = 29
    ytd_sales_col = 30

    _apply_sales_header_layout(
        ws,
        year=year,
        ytd_units_col=ytd_units_col,
        ytd_sales_col=ytd_sales_col,
    )

    by_client: dict[str, dict[date, MergedClientMonth]] = defaultdict(dict)
    client_meta = _client_meta_from_rows(rows)
    for row in rows:
        by_client[row.identity][row.month] = row

    excel_row = 5
    unit_letters = [get_column_letter(col_idx) for col_idx, _ in month_cols]
    amount_letters = [get_column_letter(col_idx + 1) for col_idx, _ in month_cols]
    ytd_units_formula_tpl = "=" + "+".join(f"{letter}{{row}}" for letter in reversed(unit_letters))
    ytd_sales_formula_tpl = "=" + "+".join(f"{letter}{{row}}" for letter in reversed(amount_letters))

    for identity, months_map in sorted(
        by_client.items(),
        key=lambda item: client_meta[item[0]]["display_name"].upper(),
    ):
        meta = client_meta[identity]
        for col_idx, value in (
            (1, meta["display_name"]),
            (2, meta.get("city") or ""),
            (3, meta.get("store_type") or ""),
            (4, meta.get("product_group") or product_group),
        ):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = LEFT

        for col_idx, month_date in month_cols:
            if month_date.month < month_from or month_date.month > month_to:
                continue
            cell = months_map.get(month_date)
            if cell is None:
                continue
            units_cell = ws.cell(row=excel_row, column=col_idx, value=float(cell.units))
            units_cell.number_format = UNITS_FORMAT
            units_cell.font = DATA_FONT
            units_cell.alignment = RIGHT
            amount_cell = ws.cell(row=excel_row, column=col_idx + 1, value=float(cell.amount))
            amount_cell.number_format = AMOUNTS_FORMAT
            amount_cell.font = DATA_FONT
            amount_cell.alignment = RIGHT

        ytd_u = ws.cell(
            row=excel_row,
            column=ytd_units_col,
            value=ytd_units_formula_tpl.format(row=excel_row),
        )
        ytd_u.font = DATA_FONT
        ytd_u.alignment = RIGHT
        ytd_s = ws.cell(
            row=excel_row,
            column=ytd_sales_col,
            value=ytd_sales_formula_tpl.format(row=excel_row),
        )
        ytd_s.font = DATA_FONT
        ytd_s.alignment = RIGHT
        ytd_s.number_format = AMOUNTS_FORMAT
        excel_row += 1

    last_data_row = excel_row - 1
    if ws.max_row >= excel_row:
        ws.delete_rows(excel_row, ws.max_row - excel_row + 1)

    sum_cols = [col for col_idx, _ in month_cols for col in (col_idx, col_idx + 1)]
    sum_cols.extend([ytd_units_col, ytd_sales_col])
    _write_row2_sums(ws, sum_cols)
    _autosize_sales_columns(
        ws,
        last_data_row=max(last_data_row, 5),
        ytd_units_col=ytd_units_col,
        ytd_sales_col=ytd_sales_col,
    )
    ws.freeze_panes = "E5"


def _ensure_monthly_links_row2(ws) -> None:
    """La hoja monthly lee totales de la fila 2 de input Licensee sales. No volcar YTD en español."""
    if not ws["B2"].value:
        ws["B2"] = "Best Sox"
    if not ws["D4"].value:
        ws["D4"] = "='input Licensee sales'!E2"


def _write_qa_sheet(
    wb: openpyxl.Workbook,
    *,
    merge_result: MergeResult,
) -> None:
    if QA_SHEET in wb.sheetnames:
        ws = wb[QA_SHEET]
        wb.remove(ws)
    ws = wb.create_sheet(QA_SHEET)
    for col, header in enumerate(QA_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _style_header_cell(cell)

    row = 2
    for superart in merge_result.qa_superarts:
        ws.cell(row=row, column=1, value=superart)
        ws.cell(row=row, column=2, value="SuperArt desconocido")
        row += 1

    for pending in merge_result.pending_clients:
        ws.cell(row=row, column=1, value="match_pendiente")
        ws.cell(row=row, column=2, value=pending.get("seed_key", ""))
        ws.cell(row=row, column=3, value=pending.get("display_name", ""))
        ws.cell(row=row, column=4, value="pending")
        row += 1

    for col_idx in range(1, 5):
        letter = get_column_letter(col_idx)
        max_len = 10.0
        for r in range(1, max(row, 2)):
            max_len = max(max_len, _display_width_for_cell(ws.cell(row=r, column=col_idx)))
        ws.column_dimensions[letter].width = min(max_len + 2.5, MAX_COL_WIDTH)


def _write_filtros_sheet(wb, filter_lines: List[Tuple[str, str]]) -> None:
    """Hoja inicial con sucursales/PV y resto de filtros aplicados."""
    if not filter_lines:
        return
    if SHEET_FILTROS in wb.sheetnames:
        wb.remove(wb[SHEET_FILTROS])
    ws = wb.create_sheet(SHEET_FILTROS, 0)
    title = ws.cell(row=1, column=1, value="Filtros aplicados")
    _style_header_cell(title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    row = 2
    for etiqueta, valor in filter_lines:
        ws.cell(row=row, column=1, value=etiqueta).font = DATA_FONT
        ws.cell(row=row, column=2, value=valor).font = DATA_FONT
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80


def export_licenciatarios_workbook(
    file_path: Path,
    *,
    pack: MonthlyReportingPack,
    merge_result: MergeResult,
    year: int,
    month_from: int,
    month_to: int,
    filter_lines: Optional[List[Tuple[str, str]]] = None,
) -> None:
    """
    Clona plantilla anual, reescribe ventas/mensual, conserva hojas auxiliares y agrega QA.
    """
    template_path = ensure_template_path(pack.pack_id)

    wb = openpyxl.load_workbook(template_path)
    preserved = {SHEET_OOH, SHEET_MINIMUM} & set(wb.sheetnames)

    if SHEET_SALES not in wb.sheetnames:
        raise ValueError(f"La plantilla no contiene hoja '{SHEET_SALES}'")
    sales_ws = wb[SHEET_SALES]
    _write_levis_sales_sheet(
        sales_ws,
        rows=merge_result.rows,
        year=year,
        month_from=month_from,
        month_to=month_to,
        product_group=str_or_default(pack.product_group, ""),
    )

    if SHEET_MONTHLY in wb.sheetnames:
        _ensure_monthly_links_row2(wb[SHEET_MONTHLY])

    _write_qa_sheet(wb, merge_result=merge_result)
    _write_filtros_sheet(wb, filter_lines or [])

    for sheet_name in preserved:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Se perdió la hoja auxiliar '{sheet_name}' durante el export")

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
