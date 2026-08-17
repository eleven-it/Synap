# -*- coding: utf-8 -*-
"""Almacén PostgreSQL historial Monthly Reporting licenciatarios."""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Optional

from django.db import transaction
from django.utils import timezone

from core.utils.administranet_types import str_or_default, to_decimal_or_none
from reports.models import (
    MonthlyReportingClientMatch,
    MonthlyReportingImportBatch,
    MonthlyReportingPack,
    MonthlyReportingSeedRow,
)
from reports.services.monthly_reporting_pack_seed import seed_monthly_reporting_packs

UNITS_QUANT = Decimal("0.000001")
AMOUNT_QUANT = Decimal("0.01")
SHEET_SALES = "input Licensee sales"


@dataclass(frozen=True)
class ImportActor:
    id_usuario: Optional[int] = None
    cod_usuario: str = ""
    nombre: str = ""


@dataclass
class ImportResult:
    batch: MonthlyReportingImportBatch
    rows_created: int = 0
    rows_updated: int = 0
    duplicate: bool = False


@dataclass
class ParsedSeedCell:
    seed_key: str
    customer_name: str
    customer_code: str = ""
    city: str = ""
    store_type: str = ""
    product_group: str = ""
    uf: str = ""
    month: date = field(default_factory=lambda: date(2026, 1, 1))
    units: Decimal = Decimal("0")
    amount: Decimal = Decimal("0")
    units_men: Decimal = Decimal("0")
    units_women: Decimal = Decimal("0")
    amount_men: Decimal = Decimal("0")
    amount_women: Decimal = Decimal("0")


def compute_file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_seed_key(customer_name: str, customer_code: str = "", **meta: str) -> str:
    code = str_or_default(customer_code, "").strip()
    if code:
        return f"code:{code}"
    normalized_name = re.sub(r"\s+", " ", str_or_default(customer_name, "").strip().lower())
    meta_blob = "|".join(
        str_or_default(meta.get(key), "").strip().lower()
        for key in ("city", "store_type", "product_group", "uf")
    )
    payload = f"{normalized_name}|{meta_blob}"
    return "name:" + hashlib.sha256(payload.encode("utf-8")).hexdigest()


def quantize_units(value: Optional[Decimal]) -> Decimal:
    dec = to_decimal_or_none(value) or Decimal("0")
    return dec.quantize(UNITS_QUANT)


def quantize_amount(value: Optional[Decimal]) -> Decimal:
    dec = to_decimal_or_none(value) or Decimal("0")
    return dec.quantize(AMOUNT_QUANT)


def _month_year_ok(year: int, default_year: int) -> bool:
    """Años de columna mes: ventana estrecha alrededor del año del pack."""
    return abs(int(year) - int(default_year)) <= 1


def _coerce_month(value: Any, default_year: int = 2026) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        if not _month_year_ok(value.year, default_year):
            return None
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        if not _month_year_ok(value.year, default_year):
            return None
        return date(value.year, value.month, 1)
    if isinstance(value, (int, float)):
        try:
            dt = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value))
            if not _month_year_ok(dt.year, default_year):
                return None
            return date(dt.year, dt.month, 1)
        except (OverflowError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    # YTD / etiquetas de fórmula: no son meses.
    if text.upper().startswith("YTD") or text.startswith("#"):
        return None
    for fmt in ("%m-%Y", "%m/%Y", "%Y-%m", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if not _month_year_ok(parsed.year, default_year):
                return None
            return date(parsed.year, parsed.month, 1)
        except ValueError:
            continue
    match = re.match(r"^(\d{1,2})[-/](\d{4})$", text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if 1 <= month <= 12 and _month_year_ok(year, default_year):
            return date(year, month, 1)
    return None


def _coerce_decimal(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text or text.startswith("#"):
        # Excel errors (#DIV/0!, #N/A, …) → 0
        return Decimal("0")
    try:
        return Decimal(text)
    except Exception:
        return Decimal("0")


def _detect_file_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".xlsb":
        return "xlsb"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    raise ValueError(f"Formato no soportado: {suffix}")


def _parse_levis_layout(rows: list[tuple[Any, ...]], default_year: int = 2026) -> list[ParsedSeedCell]:
    if len(rows) < 5:
        return []
    header = rows[3]
    month_columns: list[tuple[int, date]] = []
    col = 4
    while col < len(header) and len(month_columns) < 12:
        month_value = _coerce_month(header[col], default_year=default_year)
        if month_value is None:
            col += 1
            continue
        month_columns.append((col, month_value))
        col += 2

    parsed: list[ParsedSeedCell] = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        customer_name = str_or_default(row[0], "").strip()
        if not customer_name or customer_name.lower() == "total":
            continue
        city = str_or_default(row[1] if len(row) > 1 else "", "")
        store_type = str_or_default(row[2] if len(row) > 2 else "", "")
        product_group = str_or_default(row[3] if len(row) > 3 else "", "")
        seed_key = normalize_seed_key(
            customer_name,
            city=city,
            store_type=store_type,
            product_group=product_group,
        )
        for units_col, month in month_columns:
            amount_col = units_col + 1
            units = quantize_units(_coerce_decimal(row[units_col] if len(row) > units_col else 0))
            amount = quantize_amount(
                _coerce_decimal(row[amount_col] if len(row) > amount_col else 0)
            )
            if units == 0 and amount == 0:
                continue
            parsed.append(
                ParsedSeedCell(
                    seed_key=seed_key,
                    customer_name=customer_name,
                    city=city,
                    store_type=store_type,
                    product_group=product_group,
                    month=month,
                    units=units,
                    amount=amount,
                )
            )
    return parsed


def _parse_puma_layout(rows: list[tuple[Any, ...]], default_year: int = 2026) -> list[ParsedSeedCell]:
    if len(rows) < 5:
        return []
    header = rows[3]
    month_columns: list[tuple[int, date]] = []
    col = 10
    while col < len(header) and len(month_columns) < 12:
        month_value = _coerce_month(header[col], default_year=default_year)
        if month_value is None:
            col += 1
            continue
        month_columns.append((col, month_value))
        col += 2

    parsed: list[ParsedSeedCell] = []
    for row in rows[4:]:
        if not row:
            continue
        # Layout Puma: razón social en col 7 (idx 6); col 1 suele venir vacía.
        customer_name = str_or_default(row[6] if len(row) > 6 else "", "").strip()
        if not customer_name:
            customer_name = str_or_default(row[0], "").strip()
        if not customer_name or customer_name.lower() == "total":
            continue
        city = str_or_default(row[7] if len(row) > 7 else "", "")
        store_type = str_or_default(row[8] if len(row) > 8 else "", "")
        product_group = str_or_default(row[9] if len(row) > 9 else "", "")
        uf = ""
        seed_key = normalize_seed_key(
            customer_name,
            city=city,
            store_type=store_type,
            product_group=product_group,
            uf=uf,
        )
        for units_col, month in month_columns:
            amount_col = units_col + 1
            units = quantize_units(_coerce_decimal(row[units_col] if len(row) > units_col else 0))
            amount = quantize_amount(
                _coerce_decimal(row[amount_col] if len(row) > amount_col else 0)
            )
            if units == 0 and amount == 0:
                continue
            parsed.append(
                ParsedSeedCell(
                    seed_key=seed_key,
                    customer_name=customer_name,
                    city=city,
                    store_type=store_type,
                    product_group=product_group,
                    uf=uf,
                    month=month,
                    units=units,
                    amount=amount,
                )
            )
    return parsed


def _rows_from_openpyxl(path: Path) -> list[tuple[Any, ...]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET_SALES not in workbook.sheetnames:
        raise ValueError(f"Hoja requerida ausente: {SHEET_SALES}")
    sheet = workbook[SHEET_SALES]
    rows: list[tuple[Any, ...]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(tuple(row))
    workbook.close()
    return rows


def _rows_from_pyxlsb(path: Path) -> list[tuple[Any, ...]]:
    from pyxlsb import open_workbook

    rows: list[tuple[Any, ...]] = []
    with open_workbook(path) as workbook:
        if SHEET_SALES not in workbook.sheets:
            raise ValueError(f"Hoja requerida ausente: {SHEET_SALES}")
        with workbook.get_sheet(SHEET_SALES) as sheet:
            max_col = 0
            raw_rows: dict[int, dict[int, Any]] = {}
            for row in sheet.rows():
                for cell in row:
                    raw_rows.setdefault(cell.r, {})[cell.c] = cell.v
                    max_col = max(max_col, cell.c)
            for row_idx in sorted(raw_rows):
                row_values = tuple(
                    raw_rows[row_idx].get(col_idx)
                    for col_idx in range(max_col + 1)
                )
                rows.append(row_values)
    return rows


def parse_monthly_reporting_workbook(
    path: Path,
    pack: MonthlyReportingPack,
    *,
    default_year: int = 2026,
) -> list[ParsedSeedCell]:
    file_format = _detect_file_format(path)
    if file_format == "xlsx":
        rows = _rows_from_openpyxl(path)
    else:
        rows = _rows_from_pyxlsb(path)

    if pack.template_family == MonthlyReportingPack.TemplateFamily.PUMA:
        return _parse_puma_layout(rows, default_year=default_year)
    return _parse_levis_layout(rows, default_year=default_year)


def _get_or_create_match(cell: ParsedSeedCell) -> MonthlyReportingClientMatch:
    match, _ = MonthlyReportingClientMatch.objects.get_or_create(
        seed_key=cell.seed_key,
        defaults={
            "seed_customer_name": cell.customer_name,
            "seed_customer_code": cell.customer_code,
            "seed_city": cell.city,
            "seed_store_type": cell.store_type,
            "seed_product_group": cell.product_group,
            "seed_uf": cell.uf,
            "estado": MonthlyReportingClientMatch.Estado.PENDING,
        },
    )
    return match


def _apply_parsed_rows(
    *,
    pack: MonthlyReportingPack,
    batch: MonthlyReportingImportBatch,
    parsed_rows: Iterable[ParsedSeedCell],
    replace_mode: bool,
) -> tuple[int, int, list[dict]]:
    created = 0
    updated = 0
    skipped = 0
    audit_entries: list[dict] = []
    for cell in parsed_rows:
        match = _get_or_create_match(cell)
        existing = MonthlyReportingSeedRow.objects.filter(
            pack=pack,
            match=match,
            month=cell.month,
        ).first()
        if existing and not replace_mode:
            skipped += 1
            continue
        if existing and replace_mode:
            before = {
                "units": str(existing.units),
                "amount": str(existing.amount),
            }
            existing.units = cell.units
            existing.amount = cell.amount
            existing.units_men = cell.units_men
            existing.units_women = cell.units_women
            existing.amount_men = cell.amount_men
            existing.amount_women = cell.amount_women
            existing.city = cell.city
            existing.store_type = cell.store_type
            existing.uf = cell.uf
            existing.batch = batch
            existing.save()
            updated += 1
            audit_entries.append(
                {
                    "seed_key": cell.seed_key,
                    "month": cell.month.isoformat(),
                    "before": before,
                    "after": {"units": str(cell.units), "amount": str(cell.amount)},
                }
            )
            continue
        MonthlyReportingSeedRow.objects.create(
            pack=pack,
            match=match,
            month=cell.month,
            units=cell.units,
            amount=cell.amount,
            units_men=cell.units_men,
            units_women=cell.units_women,
            amount_men=cell.amount_men,
            amount_women=cell.amount_women,
            city=cell.city,
            store_type=cell.store_type,
            uf=cell.uf,
            batch=batch,
        )
        created += 1
    return created, updated, skipped, audit_entries


def import_monthly_reporting_file(
    pack_id: str,
    file_path: Path | str,
    *,
    replace_mode: bool = False,
    actor: Optional[ImportActor] = None,
    default_year: int = 2026,
    force_fail_after_parse: bool = False,
) -> ImportResult:
    """Importa planilla seed con idempotencia por hash SHA-256."""
    path = Path(file_path)
    pack = MonthlyReportingPack.objects.get(pack_id=pack_id)
    sha256 = compute_file_sha256(path)
    actor = actor or ImportActor()
    file_format = _detect_file_format(path)

    canonical = (
        MonthlyReportingImportBatch.objects.filter(
            pack=pack,
            file_sha256=sha256,
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
            duplicate_of__isnull=True,
        )
        .order_by("-applied_at", "-id")
        .first()
    )
    if canonical and not replace_mode:
        duplicate_batch = MonthlyReportingImportBatch.objects.create(
            pack=pack,
            file_name=path.name,
            file_size=path.stat().st_size,
            file_format=file_format,
            file_sha256=sha256,
            replace_mode=False,
            estado=MonthlyReportingImportBatch.Estado.DUPLICATE,
            duplicate_of=canonical,
            actor_id_usuario=actor.id_usuario,
            actor_cod_usuario=actor.cod_usuario,
            actor_nombre=actor.nombre,
        )
        return ImportResult(batch=duplicate_batch, duplicate=True)

    batch = MonthlyReportingImportBatch.objects.create(
        pack=pack,
        file_name=path.name,
        file_size=path.stat().st_size,
        file_format=file_format,
        file_sha256=sha256,
        replace_mode=replace_mode,
        estado=MonthlyReportingImportBatch.Estado.PENDING,
        actor_id_usuario=actor.id_usuario,
        actor_cod_usuario=actor.cod_usuario,
        actor_nombre=actor.nombre,
    )
    try:
        with transaction.atomic():
            MonthlyReportingImportBatch.objects.select_for_update().filter(pack=pack).exists()
            parsed_rows = parse_monthly_reporting_workbook(
                path,
                pack,
                default_year=default_year,
            )
            if force_fail_after_parse:
                raise RuntimeError("Fallo simulado post-parse")
            created, updated, skipped, audit_entries = _apply_parsed_rows(
                pack=pack,
                batch=batch,
                parsed_rows=parsed_rows,
                replace_mode=replace_mode,
            )
            batch.rows_created = created
            batch.rows_updated = updated
            batch.rows_skipped = skipped
            if replace_mode and audit_entries:
                batch.audit_json = {"replacements": audit_entries}
            batch.estado = MonthlyReportingImportBatch.Estado.APPLIED
            batch.applied_at = timezone.now()
            batch.save(
                update_fields=[
                    "rows_created",
                    "rows_updated",
                    "rows_skipped",
                    "audit_json",
                    "estado",
                    "applied_at",
                ]
            )
    except Exception as exc:
        MonthlyReportingImportBatch.objects.filter(pk=batch.pk).update(
            estado=MonthlyReportingImportBatch.Estado.FAILED,
            error_message=str(exc),
        )
        batch.refresh_from_db()
        raise
    return ImportResult(batch=batch, rows_created=batch.rows_created, rows_updated=batch.rows_updated)


def ensure_monthly_reporting_packs() -> list[MonthlyReportingPack]:
    return seed_monthly_reporting_packs(MonthlyReportingPack)
