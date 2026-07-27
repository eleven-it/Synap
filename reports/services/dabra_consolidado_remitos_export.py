"""
Exporter openpyxl — Informe DABRA consolidado remitos.

Consume el payload de ``get_dabra_consolidado_remitos`` (paridad preview↔Excel).
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, List, Sequence

from django.http import HttpResponse

# Headers exactos del sample DABRA 052026.xlsx (incl. espacios trailing en P904/P908/…)
REPORTE_HEADERS: List[str] = [
    "NroCUIT",
    "Fecha",
    "DocType",
    "PuntoVenta",
    "NumeroLegal",
    "Item",
    "Talle",
    "Cantidad",
    "Precio",
    "Bonificacion",
    "ImporteBonificacion",
    "Importe",
    "Iva",
    "TotalGravado",
    "DescuentoPorcentualTotal",
    "DescuentoTotal",
    "Total",
    "CompRef",
    "NumeroRef",
    "Entrega",
    "NroCAE",
    "VtoCAE",
    "Suc",
    "Categoria",
    "P901",
    "P902",
    "P903",
    "P904  ",
    "P905",
    "P906",
    "P907",
    "P908  ",
    "P909",
    "P910",
    "P911",
    "P912  ",
    "P913",
    "P914",
    "P915",
    "P916  ",
    "P917",
    "P918",
    "P919",
    "P920  ",
    "P921",
    "P922",
    "P923",
    "P924",
    "PIVA3",
]

TOTAL_FACTURAS_HEADERS = ["Fecha", "Comprobante", "Nro. Remito", "Imp Neto", "Imp Bruto"]

# Columnas Y–AW (índices 1-based 25–49) deben ser 0
_COL_Y_INDEX = 25
_COL_AW_INDEX = 49


def _nombre_archivo(mes: int, anio: int) -> str:
    return f"DABRA {mes:02d}{anio}.xlsx"


def _parse_fecha_excel(valor: str) -> Any:
    """Convierte dd/MM/yyyy a date para Excel."""
    if not valor:
        return None
    try:
        return datetime.strptime(valor, "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return valor


def _fila_reporte_excel(fila: Dict[str, Any]) -> List[Any]:
    """Una fila A–AW para hoja REPORTE (sin NombreArticulo)."""
    row: List[Any] = [None] * len(REPORTE_HEADERS)
    row[0] = int(fila["cuit_emisor"]) if fila.get("cuit_emisor") else None
    row[1] = _parse_fecha_excel(fila.get("fecha", ""))
    row[2] = fila.get("doc_type", 1)
    row[3] = fila.get("punto_venta", "")
    row[4] = fila.get("numero_legal", 0)
    row[5] = fila.get("item", "")
    row[6] = fila.get("talle", "")
    row[7] = fila.get("cantidad", 0)
    row[8] = fila.get("precio_unitario", 0)
    row[9] = fila.get("bonificacion", 0)
    row[10] = fila.get("importe_bonificacion", 0)
    row[11] = fila.get("importe", 0)
    row[12] = fila.get("importe_iva", 0)
    row[13] = fila.get("total_gravado", 0)
    # O (14) y P (15) vacías
    row[16] = fila.get("total", 0)
    row[17] = fila.get("comp_ref", "")
    row[18] = fila.get("numero_ref", "")
    row[19] = fila.get("entrega", "")
    row[20] = fila.get("cae", "")
    vto = fila.get("vto_cae", "")
    row[21] = _parse_fecha_excel(vto) if vto else None
    row[22] = fila.get("suc", "")
    row[23] = fila.get("categoria", "")
    for idx in range(_COL_Y_INDEX - 1, _COL_AW_INDEX):
        row[idx] = 0
    return row


def exportar_dabra_xlsx(
    payload: Dict[str, Any],
    *,
    mes: int,
    anio: int,
) -> HttpResponse:
    """Genera HttpResponse con workbook DABRA MMYYYY.xlsx."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws_reporte = wb.active
    ws_reporte.title = "REPORTE"
    ws_reporte.append(REPORTE_HEADERS)

    for fila in payload.get("filas") or []:
        ws_reporte.append(_fila_reporte_excel(fila))

    ws_total = wb.create_sheet("TOTAL FACTURAS")
    ws_total.append(TOTAL_FACTURAS_HEADERS)
    for tf in payload.get("totales_facturas") or []:
        ws_total.append(
            [
                _parse_fecha_excel(tf.get("fecha", "")),
                tf.get("comprobante", ""),
                tf.get("nro_remito", ""),
                tf.get("imp_neto", 0),
                tf.get("imp_bruto", 0),
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_nombre_archivo(mes, anio)}"'
    return response


def inspeccionar_workbook(content: bytes) -> Dict[str, Any]:
    """Helper de tests: devuelve hojas, headers y muestra de filas."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: Dict[str, Any] = {"sheetnames": wb.sheetnames, "headers": {}, "rows": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        out["headers"][name] = headers
        sample = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
            sample.append(list(row))
        out["rows"][name] = sample
    return out
