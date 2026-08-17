# -*- coding: utf-8 -*-
"""Tests — Ventas Mensuales Licenciatarios (fase 0 + historial híbrido Ph1-2)."""

from __future__ import annotations

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import Mock, patch

import openpyxl
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.test import SimpleTestCase, TestCase

from reports.models import (
    MonthlyReportingClientMatch,
    MonthlyReportingClientMatchAudit,
    MonthlyReportingImportBatch,
    MonthlyReportingPack,
    MonthlyReportingSeedRow,
    MonthlyReportingSuperArtCatalogEntry,
    MonthlyReportingSuperArtCatalogVersion,
    MonthlyReportingSuperArtQAPending,
    ReportDefinition,
)
from reports.services.monthly_reporting_pack_seed import (
    MONTHLY_REPORTING_PACK_DEFINITIONS,
    MONTHLY_REPORTING_TEMPLATE_FILES,
    seed_monthly_reporting_packs,
)
from reports.services.monthly_reporting_template_builder import TEMPLATE_DIR, build_all_templates
from reports.services.ventas_mensuales_licenciatarios_importer import (
    _coerce_decimal,
    _coerce_month,
    compute_file_sha256,
    import_monthly_reporting_file,
    normalize_seed_key,
    quantize_amount,
    quantize_units,
)
from reports.services.ventas_mensuales_licenciatarios_export import (
    QA_SHEET,
    SHEET_MINIMUM,
    SHEET_MONTHLY,
    SHEET_OOH,
    SHEET_SALES,
    _write_levis_sales_sheet,
    export_licenciatarios_workbook,
    resolve_template_path,
)
from reports.services.ventas_mensuales_licenciatarios_runner import (
    run_ventas_mensuales_licenciatarios,
    validate_calendar_year_range,
)
from reports.services.ventas_mensuales_licenciatarios_seed import (
    VENTAS_MENSUALES_LICENCIATARIOS_SLUG,
    _report_defaults,
)
from reports.services.ventas_marcas_mensual_rules import (
    FACTOR_DOCENAS_MAP,
    STOCK_TIPO_COMP,
    TIPOS_FAC,
    TIPOS_NC,
    apply_comprobante_sign,
    compute_units_amount_for_pack,
    factor_docenas_unimed,
    sql_base_where_clauses,
    sql_signo_imp_expr,
    sql_signo_qty_expr,
)
from reports.services.ventas_mensuales_licenciatarios_query import (
    AnetSalesRow,
    build_anet_sales_sql,
    parse_anet_sales_row,
)
from reports.services.ventas_mensuales_licenciatarios_merger import (
    CUTOVER_DATE,
    MergedClientMonth,
    anet_range_for_month,
    compare_dz_pk_parity,
    compute_ytd,
    filter_merge_result_by_clientes_excluidos,
    merge_pack_year,
    seed_months_in_range,
)
from reports.services.monthly_reporting_client_match_service import (
    MatchActor,
    apply_client_match,
    format_audit_fecha,
    match_to_aggregate_row,
    resolve_client_identity,
)
from reports.services.monthly_reporting_superart_service import (
    activate_catalog_version,
    classify_superart,
    make_classify_fn,
    register_qa_pending,
    seed_catalog_entries,
)
from reports.services.ventas_mensuales_licenciatarios_importer import ParsedSeedCell
from reports.services.ventas_mensuales_licenciatarios_reconciliation import (
    FA_NC_REFERENCE_NOTE,
    aggregate_db_seed_rows,
    aggregate_parsed_cells,
    compare_seed_aggregates,
    compute_ytd_from_aggregates,
    reconcile_pack_from_file,
    resolve_pack_source_path,
)
from reports.services.ventas_marcas_mensual_rules import TIPOS_FAC, TIPOS_NC


def _build_levis_seed_xlsx(
    path: Path,
    *,
    customer: str = "Cliente Demo",
    units: float = 12,
    amount: float = 1500.75,
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "input Licensee sales"
    ws.cell(row=4, column=1, value="Customer")
    ws.cell(row=4, column=2, value="City / Province")
    ws.cell(row=4, column=3, value="Store Type")
    ws.cell(row=4, column=4, value="Product group")
    ws.cell(row=4, column=5, value=date(2026, 1, 1))
    ws.cell(row=5, column=1, value=customer)
    ws.cell(row=5, column=2, value="CABA")
    ws.cell(row=5, column=3, value="Store")
    ws.cell(row=5, column=4, value="Bodywear")
    ws.cell(row=5, column=5, value=units)
    ws.cell(row=5, column=6, value=amount)
    wb.save(path)
    return path


class VentasMensualesLicenciatariosStubTest(SimpleTestCase):
    def test_slug_y_defaults(self):
        self.assertEqual(VENTAS_MENSUALES_LICENCIATARIOS_SLUG, "ventas-mensuales-licenciatarios")
        cfg = _report_defaults()
        self.assertEqual(cfg["name"], "Ventas Mensuales Licenciatarios")
        self.assertEqual(cfg["config"]["sibling_of"], "ventas-marcas-mensual")
        self.assertIn("levis_bw", cfg["config"]["packs"])

    def test_validate_calendar_year_rechaza_cruce_anios(self):
        with self.assertRaises(ValueError) as ctx:
            validate_calendar_year_range("2025-10-01", "2026-02-15")
        self.assertIn("mismo año calendario", str(ctx.exception).lower())


class LicenciatariosExportLayoutSimpleTests(SimpleTestCase):
    """Formato plantilla Levi’s: encabezados en inglés, city/store/group del seed, totales fila 2."""

    def test_sales_headers_ingles_city_y_sum_fila_2(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_SALES
        rows = [
            MergedClientMonth(
                identity="seed:demo",
                display_name="LEURU S.A",
                match_estado="matched",
                month=date(2026, 1, 1),
                units=Decimal("12"),
                amount=Decimal("1500.50"),
                city="Cap Fed",
                store_type="Multibrand",
                product_group="Bodywear",
            )
        ]
        _write_levis_sales_sheet(
            ws,
            rows=rows,
            year=2026,
            month_from=1,
            month_to=12,
            product_group="Bodywear",
        )
        self.assertEqual(ws.cell(row=4, column=1).value, "Customer")
        self.assertEqual(ws.cell(row=4, column=2).value, "City / Province")
        self.assertEqual(ws.cell(row=4, column=3).value, "Store Type")
        self.assertEqual(ws.cell(row=4, column=4).value, "Product group")
        self.assertEqual(ws.cell(row=5, column=2).value, "Cap Fed")
        self.assertEqual(ws.cell(row=5, column=3).value, "Multibrand")
        self.assertEqual(ws.cell(row=5, column=4).value, "Bodywear")
        self.assertEqual(ws.cell(row=2, column=5).value, "=SUM(E5:E4931)")
        self.assertEqual(ws.cell(row=2, column=6).value, "=SUM(F5:F4931)")
        self.assertEqual(ws.cell(row=4, column=29).value, "YTD_Units")
        self.assertEqual(ws.cell(row=4, column=30).value, "YTD_Sales")
        self.assertTrue(str(ws.cell(row=5, column=29).value or "").startswith("=AA5+"))
        self.assertIn("E4:F4", {str(m) for m in ws.merged_cells.ranges})
        self.assertEqual(ws.cell(row=3, column=5).value, "units")
        self.assertEqual(ws.cell(row=3, column=6).value, "amounts")
        self.assertEqual(ws.cell(row=4, column=1).fill.fgColor.rgb, "004F81BD")
        self.assertEqual(ws.cell(row=4, column=1).font.color.rgb, "00FFFFFF")
        self.assertGreaterEqual(ws.column_dimensions["A"].width or 0, 31.0)
        self.assertGreaterEqual(ws.column_dimensions["F"].width or 0, 14.0)


class LicenciatariosExportServiceArtifactsTests(SimpleTestCase):
    """ExportService lee MergeResult desde QueryResult.artifacts (no meta.extra)."""

    def test_export_pasa_skip_cache_y_lee_artifacts(self):
        from reports.services.export_service import ExportService
        from reports.services.query_runner import QueryResult
        from reports.services.ventas_mensuales_licenciatarios_merger import MergeResult

        merge = MergeResult(
            rows=[],
            ytd_by_identity={},
            pending_clients=[],
            qa_superarts=[],
        )
        qr = QueryResult(
            meta={
                "extra": {
                    "pack_id": "levis_bw",
                    "year": 2026,
                    "month_from": 1,
                    "month_to": 12,
                }
            },
            data=[],
            totals={},
            notes=[],
            artifacts={"merge_result": merge},
        )
        report = Mock()
        report.slug = "ventas-mensuales-licenciatarios"
        report.name = "VML"
        report.is_active = True

        svc = ExportService(Mock())
        captured = {}

        def _fake_export_wb(path, *, pack, merge_result, year, month_from, month_to):
            captured["merge"] = merge_result
            captured["year"] = year
            Path(path).write_bytes(b"PK")

        with patch.object(ReportDefinition.objects, "get", return_value=report), patch(
            "reports.services.export_service.QueryRunnerService"
        ) as mock_qrs, patch(
            "reports.models.MonthlyReportingPack.objects.get", return_value=Mock(pack_id="levis_bw")
        ), patch(
            "reports.services.ventas_mensuales_licenciatarios_export.export_licenciatarios_workbook",
            side_effect=_fake_export_wb,
        ), patch(
            "django.conf.settings.MEDIA_ROOT",
            new=tempfile.mkdtemp(),
        ):
            mock_qrs.return_value.run.return_value = qr
            result = svc.export(
                "ventas-mensuales-licenciatarios",
                {
                    "filters": {
                        "pack_id": "levis_bw",
                        "fecha_inicio_facturacion": "2026-01-01",
                        "fecha_fin_facturacion": "2026-12-31",
                    }
                },
                "xlsx",
            )
            run_payload = mock_qrs.return_value.run.call_args[0][1]
            self.assertTrue(run_payload.get("_skip_report_cache"))
            self.assertIs(captured["merge"], merge)
            self.assertEqual(captured["year"], 2026)
            self.assertEqual(
                result.filename,
                "Monthly Reporting Best Sox_LEVIS BW 26.xlsx",
            )

    def test_export_sin_artifacts_falla_claro(self):
        from reports.services.export_service import ExportService
        from reports.services.query_runner import QueryResult

        qr = QueryResult(
            meta={"extra": {"pack_id": "levis_bw", "year": 2026}},
            data=[],
            totals={},
            notes=[],
            artifacts={},
        )
        report = Mock()
        report.slug = "ventas-mensuales-licenciatarios"
        with patch.object(ReportDefinition.objects, "get", return_value=report), patch(
            "reports.services.export_service.QueryRunnerService"
        ) as mock_qrs:
            mock_qrs.return_value.run.return_value = qr
            with self.assertRaises(ValueError) as ctx:
                ExportService(Mock())._generate_excel_ventas_mensuales_licenciatarios(
                    Path(tempfile.mkdtemp()) / "out.xlsx",
                    report,
                    qr,
                    {"filters": {"pack_id": "levis_bw"}},
                )
            self.assertIn("artifacts", str(ctx.exception).lower())


class MonthlyReportingModelConstraintTests(TestCase):
    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.batch = MonthlyReportingImportBatch.objects.create(
            pack=self.pack,
            file_name="demo.xlsx",
            file_format="xlsx",
            file_sha256="abc123",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:test-cliente",
            seed_customer_name="Cliente Demo",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )

    def test_unique_pack_match_month(self):
        MonthlyReportingSeedRow.objects.create(
            pack=self.pack,
            match=self.match,
            month=date(2026, 1, 1),
            units=Decimal("1.000000"),
            amount=Decimal("100.00"),
            batch=self.batch,
        )
        with self.assertRaises(ValidationError):
            MonthlyReportingSeedRow.objects.create(
                pack=self.pack,
                match=self.match,
                month=date(2026, 1, 1),
                units=Decimal("2.000000"),
                amount=Decimal("200.00"),
                batch=self.batch,
            )

    def test_match_estado_requiere_anet_id(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                MonthlyReportingClientMatch.objects.create(
                    seed_key="invalid-matched",
                    seed_customer_name="Sin ID",
                    estado=MonthlyReportingClientMatch.Estado.MATCHED,
                    anet_cliente_id=None,
                )

    def test_match_pending_sin_anet_id(self):
        match = MonthlyReportingClientMatch.objects.create(
            seed_key="pending-ok",
            seed_customer_name="Pendiente OK",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
            anet_cliente_id=None,
        )
        self.assertEqual(match.estado, MonthlyReportingClientMatch.Estado.PENDING)

    def test_match_matched_con_anet_id(self):
        match = MonthlyReportingClientMatch.objects.create(
            seed_key="matched-ok",
            seed_customer_name="Matcheado OK",
            estado=MonthlyReportingClientMatch.Estado.MATCHED,
            anet_cliente_id=12345,
        )
        self.assertEqual(match.anet_cliente_id, 12345)

    def test_batch_unique_pack_sha_aplicado(self):
        MonthlyReportingImportBatch.objects.create(
            pack=self.pack,
            file_name="dup.xlsx",
            file_format="xlsx",
            file_sha256="sha-aplicado",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                MonthlyReportingImportBatch.objects.create(
                    pack=self.pack,
                    file_name="dup2.xlsx",
                    file_format="xlsx",
                    file_sha256="sha-aplicado",
                    estado=MonthlyReportingImportBatch.Estado.APPLIED,
                )

    def test_seed_row_rechaza_mes_no_primero(self):
        row = MonthlyReportingSeedRow(
            pack=self.pack,
            match=self.match,
            month=date(2026, 1, 15),
            units=Decimal("1.000000"),
            amount=Decimal("10.00"),
            batch=self.batch,
        )
        with self.assertRaises(ValidationError):
            row.save()


class MonthlyReportingPackFixtureTests(TestCase):
    def test_seed_crea_seis_packs(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        pack_ids = set(MonthlyReportingPack.objects.values_list("pack_id", flat=True))
        expected = {item["pack_id"] for item in MONTHLY_REPORTING_PACK_DEFINITIONS}
        self.assertEqual(pack_ids, expected)

    def test_plantillas_xlsx_generadas(self):
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp)
            written = build_all_templates(target)
            self.assertEqual(len(written), 6)
            for pack_id, filename in MONTHLY_REPORTING_TEMPLATE_FILES.items():
                path = target / filename
                self.assertTrue(path.exists(), msg=f"Falta plantilla {pack_id}")
                wb = openpyxl.load_workbook(path, read_only=True)
                self.assertIn("input Licensee sales", wb.sheetnames)
                self.assertIn("monthly", wb.sheetnames)
                wb.close()

    def test_directorio_plantillas_repo(self):
        build_all_templates(TEMPLATE_DIR)
        for filename in MONTHLY_REPORTING_TEMPLATE_FILES.values():
            self.assertTrue((TEMPLATE_DIR / filename).exists(), msg=filename)


class MonthlyReportingImportTests(TestCase):
    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.file_path = _build_levis_seed_xlsx(Path(self.tempdir.name) / "seed.xlsx")

    def test_hash_sha256_estable(self):
        digest = compute_file_sha256(self.file_path)
        self.assertEqual(len(digest), 64)
        self.assertEqual(digest, compute_file_sha256(self.file_path))

    def test_import_inicial_crea_filas_y_batch(self):
        result = import_monthly_reporting_file("levis_bw", self.file_path)
        self.assertFalse(result.duplicate)
        self.assertEqual(result.batch.estado, MonthlyReportingImportBatch.Estado.APPLIED)
        self.assertEqual(result.batch.rows_created, 1)
        self.assertEqual(MonthlyReportingSeedRow.objects.count(), 1)
        row = MonthlyReportingSeedRow.objects.get()
        self.assertEqual(row.units, Decimal("12.000000"))
        self.assertEqual(row.amount, Decimal("1500.75"))

    def test_reimport_idempotente_cero_altas(self):
        first = import_monthly_reporting_file("levis_bw", self.file_path)
        second = import_monthly_reporting_file("levis_bw", self.file_path)
        self.assertTrue(second.duplicate)
        self.assertEqual(second.batch.estado, MonthlyReportingImportBatch.Estado.DUPLICATE)
        self.assertEqual(second.batch.duplicate_of_id, first.batch.id)
        self.assertEqual(MonthlyReportingSeedRow.objects.count(), 1)
        self.assertEqual(
            MonthlyReportingImportBatch.objects.filter(
                estado=MonthlyReportingImportBatch.Estado.APPLIED
            ).count(),
            1,
        )

    def test_replace_actualiza_y_audita(self):
        import_monthly_reporting_file("levis_bw", self.file_path)
        updated_path = _build_levis_seed_xlsx(
            Path(self.tempdir.name) / "seed_v2.xlsx",
            units=20,
            amount=2000,
        )
        result = import_monthly_reporting_file(
            "levis_bw",
            updated_path,
            replace_mode=True,
        )
        self.assertEqual(result.batch.rows_updated, 1)
        self.assertEqual(result.batch.rows_created, 0)
        row = MonthlyReportingSeedRow.objects.get()
        self.assertEqual(row.units, Decimal("20.000000"))
        self.assertEqual(row.amount, Decimal("2000.00"))
        replacements = result.batch.audit_json.get("replacements", [])
        self.assertEqual(len(replacements), 1)
        self.assertIn("before", replacements[0])
        self.assertIn("after", replacements[0])

    def test_batch_fallido_no_persiste_filas(self):
        with self.assertRaises(RuntimeError):
            import_monthly_reporting_file(
                "levis_bw",
                self.file_path,
                force_fail_after_parse=True,
            )
        self.assertEqual(MonthlyReportingSeedRow.objects.count(), 0)
        failed = MonthlyReportingImportBatch.objects.latest("id")
        self.assertEqual(failed.estado, MonthlyReportingImportBatch.Estado.FAILED)
        self.assertIn("Fallo simulado", failed.error_message)


class MonthlyReportingImporterNormalizeTests(SimpleTestCase):
    def test_normalize_seed_key_usa_codigo(self):
        self.assertEqual(normalize_seed_key("Cliente", customer_code="C001"), "code:C001")

    def test_normalize_seed_key_colapsa_espacios(self):
        key_a = normalize_seed_key("  Cliente   Demo  ", city=" CABA ")
        key_b = normalize_seed_key("cliente demo", city="caba")
        self.assertEqual(key_a, key_b)
        self.assertTrue(key_a.startswith("name:"))

    def test_quantize_decimal_sin_float_binario(self):
        self.assertEqual(quantize_units("12.3456789"), Decimal("12.345679"))
        self.assertEqual(quantize_amount("1500.756"), Decimal("1500.76"))

    def test_coerce_decimal_ignora_errores_excel(self):
        self.assertEqual(_coerce_decimal("#DIV/0!"), Decimal("0"))
        self.assertEqual(_coerce_decimal("#N/A"), Decimal("0"))
        self.assertEqual(_coerce_decimal("1,234.5"), Decimal("1234.5"))

    def test_coerce_month_ignora_ytd_y_seriales_fuera_de_anio(self):
        self.assertIsNone(_coerce_month("YTD_Units", default_year=2026))
        self.assertIsNone(_coerce_month(1413.12, default_year=2026))
        self.assertEqual(
            _coerce_month(date(2026, 7, 1), default_year=2026),
            date(2026, 7, 1),
        )


class VentasMarcasMensualRulesExtractedTests(SimpleTestCase):
    """Phase 3.1 — reglas VMM extraídas a módulo compartido."""

    def test_tipos_fac_nc_y_stock(self):
        self.assertEqual(TIPOS_FAC, ("FA", "FB", "FC", "FE", "FM"))
        self.assertEqual(TIPOS_NC, ("NCA", "NCB", "NCC", "NCE", "NCM"))
        self.assertIn("Venta", STOCK_TIPO_COMP)
        self.assertIn("ND Anul NC", STOCK_TIPO_COMP)

    def test_sql_anulados_y_comprobantes(self):
        clauses = sql_base_where_clauses()
        joined = " ".join(clauses)
        self.assertIn("cc.Anulado = 'No'", joined)
        self.assertIn("st.Anulado = 'No'", joined)
        self.assertIn("'FA'", joined)
        self.assertIn("'NCA'", joined)
        self.assertIn("'Venta TPV'", joined)
        self.assertIn("art.tipo_art <> 'Gasto'", joined)

    def test_signo_sql_contiene_fac_y_nc(self):
        qty_sql = sql_signo_qty_expr()
        imp_sql = sql_signo_imp_expr()
        self.assertIn("'FA'", qty_sql)
        self.assertIn("-COALESCE(st.Cantidad", qty_sql)
        self.assertIn("-COALESCE(st.PrecioNetoxR", imp_sql)

    def test_apply_comprobante_sign_fac_positivo_nc_negativo(self):
        qty_fa, amt_fa = apply_comprobante_sign("FA", 10, 100)
        self.assertEqual(qty_fa, Decimal("10"))
        self.assertEqual(amt_fa, Decimal("100"))
        qty_nc, amt_nc = apply_comprobante_sign("NCA", 5, 50)
        self.assertEqual(qty_nc, Decimal("-5"))
        self.assertEqual(amt_nc, Decimal("-50"))

    def test_apply_comprobante_sign_desconocido_cero(self):
        qty, amt = apply_comprobante_sign("XX", 9, 90)
        self.assertEqual(qty, Decimal("0"))
        self.assertEqual(amt, Decimal("0"))


class VentasMensualesLicenciatariosQueryTests(SimpleTestCase):
    """Phase 3.2 — SQL query RO."""

    def test_build_anet_sales_sql_incluye_reglas_vmm(self):
        sql = build_anet_sales_sql()
        self.assertIn("cc.Anulado = 'No'", sql)
        self.assertIn("st.Anulado = 'No'", sql)
        self.assertIn("art.CodigoMarca = (", sql)
        self.assertIn("m.NombreMarca = %s", sql)
        self.assertIn("art.tipo_art <> 'Gasto'", sql)

    def test_build_anet_sales_sql_importe_post_pie(self):
        """ANET amounts MUST usar el mismo factor cabecera que VMM (SubtotalDesc/SubTotal1)."""
        sql = build_anet_sales_sql()
        self.assertIn("SubTotal1", sql)
        self.assertIn("SubtotalDesc", sql)
        self.assertIn("PrecioNetoxR", sql)
        self.assertIn("0.0001", sql)

    def test_parse_anet_sales_row_docenas(self):
        row = parse_anet_sales_row(
            {
                "codigo_cliente": 100,
                "nombre_cliente": "Cliente",
                "month_start": "2026-08-01",
                "packs_qty": 24,
                "docenas_qty": 2,
                "facturacion": 500,
            },
            unit_mode="dozens",
        )
        self.assertEqual(row.units, Decimal("2"))
        self.assertEqual(row.amount, Decimal("500"))
        self.assertEqual(row.month, date(2026, 8, 1))


class VentasMensualesLicenciatariosMergerCutoverTests(SimpleTestCase):
    """Phase 3.3 — cutover 21/22 y paridad signos."""

    def test_anet_range_julio_solo_22_31(self):
        rango = anet_range_for_month(2026, 7)
        self.assertEqual(rango, (date(2026, 7, 22), date(2026, 7, 31)))

    def test_anet_range_enero_none(self):
        self.assertIsNone(anet_range_for_month(2026, 1))

    def test_seed_months_hasta_julio_2026(self):
        self.assertEqual(seed_months_in_range(2026, 1, 12), [1, 2, 3, 4, 5, 6, 7])

    def test_compute_ytd_sin_doble_conteo_julio(self):
        identity = "seed:name:abc"
        rows = [
            MergedClientMonth(
                identity=identity,
                display_name="Demo",
                match_estado="pending",
                month=date(2026, 6, 1),
                units=Decimal("10"),
                amount=Decimal("100"),
            ),
            MergedClientMonth(
                identity=identity,
                display_name="Demo",
                match_estado="pending",
                month=date(2026, 7, 1),
                units=Decimal("5"),
                amount=Decimal("50"),
                source="seed",
            ),
            MergedClientMonth(
                identity=identity,
                display_name="Demo",
                match_estado="pending",
                month=date(2026, 7, 1),
                units=Decimal("3"),
                amount=Decimal("30"),
                source="anet",
            ),
        ]
        ytd = compute_ytd(rows)
        self.assertEqual(ytd[identity]["units"], Decimal("18"))
        self.assertEqual(ytd[identity]["amount"], Decimal("180"))

    def test_nc_signo_paridad_vmm_en_pack_units(self):
        units_fa, amt_fa = compute_units_amount_for_pack(
            cantidad=12,
            importe=120,
            nombre_unimed="P1",
            unit_mode="dozens",
            tipo_comprobante="FA",
        )
        units_nc, amt_nc = compute_units_amount_for_pack(
            cantidad=12,
            importe=120,
            nombre_unimed="P1",
            unit_mode="dozens",
            tipo_comprobante="NCA",
        )
        self.assertEqual(units_fa, Decimal("1"))
        self.assertEqual(units_nc, Decimal("-1"))
        self.assertEqual(amt_fa, Decimal("120"))
        self.assertEqual(amt_nc, Decimal("-120"))

    def test_filter_merge_excluye_anet_id_y_recalcula_ytd(self):
        identity_anet = "anet:demo:500"
        rows = [
            MergedClientMonth(
                identity=identity_anet,
                display_name="Cliente ANET",
                match_estado="anet_only",
                month=date(2026, 8, 1),
                units=Decimal("10"),
                amount=Decimal("100"),
                source="anet",
                anet_cliente_id=500,
            ),
            MergedClientMonth(
                identity="anet:demo:501",
                display_name="Otro Cliente",
                match_estado="anet_only",
                month=date(2026, 8, 1),
                units=Decimal("4"),
                amount=Decimal("40"),
                source="anet",
                anet_cliente_id=501,
            ),
        ]
        from reports.services.ventas_mensuales_licenciatarios_merger import MergeResult

        merge = MergeResult(
            rows=rows,
            ytd_by_identity=compute_ytd(rows),
            pending_clients=[],
            qa_superarts=[],
        )
        filtered = filter_merge_result_by_clientes_excluidos(
            merge,
            [500],
            base_empresa="demo",
        )
        self.assertEqual(len(filtered.rows), 1)
        self.assertEqual(filtered.rows[0].anet_cliente_id, 501)
        self.assertNotIn(identity_anet, filtered.ytd_by_identity)
        self.assertEqual(filtered.ytd_by_identity["anet:demo:501"]["units"], Decimal("4"))


class VentasMensualesLicenciatariosClientesExcluidosTests(TestCase):
    """Exclusión de clientes en merge y runner."""

    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:excluir-seed",
            seed_customer_name="Cliente Seed Matcheado",
            estado=MonthlyReportingClientMatch.Estado.MATCHED,
            anet_cliente_id=900,
            base_empresa="demo",
        )
        self.batch = MonthlyReportingImportBatch.objects.create(
            pack=self.pack,
            file_name="excl.xlsx",
            file_format="xlsx",
            file_sha256="sha-excl-seed",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        MonthlyReportingSeedRow.objects.create(
            pack=self.pack,
            match=self.match,
            month=date(2026, 3, 1),
            units=Decimal("8"),
            amount=Decimal("80"),
            batch=self.batch,
        )
        self.report = ReportDefinition.objects.create(
            slug=VENTAS_MENSUALES_LICENCIATARIOS_SLUG,
            name="Ventas Mensuales Licenciatarios",
            category="operational",
            is_active=True,
        )

    def test_merge_excluye_seed_matcheado_por_anet_id(self):
        result = merge_pack_year(
            pack=self.pack,
            year=2026,
            month_from=1,
            month_to=6,
            base_empresa="demo",
            fetch_anet_fn=lambda **kwargs: [],
        )
        self.assertEqual(len(result.rows), 1)
        filtered = filter_merge_result_by_clientes_excluidos(
            result,
            [900],
            base_empresa="demo",
        )
        self.assertEqual(filtered.rows, [])
        self.assertEqual(filtered.ytd_by_identity, {})
        self.assertEqual(filtered.pending_clients, [])

    def test_runner_aplica_clientes_excluidos(self):
        payload = {
            "filters": {
                "pack_id": "levis_bw",
                "fecha_inicio_facturacion": "2026-01-01",
                "fecha_fin_facturacion": "2026-03-31",
                "clientes_excluidos": ["900"],
            },
            "base_empresa": "demo",
        }
        result = run_ventas_mensuales_licenciatarios(
            self.report,
            payload,
            Mock(),
            fetch_anet_fn=lambda **kwargs: [],
        )
        self.assertEqual(result.data, [])
        self.assertEqual(result.totals.get("unidades"), 0)
        self.assertEqual(result.meta["filters_applied"]["clientes_excluidos"], [900])
        self.assertTrue(any("excluidos" in n.lower() for n in result.notes))


class MonthlyReportingClientMatchServiceTests(TestCase):
    """Phase 4.1 — match auditable y pendientes visibles."""

    def setUp(self):
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:cliente-x",
            seed_customer_name="Cliente X",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )

    def test_pending_identity_seed(self):
        self.assertEqual(resolve_client_identity(self.match), "seed:name:cliente-x")
        meta = match_to_aggregate_row(self.match)
        self.assertTrue(meta["pending"])

    def test_apply_match_audita_json(self):
        apply_client_match(
            self.match,
            anet_cliente_id=999,
            base_empresa="empresa_demo",
            actor=MatchActor(id_usuario=1, cod_usuario="u1", nombre="Tester"),
        )
        self.match.refresh_from_db()
        self.assertEqual(self.match.estado, MonthlyReportingClientMatch.Estado.MATCHED)
        self.assertEqual(self.match.anet_cliente_id, 999)
        audit = MonthlyReportingClientMatchAudit.objects.get(match=self.match)
        self.assertEqual(audit.before_json["estado"], "pending")
        self.assertEqual(audit.after_json["anet_cliente_id"], 999)
        self.assertRegex(format_audit_fecha(audit), r"^\d{2}/\d{2}/\d{4}$")

    def test_pendientes_visibles_en_merge(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        batch = MonthlyReportingImportBatch.objects.create(
            pack=pack,
            file_name="t.xlsx",
            file_format="xlsx",
            file_sha256="sha-pending",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        MonthlyReportingSeedRow.objects.create(
            pack=pack,
            match=self.match,
            month=date(2026, 3, 1),
            units=Decimal("4"),
            amount=Decimal("40"),
            batch=batch,
        )

        def _no_anet(**kwargs):
            return []

        result = merge_pack_year(
            pack=pack,
            year=2026,
            month_from=1,
            month_to=6,
            base_empresa="demo",
            fetch_anet_fn=_no_anet,
        )
        self.assertEqual(len(result.rows), 1)
        self.assertTrue(result.rows[0].pending)
        self.assertEqual(len(result.pending_clients), 1)
        self.assertEqual(result.pending_clients[0]["seed_key"], "name:cliente-x")

    def test_merge_anet_homonimo_une_identidad_con_seed_pendiente(self):
        """Evita filas duplicadas tipo VARTAT: seed pendiente + ANET julio mismo nombre."""
        seed_monthly_reporting_packs(MonthlyReportingPack)
        pack = MonthlyReportingPack.objects.get(pack_id="lw_propia")
        match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:vartat-dup",
            seed_customer_name="VARTAT S.A.",
            seed_city="Córdoba",
            seed_store_type="Multibrand",
            seed_product_group="LW",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        batch = MonthlyReportingImportBatch.objects.create(
            pack=pack,
            file_name="vartat.xlsx",
            file_format="xlsx",
            file_sha256="sha-vartat",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        MonthlyReportingSeedRow.objects.create(
            pack=pack,
            match=match,
            month=date(2026, 6, 1),
            units=Decimal("28"),
            amount=Decimal("100"),
            city="Córdoba",
            store_type="Multibrand",
            batch=batch,
        )

        def _anet_julio(**kwargs):
            return [
                AnetSalesRow(
                    codigo_cliente=472,
                    nombre_cliente="vartat s.a",
                    month=date(2026, 7, 1),
                    units=Decimal("19"),
                    amount=Decimal("524263.2"),
                )
            ]

        result = merge_pack_year(
            pack=pack,
            year=2026,
            month_from=1,
            month_to=7,
            base_empresa="administranet",
            fetch_anet_fn=_anet_julio,
        )
        vartat = [r for r in result.rows if "VARTAT" in (r.display_name or "").upper()]
        identities = {r.identity for r in vartat}
        self.assertEqual(len(identities), 1, msg=identities)
        self.assertTrue(next(iter(identities)).startswith("seed:"))
        months = sorted(r.month.month for r in vartat)
        self.assertEqual(months, [6, 7])
        self.assertEqual(vartat[0].city, "Córdoba")


class MonthlyReportingSuperArtServiceTests(TestCase):
    """Phase 4.2 — catálogo SuperArt Men/Women + QA."""

    def test_clasificacion_men_women(self):
        version = MonthlyReportingSuperArtCatalogVersion.objects.create(
            version=1,
            source_label="test",
            estado=MonthlyReportingSuperArtCatalogVersion.Estado.DRAFT,
        )
        seed_catalog_entries(version, [("SA-M", "men"), ("SA-W", "women")])
        activate_catalog_version(version)
        lookup = {e.superart.upper(): e.genero for e in version.entries.all()}
        self.assertEqual(classify_superart("SA-M", lookup), "men")
        self.assertEqual(classify_superart("SA-W", lookup), "women")
        self.assertIsNone(classify_superart("DESCONOCIDO", lookup))

    def test_qa_pending_incrementa_conteo(self):
        first = register_qa_pending("ART-99", {"cliente": 1})
        second = register_qa_pending("ART-99", {"cliente": 2})
        self.assertEqual(first.id, second.id)
        self.assertEqual(MonthlyReportingSuperArtQAPending.objects.count(), 1)
        self.assertEqual(second.occurrence_count, 2)

    def test_make_classify_fn_usa_catalogo_activo(self):
        version = MonthlyReportingSuperArtCatalogVersion.objects.create(
            version=2,
            source_label="activo",
            estado=MonthlyReportingSuperArtCatalogVersion.Estado.DRAFT,
        )
        MonthlyReportingSuperArtCatalogEntry.objects.create(
            version=version,
            superart="PU-1",
            genero=MonthlyReportingSuperArtCatalogEntry.Genero.MEN,
        )
        activate_catalog_version(version)
        fn = make_classify_fn()
        self.assertEqual(fn("PU-1"), "men")
        self.assertIsNone(fn("NO-EXISTE"))


class DzPkParityTests(SimpleTestCase):
    """Phase 4.3 — paridad DZ/PK misma facturación, U.M. distinta."""

    def test_misma_facturacion_cantidades_por_factor(self):
        lineas = [
            ("P1", 12, 120),
            ("P2", 6, 60),
            ("P3", 4, 40),
            ("P6", 2, 20),
            ("CU", 1, 10),
        ]
        for unimed, qty, amt in lineas:
            _, amt_dz = compute_units_amount_for_pack(
                cantidad=qty,
                importe=amt,
                nombre_unimed=unimed,
                unit_mode="dozens",
            )
            units_pk, amt_pk = compute_units_amount_for_pack(
                cantidad=qty,
                importe=amt,
                nombre_unimed=unimed,
                unit_mode="packs",
            )
            units_dz, _ = compute_units_amount_for_pack(
                cantidad=qty,
                importe=amt,
                nombre_unimed=unimed,
                unit_mode="dozens",
            )
            self.assertEqual(amt_dz, amt_pk)
            self.assertEqual(units_pk, Decimal(str(qty)))
            factor = Decimal(str(FACTOR_DOCENAS_MAP[unimed]))
            self.assertEqual(units_dz, Decimal(str(qty)) / factor)

    def test_compare_dz_pk_sin_discrepancias_amount(self):
        identity = "anet:demo:1"
        month = date(2026, 8, 1)
        base = dict(
            identity=identity,
            display_name="C",
            match_estado="matched",
            month=month,
            amount=Decimal("100"),
        )
        dz = MergedClientMonth(**base, units=Decimal("8.333333"))
        pk = MergedClientMonth(**base, units=Decimal("100"))
        disc = compare_dz_pk_parity([dz], [pk])
        self.assertEqual(disc, {})


class VentasMensualesLicenciatariosMergerIntegrationTests(TestCase):
    """Phase 3.2 GREEN — merger con fetch ANET mockeado."""

    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:merge-test",
            seed_customer_name="Merge Test",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        self.batch = MonthlyReportingImportBatch.objects.create(
            pack=self.pack,
            file_name="m.xlsx",
            file_format="xlsx",
            file_sha256="sha-merge",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        for month in (1, 2, 7):
            MonthlyReportingSeedRow.objects.create(
                pack=self.pack,
                match=self.match,
                month=date(2026, month, 1),
                units=Decimal(str(month)),
                amount=Decimal(str(month * 10)),
                batch=self.batch,
            )

    def test_merge_julio_seed_mas_anet(self):
        def _anet_julio(**kwargs):
            if kwargs["date_from"] == date(2026, 7, 22):
                return [
                    AnetSalesRow(
                        codigo_cliente=500,
                        nombre_cliente="ANET Jul",
                        month=date(2026, 7, 1),
                        units=Decimal("9"),
                        amount=Decimal("90"),
                    )
                ]
            return []

        result = merge_pack_year(
            pack=self.pack,
            year=2026,
            month_from=1,
            month_to=7,
            base_empresa="demo",
            fetch_anet_fn=_anet_julio,
        )
        jul_rows = [r for r in result.rows if r.month == date(2026, 7, 1)]
        self.assertEqual(len(jul_rows), 2)
        seed_jul = next(r for r in jul_rows if r.source == "seed")
        anet_jul = next(r for r in jul_rows if r.source == "anet")
        self.assertEqual(seed_jul.units, Decimal("7"))
        self.assertEqual(anet_jul.units, Decimal("9"))
        ytd_seed = result.ytd_by_identity[seed_jul.identity]
        self.assertEqual(ytd_seed["units"], Decimal("1") + Decimal("2") + Decimal("7"))


class VentasMensualesLicenciatariosRunnerHybridTests(TestCase):
    """Phase 5.1 — runner híbrido seed+ANET y SuperArt QA sin bloquear."""

    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:runner-hybrid",
            seed_customer_name="Runner Hybrid",
            seed_city="CABA",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        self.batch = MonthlyReportingImportBatch.objects.create(
            pack=self.pack,
            file_name="r.xlsx",
            file_format="xlsx",
            file_sha256="sha-runner",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        MonthlyReportingSeedRow.objects.create(
            pack=self.pack,
            match=self.match,
            month=date(2026, 7, 1),
            units=Decimal("5"),
            amount=Decimal("50"),
            batch=self.batch,
        )
        self.report = ReportDefinition.objects.create(
            slug=VENTAS_MENSUALES_LICENCIATARIOS_SLUG,
            name="Ventas Mensuales Licenciatarios",
            category="operational",
            is_active=True,
        )

    def test_runner_hibrido_ytd_julio_seed_mas_anet(self):
        def _anet_julio(**kwargs):
            if kwargs["date_from"] == date(2026, 7, 22):
                return [
                    AnetSalesRow(
                        codigo_cliente=700,
                        nombre_cliente="ANET Runner",
                        month=date(2026, 7, 1),
                        units=Decimal("3"),
                        amount=Decimal("30"),
                    )
                ]
            return []

        payload = {
            "filters": {
                "pack_id": "levis_bw",
                "fecha_inicio_facturacion": "2026-07-01",
                "fecha_fin_facturacion": "2026-07-31",
            },
            "base_empresa": "demo",
        }
        result = run_ventas_mensuales_licenciatarios(
            self.report,
            payload,
            Mock(),
            fetch_anet_fn=_anet_julio,
        )
        self.assertGreater(len(result.data), 0)
        extra = result.meta["extra"]
        self.assertEqual(extra["pack_id"], "levis_bw")
        self.assertEqual(extra["year"], 2026)
        jul_rows = [r for r in result.data if r["anio_mes"] == "202607"]
        self.assertEqual(len(jul_rows), 2)
        ytd_values = {r["cliente"]: r["ytd_unidades"] for r in result.data if r["anio_mes"] == "202607"}
        self.assertEqual(ytd_values["Runner Hybrid"], 5.0)
        self.assertEqual(ytd_values["ANET Runner"], 3.0)

    def test_runner_rechaza_rango_cruza_anios(self):
        payload = {
            "filters": {
                "pack_id": "levis_bw",
                "fecha_inicio_facturacion": "2025-12-01",
                "fecha_fin_facturacion": "2026-01-31",
            },
            "base_empresa": "demo",
        }
        result = run_ventas_mensuales_licenciatarios(self.report, payload, Mock())
        self.assertEqual(result.data, [])
        self.assertTrue(any("mismo año calendario" in n.lower() for n in result.notes))

    def test_superart_desconocido_qa_sin_bloquear(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        pack_puma = MonthlyReportingPack.objects.get(pack_id="puma_bw")
        version = MonthlyReportingSuperArtCatalogVersion.objects.create(
            version=10,
            source_label="runner-qa",
            estado=MonthlyReportingSuperArtCatalogVersion.Estado.ACTIVE,
        )
        MonthlyReportingSuperArtCatalogEntry.objects.create(
            version=version,
            superart="PU-KNOWN",
            genero=MonthlyReportingSuperArtCatalogEntry.Genero.MEN,
        )

        def _anet_puma(**kwargs):
            register = kwargs.get("register_unknown_superart")
            if register:
                register("PU-UNKNOWN", {"cliente": 801})
            return [
                AnetSalesRow(
                    codigo_cliente=801,
                    nombre_cliente="Puma QA",
                    month=date(2026, 8, 1),
                    units=Decimal("2"),
                    amount=Decimal("20"),
                    superart="PU-UNKNOWN",
                )
            ]

        payload = {
            "filters": {
                "pack_id": "puma_bw",
                "fecha_inicio_facturacion": "2026-08-01",
                "fecha_fin_facturacion": "2026-08-31",
            },
            "base_empresa": "demo",
        }
        result = run_ventas_mensuales_licenciatarios(
            self.report,
            payload,
            Mock(),
            fetch_anet_fn=_anet_puma,
        )
        self.assertEqual(len(result.data), 1)
        qa = result.meta["extra"].get("qa_superarts") or []
        self.assertIn("PU-UNKNOWN", qa)
        pending = MonthlyReportingSuperArtQAPending.objects.filter(superart="PU-UNKNOWN")
        self.assertTrue(pending.exists())


class VentasMensualesLicenciatariosExportTests(TestCase):
    """Phase 5.3 — export openpyxl anual con hojas plantilla y QA."""

    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        build_all_templates(TEMPLATE_DIR)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="name:export-test",
            seed_customer_name="Export Cliente",
            seed_city="Rosario",
            seed_store_type="Store",
            seed_product_group="Bodywear",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        identity = resolve_client_identity(self.match)
        self.merge_result = merge_pack_year(
            pack=self.pack,
            year=2026,
            month_from=1,
            month_to=3,
            base_empresa="demo",
            fetch_anet_fn=lambda **kwargs: [],
        )
        self.merge_result.pending_clients = [
            match_to_aggregate_row(self.match, "demo"),
        ]
        self.merge_result.qa_superarts = ["ART-X"]
        if not self.merge_result.rows:
            self.merge_result.rows = [
                MergedClientMonth(
                    identity=identity,
                    display_name="Export Cliente",
                    match_estado="pending",
                    month=date(2026, 1, 1),
                    units=Decimal("10"),
                    amount=Decimal("100"),
                    source="seed",
                    pending=True,
                    city="Rosario",
                    store_type="Store",
                    product_group="Bodywear",
                )
            ]
            self.merge_result.ytd_by_identity = compute_ytd(self.merge_result.rows)

    def test_export_conserva_hojas_plantilla_levis(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "export_levis.xlsx"
            export_licenciatarios_workbook(
                out,
                pack=self.pack,
                merge_result=self.merge_result,
                year=2026,
                month_from=1,
                month_to=3,
            )
            wb = openpyxl.load_workbook(out, read_only=True)
            self.assertIn(SHEET_SALES, wb.sheetnames)
            self.assertIn(SHEET_MONTHLY, wb.sheetnames)
            self.assertIn(QA_SHEET, wb.sheetnames)
            wb.close()

    def test_export_puma_conserva_minimum_agreed(self):
        pack_puma = MonthlyReportingPack.objects.get(pack_id="puma_bw")
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "export_puma.xlsx"
            export_licenciatarios_workbook(
                out,
                pack=pack_puma,
                merge_result=self.merge_result,
                year=2026,
                month_from=1,
                month_to=3,
            )
            wb = openpyxl.load_workbook(out, read_only=True)
            self.assertIn(SHEET_MINIMUM, wb.sheetnames)
            wb.close()

    def test_export_lw_propia_conserva_ooh(self):
        pack_lw = MonthlyReportingPack.objects.get(pack_id="lw_propia")
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "export_lw.xlsx"
            export_licenciatarios_workbook(
                out,
                pack=pack_lw,
                merge_result=self.merge_result,
                year=2026,
                month_from=1,
                month_to=3,
            )
            wb = openpyxl.load_workbook(out, read_only=True)
            self.assertIn(SHEET_OOH, wb.sheetnames)
            self.assertIn(SHEET_MINIMUM, wb.sheetnames)
            wb.close()

    def test_export_input_sales_y_qa_poblados(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "export_data.xlsx"
            export_licenciatarios_workbook(
                out,
                pack=self.pack,
                merge_result=self.merge_result,
                year=2026,
                month_from=1,
                month_to=3,
            )
            wb = openpyxl.load_workbook(out)
            sales = wb[SHEET_SALES]
            self.assertEqual(sales.cell(row=5, column=1).value, "Export Cliente")
            self.assertEqual(sales.cell(row=4, column=2).value, "City / Province")
            self.assertEqual(sales.cell(row=4, column=3).value, "Store Type")
            self.assertEqual(sales.cell(row=4, column=4).value, "Product group")
            self.assertEqual(sales.cell(row=5, column=2).value, "Rosario")
            self.assertEqual(sales.cell(row=5, column=3).value, "Store")
            self.assertEqual(sales.cell(row=5, column=4).value, "Bodywear")
            self.assertEqual(float(sales.cell(row=5, column=5).value or 0), 10.0)
            self.assertTrue(str(sales.cell(row=2, column=5).value or "").startswith("=SUM(E5:"))
            self.assertTrue(str(sales.cell(row=2, column=6).value or "").startswith("=SUM(F5:"))
            monthly = wb[SHEET_MONTHLY]
            self.assertNotIn("Unidades YTD", str(monthly["C20"].value or ""))
            self.assertIn("input Licensee sales", str(monthly["D4"].value or ""))
            qa = wb[QA_SHEET]
            self.assertEqual(qa.cell(row=2, column=1).value, "ART-X")
            self.assertIn("pending", str(qa.cell(row=3, column=4).value).lower())
            wb.close()

    def test_resolve_template_path_existe(self):
        path = resolve_template_path("levis_bw")
        self.assertTrue(path.exists())


class LicenciatariosApiTests(TestCase):
    """Phase 6.1 — API permisos, match apply/undo y rango calendario."""

    def setUp(self):
        self.match = MonthlyReportingClientMatch.objects.create(
            seed_key="api:cliente-a",
            seed_customer_name="Cliente API A",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        self.matched = MonthlyReportingClientMatch.objects.create(
            seed_key="api:cliente-b",
            seed_customer_name="Cliente API B",
            estado=MonthlyReportingClientMatch.Estado.MATCHED,
            anet_cliente_id=555,
            base_empresa="emp_test",
        )

    def _user(self, *, supervisor=False, operational=True):
        user = Mock()
        user.is_authenticated = True
        user.is_superuser = False
        user.is_admin = Mock(return_value=False)
        user.cod_usuario = "supervisor" if supervisor else "vendedor"
        user.tiene_permiso = lambda p: operational and p == "reports.view_operational"
        return user

    def _session(self):
        return {"user": {"base_empresa": "emp_test", "id_usuario": 7, "nombre": "Tester"}}

    @patch("reports.ventas_mensuales_licenciatarios_api_views.user_has_full_access", return_value=False)
    def test_get_matches_requiere_operational_y_sin_edicion(self, _full):
        from rest_framework.test import APIRequestFactory, force_authenticate

        from reports.ventas_mensuales_licenciatarios_api_views import (
            LicenciatariosClientMatchesListAPIView,
        )

        factory = APIRequestFactory()
        request = factory.get("/api/reports/licenciatarios/client-matches/")
        force_authenticate(request, user=self._user())
        request.session = self._session()
        response = LicenciatariosClientMatchesListAPIView.as_view()(request)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["can_edit"])
        self.assertGreaterEqual(response.data["pending_count"], 1)

    def test_get_sin_permiso_operational_403(self):
        from rest_framework.test import APIRequestFactory, force_authenticate

        from reports.ventas_mensuales_licenciatarios_api_views import (
            LicenciatariosClientMatchesListAPIView,
        )

        factory = APIRequestFactory()
        request = factory.get("/api/reports/licenciatarios/client-matches/")
        force_authenticate(request, user=self._user(operational=False))
        request.session = self._session()
        response = LicenciatariosClientMatchesListAPIView.as_view()(request)
        self.assertEqual(response.status_code, 403)

    @patch("reports.ventas_mensuales_licenciatarios_api_views.user_has_full_access", return_value=False)
    def test_patch_apply_rechaza_sin_full_access(self, _full):
        from rest_framework.test import APIRequestFactory, force_authenticate

        from reports.ventas_mensuales_licenciatarios_api_views import (
            LicenciatariosClientMatchDetailAPIView,
        )

        factory = APIRequestFactory()
        request = factory.patch(
            f"/api/reports/licenciatarios/client-matches/{self.match.id}/",
            {"action": "apply", "anet_cliente_id": 100},
            format="json",
        )
        force_authenticate(request, user=self._user())
        request.session = self._session()
        response = LicenciatariosClientMatchDetailAPIView.as_view()(request, match_id=self.match.id)
        self.assertEqual(response.status_code, 403)

    @patch("reports.ventas_mensuales_licenciatarios_api_views.user_has_full_access", return_value=True)
    def test_patch_apply_y_undo_supervisor(self, _full):
        from rest_framework.test import APIRequestFactory, force_authenticate

        from reports.ventas_mensuales_licenciatarios_api_views import (
            LicenciatariosClientMatchDetailAPIView,
        )

        factory = APIRequestFactory()
        request = factory.patch(
            f"/api/reports/licenciatarios/client-matches/{self.match.id}/",
            {"action": "apply", "anet_cliente_id": 1001},
            format="json",
        )
        force_authenticate(request, user=self._user(supervisor=True))
        request.session = self._session()
        response = LicenciatariosClientMatchDetailAPIView.as_view()(request, match_id=self.match.id)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["estado"], "matched")
        self.assertRegex(response.data["updated_at_display"], r"^\d{2}/\d{2}/\d{4}$")

        request_undo = factory.patch(
            f"/api/reports/licenciatarios/client-matches/{self.match.id}/",
            {"action": "undo"},
            format="json",
        )
        force_authenticate(request_undo, user=self._user(supervisor=True))
        request_undo.session = self._session()
        response_undo = LicenciatariosClientMatchDetailAPIView.as_view()(
            request_undo, match_id=self.match.id
        )
        self.assertEqual(response_undo.status_code, 200)
        self.assertEqual(response_undo.data["estado"], "pending")
        self.assertIsNone(response_undo.data["anet_cliente_id"])

    def test_runner_rechaza_rango_cruza_anios(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        report = ReportDefinition(slug=VENTAS_MENSUALES_LICENCIATARIOS_SLUG, name="Licenciatarios")
        result = run_ventas_mensuales_licenciatarios(
            report,
            {
                "filters": {
                    "pack_id": "levis_bw",
                    "fecha_inicio_facturacion": "2025-10-01",
                    "fecha_fin_facturacion": "2026-02-15",
                    "base_empresa": "emp_test",
                }
            },
            Mock(base_empresa="emp_test"),
        )
        self.assertEqual(result.data, [])
        self.assertTrue(any("mismo año calendario" in str(n).lower() for n in result.notes))


class LicenciatariosUiContractTests(TestCase):
    """Phase 6.3 — modal Synap, fechas dd/MM/yyyy, pendientes en meta."""

    def test_js_sin_dialogos_nativos(self):
        js_path = (
            Path(__file__).resolve().parents[1]
            / "static"
            / "reports"
            / "js"
            / "ventas_mensuales_licenciatarios.js"
        )
        content = js_path.read_text(encoding="utf-8")
        self.assertNotIn("window.alert", content)
        self.assertNotIn("window.confirm", content)
        self.assertNotIn("window.prompt", content)
        self.assertNotIn("alert(", content)
        self.assertNotIn("confirm(", content)

    def test_dashboard_incluye_contenedor_matriz(self):
        tpl_path = (
            Path(__file__).resolve().parents[1]
            / "templates"
            / "reports"
            / "dashboard_detail.html"
        )
        content = tpl_path.read_text(encoding="utf-8")
        self.assertIn('id="vml-matriz-container"', content)
        self.assertIn('id="vml-page"', content)
        self.assertIn("body:has(#vml-page)", content)
        vml_idx = content.find('id="vml-matriz-container"')
        window = content[max(0, vml_idx - 500) : vml_idx]
        self.assertIn("min-h-0 flex-1", window)
        self.assertNotIn("h-[min(75vh,56rem)]", window)

    def test_dashboard_incluye_boton_exportar_excel(self):
        tpl_path = (
            Path(__file__).resolve().parents[1]
            / "templates"
            / "reports"
            / "dashboard_detail.html"
        )
        content = tpl_path.read_text(encoding="utf-8")
        export_idx = content.find("data-export-excel")
        self.assertGreater(export_idx, 0)
        pre = content[max(0, export_idx - 900) : export_idx]
        self.assertIn("ventas-mensuales-licenciatarios", pre)

    def test_js_pinta_matriz_cliente_mes(self):
        js_path = (
            Path(__file__).resolve().parents[1]
            / "static"
            / "reports"
            / "js"
            / "ventas_mensuales_licenciatarios.js"
        )
        content = js_path.read_text(encoding="utf-8")
        self.assertIn("function renderMatriz", content)
        self.assertIn("vml-matriz-container", content)
        self.assertIn("pivotClientMonths", content)

    def test_dashboard_js_exporta_vml_con_pack_obligatorio(self):
        js_path = (
            Path(__file__).resolve().parents[1]
            / "static"
            / "reports"
            / "js"
            / "dashboard.js"
        )
        content = js_path.read_text(encoding="utf-8")
        self.assertIn('reportSlug === "ventas-mensuales-licenciatarios"', content)
        self.assertIn("Seleccioná un pack licenciatario antes de exportar.", content)
        self.assertIn("vml_pack_id", content)
        self.assertIn("clientes_excluidos", content)
        self.assertIn("isVentasMensualesLicenciatariosSlug(reportSlug)", content)

    def test_modal_template_sin_dialogos_nativos(self):
        tpl_path = (
            Path(__file__).resolve().parents[1]
            / "templates"
            / "reports"
            / "includes"
            / "modal_licenciatarios_match.html"
        )
        content = tpl_path.read_text(encoding="utf-8")
        self.assertIn('role="dialog"', content)
        self.assertNotIn("onclick=\"confirm", content.lower())

    @patch("reports.ventas_mensuales_licenciatarios_api_views.user_has_full_access", return_value=True)
    def test_api_auditoria_fecha_dd_mm_yyyy(self, _full):
        from rest_framework.test import APIRequestFactory, force_authenticate

        from reports.ventas_mensuales_licenciatarios_api_views import (
            LicenciatariosClientMatchDetailAPIView,
            serialize_client_match,
        )

        match = MonthlyReportingClientMatch.objects.create(
            seed_key="ui:fecha",
            seed_customer_name="Cliente Fecha",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        factory = APIRequestFactory()
        request = factory.patch(
            f"/api/reports/licenciatarios/client-matches/{match.id}/",
            {"action": "apply", "anet_cliente_id": 42, "base_empresa": "emp_test"},
            format="json",
        )
        force_authenticate(request, user=Mock(is_authenticated=True, cod_usuario="supervisor", nombre="Supervisor"))
        request.session = {"user": {"base_empresa": "emp_test", "id_usuario": 1, "nombre": "Supervisor"}}
        response = LicenciatariosClientMatchDetailAPIView.as_view()(request, match_id=match.id)
        self.assertEqual(response.status_code, 200)
        self.assertRegex(response.data["updated_at_display"], r"^\d{2}/\d{2}/\d{4}$")
        payload = serialize_client_match(match)
        self.assertRegex(payload["updated_at_display"], r"^\d{2}/\d{2}/\d{4}$")

    def test_runner_pending_clients_en_meta_extra(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        pending_match = MonthlyReportingClientMatch.objects.create(
            seed_key="ui:pendiente",
            seed_customer_name="Pendiente UI",
            estado=MonthlyReportingClientMatch.Estado.PENDING,
        )
        batch = MonthlyReportingImportBatch.objects.create(
            pack=pack,
            file_name="ui.xlsx",
            file_format="xlsx",
            file_sha256="sha-ui-pending",
            estado=MonthlyReportingImportBatch.Estado.APPLIED,
        )
        MonthlyReportingSeedRow.objects.create(
            pack=pack,
            match=pending_match,
            month=date(2026, 2, 1),
            units=Decimal("3"),
            amount=Decimal("30"),
            batch=batch,
        )
        report = ReportDefinition(slug=VENTAS_MENSUALES_LICENCIATARIOS_SLUG, name="Licenciatarios")

        def _no_anet(**kwargs):
            return []

        result = run_ventas_mensuales_licenciatarios(
            report,
            {
                "filters": {
                    "pack_id": "levis_bw",
                    "fecha_inicio_facturacion": "2026-01-01",
                    "fecha_fin_facturacion": "2026-03-31",
                    "base_empresa": "emp_test",
                }
            },
            Mock(base_empresa="emp_test"),
            fetch_anet_fn=_no_anet,
        )
        extra = result.meta.get("extra") or {}
        self.assertGreaterEqual(len(extra.get("pending_clients") or []), 1)
        self.assertTrue(any(row.get("pendiente") for row in result.data))


class MonthlyReportingReconciliationTests(TestCase):
    """Phase 7.1 — conciliación planilla vs seed PostgreSQL."""

    def setUp(self):
        seed_monthly_reporting_packs(MonthlyReportingPack)
        self.pack = MonthlyReportingPack.objects.get(pack_id="levis_bw")
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.file_path = _build_levis_seed_xlsx(Path(self.tempdir.name) / "seed.xlsx")

    def test_aggregate_parsed_cells_suma_mismo_cliente_mes(self):
        cells = [
            ParsedSeedCell(
                seed_key="name:abc",
                customer_name="Cliente",
                month=date(2026, 1, 1),
                units=Decimal("10"),
                amount=Decimal("100"),
            ),
            ParsedSeedCell(
                seed_key="name:abc",
                customer_name="Cliente",
                month=date(2026, 1, 1),
                units=Decimal("2"),
                amount=Decimal("50"),
            ),
        ]
        agg = aggregate_parsed_cells(cells)
        row = agg[("name:abc", date(2026, 1, 1))]
        self.assertEqual(row.units, Decimal("12"))
        self.assertEqual(row.amount, Decimal("150"))

    def test_reconcile_coincide_tras_import(self):
        import_monthly_reporting_file("levis_bw", self.file_path)
        result = reconcile_pack_from_file("levis_bw", self.file_path, year=2026, through_month=6)
        self.assertTrue(result.file_accessible)
        self.assertEqual(result.coincidencias, 1)
        self.assertEqual(result.discrepancias, [])
        self.assertEqual(len(result.ytd_file), 1)
        self.assertEqual(result.ytd_file[0].amount, Decimal("1500.75"))

    def test_reconcile_detecta_amount_distinto(self):
        import_monthly_reporting_file("levis_bw", self.file_path)
        row = MonthlyReportingSeedRow.objects.get()
        row.amount = Decimal("999.00")
        row.save(update_fields=["amount"])
        result = reconcile_pack_from_file("levis_bw", self.file_path, year=2026, through_month=6)
        self.assertEqual(result.coincidencias, 0)
        kinds = {m.kind for m in result.discrepancias}
        self.assertIn("amount", kinds)

    def test_reconcile_missing_in_db(self):
        result = reconcile_pack_from_file("levis_bw", self.file_path, year=2026, through_month=6)
        self.assertEqual(result.coincidencias, 0)
        self.assertTrue(any(m.kind == "missing_in_db" for m in result.discrepancias))

    def test_compute_ytd_suma_enero_a_junio(self):
        cells = [
            ParsedSeedCell(
                seed_key="k1",
                customer_name="A",
                month=date(2026, 1, 1),
                units=Decimal("1"),
                amount=Decimal("10"),
            ),
            ParsedSeedCell(
                seed_key="k1",
                customer_name="A",
                month=date(2026, 6, 1),
                units=Decimal("2"),
                amount=Decimal("20"),
            ),
        ]
        agg = aggregate_parsed_cells(cells)
        ytd = compute_ytd_from_aggregates(agg, year=2026, through_month=6)
        self.assertEqual(len(ytd), 1)
        self.assertEqual(ytd[0].units, Decimal("3"))
        self.assertEqual(ytd[0].amount, Decimal("30"))

    def test_fa_nc_reference_incluye_tipos_vmm(self):
        self.assertIn("FA", FA_NC_REFERENCE_NOTE)
        self.assertIn("NCA", FA_NC_REFERENCE_NOTE)
        self.assertTrue(TIPOS_FAC)
        self.assertTrue(TIPOS_NC)

    def test_resolve_pack_source_path_mapea_levis_bw(self):
        path = resolve_pack_source_path("levis_bw", Path("/tmp/fwdreportesjun"))
        self.assertTrue(str(path).endswith("Monthly Reporting Best Sox_LEVIS BW 26.xlsx"))

    def test_compare_seed_aggregates_db_vs_file(self):
        import_monthly_reporting_file("levis_bw", self.file_path)
        file_agg = aggregate_parsed_cells(
            [
                ParsedSeedCell(
                    seed_key="x",
                    customer_name="Demo",
                    month=date(2026, 3, 1),
                    units=Decimal("5"),
                    amount=Decimal("500"),
                )
            ]
        )
        db_agg = aggregate_db_seed_rows("levis_bw", year=2026)
        coincidencias, discrepancias = compare_seed_aggregates(file_agg, db_agg)
        self.assertEqual(coincidencias, 0)
        self.assertGreater(len(discrepancias), 0)

