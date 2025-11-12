"""
Servicio de Exportación de Reportes a Excel y PDF
Con templates corporativos
"""
import os
import json
from typing import Dict, List, Optional
from django.core.files.base import ContentFile
from django.utils import timezone
from django.conf import settings

from ..models import ReportRequest, ReportExport, ChatMessage
from core.models.models import Empresa


class ReportExportService:
    """
    Servicio para exportar reportes a Excel y PDF con formato profesional
    """
    
    def __init__(self):
        self.media_root = settings.MEDIA_ROOT
    
    def export_to_excel(
        self,
        report_request: ReportRequest,
        user,
        empresa: Optional[Empresa] = None,
        chat_message: Optional[ChatMessage] = None
    ) -> ReportExport:
        """
        Exporta reporte a Excel con formato corporativo
        
        Args:
            report_request: Solicitud de reporte a exportar
            user: Usuario que solicita la exportación
            empresa: Empresa para branding (opcional)
            chat_message: Mensaje de chat asociado (opcional)
        
        Returns:
            ReportExport con el archivo generado
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Obtener empresa del usuario si no se especificó
        if not empresa:
            empresa = user.empresa if hasattr(user, 'empresa') else None
        
        # 1. Header corporativo
        row = 1
        if empresa:
            ws[f'A{row}'] = empresa.nombre
            ws[f'A{row}'].font = Font(size=16, bold=True, color="1E40AF")
            row += 1
        
        ws[f'A{row}'] = f"Reporte: {report_request.query}"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 2
        
        # 2. Datos del reporte
        try:
            data = json.loads(report_request.result_data) if isinstance(report_request.result_data, str) else report_request.result_data
        except:
            data = []
        
        if data and len(data) > 0:
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    bottom=Side(style='thin', color='000000')
                )
            
            row += 1
            
            # Datos
            for row_data in data:
                for col, header in enumerate(headers, start=1):
                    value = row_data.get(header, '')
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Formatear números
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                
                row += 1
            
            # Ajustar anchos de columna
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = len(str(headers[col-1])) + 2
                
                for cell in ws[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # 3. Footer
        row += 2
        ws[f'A{row}'] = f"Total de registros: {len(data) if data else 0}"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = f"Synap Reports AI • {empresa.nombre if empresa else 'administraNET'}"
        ws[f'A{row}'].font = Font(size=8, italic=True, color="6B7280")
        
        # 4. Guardar archivo
        filename = f"reporte_{report_request.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('reports_ai', 'exports', timezone.now().strftime('%Y/%m'), filename)
        full_path = os.path.join(self.media_root, filepath)
        
        # Crear directorio si no existe
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        # Guardar
        wb.save(full_path)
        file_size = os.path.getsize(full_path)
        
        # Crear registro de exportación
        export = ReportExport.objects.create(
            report_request=report_request,
            chat_message=chat_message,
            user=user,
            format='excel',
            file=filepath,
            filename=filename,
            file_size=file_size,
            template_used='corporate_standard',
            export_options={
                'include_header': True,
                'include_footer': True,
                'empresa_nombre': empresa.nombre if empresa else None
            }
        )
        
        return export
    
    def export_to_pdf(
        self,
        report_request: ReportRequest,
        user,
        empresa: Optional[Empresa] = None,
        chat_message: Optional[ChatMessage] = None
    ) -> ReportExport:
        """
        Exporta reporte a PDF con template corporativo
        
        Args:
            report_request: Solicitud de reporte a exportar
            user: Usuario que solicita la exportación
            empresa: Empresa para branding (opcional)
            chat_message: Mensaje de chat asociado (opcional)
        
        Returns:
            ReportExport con el archivo generado
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        
        # Obtener empresa
        if not empresa:
            empresa = user.empresa if hasattr(user, 'empresa') else None
        
        # Crear archivo
        filename = f"reporte_{report_request.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join('reports_ai', 'exports', timezone.now().strftime('%Y/%m'), filename)
        full_path = os.path.join(self.media_root, filepath)
        
        # Crear directorio si no existe
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        # Crear PDF
        doc = SimpleDocTemplate(
            full_path,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # 1. Logo (si existe)
        if empresa and empresa.logo:
            try:
                logo_path = empresa.logo.path
                logo = Image(logo_path, width=3*inch, height=1.5*inch)
                logo.hAlign = 'CENTER'
                story.append(logo)
                story.append(Spacer(1, 0.3*inch))
            except:
                pass
        
        # 2. Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1E40AF'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        if empresa:
            story.append(Paragraph(empresa.nombre, title_style))
        
        # 3. Info del reporte
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#374151'),
            alignment=TA_CENTER,
            spaceAfter=6
        )
        story.append(Paragraph(f"<b>{report_request.query}</b>", subtitle_style))
        
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        story.append(Paragraph(
            f"Generado: {timezone.now().strftime('%d/%m/%Y a las %H:%M')}",
            info_style
        ))
        
        story.append(Spacer(1, 0.3*inch))
        
        # 4. Datos
        try:
            data = json.loads(report_request.result_data) if isinstance(report_request.result_data, str) else report_request.result_data
        except:
            data = []
        
        if data and len(data) > 0:
            # Preparar tabla
            headers = list(data[0].keys())
            table_data = [headers]
            
            # Limitar a 100 filas para PDF
            for row in data[:100]:
                table_data.append([str(row.get(h, '')) for h in headers])
            
            # Crear tabla
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Datos
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                
                # Alternado de filas
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
            ]))
            
            story.append(t)
            
            # Nota si hay más datos
            if len(data) > 100:
                story.append(Spacer(1, 0.2*inch))
                note_style = ParagraphStyle(
                    'Note',
                    parent=styles['Normal'],
                    fontSize=8,
                    textColor=colors.HexColor('#EF4444'),
                    alignment=TA_CENTER
                )
                story.append(Paragraph(
                    f"Nota: Se muestran los primeros 100 registros de {len(data)} totales",
                    note_style
                ))
        
        # 5. Footer
        story.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.HexColor('#9CA3AF'),
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f"Generado por Synap Reports AI • {empresa.nombre if empresa else 'administraNET'}",
            footer_style
        ))
        
        # Construir PDF
        doc.build(story)
        
        file_size = os.path.getsize(full_path)
        
        # Crear registro de exportación
        export = ReportExport.objects.create(
            report_request=report_request,
            chat_message=chat_message,
            user=user,
            format='pdf',
            file=filepath,
            filename=filename,
            file_size=file_size,
            template_used='corporate_standard',
            export_options={
                'include_logo': bool(empresa and empresa.logo),
                'include_header': True,
                'include_footer': True,
                'empresa_nombre': empresa.nombre if empresa else None,
                'max_rows': 100
            }
        )
        
        return export
    
    def get_download_url(self, export: ReportExport) -> str:
        """
        Obtiene la URL de descarga del export
        """
        return f"{settings.MEDIA_URL}{export.file.name}"

