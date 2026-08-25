"""Tests importación Excel pedido masivo (matriz + replace + VCM)."""

from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.test import TestCase
from openpyxl import Workbook
from rest_framework.test import APIRequestFactory, force_authenticate

from ecom.models import EcomPedidoMasivoDraft, EcomPedidoMasivoDraftCelda
from ecom.pedido_masivo_views import (
    PedidoMasivoImportarAPIView,
    PedidoMasivoPlantillaExcelAPIView,
)
from ecom.services.pedido_masivo_import import (
    HOJA_META,
    MARKER_CODIGO,
    MARKER_IDART,
    generar_plantilla_excel,
    importar_matriz_excel,
)


class _User:
    is_authenticated = True
    is_superuser = True

    def tiene_permiso(self, _c):
        return True


SUC_A = {
    "id_cliente_domicilio": 10,
    "nro": "127",
    "calle": "MORÓN - 25 de Mayo",
    "etiqueta": "MORÓN - 25 de Mayo 127",
    "nombre": "MORÓN - 25 de Mayo 127",
}
SUC_B = {
    "id_cliente_domicilio": 20,
    "nro": "736",
    "calle": "MERLO - Av. Libertador",
    "etiqueta": "MERLO - Av. Libertador 736",
    "nombre": "MERLO - Av. Libertador 736",
}

ART_OK = {
    "id_articulo": 101,
    "id_manual": "2401",
    "nombre": "Media pack",
    "codigo_marca": 5,
    "cod_art_prov": "902401-02",
    "codigo_t": "",
    "discontinuo": "No",
    "ecommerce": "Si",
    "tipo_art_fab": "Terminado",
    "multiplo_cantidad_vta": 6,
}


def _xlsx_plantilla(
    ids_suc, qtys_por_codigo, id_cliente=368, cod_viajante=30, nombres=None
):
    """ids_suc: lista id domicilio. qtys: {codigo: [q1, q2, ...]}."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"
    ws.cell(1, 1, MARKER_CODIGO)
    ws.cell(1, 2, id_cliente)
    ws.cell(2, 1, "Código")
    ws.cell(2, 2, "Artículo")
    for i, idd in enumerate(ids_suc):
        ws.cell(1, 3 + i, idd)
        ws.cell(2, 3 + i, f"Suc {idd}")
    fila = 3
    for codigo, qtys in qtys_por_codigo.items():
        ws.cell(fila, 1, codigo)
        ws.cell(fila, 2, (nombres or {}).get(codigo, ""))
        for i, q in enumerate(qtys):
            if q is not None:
                ws.cell(fila, 3 + i, q)
        fila += 1
    ws_m = wb.create_sheet(HOJA_META)
    ws_m.append(["id_cliente", id_cliente])
    ws_m.append(["cod_viajante", cod_viajante])
    ws_m.append(["plantilla_version", 2])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


SUC_HEADER_V5 = {
    10: "127\nMORÓN - 25 de Mayo",
    20: "736\nMERLO - Av. Libertador",
}


def _xlsx_v5(ids_suc, filas, id_cliente=368, cod_viajante=30):
    """filas: {codigo: {"id_articulo": int, "nombre": str, "precio": n|None, "qtys": [...]}}."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"
    ws.cell(1, 1, "Código")
    ws.cell(1, 2, "Artículo")
    ws.cell(1, 3, MARKER_IDART)
    ws.cell(1, 4, "Precio")
    for i, idd in enumerate(ids_suc):
        ws.cell(1, 5 + i, SUC_HEADER_V5.get(idd, f"{idd}\nSuc {idd}"))
    fila = 2
    for codigo, data in filas.items():
        ws.cell(fila, 1, codigo)
        ws.cell(fila, 2, data.get("nombre") or "x")
        ws.cell(fila, 3, data["id_articulo"])
        if data.get("precio") is not None:
            ws.cell(fila, 4, data["precio"])
        for i, q in enumerate(data.get("qtys") or []):
            if q is not None:
                ws.cell(fila, 5 + i, q)
        fila += 1
    ws_m = wb.create_sheet(HOJA_META)
    ws_m.append(["id_cliente", id_cliente])
    ws_m.append(["cod_viajante", cod_viajante])
    ws_m.append(["draft_id", 1])
    ws_m.append(["plantilla_version", 5])
    ws_m.append(["sucursal_ids", *ids_suc])
    ws_m.append(["col_primera_sucursal", 5])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _draft(**kwargs):
    defaults = dict(
        base_empresa="emp_m",
        id_usuario=1,
        id_cliente=368,
        cod_viajante=30,
        estado=EcomPedidoMasivoDraft.ESTADO_BORRADOR,
        modo=EcomPedidoMasivoDraft.MODO_MASIVO,
    )
    defaults.update(kwargs)
    return EcomPedidoMasivoDraft.objects.create(**defaults)


class TestImportarMatrizExcel(TestCase):
    def setUp(self):
        self.p_suc = patch(
            "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
            return_value=[SUC_A, SUC_B],
        )
        self.p_marcas = patch(
            "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
            side_effect=lambda *a, **k: (
                [5] if k.get("id_cliente_domicilio") == 10 else [5, 8]
            ),
        )
        self.p_desc = patch(
            "ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo"
        )
        self.p_precio = patch(
            "ecom.services.pedido_masivo_import.asegurar_precio_fila_articulo"
        )
        self.p_pie = patch(
            "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
            return_value={
                "descRenglon": Decimal("0"),
                "descPie": Decimal("5"),
                "lista_id": 1,
            },
        )
        self.p_nom = patch(
            "ecom.services.pedido_masivo_import._nombre_cliente",
            return_value="Dabra S.A.",
        )
        self.p_suc.start()
        self.p_marcas.start()
        self.p_desc.start()
        self.p_precio.start()
        self.p_pie.start()
        self.p_nom.start()

    def tearDown(self):
        self.p_suc.stop()
        self.p_marcas.stop()
        self.p_desc.stop()
        self.p_precio.stop()
        self.p_pie.stop()
        self.p_nom.stop()

    def _lookup(self, _base, codigos):
        out = {}
        for c in codigos:
            if c in ("2401", "902401-02"):
                out[c] = [dict(ART_OK)]
            else:
                out[c] = []
        return out

    def test_importa_y_reemplaza_celdas(self):
        d = _draft()
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=999, id_cliente_domicilio=10, cantidad_packs=Decimal("99")
        )
        raw = _xlsx_plantilla([10, 20], {"2401": [12, 6]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.count(), 2)
        self.assertFalse(d.celdas.filter(id_articulo=999).exists())
        self.assertEqual(
            d.celdas.get(id_articulo=101, id_cliente_domicilio=10).cantidad_packs,
            Decimal("12"),
        )
        self.assertEqual(
            d.celdas.get(id_articulo=101, id_cliente_domicilio=20).cantidad_packs,
            Decimal("6"),
        )
        d.refresh_from_db()
        self.assertEqual(d.descuento_pie_pct, Decimal("5"))

    def _lookup_ids(self, _base, ids):
        out = {}
        for i in ids:
            if int(i) == 101:
                out[int(i)] = dict(ART_OK)
        return out

    def test_importa_v5_cantidad_y_precio(self):
        d = _draft()
        raw = _xlsx_v5(
            [10, 20],
            {
                "2401": {
                    "id_articulo": 101,
                    "nombre": "Media pack",
                    "precio": 125.5,
                    "qtys": [12, 6],
                }
            },
        )
        res = importar_matriz_excel(
            d, raw, consultar_arts=self._lookup, consultar_ids=self._lookup_ids
        )
        self.assertTrue(res["ok"], res)
        d.refresh_from_db()
        self.assertEqual(d.precios_fila.get("101"), 125.5)
        self.assertEqual(
            d.celdas.get(id_articulo=101, id_cliente_domicilio=10).cantidad_packs,
            Decimal("12"),
        )

    def test_rechaza_v5_precio_cero(self):
        d = _draft()
        raw = _xlsx_v5(
            [10],
            {
                "2401": {
                    "id_articulo": 101,
                    "nombre": "Media pack",
                    "precio": 0,
                    "qtys": [6],
                }
            },
        )
        res = importar_matriz_excel(
            d, raw, consultar_arts=self._lookup, consultar_ids=self._lookup_ids
        )
        self.assertFalse(res["ok"])
        self.assertEqual(res["errores"][0]["code"], "precio_cero")
        self.assertEqual(d.celdas.count(), 0)

    def test_rechaza_sucursal_fuera_territorio(self):
        d = _draft()
        raw = _xlsx_plantilla([10, 99], {"2401": [6, 6]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        codes = {e["code"] for e in res["errores"]}
        self.assertIn("sucursal_fuera_territorio", codes)
        self.assertEqual(d.celdas.count(), 0)

    def test_rechaza_marca_fuera_territorio(self):
        d = _draft()
        art_otra = dict(ART_OK, codigo_marca=99)

        def lookup(_b, codigos):
            return {c: [art_otra] for c in codigos}

        raw = _xlsx_plantilla([10], {"2401": [6]})
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["errores"][0]["code"], "marca_fuera_territorio")
        self.assertEqual(d.celdas.count(), 0)

    def test_rechaza_articulo_inexistente(self):
        d = _draft()
        raw = _xlsx_plantilla([10], {"NO-EXISTE": [6]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["errores"][0]["code"], "articulo_no_encontrado")

    def test_rechaza_multiplo(self):
        d = _draft()
        raw = _xlsx_plantilla([10], {"2401": [7]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["errores"][0]["code"], "multiplo_empaque")
        self.assertEqual(d.celdas.count(), 0)

    def test_all_or_nothing_conserva_celdas_previas(self):
        d = _draft()
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=5, id_cliente_domicilio=10, cantidad_packs=Decimal("3")
        )
        raw = _xlsx_plantilla([10], {"2401": [7]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(d.celdas.count(), 1)
        self.assertEqual(d.celdas.get().id_articulo, 5)

    def test_replace_completa_descuento_pie_desde_cliente(self):
        d = _draft(descuento_pie_pct=Decimal("7"))
        raw = _xlsx_plantilla([10], {"2401": [6]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertTrue(res["ok"], res)
        d.refresh_from_db()
        self.assertEqual(d.descuento_pie_pct, Decimal("5"))

    def test_error_conserva_descuento_pie_previo(self):
        d = _draft(descuento_pie_pct=Decimal("7"))
        raw = _xlsx_plantilla([10], {"2401": [7]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        d.refresh_from_db()
        self.assertEqual(d.descuento_pie_pct, Decimal("7"))

    def test_rechaza_cliente_distinto(self):
        d = _draft()
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=5, id_cliente_domicilio=10, cantidad_packs=Decimal("3")
        )
        raw = _xlsx_plantilla([10], {"2401": [6]}, id_cliente=999)
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["errores"][0]["code"], "cliente_no_coincide")
        self.assertEqual(d.celdas.count(), 1)

    def test_rechaza_vendedor_distinto(self):
        d = _draft()
        raw = _xlsx_plantilla([10], {"2401": [6]}, cod_viajante=99)
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertTrue(any(e["code"] == "vendedor_no_coincide" for e in res["errores"]))

    def test_rechaza_sin_cantidades(self):
        d = _draft()
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=5, id_cliente_domicilio=10, cantidad_packs=Decimal("3")
        )
        raw = _xlsx_plantilla([10], {"2401": [None]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["code"], "sin_cantidades")
        self.assertEqual(d.celdas.count(), 1)

    def test_rechaza_plantilla_sin_cliente(self):
        d = _draft()
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedido"
        ws.cell(1, 1, MARKER_CODIGO)
        ws.cell(1, 2, "nombre")
        ws.cell(1, 3, 10)
        ws.cell(2, 1, "Código")
        ws.cell(2, 3, "Suc 10")
        ws.cell(3, 1, "2401")
        ws.cell(3, 3, 6)
        bio = BytesIO()
        wb.save(bio)
        res = importar_matriz_excel(d, bio.getvalue(), consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertTrue(any(e["code"] == "plantilla_sin_cliente" for e in res["errores"]))

    def test_modo_simple_solo_domicilio_fijo(self):
        d = _draft(
            modo=EcomPedidoMasivoDraft.MODO_SIMPLE,
            id_domicilio_fijo=10,
        )
        raw = _xlsx_plantilla([10, 20], {"2401": [6, 6]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertTrue(
            any(e["code"] == "sucursal_fuera_territorio" for e in res["errores"])
        )

    def test_columna_extra_vacia_fuera_territorio(self):
        d = _draft()
        raw = _xlsx_plantilla([10, 99], {"2401": [6, None]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertTrue(
            any(e["code"] == "sucursal_fuera_territorio" for e in res["errores"])
        )
        self.assertEqual(d.celdas.count(), 0)

    def test_acepta_codigo_barras(self):
        d = _draft()

        def lookup(_b, codigos):
            out = {}
            for c in codigos:
                if c in ("2401", "7790001112223"):
                    out[c] = [dict(ART_OK)]
                else:
                    out[c] = []
            return out

        raw = _xlsx_plantilla([10], {"7790001112223": [6]})
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.get().id_articulo, 101)

    def test_superart_ambiguo_se_desambigua_por_nombre(self):
        d = _draft()
        art_a = dict(
            ART_OK,
            id_articulo=203,
            id_manual="3766",
            nombre="3766 T5 Puma Lifestyle Sock CAI Blanco Logo Negro 1P",
        )
        art_b = dict(
            ART_OK,
            id_articulo=204,
            id_manual="3766",
            nombre="3766 T5 Puma Lifestyle Sock CAI Negro 1P",
        )

        def lookup(_b, codigos):
            return {c: [art_a, art_b] for c in codigos}

        raw = _xlsx_plantilla(
            [10],
            {"3766": [6]},
            nombres={"3766": art_a["nombre"]},
        )
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.get().id_articulo, 203)

    def test_superart_desambigua_talle_por_tokens_nombre(self):
        d = _draft()
        art_t4 = dict(
            ART_OK,
            id_articulo=301,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T4 Puma Invisible Sneaker Blanco 3P",
        )
        art_t5 = dict(
            ART_OK,
            id_articulo=302,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T5 Puma Invisible Sneaker Blanco 3P",
        )

        def lookup(_b, codigos):
            return {c: [art_t4, art_t5] for c in codigos}

        raw = _xlsx_plantilla(
            [10],
            {"906807": [36]},
            nombres={"906807": art_t4["nombre"]},
        )
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.get().id_articulo, 301)

    def test_dos_filas_mismo_superart_distinto_talle_importan_ambas(self):
        d = _draft()
        art_t4 = dict(
            ART_OK,
            id_articulo=401,
            id_manual="906807-15",
            cod_art_prov="906807-15",
            nombre="906807-15 T4 Puma Invisible Sneaker Bl/Ne/Gm 3P",
        )
        art_t5 = dict(
            ART_OK,
            id_articulo=402,
            id_manual="906807-15",
            cod_art_prov="906807-15",
            nombre="906807-15 T5 Puma Invisible Sneaker Bl/Ne/Gm 3P",
        )

        def lookup(_b, codigos):
            return {c: [art_t4, art_t5] for c in codigos}

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedido"
        ws.cell(1, 1, "Código")
        ws.cell(1, 2, "Artículo")
        ws.cell(1, 3, "id_articulo")
        ws.cell(1, 4, "SUC 14")
        ws.cell(1, 5, "SUC 20")
        ws.cell(2, 1, "906807")
        ws.cell(2, 2, art_t4["nombre"])
        ws.cell(2, 4, 24)
        ws.cell(2, 5, 0)
        ws.cell(3, 1, "906807")
        ws.cell(3, 2, art_t5["nombre"])
        ws.cell(3, 4, 12)
        ws.cell(3, 5, 6)
        ws_m = wb.create_sheet("_Synap")
        ws_m.append(["draft_id", d.id])
        ws_m.append(["id_cliente", d.id_cliente])
        ws_m.append(["cod_viajante", d.cod_viajante])
        ws_m.append(["sucursal_ids", 14, 20])
        ws_m.append(["col_primera_sucursal", 4])
        bio = BytesIO()
        wb.save(bio)

        with patch(
            "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
            return_value=[
                {"id_cliente_domicilio": 14, "nro": "14"},
                {"id_cliente_domicilio": 20, "nro": "20"},
            ],
        ), patch(
            "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
            return_value=[5],
        ), patch(
            "ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo"
        ), patch(
            "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
            return_value={},
        ), patch(
            "ecom.services.pedido_masivo_import._nombre_cliente",
            return_value="Cliente",
        ):
            res = importar_matriz_excel(d, bio.getvalue(), consultar_arts=lookup)
        self.assertTrue(res["ok"], res.get("errores"))
        celdas = {
            (c.id_articulo, c.id_cliente_domicilio): float(c.cantidad_packs)
            for c in d.celdas.all()
        }
        self.assertEqual(celdas.get((401, 14)), 24.0)
        self.assertEqual(celdas.get((402, 14)), 12.0)
        self.assertEqual(celdas.get((402, 20)), 6.0)

    def test_superart_usa_prefijo_nombre_si_codigo_padre_vacio(self):
        d = _draft()
        art_t5 = dict(
            ART_OK,
            id_articulo=601,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T5 Puma Invisible Sneaker Blanco 3P",
        )
        art_t4 = dict(
            ART_OK,
            id_articulo=602,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T4 Puma Invisible Sneaker Blanco 3P",
        )

        def lookup(_b, codigos):
            # Simula MySQL real: SuperArt 906807 no matchea exacto; el SKU sí.
            out = {}
            for c in codigos:
                if c == "906807-03":
                    out[c] = [art_t4, art_t5]
                else:
                    out[c] = []
            return out

        raw = _xlsx_plantilla(
            [10],
            {"906807": [12]},
            nombres={"906807": art_t5["nombre"]},
        )
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertTrue(res["ok"], res.get("errores"))
        self.assertEqual(d.celdas.get().id_articulo, 601)

    def test_v4_idart_erroneo_nombre_correcto_importa_t5(self):
        """Fila con IDArt de otra variante pero nombre T5: resuelve por catálogo."""
        d = _draft()
        art_t5 = dict(
            ART_OK,
            id_articulo=601,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T5 Puma Invisible Sneaker Blanco 3P",
        )
        art_t4 = dict(
            ART_OK,
            id_articulo=602,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T4 Puma Invisible Sneaker Blanco 3P",
        )

        def lookup(_b, codigos):
            # Solo existe T5 (T4 no está en la base del cliente)
            out = {}
            for c in codigos:
                if c in ("906807", "906807-03"):
                    out[c] = [art_t5]
                else:
                    out[c] = []
            return out

        def lookup_ids(_b, ids):
            return {602: art_t4}

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedido"
        ws.cell(1, 1, "Código")
        ws.cell(1, 2, "Artículo")
        ws.cell(1, 3, MARKER_IDART)
        ws.cell(1, 4, "SUC 14")
        ws.cell(1, 5, "SUC 20")
        # Fila 11: T4 no existe en catálogo
        ws.cell(2, 1, "906807")
        ws.cell(2, 2, art_t4["nombre"])
        ws.cell(2, 3, 602)
        ws.cell(2, 4, 6)
        # Fila 12: T5 existe; columna C con IDArt de T4 (fila copiada)
        ws.cell(3, 1, "906807")
        ws.cell(3, 2, art_t5["nombre"])
        ws.cell(3, 3, 602)
        ws.cell(3, 4, 12)
        ws.cell(3, 5, 6)
        ws_m = wb.create_sheet("_Synap")
        ws_m.append(["draft_id", d.id])
        ws_m.append(["id_cliente", d.id_cliente])
        ws_m.append(["cod_viajante", d.cod_viajante])
        ws_m.append(["sucursal_ids", 14, 20])
        ws_m.append(["col_primera_sucursal", 4])
        bio = BytesIO()
        wb.save(bio)

        with patch(
            "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
            return_value=[
                {"id_cliente_domicilio": 14, "nro": "14"},
                {"id_cliente_domicilio": 20, "nro": "20"},
            ],
        ), patch(
            "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
            return_value=[5],
        ), patch(
            "ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo"
        ), patch(
            "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
            return_value={},
        ), patch(
            "ecom.services.pedido_masivo_import._nombre_cliente",
            return_value="Cliente",
        ):
            res = importar_matriz_excel(
                d,
                bio.getvalue(),
                consultar_arts=lookup,
                consultar_ids=lookup_ids,
            )
        self.assertFalse(res["ok"])
        err_filas = {e["fila"]: e["code"] for e in res["errores"]}
        self.assertEqual(err_filas.get(2), "articulo_nombre_no_coincide")
        self.assertNotIn(3, err_filas)
        self.assertEqual(d.celdas.count(), 0)

    def test_v4_fila_solo_t5_idart_erroneo_importa(self):
        """Misma fila con IDArt de T4 pero nombre T5: resuelve por catálogo."""
        d = _draft()
        art_t5 = dict(
            ART_OK,
            id_articulo=601,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T5 Puma Invisible Sneaker Blanco 3P",
        )
        art_t4 = dict(
            ART_OK,
            id_articulo=602,
            id_manual="906807-03",
            cod_art_prov="906807-03",
            nombre="906807-03 T4 Puma Invisible Sneaker Blanco 3P",
        )

        def lookup(_b, codigos):
            out = {}
            for c in codigos:
                if c in ("906807", "906807-03"):
                    out[c] = [art_t5]
                else:
                    out[c] = []
            return out

        def lookup_ids(_b, ids):
            return {602: art_t4}

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedido"
        ws.cell(1, 1, "Código")
        ws.cell(1, 2, "Artículo")
        ws.cell(1, 3, MARKER_IDART)
        ws.cell(1, 4, "SUC 14")
        ws.cell(2, 1, "906807")
        ws.cell(2, 2, art_t5["nombre"])
        ws.cell(2, 3, 602)
        ws.cell(2, 4, 12)
        ws_m = wb.create_sheet("_Synap")
        ws_m.append(["draft_id", d.id])
        ws_m.append(["id_cliente", d.id_cliente])
        ws_m.append(["cod_viajante", d.cod_viajante])
        ws_m.append(["sucursal_ids", 14])
        ws_m.append(["col_primera_sucursal", 4])
        bio = BytesIO()
        wb.save(bio)

        with patch(
            "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
            return_value=[{"id_cliente_domicilio": 14, "nro": "14"}],
        ), patch(
            "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
            return_value=[5],
        ), patch(
            "ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo"
        ), patch(
            "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
            return_value={},
        ), patch(
            "ecom.services.pedido_masivo_import._nombre_cliente",
            return_value="Cliente",
        ):
            res = importar_matriz_excel(
                d,
                bio.getvalue(),
                consultar_arts=lookup,
                consultar_ids=lookup_ids,
            )
        self.assertTrue(res["ok"], res.get("errores"))
        self.assertEqual(d.celdas.get().id_articulo, 601)

    def test_nombre_excel_debe_coincidir_exacto(self):
        d = _draft()
        art_t4 = dict(
            ART_OK,
            id_articulo=501,
            id_manual="906807-15",
            nombre="906807-15 T4 Puma Invisible Sneaker Bl/Ne/Gm 3P",
        )
        art_t5 = dict(
            ART_OK,
            id_articulo=502,
            id_manual="906807-15",
            nombre="906807-15 T5 Puma Invisible Sneaker Bl/Ne/Gm 3P",
        )

        def lookup(_b, codigos):
            return {c: [art_t4, art_t5] for c in codigos}

        raw = _xlsx_plantilla(
            [10],
            {"906807": [24]},
            nombres={"906807": "906807-15 T4 Puma Invisible Sneaker Bl/Ne/Gm 3P ERRÓNEO"},
        )
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertFalse(res["ok"])
        err = next(e for e in res["errores"] if e["code"] == "articulo_nombre_no_coincide")
        self.assertEqual(err["fila"], 3)
        self.assertEqual(err["codigo_articulo"], "906807")
        self.assertIn("ERRÓNEO", err["nombre_articulo"])
        self.assertEqual(d.celdas.count(), 0)

    def test_codigo_numerico_prioriza_id_manual_sobre_idart(self):
        d = _draft()
        por_manual = dict(
            ART_OK,
            id_articulo=500,
            id_manual="64",
            nombre="0458 Suede Crew Sock 2P 005 Blanco",
        )
        por_idart = dict(
            ART_OK,
            id_articulo=64,
            id_manual="9999",
            nombre="Otro artículo IDArt 64",
        )

        def lookup(_b, codigos):
            return {c: [por_idart, por_manual] for c in codigos}

        raw = _xlsx_plantilla([10], {"64": [6]})
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.get().id_articulo, 500)

    def test_superart_sin_nombre_distinto_sigue_ambiguo(self):
        d = _draft()
        art_a = dict(ART_OK, id_articulo=203, id_manual="3766", nombre="Color A")
        art_b = dict(ART_OK, id_articulo=204, id_manual="3766", nombre="Color B")

        def lookup(_b, codigos):
            return {c: [art_a, art_b] for c in codigos}

        raw = _xlsx_plantilla([10], {"3766": [6]})
        res = importar_matriz_excel(d, raw, consultar_arts=lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["errores"][0]["code"], "articulo_ambiguo")
        self.assertEqual(d.celdas.count(), 0)

    def test_qty_cero_no_crea_celda(self):
        d = _draft()
        raw = _xlsx_plantilla([10, 20], {"2401": [6, 0]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.count(), 1)
        self.assertEqual(d.celdas.get().id_cliente_domicilio, 10)

    def test_agrega_varios_errores_sin_tocar_borrador(self):
        d = _draft()
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=5, id_cliente_domicilio=10, cantidad_packs=Decimal("3")
        )
        raw = _xlsx_plantilla([10], {"NO-EXISTE": [6], "2401": [7]})
        res = importar_matriz_excel(d, raw, consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        codes = {e["code"] for e in res["errores"]}
        self.assertIn("articulo_no_encontrado", codes)
        self.assertIn("multiplo_empaque", codes)
        self.assertEqual(d.celdas.count(), 1)
        self.assertEqual(d.celdas.get().id_articulo, 5)

    def test_xlsx_corrupto_no_toca_celdas(self):
        d = _draft()
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=5, id_cliente_domicilio=10, cantidad_packs=Decimal("3")
        )
        res = importar_matriz_excel(d, b"no-es-xlsx", consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["code"], "archivo_invalido")
        self.assertEqual(d.celdas.count(), 1)

    def test_sin_hoja_pedido(self):
        d = _draft()
        wb = Workbook()
        wb.active.title = "Otra"
        bio = BytesIO()
        wb.save(bio)
        res = importar_matriz_excel(d, bio.getvalue(), consultar_arts=self._lookup)
        self.assertFalse(res["ok"])
        self.assertEqual(res["code"], "archivo_invalido")
        self.assertIn("Pedido", res["error"])

    def test_v4_mapea_cantidad_por_nro_sucursal_en_encabezado(self):
        d = _draft()
        raw = _xlsx_plantilla_v4(
            [10, 20],
            ["127\nMORÓN - 25 de Mayo", "736\nMERLO - Av. Libertador"],
            {"2401": [12, 6]},
        )

        def lookup_ids(_b, ids):
            return {i: dict(ART_OK) for i in ids if i == 101}

        res = importar_matriz_excel(
            d,
            raw,
            consultar_arts=lambda *_a, **_k: {},
            consultar_ids=lookup_ids,
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(
            d.celdas.get(id_cliente_domicilio=10).cantidad_packs, Decimal("12")
        )
        self.assertEqual(
            d.celdas.get(id_cliente_domicilio=20).cantidad_packs, Decimal("6")
        )

    def test_v4_importa_columnas_reordenadas_por_encabezado(self):
        d = _draft()
        raw = _xlsx_plantilla_v4(
            [10, 20],
            ["736\nMERLO - Av. Libertador", "127\nMORÓN - 25 de Mayo"],
            {"2401": [6, 12]},
        )

        def lookup_ids(_b, ids):
            return {i: dict(ART_OK) for i in ids if i == 101}

        res = importar_matriz_excel(
            d,
            raw,
            consultar_arts=lambda *_a, **_k: {},
            consultar_ids=lookup_ids,
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(
            d.celdas.get(id_cliente_domicilio=20).cantidad_packs, Decimal("6")
        )
        self.assertEqual(
            d.celdas.get(id_cliente_domicilio=10).cantidad_packs, Decimal("12")
        )

    def test_v4_importa_subset_columnas_eliminadas(self):
        d = _draft()
        raw = _xlsx_plantilla_v4(
            [10, 20, 99],
            ["127\nMORÓN - 25 de Mayo", "736\nMERLO - Av. Libertador"],
            {"2401": [6, 12]},
        )

        def lookup_ids(_b, ids):
            return {i: dict(ART_OK) for i in ids if i == 101}

        res = importar_matriz_excel(
            d,
            raw,
            consultar_arts=lambda *_a, **_k: {},
            consultar_ids=lookup_ids,
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.count(), 2)

    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
    )
    def test_v4_desambigua_mismo_nro_por_calle_encabezado(self, mock_suc, _m):
        mock_suc.return_value = [
            {
                "id_cliente_domicilio": 30,
                "nro": "142",
                "calle": "MONTE AGUDO 2323/27",
                "etiqueta": "MONTE AGUDO 2323/27",
                "nombre": "MONTE AGUDO 2323/27",
            },
            {
                "id_cliente_domicilio": 31,
                "nro": "142",
                "calle": "OTRA CALLE DISTINTA",
                "etiqueta": "OTRA CALLE DISTINTA",
                "nombre": "OTRA CALLE DISTINTA",
            },
        ]
        d = _draft()
        raw = _xlsx_plantilla_v4(
            [30, 31],
            ["142\n142 - MONTE AGUDO 2323/27 - FLORENCIO VA"],
            {"2401": [6]},
        )

        def lookup_ids(_b, ids):
            return {i: dict(ART_OK) for i in ids if i == 101}

        res = importar_matriz_excel(
            d,
            raw,
            consultar_arts=lambda *_a, **_k: {},
            consultar_ids=lookup_ids,
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.get(id_cliente_domicilio=30).cantidad_packs, Decimal("6"))
        self.assertFalse(d.celdas.filter(id_cliente_domicilio=31).exists())

    def test_v4_avisa_cantidades_en_columna_sin_encabezado(self):
        d = _draft()
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedido"
        ws.cell(1, 1, "Código")
        ws.cell(1, 2, "Artículo")
        ws.cell(1, 3, MARKER_IDART)
        ws.cell(1, 4, "127\nMORÓN - 25 de Mayo")
        ws.cell(2, 1, "2401")
        ws.cell(2, 2, "nombre")
        ws.cell(2, 3, 101)
        ws.cell(2, 4, 6)
        ws.cell(2, 5, 99)
        ws.cell(3, 5, 12)
        ws_m = wb.create_sheet(HOJA_META)
        ws_m.append(["id_cliente", 368])
        ws_m.append(["cod_viajante", 30])
        ws_m.append(["plantilla_version", 4])
        ws_m.append(["sucursal_ids", 10])
        ws_m.append(["col_primera_sucursal", 4])
        bio = BytesIO()
        wb.save(bio)

        def lookup_ids(_b, ids):
            return {i: dict(ART_OK) for i in ids if i == 101}

        res = importar_matriz_excel(
            d,
            bio.getvalue(),
            consultar_arts=lambda *_a, **_k: {},
            consultar_ids=lookup_ids,
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.count(), 1)
        self.assertEqual(len(res.get("avisos") or []), 1)
        self.assertEqual(res["avisos"][0]["code"], "columna_sin_encabezado")
        self.assertEqual(res["avisos"][0]["columna"], "E")
        self.assertIn("no se importaron", res["avisos"][0]["mensaje"].lower())


def _xlsx_plantilla_v4(ids_suc, header_labels, qtys_por_codigo, id_cliente=368, cod_viajante=30):
    """Plantilla v4 con encabezados de sucursal personalizados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"
    ws.cell(1, 1, "Código")
    ws.cell(1, 2, "Artículo")
    ws.cell(1, 3, MARKER_IDART)
    for i, lbl in enumerate(header_labels):
        ws.cell(1, 4 + i, lbl)
    fila = 2
    for codigo, qtys in qtys_por_codigo.items():
        ws.cell(fila, 1, codigo)
        ws.cell(fila, 2, "nombre")
        ws.cell(fila, 3, 101)
        for i, q in enumerate(qtys):
            if q is not None:
                ws.cell(fila, 4 + i, q)
        fila += 1
    ws_m = wb.create_sheet(HOJA_META)
    ws_m.append(["id_cliente", id_cliente])
    ws_m.append(["cod_viajante", cod_viajante])
    ws_m.append(["plantilla_version", 4])
    ws_m.append(["sucursal_ids", *ids_suc])
    ws_m.append(["col_primera_sucursal", 4])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class TestPlantillaExcel(TestCase):
    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
        return_value=[SUC_A, SUC_B],
    )
    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    @patch(
        "ecom.services.pedido_masivo_import._nombre_cliente",
        return_value="Dabra S.A.",
    )
    @patch(
        "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
        return_value={
            "descRenglon": Decimal("0"),
            "descPie": Decimal("0"),
            "lista_id": 1,
        },
    )
    @patch(
        "ecom.services.pedido_masivo_import._precio_real_articulo",
        return_value=Decimal("80"),
    )
    def test_plantilla_incluye_marker_e_ids(self, _p, _ctx, _n, _m, _s):
        d = _draft()
        raw = generar_plantilla_excel(
            d,
            articulos=[
                {"id_articulo": 101, "id_manual": "2401", "nombre": "Media pack"}
            ],
        )
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(raw))
        self.assertIn("Pedido", wb.sheetnames)
        self.assertIn(HOJA_META, wb.sheetnames)
        ws = wb["Pedido"]
        self.assertEqual(ws.cell(1, 1).value, "Código")
        self.assertEqual(ws.cell(1, 2).value, "Artículo")
        self.assertEqual(ws.cell(1, 3).value, MARKER_IDART)
        self.assertEqual(ws.cell(1, 4).value, "Precio")
        self.assertTrue(bool(ws.column_dimensions["C"].hidden))
        self.assertFalse(bool(ws.row_dimensions[1].hidden))
        self.assertGreaterEqual(ws.row_dimensions[1].height or 0, 45)
        self.assertFalse(bool(ws.protection.sheet))
        self.assertEqual(ws.freeze_panes, "E2")
        self.assertIn("127", str(ws.cell(1, 5).value or ""))
        self.assertIn("MORÓN", str(ws.cell(1, 5).value or "").replace("\n", " "))
        self.assertIn(
            "precio",
            " ".join(str(ws.cell(1, c).value or "").lower() for c in range(1, 7)),
        )
        self.assertEqual(ws.cell(2, 1).value, "2401")
        self.assertEqual(ws.cell(2, 2).value, "Media pack")
        self.assertEqual(ws.cell(2, 3).value, 101)
        ws_m = wb[HOJA_META]
        claves = {ws_m.cell(i, 1).value: ws_m.cell(i, 2).value for i in range(1, 8)}
        self.assertEqual(claves.get("id_cliente"), 368)
        self.assertEqual(claves.get("cod_viajante"), 30)
        self.assertEqual(claves.get("plantilla_version"), 5)
        self.assertEqual(claves.get("col_primera_sucursal"), 5)
        self.assertEqual(ws_m.cell(6, 1).value, "sucursal_ids")
        self.assertEqual(ws_m.cell(6, 2).value, 10)
        self.assertEqual(ws_m.cell(6, 3).value, 20)

    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
        return_value=[SUC_A, SUC_B],
    )
    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    @patch(
        "ecom.services.pedido_masivo_import._nombre_cliente",
        return_value="Dabra S.A.",
    )
    @patch(
        "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
        return_value={
            "descRenglon": Decimal("0"),
            "descPie": Decimal("5"),
            "lista_id": 1,
        },
    )
    @patch("ecom.services.pedido_masivo_import._precio_real_articulo", return_value=Decimal("80"))
    @patch("ecom.services.pedido_masivo_import.asegurar_precio_fila_articulo")
    @patch("ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo")
    def test_importa_plantilla_v4_generada(self, _d, _pr, _px, _pie, _n, _m, _s):
        d = _draft()
        raw = generar_plantilla_excel(
            d,
            articulos=[
                {"id_articulo": 101, "id_manual": "2401", "nombre": "Media pack"}
            ],
        )
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(raw))
        wb["Pedido"]["E2"] = 6
        bio = BytesIO()
        wb.save(bio)

        def lookup_ids(_b, ids):
            return {i: dict(ART_OK) for i in ids if i == 101}

        res = importar_matriz_excel(
            d, bio.getvalue(), consultar_arts=lambda *_a, **_k: {}, consultar_ids=lookup_ids
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.count(), 1)
        self.assertEqual(d.celdas.get().id_articulo, 101)
        self.assertEqual(d.celdas.get().id_cliente_domicilio, 10)

    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
        return_value=[SUC_A, SUC_B],
    )
    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    @patch(
        "ecom.services.pedido_masivo_import._nombre_cliente",
        return_value="Dabra S.A.",
    )
    @patch(
        "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
        return_value={
            "descRenglon": Decimal("0"),
            "descPie": Decimal("5"),
            "lista_id": 1,
        },
    )
    @patch("ecom.services.pedido_masivo_import._precio_real_articulo", return_value=Decimal("80"))
    @patch("ecom.services.pedido_masivo_import.asegurar_precio_fila_articulo")
    @patch("ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo")
    def test_plantilla_v4_dos_sku_mismo_superart(self, _d, _pr, _px, _pie, _n, _m, _s):
        d = _draft()
        art_a = dict(ART_OK, id_articulo=203, id_manual="3766", nombre="Blanco")
        art_b = dict(ART_OK, id_articulo=204, id_manual="3766", nombre="Negro")
        raw = generar_plantilla_excel(
            d,
            articulos=[
                {"id_articulo": 203, "id_manual": "3766", "nombre": "Blanco"},
                {"id_articulo": 204, "id_manual": "3766", "nombre": "Negro"},
            ],
        )
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(raw))
        ws = wb["Pedido"]
        self.assertEqual(ws.cell(2, 1).value, "3766")
        self.assertEqual(ws.cell(3, 1).value, "3766")
        self.assertEqual(ws.cell(2, 3).value, 203)
        self.assertEqual(ws.cell(3, 3).value, 204)
        ws["E2"] = 6
        ws["E3"] = 12
        bio = BytesIO()
        wb.save(bio)

        def lookup_ids(_b, ids):
            catalog = {203: art_a, 204: art_b}
            return {i: catalog[i] for i in ids if i in catalog}

        res = importar_matriz_excel(
            d,
            bio.getvalue(),
            consultar_arts=lambda *_a, **_k: {},
            consultar_ids=lookup_ids,
        )
        self.assertTrue(res["ok"], res)
        self.assertEqual(d.celdas.count(), 2)
        self.assertEqual(
            d.celdas.get(id_articulo=203).cantidad_packs, Decimal("6")
        )
        self.assertEqual(
            d.celdas.get(id_articulo=204).cantidad_packs, Decimal("12")
        )


class TestApiImportExcel(TestCase):
    @patch("ecom.pedido_masivo_views._session_base_empresa", return_value="emp_m")
    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
        return_value=[SUC_A],
    )
    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    @patch("ecom.services.pedido_masivo_import.asegurar_precio_fila_articulo")
    @patch("ecom.services.pedido_masivo_import.asegurar_descuento_fila_articulo")
    @patch(
        "ecom.services.pedido_masivo_import.leer_contexto_cliente_masivo",
        return_value={
            "descRenglon": Decimal("0"),
            "descPie": Decimal("5"),
            "lista_id": 1,
        },
    )
    @patch(
        "ecom.pedido_masivo_views._flags_cabecera_masivo",
        return_value={"puede_editar_precio_linea": True, "es_supervisor": False},
    )
    def test_post_importar_ok(self, _f, _pie, _d, _pr, _m, _s, _b):
        from django.core.files.uploadedfile import SimpleUploadedFile

        d = _draft(id_usuario=55)

        def lookup(_base, codigos):
            return {c: [dict(ART_OK)] for c in codigos}

        raw = _xlsx_plantilla([10], {"2401": [6]})
        factory = APIRequestFactory()
        req = factory.post(
            "/ecom/api/mayoristapp/pedido-masivo/importar/",
            {
                "draft_id": str(d.pk),
                "archivo": SimpleUploadedFile(
                    "p.xlsx",
                    raw,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            format="multipart",
        )
        req.session = {"user": {"base_empresa": "emp_m", "id_usuario": 55}}
        force_authenticate(req, user=_User())
        with patch(
            "ecom.services.pedido_masivo_import.consultar_articulos_por_codigos",
            side_effect=lookup,
        ):
            with patch(
                "ecom.pedido_masivo_views._serializar_matriz_ui",
                return_value={"draft_id": d.pk, "celdas": {}},
            ):
                resp = PedidoMasivoImportarAPIView.as_view()(req)
        self.assertEqual(resp.status_code, 200, getattr(resp, "data", None))
        self.assertTrue(resp.data["ok"])
        self.assertEqual(d.celdas.count(), 1)

    @patch("ecom.pedido_masivo_views._session_base_empresa", return_value="emp_m")
    def test_post_csv_rechaza(self, _b):
        from django.core.files.uploadedfile import SimpleUploadedFile

        d = _draft(id_usuario=55)
        factory = APIRequestFactory()
        req = factory.post(
            "/ecom/api/mayoristapp/pedido-masivo/importar/",
            {
                "draft_id": str(d.pk),
                "archivo": SimpleUploadedFile(
                    "p.csv", b"a,b\n1,2", content_type="text/csv"
                ),
            },
            format="multipart",
        )
        req.session = {"user": {"base_empresa": "emp_m", "id_usuario": 55}}
        force_authenticate(req, user=_User())
        resp = PedidoMasivoImportarAPIView.as_view()(req)
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.data["code"], "archivo_invalido")
        self.assertEqual(d.celdas.count(), 0)

    @patch("ecom.pedido_masivo_views._session_base_empresa", return_value="emp_m")
    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
        return_value=[SUC_A],
    )
    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    def test_post_multiplo_invalido_409(self, _m, _s, _b):
        from django.core.files.uploadedfile import SimpleUploadedFile

        d = _draft(id_usuario=55)
        EcomPedidoMasivoDraftCelda.objects.create(
            draft=d, id_articulo=5, id_cliente_domicilio=10, cantidad_packs=Decimal("3")
        )

        def lookup(_base, codigos):
            return {c: [dict(ART_OK)] for c in codigos}

        raw = _xlsx_plantilla([10], {"2401": [7]})
        factory = APIRequestFactory()
        req = factory.post(
            "/ecom/api/mayoristapp/pedido-masivo/importar/",
            {
                "draft_id": str(d.pk),
                "archivo": SimpleUploadedFile(
                    "p.xlsx",
                    raw,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            format="multipart",
        )
        req.session = {"user": {"base_empresa": "emp_m", "id_usuario": 55}}
        force_authenticate(req, user=_User())
        with patch(
            "ecom.services.pedido_masivo_import.consultar_articulos_por_codigos",
            side_effect=lookup,
        ):
            resp = PedidoMasivoImportarAPIView.as_view()(req)
        self.assertEqual(resp.status_code, 409, getattr(resp, "data", None))
        self.assertEqual(resp.data["code"], "validacion")
        self.assertEqual(d.celdas.count(), 1)
        self.assertEqual(d.celdas.get().id_articulo, 5)

    @patch("ecom.pedido_masivo_views._session_base_empresa", return_value="emp_m")
    @patch(
        "ecom.services.pedido_masivo_import.listar_sucursales_cliente",
        return_value=[SUC_A],
    )
    @patch(
        "ecom.services.pedido_masivo_import.marcas_asignadas_viajante_cliente",
        return_value=[5],
    )
    @patch(
        "ecom.services.pedido_masivo_import.listar_articulos_plantilla_vcm",
        return_value=[],
    )
    @patch(
        "ecom.services.pedido_masivo_import._nombre_cliente",
        return_value="Dabra",
    )
    def test_get_plantilla(self, _n, _a, _m, _s, _b):
        d = _draft(id_usuario=55)
        factory = APIRequestFactory()
        req = factory.get(
            f"/ecom/api/mayoristapp/pedido-masivo/plantilla-excel/?draft_id={d.pk}"
        )
        req.session = {"user": {"base_empresa": "emp_m", "id_usuario": 55}}
        force_authenticate(req, user=_User())
        resp = PedidoMasivoPlantillaExcelAPIView.as_view()(req)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn(".xlsx", resp["Content-Disposition"])
