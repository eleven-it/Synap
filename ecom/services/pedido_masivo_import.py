"""Importación Excel de pedido masivo (matriz artículo × sucursal).

Formato plantilla v4: hoja ``Pedido`` con A=SuperArt, B=nombre, C oculta=IDArt,
D+=packs. Una fila por SKU (el SuperArt/`id_manual` no es único).
v3 (sin columna IDArt): se desambigua por nombre / CodArtProv y se prioriza
``id_manual`` frente a colisión con ``IDArt``.
Hoja oculta ``_Synap`` identifica cliente y vendedor.

Modo: **reemplazo total** del borrador (se vacían celdas y queda solo el Excel).
Validación de territorio: cuaterna VCM vendedor → cliente → sucursal → marca.
"""

from __future__ import annotations

import io
import logging
import re
from decimal import Decimal
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from core.mysql_pool import get_mysql_pool
from core.utils.administranet_types import str_or_default, to_decimal_or_none, to_int_or_none
from ecom.models import EcomPedidoMasivoDraft, EcomPedidoMasivoDraftCelda
from ecom.services.multiplo_empaque import (
    cantidad_respeta_multiplo,
    mensaje_multiplo_invalido,
    multiplo_empaque_venta,
)
from ecom.services.pedido_masivo_matriz import (
    _clamp_pct,
    _clave_orden_nro_sucursal,
    _nombre_cliente,
    asegurar_descuento_fila_articulo,
    leer_contexto_cliente_masivo,
    listar_sucursales_cliente,
    marcas_asignadas_viajante_cliente,
)

logger = logging.getLogger(__name__)

HOJA_PEDIDO = "Pedido"
HOJA_SUCURSALES = "Sucursales"
HOJA_INSTRUCCIONES = "Instrucciones"
HOJA_META = "_Synap"
MARKER_CODIGO = "codigo_articulo"
PLANTILLA_VERSION = 4
MARKER_IDART = "id_articulo"
MAX_BYTES = 8 * 1024 * 1024
MAX_ERRORES = 200
MAX_ARTICULOS_PLANTILLA = 5000  # red de seguridad; administranet prod 13/08/2026: 310 ecommerce Terminado

_FILL_ID = PatternFill("solid", fgColor="E2E8F0")
_FILL_HDR = PatternFill("solid", fgColor="0F172A")
_FILL_LOCK = PatternFill("solid", fgColor="F1F5F9")
_FILL_QTY = PatternFill("solid", fgColor="FFFBEB")
_FONT_HDR = Font(color="FFFFFF", bold=True, size=9)
_FONT_HDR_SUC = Font(color="FFFFFF", bold=True, size=8)
_FONT_ID = Font(color="64748B", size=8)
_FONT_ART = Font(size=10)


def _err(
    mensaje: str,
    *,
    code: str,
    fila: Optional[int] = None,
    columna: Optional[str] = None,
    codigo_articulo: str = "",
    sucursal: str = "",
) -> Dict[str, Any]:
    return {
        "fila": fila,
        "columna": columna or "",
        "codigo_articulo": codigo_articulo or "",
        "sucursal": sucursal or "",
        "mensaje": mensaje,
        "code": code,
    }


def _celda_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _qty_celda(val: Any) -> Tuple[Optional[Decimal], Optional[str]]:
    """Vacío → (None, None). Inválido → (None, mensaje)."""
    if val is None or val == "":
        return None, None
    if isinstance(val, str) and not val.strip():
        return None, None
    if isinstance(val, str):
        val = val.strip().replace(" ", "").replace(",", ".")
    qty = to_decimal_or_none(val)
    if qty is None:
        return None, "Cantidad inválida."
    if qty < 0:
        return None, "La cantidad no puede ser negativa."
    if qty == 0:
        return None, None
    return qty, None


def consultar_articulos_por_codigos(
    base_empresa: str, codigos: Sequence[str]
) -> Dict[str, List[Dict[str, Any]]]:
    """Resuelve códigos Excel → candidatos de ``articulo`` (exacto)."""
    out: Dict[str, List[Dict[str, Any]]] = {}
    unicos = []
    seen: Set[str] = set()
    for c in codigos:
        k = (c or "").strip()
        if not k or k in seen:
            continue
        seen.add(k)
        unicos.append(k)
    if not unicos:
        return out
    sql = """
        SELECT
            articulo.IDArt,
            COALESCE(articulo.id_manual, '') AS id_manual,
            COALESCE(articulo.NombreArticulo, '') AS nombre,
            articulo.CodigoMarca,
            COALESCE(articulo.CodArtProv, '') AS cod_art_prov,
            COALESCE(articulo.CodigoArticuloT, '') AS codigo_t,
            COALESCE(articulo.Discontinuo, 'No') AS discontinuo,
            COALESCE(articulo.ecommerce, 'No') AS ecommerce,
            COALESCE(TRIM(articulo.tipo_art_fab), '') AS tipo_art_fab,
            articulo.multiplo_cantidad_vta
        FROM articulo
        WHERE articulo.id_manual = %s
           OR CAST(articulo.IDArt AS CHAR) = %s
           OR articulo.CodigoArticuloT = %s
           OR articulo.NroCodBarra = %s
           OR articulo.NroCodBarraF = %s
           OR articulo.CodArtProv = %s
        LIMIT 40
    """
    try:
        pool = get_mysql_pool()
        with pool.get_connection(base_empresa.strip()) as conn:
            cursor = conn.cursor()
            try:
                for codigo in unicos:
                    cursor.execute(sql, [codigo] * 6)
                    cols = [d[0] for d in cursor.description] if cursor.description else []
                    rows = [dict(zip(cols, r)) for r in cursor.fetchall()]
                    seen_ids: Set[int] = set()
                    arts: List[Dict[str, Any]] = []
                    for row in rows:
                        aid = to_int_or_none(row.get("IDArt"))
                        if aid is None or aid in seen_ids:
                            continue
                        seen_ids.add(aid)
                        arts.append(
                            {
                                "id_articulo": aid,
                                "id_manual": str_or_default(row.get("id_manual"), ""),
                                "nombre": str_or_default(row.get("nombre"), ""),
                                "codigo_marca": to_int_or_none(row.get("CodigoMarca")),
                                "cod_art_prov": str_or_default(row.get("cod_art_prov"), ""),
                                "codigo_t": str_or_default(row.get("codigo_t"), ""),
                                "discontinuo": str_or_default(row.get("discontinuo"), "No"),
                                "ecommerce": str_or_default(row.get("ecommerce"), "No"),
                                "tipo_art_fab": str_or_default(row.get("tipo_art_fab"), ""),
                                "multiplo_cantidad_vta": row.get("multiplo_cantidad_vta"),
                            }
                        )
                    out[codigo] = arts
            finally:
                cursor.close()
    except Exception as e:
        logger.warning("consultar_articulos_por_codigos: %s", e)
        for codigo in unicos:
            out.setdefault(codigo, [])
    return out


def _art_desde_row(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    aid = to_int_or_none(row.get("IDArt") or row.get("id_articulo"))
    if aid is None:
        return None
    return {
        "id_articulo": aid,
        "id_manual": str_or_default(row.get("id_manual"), ""),
        "nombre": str_or_default(row.get("nombre"), ""),
        "codigo_marca": to_int_or_none(row.get("CodigoMarca") or row.get("codigo_marca")),
        "cod_art_prov": str_or_default(row.get("cod_art_prov"), ""),
        "codigo_t": str_or_default(row.get("codigo_t"), ""),
        "discontinuo": str_or_default(row.get("discontinuo"), "No"),
        "ecommerce": str_or_default(row.get("ecommerce"), "No"),
        "tipo_art_fab": str_or_default(row.get("tipo_art_fab"), ""),
        "multiplo_cantidad_vta": row.get("multiplo_cantidad_vta"),
    }


def consultar_articulos_por_ids(
    base_empresa: str, ids: Sequence[int]
) -> Dict[int, Dict[str, Any]]:
    """Resuelve IDArt de plantilla → artículo."""
    out: Dict[int, Dict[str, Any]] = {}
    ids_ok = []
    seen: Set[int] = set()
    for i in ids:
        aid = to_int_or_none(i)
        if aid is None or aid in seen:
            continue
        seen.add(aid)
        ids_ok.append(aid)
    if not ids_ok:
        return out
    ph = ",".join(["%s"] * len(ids_ok))
    sql = f"""
        SELECT
            articulo.IDArt,
            COALESCE(articulo.id_manual, '') AS id_manual,
            COALESCE(articulo.NombreArticulo, '') AS nombre,
            articulo.CodigoMarca,
            COALESCE(articulo.CodArtProv, '') AS cod_art_prov,
            COALESCE(articulo.CodigoArticuloT, '') AS codigo_t,
            COALESCE(articulo.Discontinuo, 'No') AS discontinuo,
            COALESCE(articulo.ecommerce, 'No') AS ecommerce,
            COALESCE(TRIM(articulo.tipo_art_fab), '') AS tipo_art_fab,
            articulo.multiplo_cantidad_vta
        FROM articulo
        WHERE articulo.IDArt IN ({ph})
    """
    try:
        pool = get_mysql_pool()
        with pool.get_connection(base_empresa.strip()) as conn:
            cursor = conn.cursor()
            try:
                cursor.execute(sql, ids_ok)
                cols = [d[0] for d in cursor.description] if cursor.description else []
                for r in cursor.fetchall():
                    art = _art_desde_row(dict(zip(cols, r)))
                    if art:
                        out[int(art["id_articulo"])] = art
            finally:
                cursor.close()
    except Exception as e:
        logger.warning("consultar_articulos_por_ids: %s", e)
    return out


def _territorio(
    draft: EcomPedidoMasivoDraft,
) -> Tuple[List[Dict[str, Any]], Dict[int, Set[int]]]:
    sucursales = listar_sucursales_cliente(
        draft.base_empresa,
        draft.id_cliente,
        draft.cod_viajante,
    )
    sucursales = sorted(sucursales, key=_clave_orden_nro_sucursal)
    if draft.modo == EcomPedidoMasivoDraft.MODO_SIMPLE:
        fijo = to_int_or_none(draft.id_domicilio_fijo)
        if fijo:
            sucursales = [
                s
                for s in sucursales
                if to_int_or_none(s.get("id_cliente_domicilio")) == fijo
            ]
    marcas_map: Dict[int, Set[int]] = {}
    cv = to_int_or_none(draft.cod_viajante)
    for s in sucursales:
        idd = to_int_or_none(s.get("id_cliente_domicilio"))
        if idd is None:
            continue
        if cv is None:
            marcas_map[idd] = set()
            continue
        marcas_map[idd] = set(
            marcas_asignadas_viajante_cliente(
                draft.base_empresa,
                cv,
                draft.id_cliente,
                id_cliente_domicilio=idd,
            )
        )
    return sucursales, marcas_map


def listar_articulos_plantilla_vcm(draft: EcomPedidoMasivoDraft) -> List[Dict[str, Any]]:
    """Artículos Terminado/ecommerce de la unión de marcas VCM (sin precio ni stock)."""
    _sucursales, marcas_map = _territorio(draft)
    marcas: Set[int] = set()
    for ms in marcas_map.values():
        marcas.update(ms)
    marcas_l = sorted(m for m in marcas if m is not None)
    if not marcas_l:
        return []
    ph = ",".join(["%s"] * len(marcas_l))
    sql = f"""
        SELECT
            articulo.IDArt,
            COALESCE(articulo.id_manual, '') AS id_manual,
            COALESCE(articulo.NombreArticulo, '') AS nombre
        FROM articulo
        WHERE articulo.Discontinuo = 'No'
          AND articulo.ecommerce = 'Si'
          AND COALESCE(TRIM(articulo.tipo_art_fab), '') = 'Terminado'
          AND articulo.CodigoMarca IN ({ph})
        ORDER BY
            CASE WHEN articulo.id_manual IS NULL OR articulo.id_manual = '' THEN 1 ELSE 0 END,
            articulo.id_manual,
            articulo.NombreArticulo
        LIMIT %s
    """
    out: List[Dict[str, Any]] = []
    seen_id: Set[int] = set()
    try:
        pool = get_mysql_pool()
        with pool.get_connection(draft.base_empresa.strip()) as conn:
            cursor = conn.cursor()
            try:
                cursor.execute(sql, [*marcas_l, MAX_ARTICULOS_PLANTILLA])
                for row in cursor.fetchall():
                    aid = to_int_or_none(row[0])
                    if aid is None or aid in seen_id:
                        continue
                    codigo = str_or_default(row[1], "") or str(aid)
                    seen_id.add(aid)
                    out.append(
                        {
                            "id_articulo": aid,
                            "id_manual": codigo,
                            "nombre": str_or_default(row[2], ""),
                        }
                    )
            finally:
                cursor.close()
    except Exception as e:
        logger.warning("listar_articulos_plantilla_vcm: %s", e)
        return []
    return out


def _etiqueta_suc(s: Dict[str, Any]) -> str:
    return str_or_default(s.get("etiqueta") or s.get("nombre") or s.get("calle"), "")


def _rotulo_columna_suc(s: Dict[str, Any]) -> str:
    """Rótulo corto de columna: nro + calle (legible con ajuste de texto)."""
    nro = str_or_default(s.get("nro"), "").strip()
    calle = str_or_default(s.get("calle"), "").strip()
    if not calle:
        calle = _etiqueta_suc(s)
    if len(calle) > 42:
        calle = calle[:40].rstrip() + "…"
    if nro and nro != "-":
        return f"{nro}\n{calle}"
    return calle


def generar_plantilla_excel(
    draft: EcomPedidoMasivoDraft,
    *,
    articulos: Optional[Sequence[Dict[str, Any]]] = None,
) -> bytes:
    """Plantilla VCM: SuperArt + nombre + cantidades. Sin precios. Identifica al cliente."""
    sucursales, _marcas = _territorio(draft)
    if articulos is None:
        articulos = listar_articulos_plantilla_vcm(draft)
    nombre_cli = _nombre_cliente(draft.base_empresa, draft.id_cliente)
    id_cli = to_int_or_none(draft.id_cliente) or 0
    cv = to_int_or_none(draft.cod_viajante)

    qty_map: Dict[Tuple[int, int], Decimal] = {}
    for cel in draft.celdas.all():
        aid = to_int_or_none(cel.id_articulo)
        idd = to_int_or_none(cel.id_cliente_domicilio)
        if aid is None or idd is None:
            continue
        qty_map[(aid, idd)] = cel.cantidad_packs

    wb = Workbook()

    ws_i = wb.active
    ws_i.title = HOJA_INSTRUCCIONES
    instrucciones = [
        "Plantilla de pedido masivo Synap — solo completá cantidades (packs).",
        "No agregues columnas, ni precios, ni descuentos: eso sale del cliente.",
        "",
        f"Cliente: {id_cli} — {nombre_cli or '—'}",
        f"Vendedor: {cv if cv is not None else '—'}",
        f"Borrador: #{draft.pk}",
        "",
        "Al importar se reemplaza todo el borrador. Usá la plantilla de este mismo pedido.",
        "Hay una fila por color/SKU. No borres la columna oculta (identifica el artículo).",
    ]
    for i, texto in enumerate(instrucciones, start=1):
        ws_i.cell(i, 1, texto)
    ws_i.column_dimensions["A"].width = 110
    ws_i.sheet_state = "hidden"

    ws_s = wb.create_sheet(HOJA_SUCURSALES)
    ws_s.append(["id_cliente_domicilio", "nro", "calle", "etiqueta"])
    for col in range(1, 5):
        c = ws_s.cell(1, col)
        c.fill = _FILL_HDR
        c.font = _FONT_HDR
    for s in sucursales:
        ws_s.append(
            [
                int(s["id_cliente_domicilio"]),
                str_or_default(s.get("nro"), ""),
                str_or_default(s.get("calle"), ""),
                _etiqueta_suc(s),
            ]
        )
    for col, w in enumerate((22, 14, 55, 55), start=1):
        ws_s.column_dimensions[get_column_letter(col)].width = w
    ws_s.sheet_state = "hidden"

    ws_m = wb.create_sheet(HOJA_META)
    ids_suc = [
        int(s["id_cliente_domicilio"])
        for s in sucursales
        if to_int_or_none(s.get("id_cliente_domicilio")) is not None
    ]
    ws_m.append(["id_cliente", id_cli])
    ws_m.append(["nombre_cliente", nombre_cli])
    ws_m.append(["cod_viajante", cv if cv is not None else ""])
    ws_m.append(["draft_id", draft.pk or 0])
    ws_m.append(["plantilla_version", PLANTILLA_VERSION])
    ws_m.append(["sucursal_ids", *ids_suc])
    ws_m.append(["col_primera_sucursal", 4])
    ws_m.sheet_state = "veryHidden"

    ws = wb.create_sheet(HOJA_PEDIDO, 0)
    n_suc = len(sucursales)
    ws.cell(1, 1, "Código")
    ws.cell(1, 2, "Artículo")
    ws.cell(1, 3, MARKER_IDART)
    for idx, s in enumerate(sucursales):
        col = 4 + idx
        ws.cell(1, col, _rotulo_columna_suc(s))
    for col in range(1, 4 + max(n_suc, 1)):
        c1 = ws.cell(1, col)
        c1.fill = _FILL_HDR
        c1.font = _FONT_HDR_SUC if col >= 4 else _FONT_HDR
        c1.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for ridx, art in enumerate(articulos):
        fila = 2 + ridx
        aid = int(art["id_articulo"])
        codigo = str_or_default(art.get("id_manual"), "") or str(aid)
        ca = ws.cell(fila, 1, codigo)
        ca.fill = _FILL_LOCK
        ca.font = _FONT_ART
        ca.alignment = Alignment(vertical="center")
        cn = ws.cell(fila, 2, str_or_default(art.get("nombre"), ""))
        cn.fill = _FILL_LOCK
        cn.font = _FONT_ART
        cn.alignment = Alignment(wrap_text=True, vertical="center")
        cid = ws.cell(fila, 3, aid)
        cid.fill = _FILL_LOCK
        cid.number_format = "0"
        ws.row_dimensions[fila].height = 20
        for idx, s in enumerate(sucursales):
            col = 4 + idx
            idd = to_int_or_none(s.get("id_cliente_domicilio"))
            qty = qty_map.get((aid, idd)) if idd is not None else None
            cell = ws.cell(fila, col, qty if qty is not None else None)
            cell.fill = _FILL_QTY
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0"

    n_arts = len(articulos)
    last_row = 1 + max(n_arts, 1)
    last_col = 3 + max(n_suc, 1)
    if n_arts == 0:
        for idx in range(n_suc):
            cell = ws.cell(2, 4 + idx)
            cell.fill = _FILL_QTY

    ws.row_dimensions[1].height = 56
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["C"].width = 10
    for idx in range(n_suc):
        ws.column_dimensions[get_column_letter(4 + idx)].width = 16
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100
    if n_suc:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
        dv = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Cantidad",
            error="Solo números mayores o iguales a 0 (packs).",
        )
        dv.add(f"D2:{get_column_letter(last_col)}{max(last_row, 2)}")
        ws.add_data_validation(dv)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _leer_workbook_import(raw: bytes) -> Tuple[List[Tuple[Any, ...]], Dict[str, Any]]:
    if len(raw) > MAX_BYTES:
        raise ValueError("El archivo supera el tamaño máximo (8 MB).")
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el Excel: {e}") from e
    try:
        meta: Dict[str, Any] = {}
        if HOJA_META in wb.sheetnames:
            ws_m = wb[HOJA_META]
            for row in ws_m.iter_rows(min_row=1, values_only=True):
                if not row:
                    continue
                k = _celda_str(row[0]).strip().lower()
                if not k:
                    continue
                if k == "sucursal_ids":
                    meta[k] = [
                        i
                        for i in (to_int_or_none(x) for x in row[1:])
                        if i is not None
                    ]
                else:
                    meta[k] = row[1] if len(row) > 1 else None
        if HOJA_PEDIDO not in wb.sheetnames:
            raise ValueError(
                "El Excel no tiene la hoja Pedido. Descargá de nuevo la plantilla."
            )
        ws = wb[HOJA_PEDIDO]
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        return rows, meta
    finally:
        wb.close()


def _extraer_nro_sucursal_header(raw: Any) -> Optional[str]:
    """Número de sucursal (NroCalle) desde el encabezado de columna de la plantilla."""
    s = _celda_str(raw).strip()
    if not s:
        return None
    primera = s.split("\n", 1)[0].strip()
    if primera.isdigit():
        return primera
    m = re.match(r"^(\d+)", primera)
    if m:
        return m.group(1)
    m = re.match(r"^suc\s+(\d+)", primera, re.I)
    return m.group(1) if m else None


def _rotulo_sucursal_id(
    idd: int, by_id: Dict[int, Dict[str, Any]]
) -> str:
    s = by_id.get(idd)
    if not s:
        return f"id {idd}"
    nro = str_or_default(s.get("nro"), "").strip()
    if nro and nro != "-":
        return f"SUC {nro}"
    return _etiqueta_suc(s)


def _mapear_columnas_sucursal(
    headers_id: Sequence[Any],
    headers_nom: Sequence[Any],
    sucursales: Sequence[Dict[str, Any]],
    errores: List[Dict[str, Any]],
    *,
    col_primera: int = 3,
    ids_meta: Optional[Sequence[int]] = None,
) -> List[Tuple[int, Optional[int], str]]:
    """Lista (col_1based, id_domicilio|None, etiqueta_header) desde ``col_primera``.

    La sucursal se resuelve por el encabezado visible (número/calle), no por la
    posición de columna. Si hay ``ids_meta`` de ``_Synap``, se valida que cada
    columna coincida con la sucursal esperada en esa posición (detecta reorden).
    """
    by_id = {
        to_int_or_none(s.get("id_cliente_domicilio")): s
        for s in sucursales
        if to_int_or_none(s.get("id_cliente_domicilio")) is not None
    }
    by_nro: Dict[str, List[int]] = {}
    by_etiq: Dict[str, List[int]] = {}
    for s in sucursales:
        idd = to_int_or_none(s.get("id_cliente_domicilio"))
        if idd is None:
            continue
        nro = str_or_default(s.get("nro"), "").strip()
        calle = str_or_default(s.get("calle"), "").strip()
        if nro and nro != "-":
            by_nro.setdefault(nro, []).append(idd)
        keys = {
            _etiqueta_suc(s).strip().lower(),
            calle.lower(),
            nro.lower(),
            f"suc {nro}".strip().lower(),
        }
        if nro and calle:
            keys.add(f"{nro}\n{calle}".lower())
            keys.add(_rotulo_columna_suc(s).strip().lower())
        for k in keys:
            if k:
                by_etiq.setdefault(k, []).append(idd)

    n = max(len(headers_id), len(headers_nom))
    inicio = max(2, int(col_primera) - 1)
    out: List[Tuple[int, Optional[int], str]] = []
    for i in range(inicio, n):
        col = i + 1
        raw_id = headers_id[i] if i < len(headers_id) else None
        raw_nom = headers_nom[i] if i < len(headers_nom) else None
        letra = get_column_letter(col)
        etiqueta = _celda_str(raw_nom) or _celda_str(raw_id)
        if not etiqueta:
            continue
        idd = to_int_or_none(raw_id)
        if idd is not None and idd in by_id:
            out.append((col, idd, etiqueta or _etiqueta_suc(by_id[idd])))
            continue
        key = etiqueta.strip().lower()
        primera_linea = etiqueta.split("\n", 1)[0].strip().lower()
        candidatos = by_etiq.get(key) or by_etiq.get(primera_linea) or []
        if len(candidatos) == 1:
            out.append((col, candidatos[0], etiqueta))
            continue
        if len(candidatos) > 1:
            errores.append(
                _err(
                    "Sucursal ambigua (mismo nombre en más de un domicilio). Usá la plantilla descargada.",
                    code="sucursal_ambigua",
                    fila=1,
                    columna=letra,
                    sucursal=etiqueta,
                )
            )
            out.append((col, None, etiqueta))
            continue
        nro_hdr = _extraer_nro_sucursal_header(etiqueta)
        if nro_hdr:
            cand_nro = by_nro.get(nro_hdr) or []
            if len(cand_nro) == 1:
                out.append((col, cand_nro[0], etiqueta))
                continue
            if len(cand_nro) > 1:
                errores.append(
                    _err(
                        (
                            f"El número de sucursal «{nro_hdr}» coincide con más de un domicilio. "
                            "Usá la plantilla descargada sin modificar encabezados."
                        ),
                        code="sucursal_ambigua",
                        fila=1,
                        columna=letra,
                        sucursal=etiqueta,
                    )
                )
                out.append((col, None, etiqueta))
                continue
            errores.append(
                _err(
                    (
                        f"La sucursal «SUC {nro_hdr}» del encabezado no está en el territorio "
                        "del vendedor para este cliente."
                    ),
                    code="sucursal_fuera_territorio",
                    fila=1,
                    columna=letra,
                    sucursal=etiqueta,
                )
            )
            out.append((col, None, etiqueta))
            continue
        if idd is not None:
            errores.append(
                _err(
                    "La sucursal no está en el territorio del vendedor para este cliente.",
                    code="sucursal_fuera_territorio",
                    fila=1,
                    columna=letra,
                    sucursal=etiqueta or str(idd),
                )
            )
            out.append((col, None, etiqueta))
            continue
        errores.append(
            _err(
                "Sucursal no reconocida en el encabezado de columna. Usá la plantilla descargada.",
                code="sucursal_desconocida",
                fila=1,
                columna=letra,
                sucursal=etiqueta,
            )
        )
        out.append((col, None, etiqueta))

    if ids_meta:
        for idx, (col, idd, etiqueta) in enumerate(out):
            if idx >= len(ids_meta):
                break
            meta_id = to_int_or_none(ids_meta[idx])
            if meta_id is None or idd is None:
                continue
            if idd != meta_id:
                letra = get_column_letter(col)
                errores.append(
                    _err(
                        (
                            f"La columna {letra} indica la sucursal "
                            f"«{_rotulo_sucursal_id(idd, by_id)}» pero en la posición {idx + 1} "
                            f"la plantilla Synap esperaba «{_rotulo_sucursal_id(meta_id, by_id)}». "
                            "No reordene, inserte ni elimine columnas; descargue de nuevo "
                            "la plantilla de este pedido."
                        ),
                        code="columna_sucursal_desalineada",
                        fila=1,
                        columna=letra,
                        sucursal=etiqueta,
                    )
                )
        if len(ids_meta) > len(out):
            errores.append(
                _err(
                    (
                        f"Faltan columnas de sucursal: la plantilla Synap define {len(ids_meta)} "
                        f"pero el Excel solo tiene {len(out)} reconocibles en la fila de encabezados."
                    ),
                    code="columnas_sucursal_incompletas",
                    fila=1,
                )
            )
    return out


def _norm_txt(val: Any) -> str:
    return " ".join(str_or_default(val, "").strip().lower().split())


def _articulo_vendible(art: Dict[str, Any]) -> bool:
    disc = str_or_default(art.get("discontinuo"), "No").strip().lower()
    ecom = str_or_default(art.get("ecommerce"), "No").strip().lower()
    tipo = str_or_default(art.get("tipo_art_fab"), "").strip()
    return disc in ("no", "") and ecom in ("si", "sí") and tipo == "Terminado"


def _puntuar_candidato(codigo: str, art: Dict[str, Any], nombre_excel: str) -> int:
    """Mayor puntaje gana. IDArt solo suma 1: un SuperArt numérico no debe perder contra otro SKU."""
    codigo_n = (codigo or "").strip().lower()
    nombre_n = _norm_txt(nombre_excel)
    score = 0
    idm = str_or_default(art.get("id_manual"), "").strip().lower()
    cap = str_or_default(art.get("cod_art_prov"), "").strip().lower()
    ct = str_or_default(art.get("codigo_t"), "").strip().lower()
    aid = to_int_or_none(art.get("id_articulo"))
    if idm and idm == codigo_n:
        score += 8
    if cap and cap == codigo_n:
        score += 8
    if ct and ct == codigo_n:
        score += 8
    if aid is not None and str(aid) == (codigo or "").strip():
        score += 1
    nom = _norm_txt(art.get("nombre"))
    if nombre_n and nom and nombre_n == nom:
        score += 16
    if cap and nombre_n and (nombre_n == cap or nombre_n.startswith(cap + " ")):
        score += 12
    return score


def _elegir_articulo(
    codigo: str,
    candidatos: Sequence[Dict[str, Any]],
    fila: int,
    errores: List[Dict[str, Any]],
    *,
    nombre_excel: str = "",
) -> Optional[Dict[str, Any]]:
    if not candidatos:
        errores.append(
            _err(
                "Artículo no encontrado.",
                code="articulo_no_encontrado",
                fila=fila,
                columna="A",
                codigo_articulo=codigo,
            )
        )
        return None
    vendibles = [a for a in candidatos if _articulo_vendible(a)]
    pool = vendibles or list(candidatos)
    elegido: Optional[Dict[str, Any]] = None
    if len(pool) == 1:
        elegido = pool[0]
    else:
        ranked = sorted(
            vendibles,
            key=lambda a: _puntuar_candidato(codigo, a, nombre_excel),
            reverse=True,
        )
        if ranked:
            best = _puntuar_candidato(codigo, ranked[0], nombre_excel)
            winners = [
                a
                for a in ranked
                if _puntuar_candidato(codigo, a, nombre_excel) == best
            ]
            if len(winners) == 1 and best > 0:
                elegido = winners[0]
        if elegido is None:
            errores.append(
                _err(
                    "Código de artículo ambiguo: coincide con más de un artículo.",
                    code="articulo_ambiguo",
                    fila=fila,
                    columna="A",
                    codigo_articulo=codigo,
                )
            )
            return None
    if not _articulo_vendible(elegido):
        errores.append(
            _err(
                "El artículo no está activo para venta (Terminado / ecommerce).",
                code="articulo_inactivo",
                fila=fila,
                columna="A",
                codigo_articulo=codigo,
            )
        )
        return None
    return elegido


def importar_matriz_excel(
    draft: EcomPedidoMasivoDraft,
    archivo_bytes: bytes,
    *,
    consultar_arts=None,
    consultar_ids=None,
) -> Dict[str, Any]:
    """Parsea, valida VCM y reemplaza celdas. All-or-nothing si hay errores."""
    if draft.estado not in (
        EcomPedidoMasivoDraft.ESTADO_BORRADOR,
        EcomPedidoMasivoDraft.ESTADO_CONFIRMANDO,
    ):
        return {
            "ok": False,
            "error": "El borrador no es editable.",
            "code": "draft_no_editable",
            "errores": [
                _err("El borrador no es editable.", code="draft_no_editable")
            ],
        }

    errores: List[Dict[str, Any]] = []
    try:
        rows, meta = _leer_workbook_import(archivo_bytes)
    except ValueError as e:
        return {
            "ok": False,
            "error": str(e),
            "code": "archivo_invalido",
            "errores": [_err(str(e), code="archivo_invalido")],
        }
    if len(rows) < 2:
        return {
            "ok": False,
            "error": "El Excel no tiene el formato de plantilla (faltan encabezados).",
            "code": "archivo_invalido",
            "errores": [
                _err(
                    "El Excel no tiene el formato de plantilla (faltan encabezados).",
                    code="archivo_invalido",
                )
            ],
        }

    sucursales, marcas_map = _territorio(draft)
    a1 = _celda_str(rows[0][0] if rows[0] else "").lower()
    c1 = _celda_str(rows[0][2] if rows[0] and len(rows[0]) > 2 else "").lower()
    es_plantilla_v2 = a1.replace(" ", "_") == MARKER_CODIGO
    es_v4 = c1.replace(" ", "_") == MARKER_IDART
    ids_meta = meta.get("sucursal_ids") if isinstance(meta.get("sucursal_ids"), list) else []
    col_suc = to_int_or_none(meta.get("col_primera_sucursal")) or (4 if es_v4 else 3)
    if es_plantilla_v2:
        headers_id = list(rows[0])
        headers_nom = list(rows[1]) if len(rows) > 1 else []
        data_rows = rows[2:]
        fila_base = 3
        col_suc = 3
    elif ids_meta:
        # Encabezados visibles (nro + calle); ids_meta solo valida alineación.
        headers_id = list(rows[0])
        headers_nom = list(rows[0])
        data_rows = rows[1:]
        fila_base = 2
    else:
        headers_id = list(rows[0])
        headers_nom = list(rows[0])
        data_rows = rows[1:]
        fila_base = 2
        col_suc = 4 if es_v4 else 3

    id_cli_excel = to_int_or_none(meta.get("id_cliente"))
    if id_cli_excel is None and es_plantilla_v2 and rows and len(rows[0]) > 1:
        id_cli_excel = to_int_or_none(rows[0][1])
    id_cli_draft = to_int_or_none(draft.id_cliente)
    if id_cli_excel is None:
        errores.append(
            _err(
                "Falta la identificación del cliente. Descargá de nuevo la plantilla de este pedido.",
                code="plantilla_sin_cliente",
                fila=1,
            )
        )
    elif id_cli_draft is not None and id_cli_excel != id_cli_draft:
        nom = _nombre_cliente(draft.base_empresa, id_cli_draft) or str(id_cli_draft)
        errores.append(
            _err(
                (
                    f"Esta plantilla es del cliente {id_cli_excel} y el pedido abierto "
                    f"es {id_cli_draft} ({nom}). Descargá la plantilla de este cliente."
                ),
                code="cliente_no_coincide",
                fila=1,
            )
        )
    cv_excel = to_int_or_none(meta.get("cod_viajante"))
    cv_draft = to_int_or_none(draft.cod_viajante)
    if (
        cv_excel is not None
        and cv_draft is not None
        and cv_excel != cv_draft
    ):
        errores.append(
            _err(
                "Esta plantilla es de otro vendedor. Descargá la plantilla de este pedido.",
                code="vendedor_no_coincide",
                fila=1,
            )
        )

    cols = _mapear_columnas_sucursal(
        headers_id,
        headers_nom,
        sucursales,
        errores,
        col_primera=col_suc,
        ids_meta=ids_meta if ids_meta else None,
    )
    if not cols:
        errores.append(
            _err(
                "No hay columnas de sucursal. Descargá la plantilla del cliente.",
                code="sin_columnas_sucursal",
                fila=1,
            )
        )

    codigos_filas: List[Tuple[int, str, Any]] = []
    for offset, row in enumerate(data_rows):
        fila = fila_base + offset
        if not row:
            continue
        codigo = _celda_str(row[0] if len(row) > 0 else None)
        if not codigo:
            hay_qty = False
            for col, _idd, _et in cols:
                idx = col - 1
                if idx < len(row):
                    qty, qerr = _qty_celda(row[idx])
                    if qty is not None or qerr:
                        hay_qty = True
                        break
            if hay_qty:
                errores.append(
                    _err(
                        "Falta el código de artículo.",
                        code="articulo_sin_codigo",
                        fila=fila,
                        columna="A",
                    )
                )
            continue
        codigos_filas.append((fila, codigo, row))

    lookup_fn = consultar_arts or consultar_articulos_por_codigos
    catalogo = lookup_fn(draft.base_empresa, [c for _f, c, _r in codigos_filas])
    ids_filas = []
    if es_v4:
        for _f, _c, row in codigos_filas:
            if len(row) > 2:
                ida = to_int_or_none(row[2])
                if ida is not None:
                    ids_filas.append(ida)
    lookup_ids = consultar_ids or consultar_articulos_por_ids
    arts_por_id = lookup_ids(draft.base_empresa, ids_filas) if es_v4 else {}

    vistos: Dict[str, int] = {}
    vistos_id: Dict[int, int] = {}
    celdas_ok: List[Tuple[int, int, Decimal]] = []
    arts_ok: Dict[int, Dict[str, Any]] = {}

    for fila, codigo, row in codigos_filas:
        id_art_fila = to_int_or_none(row[2]) if es_v4 and len(row) > 2 else None
        nombre_excel = _celda_str(row[1]) if len(row) > 1 else ""
        prev = vistos_id.get(id_art_fila) if id_art_fila is not None else vistos.get(codigo.lower())
        if prev:
            hay_qty_dup = False
            for col, _idd, _et in cols:
                idx = col - 1
                if idx < len(row):
                    qty, qerr = _qty_celda(row[idx])
                    if qty is not None or qerr:
                        hay_qty_dup = True
                        break
            if hay_qty_dup:
                errores.append(
                    _err(
                        f"Artículo repetido (ya aparece en la fila {prev}).",
                        code="articulo_duplicado",
                        fila=fila,
                        columna="A",
                        codigo_articulo=codigo,
                    )
                )
            continue
        if id_art_fila is not None:
            vistos_id[id_art_fila] = fila
        else:
            vistos[codigo.lower()] = fila
        if id_art_fila is not None:
            art = arts_por_id.get(id_art_fila)
            if not art:
                errores.append(
                    _err(
                        "Artículo no encontrado.",
                        code="articulo_no_encontrado",
                        fila=fila,
                        columna="A",
                        codigo_articulo=codigo,
                    )
                )
                continue
            vendible = _elegir_articulo(
                codigo, [art], fila, errores, nombre_excel=nombre_excel
            )
            if not vendible:
                continue
            art = vendible
        else:
            art = _elegir_articulo(
                codigo,
                catalogo.get(codigo) or [],
                fila,
                errores,
                nombre_excel=nombre_excel,
            )
            if not art:
                continue
        aid = int(art["id_articulo"])
        arts_ok[aid] = art
        marca = to_int_or_none(art.get("codigo_marca"))
        multiplo = multiplo_empaque_venta(art.get("multiplo_cantidad_vta"))
        for col, idd, etiqueta in cols:
            letra = get_column_letter(col)
            idx = col - 1
            val = row[idx] if idx < len(row) else None
            qty, qerr = _qty_celda(val)
            if qerr:
                errores.append(
                    _err(
                        qerr,
                        code="cantidad_invalida",
                        fila=fila,
                        columna=letra,
                        codigo_articulo=codigo,
                        sucursal=etiqueta,
                    )
                )
                continue
            if qty is None:
                continue
            if idd is None:
                continue
            if not cantidad_respeta_multiplo(qty, multiplo):
                errores.append(
                    _err(
                        mensaje_multiplo_invalido(multiplo),
                        code="multiplo_empaque",
                        fila=fila,
                        columna=letra,
                        codigo_articulo=codigo,
                        sucursal=etiqueta,
                    )
                )
                continue
            permitidas = marcas_map.get(idd) or set()
            if marca is None or marca not in permitidas:
                errores.append(
                    _err(
                        "La marca del artículo no está asignada a esta sucursal para el vendedor.",
                        code="marca_fuera_territorio",
                        fila=fila,
                        columna=letra,
                        codigo_articulo=codigo,
                        sucursal=etiqueta,
                    )
                )
                continue
            celdas_ok.append((aid, idd, qty))

    if errores:
        return {
            "ok": False,
            "error": f"La importación tiene {len(errores)} error(es). No se modificó el borrador.",
            "code": "validacion",
            "errores": errores[:MAX_ERRORES],
            "errores_total": len(errores),
        }

    if not celdas_ok:
        return {
            "ok": False,
            "error": "No hay cantidades para importar. Completá packs en la plantilla.",
            "code": "sin_cantidades",
            "errores": [
                _err(
                    "No hay cantidades para importar. Completá packs en la plantilla.",
                    code="sin_cantidades",
                )
            ],
        }

    if draft.estado == EcomPedidoMasivoDraft.ESTADO_CONFIRMANDO:
        draft.estado = EcomPedidoMasivoDraft.ESTADO_BORRADOR

    with transaction.atomic():
        EcomPedidoMasivoDraftCelda.objects.filter(draft=draft).delete()
        draft.descuentos_fila = {}
        draft.ultimo_error = {}
        if celdas_ok:
            EcomPedidoMasivoDraftCelda.objects.bulk_create(
                [
                    EcomPedidoMasivoDraftCelda(
                        draft=draft,
                        id_articulo=aid,
                        id_cliente_domicilio=idd,
                        cantidad_packs=qty,
                    )
                    for aid, idd, qty in celdas_ok
                ],
                batch_size=500,
            )
        for aid in arts_ok:
            asegurar_descuento_fila_articulo(draft, aid, draft.base_empresa)
        ctx_cli = leer_contexto_cliente_masivo(draft.base_empresa, draft.id_cliente)
        draft.descuento_pie_pct = _clamp_pct(ctx_cli.get("descPie"))
        draft.save(
            update_fields=[
                "estado",
                "descuentos_fila",
                "descuento_pie_pct",
                "ultimo_error",
                "updated_at",
            ]
        )

    n_suc = len({idd for _a, idd, _q in celdas_ok})
    return {
        "ok": True,
        "message": (
            f"Se importaron {len(celdas_ok)} cantidad(es) "
            f"en {len(arts_ok)} artículo(s) y {n_suc} sucursal(es)."
        ),
        "celdas": len(celdas_ok),
        "articulos": len(arts_ok),
        "sucursales": n_suc,
        "errores": [],
    }


def nombre_archivo_plantilla(draft: EcomPedidoMasivoDraft) -> str:
    cid = to_int_or_none(draft.id_cliente) or 0
    crudo = _nombre_cliente(draft.base_empresa, draft.id_cliente)
    slug = re.sub(r"[^A-Za-z0-9_-]+", "_", (crudo or "").strip())[:40].strip("_")
    if slug:
        return f"pedido_masivo_{cid}_{slug}.xlsx"
    return f"pedido_masivo_cliente_{cid}.xlsx"
