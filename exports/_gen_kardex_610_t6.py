#!/usr/bin/env python3
"""Genera Excel kardex 610 T6 con OPA + Remitos + reconstrucción Terminado."""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import MySQLdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path("/tmp/kardex_610_t6_terminado.xlsx")

PACKS = {
    1398: "610 T6 Kamp Tripack Bl/Gm/Ne 3P",
    1399: "610 T6 Kamp Tripack Blanco 3P",
    1400: "610 T6 Kamp Tripack Negro 3P",
}
TERM = 6

thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
hdr_fill = PatternFill("solid", fgColor="4C1D95")
hdr_font = Font(color="FFFFFF", bold=True, size=11)
opa_fill = PatternFill("solid", fgColor="EDE9FE")
rem_fill = PatternFill("solid", fgColor="DBEAFE")
fa_fill = PatternFill("solid", fgColor="FFF7ED")
inv_fill = PatternFill("solid", fgColor="F1F5F9")
neg_font = Font(color="B91C1C", bold=True)
neg_fill = PatternFill("solid", fgColor="FEF2F2")


def fmt_fecha(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    return str(v) if v else ""


def style_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, r, vals, fill=None):
    for col, v in enumerate(vals, 1):
        cell = ws.cell(r, col, v)
        cell.border = thin
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill


def afecta_deposito(comp: str) -> bool:
    return (comp or "").upper() != "FA"


def main() -> None:
    conn = MySQLdb.connect(
        host="181.174.198.194",
        port=30804,
        user="administranet",
        passwd="a7v8xx0805",
        db="administranet",
        charset="utf8mb4",
    )
    c = conn.cursor(MySQLdb.cursors.DictCursor)
    wb = Workbook()

    # --- Resumen / Demanda PED ---
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Kardex demanda 610 T6 — OPA + Remitos + reconstrucción Terminado"
    ws["A1"].font = Font(bold=True, size=14, color="4C1D95")
    ws["A2"] = (
        "Base administranet Bestsox (181.174) · Generado "
        + datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    )
    ws["A3"] = (
        "Hojas: Resumen (PED) | OPA y Remitos (cronológico pack+componentes) | "
        "Kardex Terminado por pack"
    )
    ws["A4"] = (
        "OPA = armado MPR (entrada pack Terminado + salida componentes Semi). "
        "REM = egreso cliente. FA se lista pero no mueve stock_deposito."
    )

    ws["A6"] = "Pedidos con demanda comercial (no Cerrado/Facturado)"
    ws["A6"].font = Font(bold=True)
    ph = ",".join(["%s"] * len(PACKS))
    c.execute(
        f"""
        SELECT cp.NroComprobante, cp.CodigoMovimiento, cp.Fecha, cp.Estado,
               cp.estado_pedido_opt, cp.tipo_pedido_opt, cp.Codigo AS id_cliente,
               COALESCE(cli.nombre_cliente, '') AS cliente,
               sp.IDArt, a.NombreArticulo, sp.Cantidad, sp.cantidad_pendiente,
               sp.cantidad_entregada,
               GREATEST(
                   0,
                   COALESCE(sp.cantidad_pendiente, 0),
                   COALESCE(sp.Cantidad, 0) - COALESCE(sp.cantidad_entregada, 0)
               ) AS qty_comercial
        FROM stockp sp
        JOIN comp_ped cp ON cp.CodigoMovimiento = sp.CodigoMovimiento
        JOIN articulo a ON a.IDArt = sp.IDArt
        LEFT JOIN cliente cli ON cli.Codigo = cp.Codigo
        WHERE sp.IDArt IN ({ph})
          AND COALESCE(cp.Anulado, 'No') = 'No'
          AND COALESCE(cp.TipoComprobante, '') = 'PED'
          AND COALESCE(cp.Estado, '') NOT IN ('Facturado', 'Cerrado')
          AND COALESCE(cp.estado_pedido_opt, '') IN ('Pendiente', 'Parcial')
        ORDER BY cp.Fecha, sp.IDArt
        """,
        list(PACKS),
    )
    ped_rows = list(c.fetchall() or [])
    h = [
        "Pedido",
        "Cód.mov",
        "Fecha",
        "Estado",
        "Estado OPT",
        "Cliente",
        "IDArt",
        "Artículo",
        "Cant.",
        "Pendiente",
        "Qty comercial",
    ]
    style_header(ws, 7, h)
    r = 8
    for m in ped_rows:
        write_row(
            ws,
            r,
            [
                m.get("NroComprobante") or "",
                int(m.get("CodigoMovimiento") or 0),
                fmt_fecha(m.get("Fecha")),
                m.get("Estado") or "",
                m.get("estado_pedido_opt") or "",
                f"{m.get('id_cliente') or ''} {m.get('cliente') or ''}".strip(),
                int(m.get("IDArt") or 0),
                m.get("NombreArticulo") or "",
                float(m.get("Cantidad") or 0),
                float(m.get("cantidad_pendiente") or 0),
                float(m.get("qty_comercial") or 0),
            ],
        )
        r += 1
    autosize(ws, [16, 10, 12, 14, 12, 28, 8, 42, 10, 10, 12])

    # --- OPA y Remitos ---
    ws2 = wb.create_sheet("OPA y Remitos")
    ws2["A1"] = (
        "Movimientos por demanda: OPA (armado pack+componentes) y REM/FA (egreso packs)."
    )
    ws2["A1"].font = Font(bold=True, size=12, color="4C1D95")
    ws2["A2"] = "Orden cronológico por FechaControl."

    c.execute(
        """
        SELECT
          m.fecha AS fecha_comp,
          COALESCE(m.fecha_control, s.FechaControl) AS fecha_hora,
          m.tipo_mov,
          m.nro_comprobante,
          m.codigo_movimiento,
          m.detalle,
          m.motivo_movimiento,
          s.IDArt,
          a.NombreArticulo,
          s.CodDeposito,
          d.NombreDeposito,
          s.Entrada,
          s.Salida,
          s.Comprobante,
          s.TipoComp,
          s.NroPedido,
          s.codmov_pedido,
          s.NroRemito,
          s.CodigoCP,
          COALESCE(cli.nombre_cliente, '') AS cliente
        FROM movimiento_stock m
        JOIN stock s ON s.CodigoMovimiento = m.codigo_movimiento
        JOIN articulo a ON a.IDArt = s.IDArt
        LEFT JOIN deposito d ON d.CodDeposito = s.CodDeposito
        LEFT JOIN cliente cli ON cli.Codigo = s.CodigoCP
        WHERE m.tipo_mov = 'OPA'
          AND COALESCE(m.anulado, 'No') <> 'Si'
          AND COALESCE(s.Anulado, 'No') <> 'Si'
          AND (
            m.detalle LIKE %s
            OR m.detalle LIKE %s
            OR m.detalle LIKE %s
          )
        ORDER BY COALESCE(m.fecha_control, s.FechaControl),
                 m.codigo_movimiento, s.IDArt
        """,
        ["%pack 1398%", "%pack 1399%", "%pack 1400%"],
    )
    opa_rows = list(c.fetchall() or [])

    c.execute(
        """
        SELECT
          s.Fecha AS fecha_comp,
          s.FechaControl AS fecha_hora,
          s.Comprobante AS tipo_mov,
          s.NroComprobante AS nro_comprobante,
          s.CodigoMovimiento AS codigo_movimiento,
          CONCAT(COALESCE(s.TipoComp, ''), ' ', COALESCE(s.Descripcion, '')) AS detalle,
          '' AS motivo_movimiento,
          s.IDArt,
          a.NombreArticulo,
          s.CodDeposito,
          d.NombreDeposito,
          s.Entrada,
          s.Salida,
          s.Comprobante,
          s.TipoComp,
          s.NroPedido,
          s.codmov_pedido,
          s.NroRemito,
          s.CodigoCP,
          COALESCE(cli.nombre_cliente, '') AS cliente
        FROM stock s
        JOIN articulo a ON a.IDArt = s.IDArt
        LEFT JOIN deposito d ON d.CodDeposito = s.CodDeposito
        LEFT JOIN cliente cli ON cli.Codigo = s.CodigoCP
        WHERE s.IDArt IN (1398, 1399, 1400)
          AND s.Comprobante IN ('REM', 'FA')
          AND COALESCE(s.Anulado, 'No') <> 'Si'
        ORDER BY s.FechaControl, s.CodigoMovimiento, s.IDArt
        """
    )
    rem_rows = list(c.fetchall() or [])

    combined = []
    for m in opa_rows:
        combined.append({**m, "tipo_demanda": "OPA"})
    for m in rem_rows:
        tipo = "REM" if (m.get("Comprobante") or "") == "REM" else "FA"
        combined.append({**m, "tipo_demanda": tipo})

    def sort_key(m):
        fh = m.get("fecha_hora") or datetime.min
        if isinstance(fh, date) and not isinstance(fh, datetime):
            fh = datetime.combine(fh, datetime.min.time())
        return (fh, int(m.get("codigo_movimiento") or 0), int(m.get("IDArt") or 0))

    combined.sort(key=sort_key)

    headers2 = [
        "Fecha comprobante",
        "Fecha/hora",
        "Tipo demanda",
        "Comprobante",
        "TipoComp",
        "Nro",
        "Cód.mov",
        "IDArt",
        "Artículo",
        "Depósito",
        "Entrada",
        "Salida",
        "Cliente",
        "Nro pedido",
        "Cód.mov pedido",
        "Nro remito",
        "Detalle",
    ]
    style_header(ws2, 4, headers2)
    r = 5
    for m in combined:
        tipo = m.get("tipo_demanda") or ""
        fill = opa_fill if tipo == "OPA" else (rem_fill if tipo == "REM" else fa_fill)
        codmov_ped = m.get("codmov_pedido")
        write_row(
            ws2,
            r,
            [
                fmt_fecha(m.get("fecha_comp"))[:10] if m.get("fecha_comp") else "",
                fmt_fecha(m.get("fecha_hora")),
                tipo,
                m.get("Comprobante") or "",
                m.get("TipoComp") or "",
                m.get("nro_comprobante") or "",
                int(m.get("codigo_movimiento") or 0),
                int(m.get("IDArt") or 0),
                m.get("NombreArticulo") or "",
                m.get("NombreDeposito") or str(m.get("CodDeposito") or ""),
                float(m.get("Entrada") or 0),
                float(m.get("Salida") or 0),
                f"{m.get('CodigoCP') or ''} {m.get('cliente') or ''}".strip(),
                m.get("NroPedido") or "",
                int(codmov_ped) if codmov_ped else "",
                m.get("NroRemito") or "",
                (m.get("detalle") or m.get("motivo_movimiento") or "")[:120],
            ],
            fill=fill,
        )
        r += 1
    ws2.freeze_panes = "A5"
    if r > 5:
        ws2.auto_filter.ref = f"A4:Q{r - 1}"
    autosize(ws2, [14, 20, 12, 12, 14, 16, 10, 8, 42, 14, 10, 10, 22, 12, 12, 12, 50])

    # --- Kardex Terminado por pack ---
    for aid, nombre in PACKS.items():
        short = {1398: "Kardex Mix", 1399: "Kardex Blanco", 1400: "Kardex Negro"}[aid]
        wsk = wb.create_sheet(short)
        c.execute(
            "SELECT saldo FROM stock_deposito WHERE id_articulo=%s AND id_deposito=%s",
            [aid, TERM],
        )
        sd = c.fetchone()
        sd_val = Decimal(str(sd["saldo"])) if sd else Decimal("0")
        wsk["A1"] = nombre
        wsk["A1"].font = Font(bold=True, size=13, color="4C1D95")
        wsk["A2"] = (
            f"IDArt {aid} · Terminado · stock_deposito={float(sd_val):.0f} · "
            "OPA/REM/FA/Inventario; saldo depósito excluye FA"
        )
        c.execute(
            """
            SELECT s.Fecha, s.FechaControl, s.Comprobante, s.TipoComp, s.Tipo,
                   s.NroComprobante, s.CodigoMovimiento, s.CodigoCP,
                   COALESCE(cli.nombre_cliente, '') AS cliente,
                   s.Entrada, s.Salida, s.NroPedido, s.codmov_pedido, s.NroRemito,
                   COALESCE(m.tipo_mov, '') AS tipo_mov,
                   COALESCE(m.detalle, s.Descripcion, '') AS detalle
            FROM stock s
            LEFT JOIN movimiento_stock m ON m.codigo_movimiento = s.CodigoMovimiento
            LEFT JOIN cliente cli ON cli.Codigo = s.CodigoCP
            WHERE s.IDArt=%s AND s.CodDeposito=%s AND COALESCE(s.Anulado,'No')<>'Si'
            ORDER BY COALESCE(s.FechaControl, CAST(s.Fecha AS DATETIME)),
                     s.CodigoMovimiento, s.id_stock
            """,
            [aid, TERM],
        )
        movs = list(c.fetchall() or [])
        headers = [
            "Fecha comprobante",
            "Fecha/hora",
            "Tipo demanda",
            "Comprobante",
            "TipoComp",
            "Nro",
            "Cód.mov",
            "Cliente",
            "Nro pedido",
            "Entrada",
            "Salida",
            "Afecta depósito",
            "Saldo depósito corrido",
            "Detalle",
        ]
        style_header(wsk, 4, headers)
        saldo = Decimal("0")
        rr = 5
        for m in movs:
            comp = m.get("Comprobante") or ""
            tipo_mov = (m.get("tipo_mov") or "").upper()
            if tipo_mov == "OPA" or (
                comp == "MSTOCK" and (m.get("TipoComp") or "") == "Armado"
            ):
                tipo_dem = "OPA"
                fill = opa_fill
            elif comp == "REM":
                tipo_dem = "REM"
                fill = rem_fill
            elif comp == "FA":
                tipo_dem = "FA"
                fill = fa_fill
            else:
                tipo_dem = m.get("TipoComp") or "OTRO"
                fill = inv_fill
            ent = Decimal(str(m.get("Entrada") or 0))
            sal = Decimal(str(m.get("Salida") or 0))
            af = afecta_deposito(comp)
            if af:
                saldo += ent - sal
            write_row(
                wsk,
                rr,
                [
                    fmt_fecha(m.get("Fecha"))[:10] if m.get("Fecha") else "",
                    fmt_fecha(m.get("FechaControl")),
                    tipo_dem,
                    comp,
                    m.get("TipoComp") or "",
                    m.get("NroComprobante") or "",
                    int(m.get("CodigoMovimiento") or 0),
                    f"{m.get('CodigoCP') or ''} {m.get('cliente') or ''}".strip(),
                    m.get("NroPedido") or "",
                    float(ent),
                    float(sal),
                    "Sí" if af else "No (FA)",
                    float(saldo),
                    (m.get("detalle") or "")[:100],
                ],
                fill=fill,
            )
            if saldo < 0:
                wsk.cell(rr, 13).font = neg_font
                wsk.cell(rr, 13).fill = neg_fill
            rr += 1
        wsk.freeze_panes = "A5"
        if rr > 5:
            wsk.auto_filter.ref = f"A4:N{rr - 1}"
        autosize(wsk, [14, 20, 12, 12, 14, 16, 10, 22, 12, 10, 10, 14, 16, 45])

    wb.save(OUT)
    print(
        f"OK {OUT} sheets={wb.sheetnames} opa={len(opa_rows)} "
        f"rem={len(rem_rows)} combined={len(combined)} size={OUT.stat().st_size}"
    )
    conn.close()


if __name__ == "__main__":
    main()
